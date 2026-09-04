"""
app.py — Servidor principal
Recebe webhooks do Baileys, parseia mensagens de falha
e grava automaticamente no Google Sheets.

Dois fluxos de entrada:
  1. POST /webhook  — mensagens em tempo real enviadas pelo server.js
  2. POST /rondas   — chamado pelo botão do dashboard; busca as últimas
                      6 horas de histórico em cada grupo via server.js
                      e processa as mensagens encontradas

Suporta:
- Mensagens individuais de ocorrência (🔴/🟡/🟢/🟠)
- Mensagens de normalização (✅ + "NORMALIZADO")
- Rondas diárias completas (múltiplas ocorrências em uma mensagem)
- Formato Cos Grid com bullets (·) sem emojis
"""

import os, re, json, logging, time, random, base64, uuid, sqlite3, threading
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import gspread
from google.oauth2.service_account import Credentials
from relatorio_semanal import (coletar_atividades_e_desligamentos_por_usina, gerar_relatorio_pptx,
                                listar_usinas_cliente, montar_status_zeladoria_por_usina,
                                coletar_chamados_fabricante_por_usina)
from relatorio_handover import gerar_handover_docx
from relatorio_handover_usina import montar_relatorio_handover_usina
from relatorio_handover_usina_docx import gerar_handover_usina_completo
from relatorio_ata_reuniao import gerar_ata_reuniao_docx
import docx as _docx_lib  # leitura de transcrições .docx (Teams) enviadas pelo usuário
import pdfplumber
from pdf2image import convert_from_bytes
from pypdf import PdfReader, PdfWriter
from io import BytesIO

# Push notifications (pywebpush)
try:
    from pywebpush import webpush, WebPushException
    PUSH_ENABLED = True
except ImportError:
    PUSH_ENABLED = False
    log_push = logging.getLogger(__name__)
    log_push.warning("pywebpush não instalado — notificações push desabilitadas")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# ── Fuso horário ─────────────────────────────────────────────────────────
# O Render roda o servidor em UTC (sem TZ configurada). datetime.now() puro
# retorna UTC, mas todo timestamp gravado no histórico/planilha é lido por
# humanos no Brasil (GMT-3). Sem essa conversão, todo horário registrado no
# sistema aparece 3h à frente do horário real de Brasília. Use agora_br()
# em vez de datetime.now() em qualquer lugar que grave/exiba horário local.
_TZ_BR = ZoneInfo("America/Sao_Paulo")


def agora_br():
    """Retorna o datetime atual já convertido para o horário de Brasília (GMT-3)."""
    return datetime.now(_TZ_BR)


log = logging.getLogger(__name__)

app = Flask(__name__)

# Permite requisições do GitHub Pages e de qualquer origem
# (o dashboard fica em fred-alexandrino.github.io)
CORS(app, resources={r"/*": {"origins": "*", "expose_headers": ["Content-Disposition"]}})


@app.errorhandler(Exception)
def _tratar_erro_nao_previsto(e):
    """Rede de segurança global: sem isso, qualquer exceção não tratada
    em qualquer endpoint vira a página de erro HTML padrão do Flask/
    Werkzeug — e o frontend, que sempre espera JSON, quebra com
    'Unexpected token '<'' em vez de mostrar o erro real. Preserva o
    código HTTP de erros conhecidos (ex.: HTTPException do Werkzeug,
    como 404) e usa 500 pra qualquer outra coisa inesperada."""
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return jsonify({"ok": False, "error": e.description or str(e)}), e.code
    log.error(f"[erro-nao-tratado] {request.method} {request.path}: {e}")
    return jsonify({"ok": False, "error": f"Erro interno inesperado: {e}"}), 500

# ── Configuração ──────────────────────────────────────────────────────────────
SHEET_ID       = os.environ.get("SHEET_ID", "1VLo8__wxSJVWiUIFd_JTcOnadJlUt440i1M1pC0ehTs")
SHEET_NAME     = os.environ.get("SHEET_NAME", "Painel de Falhas - Fred Alexandrino")
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET", "")
SHEET_EDIT_SECRET = os.environ.get("SHEET_EDIT_SECRET", "")
GRUPOS_FILTRO  = os.environ.get("GRUPOS_IDS", "").split(",")

# ── Configuração VAPID para notificações push ────────────────────────────────
VAPID_PUBLIC_KEY  = os.environ.get("VAPID_PUBLIC_KEY", "BJyGD9Lno29xj3_a6i5MjSHoZhHwfev7bRJRCqjnyL-o1vo9Hbf2zmrNtoONHtA92F59LGLc52HNE7oUkKqs5Yk")
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
VAPID_CLAIMS      = {"sub": "mailto:fred@gridco.com.br"}

# Subscriptions em memória (persistidas na planilha em produção)
# { endpoint: subscription_json }
_push_subscriptions = {}

# ── URL do servidor WhatsApp (Baileys) — usado pelo endpoint /rondas (Baileys) — usado pelo endpoint /rondas
WPP_SERVER_URL = os.environ.get("WPP_SERVER_URL", "").rstrip("/")

# Nome da aba de log de mensagens
LOG_SHEET_NAME = "Log de Mensagens"

# Nome da aba onde as subscriptions de push são persistidas (sobrevive a reinícios do Render)
PUSH_SHEET_NAME = "Push Subscriptions"

# ── Cache de credenciais Google (reutiliza a conexão) ────────────────────────
_gc_cache = None

def get_gc():
    global _gc_cache
    if _gc_cache is None:
        creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
        if not creds_json:
            raise ValueError("GOOGLE_CREDENTIALS_JSON não configurado")
        creds_dict = json.loads(creds_json)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        from google.oauth2.service_account import Credentials as _Creds
        creds = _Creds.from_service_account_info(creds_dict, scopes=scopes)
        _gc_cache = gspread.authorize(creds)
    return _gc_cache


def _gspread_retry(fn, tentativas=3, esperas=(4, 10, 20)):
    """Executa uma chamada ao Google Sheets com retry exponencial em caso
    de erro 429 (cota de leitura/escrita por minuto excedida — limite
    padrão do Google é 60 requisições/min por usuário). Ficou comum
    depois que as funcionalidades de fotos/graus de Zeladoria passaram a
    fazer várias chamadas em sequência rápida (uma leitura + escrita por
    usina, por lote). Sem isso, o erro sobe cru como um 500 genérico pro
    frontend. `fn` deve ser uma função sem argumentos (lambda ou closure)."""
    ultima_excecao = None
    for tentativa in range(tentativas):
        try:
            return fn()
        except gspread.exceptions.APIError as e:
            ultima_excecao = e
            corpo = str(e)
            if ("429" in corpo or "Quota exceeded" in corpo) and tentativa < tentativas - 1:
                time.sleep(esperas[tentativa])
                continue
            raise
    raise ultima_excecao


def get_log_sheet():
    """Retorna a aba 'Log de Mensagens' da planilha."""
    gc = get_gc()
    return gc.open_by_key(SHEET_ID).worksheet(LOG_SHEET_NAME)

def gravar_log_mensagem(grupo_id, grupo_nome, texto):
    """
    Grava uma mensagem recebida na aba 'Log de Mensagens'.
    Colunas: Timestamp | GrupoId | GrupoNome | Texto | Processado
    """
    try:
        ws_log = get_log_sheet()
        ts = agora_br().strftime("%d/%m/%Y %H:%M:%S")
        ws_log.append_row([ts, grupo_id, grupo_nome, texto, ""])
        log.info(f"📝 [Log] Mensagem gravada: {grupo_id}")
    except Exception as e:
        log.error(f"❌ [Log] Erro ao gravar mensagem: {e}")

_log_cache = {"ts": 0, "rows": None}

def ler_log_mensagens(horas=6):
    """
    Lê mensagens do log das últimas N horas.
    Usa cache de 60s para não estourar quota da API do Google.
    Retorna lista de dicts com grupo_id, texto, timestamp.
    """
    import time
    try:
        ws_log = get_log_sheet()
        # Cache de 60s para evitar quota exceeded
        agora = time.time()
        if agora - _log_cache["ts"] > 60 or _log_cache["rows"] is None:
            _log_cache["rows"] = ws_log.get_all_values()
            _log_cache["ts"]   = agora
            log.info("[Log] Cache atualizado")
        rows = _log_cache["rows"]
        if len(rows) < 2:
            return []

        desde = agora_br().timestamp() - (horas * 3600)
        mensagens = []

        for row in rows[1:]:  # pula cabeçalho
            if len(row) < 4: continue
            ts_str     = row[0].strip()
            grupo_id   = row[1].strip()
            texto      = row[3].strip()
            processado = row[4].strip() if len(row) > 4 else ""
            # Pula mensagens já processadas pelo botão Verificar Rondas
            if processado == "✅": continue
            if not texto or not grupo_id: continue

            # Converte timestamp
            try:
                from datetime import datetime as _dt
                dt = _dt.strptime(ts_str, "%d/%m/%Y %H:%M:%S")
                ts = dt.timestamp()
            except:
                continue

            if ts < desde: continue
            mensagens.append({"grupo_id": grupo_id, "texto": texto, "timestamp": ts_str, "linha_idx": rows[1:].index(row) + 2})

        log.info(f"[Log] {len(mensagens)} mensagens nas últimas {horas}h")
        return mensagens
    except Exception as e:
        log.error(f"❌ [Log] Erro ao ler mensagens: {e}")
        return []

def marcar_processado(ws_log, linha_idx):
    """Marca uma linha do log como processada (coluna E)."""
    try:
        ws_log.update_cell(linha_idx, 5, "✅")
    except:
        pass

def ler_log_historico(horas=24):
    """
    Lê TODAS as mensagens do log das últimas N horas — incluindo já processadas.
    Usada pelo endpoint /rondas/grupos para exibição histórica (somente leitura).
    """
    import time
    try:
        ws_log = get_log_sheet()
        agora = time.time()
        if agora - _log_cache["ts"] > 60 or _log_cache["rows"] is None:
            _log_cache["rows"] = ws_log.get_all_values()
            _log_cache["ts"]   = agora
        rows = _log_cache["rows"]
        if len(rows) < 2:
            return []

        desde = agora_br().timestamp() - (horas * 3600)
        mensagens = []
        for row in rows[1:]:
            if len(row) < 4: continue
            ts_str   = row[0].strip()
            grupo_id = row[1].strip()
            texto    = row[3].strip()
            processado = row[4].strip() if len(row) > 4 else ""
            if not texto or not grupo_id: continue
            try:
                from datetime import datetime as _dt
                dt = _dt.strptime(ts_str, "%d/%m/%Y %H:%M:%S")
                ts = dt.timestamp()
            except:
                continue
            if ts < desde: continue
            mensagens.append({
                "grupo_id":   grupo_id,
                "texto":      texto,
                "timestamp":  ts_str,
                "processado": processado == "✅",
                "linha_idx":  rows[1:].index(row) + 2
            })
        return mensagens
    except Exception as e:
        log.error(f"❌ [Log] Erro ao ler histórico: {e}")
        return []

def limpar_log_antigo():
    """
    Remove linhas do 'Log de Mensagens' com mais de 5 dias.
    Chamado automaticamente no endpoint /rondas.
    """
    import time
    try:
        ws_log = get_log_sheet()
        rows = ws_log.get_all_values()
        if len(rows) < 2:
            return 0

        limite = agora_br().timestamp() - (5 * 24 * 3600)
        linhas_deletar = []

        for i, row in enumerate(rows[1:], start=2):  # pula cabeçalho
            if len(row) < 1: continue
            ts_str = row[0].strip()
            try:
                from datetime import datetime as _dt
                dt = _dt.strptime(ts_str, "%d/%m/%Y %H:%M:%S")
                if dt.timestamp() < limite:
                    linhas_deletar.append(i)
            except:
                continue

        if not linhas_deletar:
            return 0

        # Deleta de baixo para cima (evita deslocamento de índices)
        for idx in reversed(linhas_deletar):
            ws_log.delete_rows(idx)

        # Invalida cache após limpeza
        _log_cache["ts"]   = 0
        _log_cache["rows"] = None

        log.info(f"🧹 [Log] {len(linhas_deletar)} linha(s) antigas removidas (>5 dias)")
        return len(linhas_deletar)
    except Exception as e:
        log.error(f"❌ [Log] Erro ao limpar log antigo: {e}")
        return 0

# ══════════════════════════════════════════════════════════════════════════════
# CATÁLOGO CANÔNICO DE USINAS
#
# Estrutura: nome_oficial → { cliente, aliases: [lista de variações] }
#
# Regras gerais aplicadas automaticamente pela função canonizar_usina():
#   - Remove prefixos "UFV ", "Usina ", "UFV Usina "
#   - Normaliza acentos para comparação (ç→c, ã→a, etc.)
#   - Trata 1/I/A/1A/IA como sufixo "1" e 2/II/B/1B/IB como sufixo "2"
#   - Usinas sem alias explícito são reconhecidas pelo nome base
# ══════════════════════════════════════════════════════════════════════════════

CATALOGO_USINAS = {
    # ── RENOGRID ──────────────────────────────────────────────────────────────
    "Nova Xavantina I": {
        "cliente": "Renogrid",
        "aliases": [
            "nova xavantina 1", "nova xavantina i",
            "xavantina 1", "xavantina i",
            "nova xavantina 1a", "nova xavantina ia",
            "xavantina 1a", "xavantina ia",
        ],
    },
    "Nova Xavantina II": {
        "cliente": "Renogrid",
        "aliases": [
            "nova xavantina 2", "nova xavantina ii",
            "xavantina 2", "xavantina ii",
            "nova xavantina 1b", "nova xavantina ib",
            "xavantina 1b", "xavantina ib",
        ],
    },
    "Colíder I": {
        "cliente": "Renogrid",
        "aliases": [
            "colider i", "colider 1", "colíder 1", "colíder i",
            "colider 1a", "colider ia", "colíder 1a", "colíder ia",
        ],
    },
    "Colíder II": {
        "cliente": "Renogrid",
        "aliases": [
            "colider ii", "colider 2", "colíder 2", "colíder ii",
            "colider 1b", "colider ib", "colíder 1b", "colíder ib",
        ],
    },
    "Nobres": {
        "cliente": "Renogrid",
        "aliases": ["nobres"],
    },
    "Elias Fausto": {
        "cliente": "Renogrid",
        "aliases": ["elias fausto"],
    },
    "Crateús": {
        "cliente": "Renogrid",
        "aliases": ["crateus", "crateús", "cratéus"],
    },

    # ── THOPEN ────────────────────────────────────────────────────────────────
    "Boa Esperança do Sul I": {
        "cliente": "Thopen",
        "aliases": [
            "boa esperanca do sul i", "boa esperanca do sul 1",
            "boa esperanca do sul a", "boa esperanca do sul 1a",
            "boa esperanca do sul ia",
            "boa esperança do sul i", "boa esperança do sul 1",
            "boa esperança do sul a", "boa esperança do sul 1a",
            "boa esperança do sul ia",
            "boa esperanca i", "boa esperanca 1",
            "boa esperança i", "boa esperança 1",
        ],
    },
    "Boa Esperança do Sul II": {
        "cliente": "Thopen",
        "aliases": [
            "boa esperanca do sul ii", "boa esperanca do sul 2",
            "boa esperanca do sul b", "boa esperanca do sul 1b",
            "boa esperanca do sul ib",
            "boa esperança do sul ii", "boa esperança do sul 2",
            "boa esperança do sul b", "boa esperança do sul 1b",
            "boa esperança do sul ib",
            "boa esperanca ii", "boa esperanca 2",
            "boa esperança ii", "boa esperança 2",
        ],
    },
    "Ibaté I": {
        "cliente": "Thopen",
        "aliases": [
            "ibate i", "ibate 1", "ibate 1a", "ibate ia", "ibate a",
            "ibaté i", "ibaté 1", "ibaté 1a", "ibaté ia", "ibaté a",
        ],
    },
    "Ibaté II": {
        "cliente": "Thopen",
        "aliases": [
            "ibate ii", "ibate 2", "ibate 1b", "ibate ib", "ibate b",
            "ibaté ii", "ibaté 2", "ibaté 1b", "ibaté ib", "ibaté b",
        ],
    },
    "Matão I": {
        "cliente": "Thopen",
        "aliases": [
            "matao 1", "matao i", "matao 1a", "matao ia", "matao a",
            "matão 1", "matão i", "matão 1a", "matão ia", "matão a",
        ],
    },
    "Matão II - Topázio": {
        "cliente": "Thopen",
        "aliases": [
            "matao 2", "matao ii", "matao 1b", "matao ib", "matao b",
            "matão 2", "matão ii", "matão 1b", "matão ib", "matão b",
            "matao 2 topazio", "matão 2 topázio",
            "topazio", "topázio",
        ],
    },
    "Sítio Bonfim": {
        "cliente": "Thopen",
        "aliases": [
            "sitio bonfim", "sítio bonfim",
            "bonfim",
        ],
    },
    "Poconé": {
        "cliente": "Thopen",
        "aliases": ["pocone", "poconé", "poconé"],
    },
    "Diamantino": {
        "cliente": "Thopen",
        "aliases": ["diamantino"],
    },
    "Canarana I": {
        "cliente": "Thopen",
        "aliases": [
            "canarana i", "canarana 1", "canarana 1a", "canarana ia", "canarana a",
        ],
    },
    "Canarana II": {
        "cliente": "Thopen",
        "aliases": [
            "canarana ii", "canarana 2", "canarana 1b", "canarana ib", "canarana b",
        ],
    },
    "Ribeirão Cascalheira": {
        "cliente": "Thopen",
        "aliases": [
            "ribeirao cascalheira", "ribeirão cascalheira",
            "ribeirao", "cascalheira",
        ],
    },

    # ── 2C ───────────────────────────────────────────────────────────────────
    "Araputanga": {
        "cliente": "2C",
        "aliases": ["araputanga"],
    },
    "Sete Lagoas": {
        "cliente": "2C",
        "aliases": ["sete lagoas"],
    },

    # ── GD Energy ─────────────────────────────────────────────────────────────
    "Guajirú": {
        "cliente": "GD Energy",
        "aliases": ["guajiru", "guajirú", "guajiru"],
    },
    "Sol do Norte I": {
        "cliente": "GD Energy",
        "aliases": [
            "sol do norte i", "sol do norte 1",
            "sol do norte 1a", "sol do norte ia", "sol do norte a",
        ],
    },
    "Sol do Norte II": {
        "cliente": "GD Energy",
        "aliases": [
            "sol do norte ii", "sol do norte 2",
            "sol do norte 1b", "sol do norte ib", "sol do norte b",
        ],
    },

    # ── Alves Lima ────────────────────────────────────────────────────────────
    "ABC Morada Nova": {
        "cliente": "Alves Lima",
        "aliases": ["abc morada nova", "morada nova"],
    },

    # ── Sal Energia ───────────────────────────────────────────────────────────
    # Cliente confirmado pelo Fred em 27/07/2026, mas nunca tinha sido
    # adicionado ao catálogo — por isso nenhuma OS da Fracttal desse
    # cliente era reconhecida automaticamente (nem na descoberta, nem na
    # auditoria), sempre caindo em revisão manual. Nomes oficiais no
    # formato "Codinome (Cidade)", igual usado na tabela de localizações.
    "SunPower (Cascavel)": {
        "cliente": "Sal Energia",
        "aliases": ["sunpower", "sunpower cascavel", "cascavel"],
    },
    "Hortina (Quixadá I)": {
        "cliente": "Sal Energia",
        "aliases": ["hortina", "quixada i", "quixada 1", "quixadá i", "quixadá 1"],
    },
    "Vitesse (Quixadá II)": {
        "cliente": "Sal Energia",
        "aliases": ["vitesse", "quixada ii", "quixada 2", "quixadá ii", "quixadá 2"],
    },
    "Salvales (Aquiraz I)": {
        "cliente": "Sal Energia",
        "aliases": ["salvales", "aquiraz i", "aquiraz 1"],
    },
    "Carosa (Aquiraz II)": {
        "cliente": "Sal Energia",
        # Corrigido 24/08/2026: o par estava invertido (Carosa/Aquiraz I,
        # Salvales/Aquiraz II) desde a criação do catálogo em 27/07/2026,
        # nunca atualizado após a correção de nomenclatura confirmada com
        # o Fred em 18/08/2026. Isso fazia canonizar_usina() nunca achar
        # match exato pro texto real vindo do groups_1_description da
        # Fracttal ("Carosa (Aquiraz II)"), caindo no fallback de busca
        # parcial (passo 2) — que score por tamanho de alias e escolhia
        # "aquiraz ii" (10 chars, olhando o cadastro errado de Salvales)
        # em vez de "carosa" (6 chars), classificando incorretamente como
        # Salvales. Ex.: OS 11713 (ativo SALE-CRS200-TRFR1, prefixo CRS =
        # Carosa) caiu como "Salvales (Aquiraz II)" por causa disso.
        "aliases": ["carosa", "aquiraz ii", "aquiraz 2"],
    },
}

# ── Índice invertido: alias_normalizado → nome_oficial ────────────────────────
import unicodedata as _ud_usina

def _norm_usina(s):
    """Normaliza string de usina para lookup: sem acento, minúsculo, sem espaços duplos."""
    s = _ud_usina.normalize("NFKD", (s or "").lower())
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s).strip()
    return s

# Constrói índice na inicialização
_ALIAS_INDEX = {}   # alias_norm → nome_oficial
_CLIENTE_INDEX = {} # nome_oficial → cliente

for _nome_oficial, _info in CATALOGO_USINAS.items():
    _CLIENTE_INDEX[_nome_oficial] = _info["cliente"]
    # Adiciona o próprio nome oficial como alias
    _ALIAS_INDEX[_norm_usina(_nome_oficial)] = _nome_oficial
    for _alias in _info["aliases"]:
        _ALIAS_INDEX[_norm_usina(_alias)] = _nome_oficial

# Prefixos a remover antes de lookup
_PREFIXOS_USINA = re.compile(
    r"^(?:ufv\s+)?(?:usina\s+)?(?:ufv\s+)?",
    re.IGNORECASE
)
# Sufixos a remover (lixo que pode vir junto)
_SUFIXOS_USINA = re.compile(
    r"\s*[-–|]\s*(?:normaliz\w+|ok|trip\s*\w*|desvio\w*).*$",
    re.IGNORECASE
)

def canonizar_usina(texto_bruto):
    """
    Recebe qualquer variação de nome de usina e retorna o nome oficial canônico.
    Retorna None se a usina não estiver no catálogo (outro supervisor) NEM
    na lista de usinas emprestadas temporariamente (ver
    _usinas_temporarias, seção "Supervisão Temporária" — usinas de outro
    supervisor que o Fred assumiu por período de férias/ausência).

    Exemplos:
      "UFV Xavantina 1"         → "Nova Xavantina I"
      "Boa Esperança do Sul IB" → "Boa Esperança do Sul II"
      "Usina Crateus"           → "Crateús"
      "UFV Topázio"             → "Matão II - Topázio"
      "Fazenda XYZ"             → None  (fora do catálogo)
    """
    if not texto_bruto:
        return None

    # Remove emojis e caracteres especiais comuns
    s = re.sub(r"[🔴🟡🟢🟠✅⏸️🔧⚠️*]", "", texto_bruto).strip()
    # Remove sufixos como "| NORMALIZADA | Trip 59B"
    s = _SUFIXOS_USINA.sub("", s).strip()
    # Remove prefixos "UFV ", "Usina ", etc.
    s = _PREFIXOS_USINA.sub("", s).strip()
    # Remove pontuação final
    s = s.rstrip(".,:-|").strip()

    # Normaliza para lookup
    s_norm = _norm_usina(s)

    # 1. Lookup direto no índice
    if s_norm in _ALIAS_INDEX:
        return _ALIAS_INDEX[s_norm]

    # 1b. Lookup no índice de usinas emprestadas temporariamente (dinâmico,
    # recarregado a cada poucos minutos — ver _indices_temporarios)
    alias_temp, _ = _indices_temporarios()
    if s_norm in alias_temp:
        return alias_temp[s_norm]

    # 2. Busca parcial — útil para variações não previstas
    # Tenta encontrar qual usina tem maior sobreposição com o texto
    melhor = None
    melhor_score = 0
    for alias_norm, nome_oficial in {**_ALIAS_INDEX, **alias_temp}.items():
        # Match se o alias está contido no texto ou vice-versa
        if alias_norm in s_norm or s_norm in alias_norm:
            score = len(alias_norm)  # prefere matches mais longos
            if score > melhor_score:
                melhor_score = score
                melhor = nome_oficial

    if melhor and melhor_score >= 4:  # evita matches em strings muito curtas
        return melhor

    return None  # usina não reconhecida — ignorar


def inferir_cliente(usina_canonical):
    """Retorna o cliente dado o nome canônico da usina."""
    if usina_canonical in _CLIENTE_INDEX:
        return _CLIENTE_INDEX[usina_canonical]
    _, cliente_temp = _indices_temporarios()
    return cliente_temp.get(usina_canonical, "")


def usina_permitida(texto):
    """Retorna True se a usina for reconhecida no catálogo."""
    return canonizar_usina(texto) is not None



# Mantém compatibilidade com código legado que usava CLIENTE_POR_USINA
CLIENTE_POR_USINA = {
    _norm_usina(nome): info["cliente"]
    for nome, info in CATALOGO_USINAS.items()
}
USINAS_PERMITIDAS = set(CATALOGO_USINAS.keys())

STATUS_VALIDOS = {
    "em aberto": "Em Aberto", "aberto": "Em Aberto",
    "concluído": "Concluído", "concluido": "Concluído", "resolvido": "Concluído",
    "aguardando cliente": "Aguardando Cliente",
    "aguardando fabricante": "Aguardando Fabricante",
    "aguardando equipamento": "Aguardando Equipamento",
    "em andamento": "Em Andamento",
    "corrigir ronda": "Corrigir Ronda - COS",
    "corrigir ronda - cos": "Corrigir Ronda - COS",
    "fechado": "Fechado",
}

# ── Padrões de extração ───────────────────────────────────────────────────────
_P = r"^[\s*·\-–]*"

PADROES = {
    "usina": re.compile(
        r"^(?:(?:🔴|🟡|🟢|🟠|✅|⏸️|🔧)[\s]*)?(?:DESVIO:[\s]*|UFV[\s]+DESVIO:[\s]*)?(?:UFV[\s]+)?Usina:?[\s]*([^\n\r*·:]{2,60}?)\s*(?:\*[^\n\r]*)?$",
        re.IGNORECASE | re.MULTILINE
    ),
    "problema": re.compile(_P + r"Probl[eo]ma[s]?(?:\s+do\s+\w+)?:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "descricao": re.compile(_P + r"Descri(?:ção|cao|çao|ção|c[aã]o)?(?:\s+d[oa]s?\s+\w+)?:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "acao": re.compile(_P + r"A[çc][aã]o(?:es)?:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "equipe": re.compile(_P + r"(?:Equipe[:\s]+(?:Acionada:?)?|T[eé]cnico\s+Acionado:)[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "supervisor": re.compile(_P + r"Supervisor[:\s]+(?:Acionado:?)?[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "inicio": re.compile(_P + r"In[ií]ci[oo](?:[\s]+(?:d[ao][\s]+)?[Oo]corrên?cia)?:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "fim": re.compile(_P + r"(?:Fim|T[eé]rmino)(?:[\s]+(?:d[ao][\s]+)?[Oo]corrên?cia)?:[ \t]*([^\n\r]*)", re.IGNORECASE | re.MULTILINE),
    "os": re.compile(_P + r"N[ºo°]?\.?[\s]*(?:da[\s]+)?OS:?[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "impacto": re.compile(_P + r"Impacto[s]?:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "equipamento": re.compile(_P + r"Equipamento[s]?[^:\n]*:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "causa": re.compile(_P + r"Causa[^:\n]*:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "chamado_conc": re.compile(_P + r"Chamado\s+Concession[aá]ria:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "tipo_manut": re.compile(_P + r"Tipo\s+Manuten[çc][aã]o[^:]*:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "identificacao": re.compile(_P + r"[Ii]dentifica[çc][aã]o:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "equip_problema": re.compile(_P + r"Equipamentos\s+com\s+Problema:[ \t]*([^\n\r]+)", re.IGNORECASE | re.MULTILINE),
    "cos_problema":   re.compile(r"·\s*Probl[eo]ma[s]?:[ \t]*([^\n\r]+)", re.IGNORECASE),
    "cos_descricao":  re.compile(r"·\s*Descri[çc][aã]o[^:]*:[ \t]*([^\n\r]+)", re.IGNORECASE),
    "cos_impacto":    re.compile(r"·\s*Impacto[s]?:[ \t]*([^\n\r]+)", re.IGNORECASE),
    "cos_acao":       re.compile(r"·\s*A[çc][aã]o(?:es)?:[ \t]*([^\n\r]+)", re.IGNORECASE),
    "cos_equipe":     re.compile(r"[·*]\s*(?:Equipe\s+Acionada|T[eé]cnico\s+Acionado):[ \t]*([^\n\r]+)", re.IGNORECASE),
    "cos_supervisor": re.compile(r"[·*]\s*Supervisor(?:\s+Acionado)?:[ \t]*([^\n\r]+)", re.IGNORECASE),
    "cos_inicio":     re.compile(r"[·*]\s*In[ií]ci[oo](?:\s+da\s+[Oo]corrência)?:[ \t]*([^\n\r]+)", re.IGNORECASE),
    "cos_fim":        re.compile(r"[·*]\s*(?:Fim|T[eé]rmino)(?:\s+da\s+[Oo]corrência)?:[ \t]*([^\n\r]*)", re.IGNORECASE),
    "cos_os":         re.compile(r"[·*]\s*N[ºo°]\.?[\s]*(?:da[\s]+)?OS:?[ \t]*([^\n\r]+)", re.IGNORECASE),
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def eh_formato_cos_grid(texto):
    tem_bullet = bool(re.search(r"[·*]\s*(?:Problema|Descrição|Impacto|Ação|Equipe|Supervisor|Início|Fim|Nº)", texto, re.IGNORECASE))
    tem_usina  = bool(re.search(r"Usina:", texto, re.IGNORECASE))
    return tem_bullet and tem_usina

def extrair(texto, padrao):
    m = padrao.search(texto)
    return m.group(1).strip().lstrip("*·").strip() if m else ""

def vazio(v):
    return not v or str(v).strip() in ("", "--", "-", "N/A", "n/a", "não", "nao", "Não")


def gravar_data_se_vazia(ws, num_linha, coluna_idx, row, label=""):
    """
    Grava a data/hora atual (agora_br()) na coluna informada, SOMENTE se a
    célula ainda estiver vazia. Nunca sobrescreve um valor já existente —
    seja ele gravado pelo robô antes, seja uma correção manual feita por Fred
    direto na planilha.

    coluna_idx: índice 0-based da coluna dentro de `row` (M=12, N=13, O=14).
    A escrita usa update_cell com o índice 1-based correspondente.
    """
    valor_atual = (row[coluna_idx] if len(row) > coluna_idx else "").strip()
    if not vazio(valor_atual):
        return False  # já preenchido (robô ou manual) — não mexe

    agora = agora_br().strftime("%d/%m/%Y %H:%M:%S")
    ws.update_cell(num_linha, coluna_idx + 1, agora)
    log.info(f"   → {label} gravada automaticamente: linha {num_linha} = {agora}")
    return True


def anexar_mensagem_original(ws, num_linha, coluna_idx, row, texto_bruto):
    """
    Anexa o texto bruto da mensagem do WhatsApp (já segmentado por ocorrência,
    vindo de parse_bloco) na coluna de 'Mensagens Originais' (V), no mesmo
    padrão do Histórico Cronológico: timestamp + texto, separado por linha
    em branco do conteúdo anterior. Sempre acrescenta, nunca substitui.

    Evita duplicar a mesma mensagem se o webhook reprocessar o mesmo texto
    (reenvio, retry de rede) — mesma proteção usada no Histórico Cronológico.
    """
    if vazio(texto_bruto):
        return
    texto_limpo = texto_bruto.strip()
    atual = (row[coluna_idx] if len(row) > coluna_idx else "").strip()
    if texto_limpo in atual:
        return  # mensagem idêntica já registrada — não duplica
    agora = agora_br().strftime("%d/%m %H:%M")
    entrada = f"{agora} - {texto_limpo}"
    novo = (atual + "\n\n" + entrada).strip() if atual else entrada
    ws.update_cell(num_linha, coluna_idx + 1, novo)


def normalizar_texto(t):
    import unicodedata
    return unicodedata.normalize("NFKD", t.lower()).encode("ascii", "ignore").decode("ascii").strip()

# inferir_cliente e usina_permitida definidas acima via canonizar_usina()

def extrair_tecnico(s):
    m = re.search(r"@([\w\s]+?)(?:\s*[-–]\s*[\w-]+)?\s*$", s)
    if m:
        return m.group(1).strip()
    s = re.sub(r"^[Ss]im[,\s]*", "", s).strip()
    return re.sub(r"@", "", s).strip()

def limpar_nome(s):
    s = re.sub(r"^[Ss]im[,\s]+", "", s).strip()
    s = re.sub(r"[@~]", "", s).strip()
    s = re.sub(r"\s*\|.*$", "", s).strip()
    s = re.sub(r"^[Tt][eé]cnico\s+", "", s).strip()
    return s

_REGEX_EQUIP = re.compile(
    r"(?<![\w-])("
    r"INV-\d+|"
    r"Inversor(?:es)?\s+\d+(?:[,\s]+\d+)*(?:\s+e\s+\d+)*|"
    r"Tracker(?:s)?\s+\d+(?:[,\s]+\d+)*(?:\s+e\s+\d+)*|"
    r"Tck(?:s)?\s+\d+(?:[,\s]+\d+)*|"
    r"Motor(?:[\w\s/]*Tracker)?\s*\d*|"
    r"TCU(?:[\w\s]*Tracker)?\s*\d*|"
    r"Fieldlogger|Smartlogger|"
    r"Rel[eé](?:\s+(?:UPR|EP\d+|de\s+[Pp]roteção|de\s+[Tt]emperatura|[A-Z0-9]+))?|"
    r"ETM|NVR|GCU|RSU|NCU|DPS|"
    r"Nobreak(?:\s+[\w]+)?|"
    r"EP\d+|Igate(?:[\w\s]*)?|"
    r"Câmera(?:s)?(?:[\w\s]*)?|"
    r"Piranometro(?:[\w\s]*)?|Anemômetro|"
    r"Exaustor(?:[\w\s]*)?|"
    r"Otimizador(?:es)?(?:[\w\s]*)?\d*|"
    r"Chave\s+Seccionadora(?:[\w\s]*)?|"
    r"Stringbox|Combiner(?:[\w\s]*)?\d*|"
    r"Transformador(?:[\w\s]*)?\d*|"
    r"Ventilador(?:[\w\s]*)?|Switch(?:[\w\s]*)?|"
    r"Bateria(?:[\w\s-]*)?(?:Tracker\s+\d+)?"
    r")(?![\w-])",
    re.IGNORECASE
)

def _limpar_equipamento(equip):
    equip = equip.strip()
    equip = re.sub(r"Tck\s*", "Tracker ", equip, flags=re.IGNORECASE)
    equip = re.sub(r"(?<=[Tt]racker\s)0+(\d)", r"\1", equip)
    equip = re.sub(r"(?<=[Mm]otor\s)0+(\d)", r"\1", equip)
    # Normaliza Inversor N e INV-N → INV-NN
    equip = re.sub(r"(?i)\bInversor(?:es)?\s+(\d+)\b", lambda m: f"INV-{int(m.group(1)):02d}", equip)
    equip = re.sub(r"\bINV-?(\d+)\b", lambda m: f"INV-{int(m.group(1)):02d}", equip)
    if equip and not equip.startswith("INV-"):
        equip = equip[0].upper() + equip[1:]
    return equip.strip()

def inferir_equipamento(problema="", descricao="", identificacao="", equip_problema="", acao="", impacto=""):
    for texto in [identificacao, problema, descricao, impacto, acao, equip_problema]:
        if not texto or str(texto).strip() in ("", "--", "-", "N/A"):
            continue
        m = _REGEX_EQUIP.search(texto)
        if m:
            return _limpar_equipamento(m.group(0))
    fonte = problema or descricao or ""
    return fonte[:60] if fonte else ""

def normalizar_inversores(texto):
    """
    Padroniza nomenclatura de inversores para INV-XX.
    Exemplos:
      "inversor 4"    → "INV-04"
      "inversor 04"   → "INV-04"
      "Inversor 14"   → "INV-14"
      "INV-4"         → "INV-04"
    """
    if not texto:
        return texto
    def fmt(n):
        return f"INV-{int(n):02d}"
    # inversor N → INV-NN
    texto = re.sub(
        r'\bInversor(?:es)?\s+(\d+)\b',
        lambda m: fmt(m.group(1)),
        texto, flags=re.IGNORECASE
    )
    # INV-N → INV-NN (sem zero à esquerda)
    texto = re.sub(
        r'\bINV-(\d+)\b',
        lambda m: fmt(m.group(1)),
        texto
    )
    return texto


def extrair_inversores_multiplos(bloco, dados_base):
    """
    Detecta mensagens com múltiplos inversores (ex: "Inversores 6 e 7")
    e retorna lista de dados individuais, um por inversor.

    Se houver ações/causas individuais por inversor no texto, distribui.
    Caso contrário, replica as mesmas informações para cada um.

    Retorna [] se não houver múltiplos inversores (processamento normal).
    """
    falha = dados_base.get("falha", "")
    acao  = dados_base.get("acao_texto", "") or dados_base.get("acao", "")

    # Detecta padrão: "inversores N e M" ou "inversores N, M e K"
    # Exemplos: "Inversores 6 e 7", "Inversores 06, 07 e 08"
    m = re.search(
        r'\bInversores?\s+((?:\d+(?:\s*[,e]\s*)?)+)',
        falha + " " + acao,
        re.IGNORECASE
    )
    if not m:
        return []

    nums_raw = re.findall(r'\d+', m.group(1))
    if len(nums_raw) < 2:
        return []  # só um inversor — processamento normal

    nums = [f"{int(n):02d}" for n in nums_raw]
    log.info(f"[Multi-INV] Detectados {len(nums)} inversores: {nums}")

    # Tenta extrair ações individuais por inversor no texto completo
    # Padrão: "INV-06: texto... INV-07: texto..."
    acoes_individuais = {}
    causas_individuais = {}

    for num in nums:
        inv_tag = f"INV-{num}"
        # Busca padrão "INV-XX: ..." ou "Inversor XX: ..."
        m_acao = re.search(
            rf'(?:INV-{num}|[Ii]nversor\s+0*{int(num)})\s*[:\-–]\s*([^\n\.]+)',
            acao
        )
        if m_acao:
            acoes_individuais[num] = m_acao.group(1).strip()

        m_causa = re.search(
            rf'(?:INV-{num}|[Ii]nversor\s+0*{int(num)})\s*[:\-–]\s*([^\n\.]+)',
            dados_base.get("causa", "")
        )
        if m_causa:
            causas_individuais[num] = m_causa.group(1).strip()

    # Gera lista de dados individuais
    lista = []
    for num in nums:
        inv_nome = f"INV-{num}"
        # Falha: substitui referência genérica pelo inversor específico
        # Ex: "Falha nos inversores 6 e 7" → "Falha no INV-06"
        falha_ind = re.sub(
            r'(?:nos\s+|no\s+)?\bInversores?\s+[\d,\s]+(?:e\s+\d+)?',
            f"no {inv_nome}",
            falha, flags=re.IGNORECASE
        ).strip() or falha

        dados_ind = {
            **dados_base,
            "equipamento":  inv_nome,
            "equip_impact": inv_nome,
            "falha":        falha_ind,
            "acao_texto":   acoes_individuais.get(num, dados_base.get("acao_texto", "")),
            "causa":        causas_individuais.get(num, dados_base.get("causa", "")),
        }
        # Recalcula ação composta
        partes = []
        if dados_ind["acao_texto"]:
            partes.append(dados_ind["acao_texto"])
        dados_ind["acao"] = " | ".join(partes) if partes else dados_base.get("acao", "")
        lista.append(dados_ind)

    return lista



def eh_normalizacao(texto):
    """
    Detecta se um bloco/texto indica normalização de ocorrência.
    Cobre:
      - ✅ + NORMALIZADO (qualquer posição)
      - Palavra NORMALIZADO/NORMALIZADA no campo usina (ex: 'Colider 1 - NORMALIZADO')
      - Fim da Ocorrência preenchido
      - Termos como 'ocorrência normalizada', 'usina normalizada'
    """
    return bool(re.search(
        r'normalizado|normalizada|✅.*normal|normal.*✅|ocorr[êe]ncia\s+encerrada',
        texto, re.IGNORECASE
    ))


def detectar_status_emoji(bloco):
    if re.search(r"✅", bloco):
        if eh_normalizacao(bloco):
            return "normalizado"
        return "Em Aberto"
    if re.search(r"🔴|🟡|🟠|⏸️", bloco): return "Em Aberto"
    return "Em Aberto"

def extrair_data_fmt(texto_data, fallback):
    if vazio(texto_data):
        return fallback
    m = re.search(r"(\d{2}/\d{2})", texto_data)
    return m.group(1) if m else fallback

def similaridade_falha(falha1, falha2):
    n1 = normalizar_texto(falha1)
    n2 = normalizar_texto(falha2)
    palavras1 = set(p for p in n1.split() if len(p) > 3)
    palavras2 = set(p for p in n2.split() if len(p) > 3)
    if not palavras1 or not palavras2:
        return False
    intersecao = palavras1 & palavras2
    menor = min(len(palavras1), len(palavras2))
    return len(intersecao) / menor >= 0.5


# ── Parse formato Cos Grid ────────────────────────────────────────────────────

def parse_bloco_cos_grid(bloco):
    usina_raw = extrair(bloco, PADROES["usina"])
    if not usina_raw:
        return None

    # Canoniza usando o catálogo oficial — resolve qualquer variação de nome
    usina_canonical = canonizar_usina(usina_raw)
    if not usina_canonical:
        log.info(f"Usina não reconhecida (Cos Grid): {usina_raw!r}")
        return None
    usina = usina_canonical

    normalizar_usina = bool(re.search(r"NORMALIZ", usina_raw, re.IGNORECASE))

    problema    = extrair(bloco, PADROES["cos_problema"])
    descricao   = extrair(bloco, PADROES["cos_descricao"])
    impacto     = extrair(bloco, PADROES["cos_impacto"])
    acao_txt    = extrair(bloco, PADROES["cos_acao"])
    equipe_raw  = extrair(bloco, PADROES["cos_equipe"])
    superv_raw  = extrair(bloco, PADROES["cos_supervisor"])
    inicio_txt  = extrair(bloco, PADROES["cos_inicio"])
    fim_txt     = extrair(bloco, PADROES["cos_fim"])
    os_txt      = extrair(bloco, PADROES["cos_os"])

    if not problema:  problema  = extrair(bloco, PADROES["problema"])
    if not descricao: descricao = extrair(bloco, PADROES["descricao"])
    if not acao_txt:  acao_txt  = extrair(bloco, PADROES["acao"])
    if not os_txt:    os_txt    = extrair(bloco, PADROES["os"])

    falha = problema or descricao or impacto or ""

    # Não cria ocorrência se não há falha identificada E não é normalização
    if vazio(falha) and not normalizar_usina:
        log.info(f"[COS Grid] Sem falha/problema identificado para {usina} — ignorando")
        return None

    equip = inferir_equipamento(problema=problema, descricao=descricao, acao=acao_txt, impacto=impacto)
    if not equip:
        equip = "Usina / Sistema Geral"

    tec = limpar_nome(equipe_raw) if not vazio(equipe_raw) else ""
    sup = limpar_nome(superv_raw) if not vazio(superv_raw) else ""

    partes_acao = []
    if not vazio(acao_txt): partes_acao.append(acao_txt)
    if not vazio(tec):      partes_acao.append(f"Técnico: {tec}")
    if not vazio(sup):      partes_acao.append(f"Supervisor: {sup}")
    if not partes_acao:     partes_acao.append("Inspeção em campo")

    os_num = ""
    if not vazio(os_txt):
        m_os = re.search(r"[\d]+", os_txt)
        os_num = m_os.group() if m_os else ""

    fim_preenchido = not vazio(fim_txt) and fim_txt.strip() not in ("", "-", "--")
    normalizar = normalizar_usina or fim_preenchido or eh_normalizacao(bloco)

    hoje     = agora_br().strftime("%d/%m")
    data_ini = extrair_data_fmt(inicio_txt, hoje)
    hist     = [f"{data_ini} - Registro inicial"]
    if not vazio(acao_txt):
        hist.append(f"{hoje} - {acao_txt}")
    if not vazio(tec):
        hist.append(f"{hoje} - Técnico em campo: {tec}")
    if normalizar:
        data_fim = extrair_data_fmt(fim_txt, hoje)
        hist.append(f"{data_fim} - Ocorrência normalizada")

    # "Descrição dos Problemas" → causa (motivo técnico da falha)
    # "Impacto"               → equipamentos impactados
    causa_final = descricao or ""
    equip_impact = impacto or equip

    # Status: se equipe foi acionada = "Em Andamento", senão "Em Aberto"
    equipe_acionada = not vazio(equipe_raw)
    if normalizar:
        status_calc = "Concluído"
    elif equipe_acionada:
        status_calc = "Em Andamento"
    else:
        status_calc = "Em Aberto"

    return {
        "usina":       usina,
        "cliente":     inferir_cliente(usina),
        "equipamento": equip,
        "falha":       falha,
        "causa":       causa_final,
        "equip_impact":equip_impact,
        "acao":        " | ".join(partes_acao),
        "status":      status_calc,
        "historico":   "\n".join(hist),
        "os":          os_num,
        "normalizar":  normalizar,
        "acao_texto":  acao_txt,
        "mensagem_bruta": bloco.strip(),
    }


# ── Parse de blocos (formato original) ───────────────────────────────────────

def normalizar_num(num_str):
    try:
        return str(int(num_str))
    except:
        return num_str

def extrair_atualizacoes_por_ativo(texto_acao):
    PRIORIDADE = {"normalizado": 3, "tratativa fabricante": 2, "garantia": 1, "outro": 0}
    padroes_ativo = [
        (re.compile(r"(Tracker[s]?\s+[\d,\s]+(?:e\s+\d+)?)\s+normalizado[s]?", re.IGNORECASE), "normalizado"),
        (re.compile(r"(Tracker[s]?\s+[\d,\s]+(?:e\s+\d+)?)\s+em\s+operação", re.IGNORECASE), "normalizado"),
        (re.compile(r"TCU\s+dos\s+(Tracker[s]?\s+[\d,\s]+(?:e\s+\d+)?)\s+em\s+garantia", re.IGNORECASE), "garantia"),
        (re.compile(r"(Tracker[s]?\s+[\d,\s]+(?:e\s+\d+)?)\s+permanece\s+em\s+garantia", re.IGNORECASE), "garantia"),
        (re.compile(r"(Tracker[s]?\s+[\d,\s]+(?:e\s+\d+)?)\s+em\s+garantia", re.IGNORECASE), "garantia"),
        (re.compile(r"(Tracker[s]?\s+[\d,\s]+(?:e\s+\d+)?)\s+em\s+tratativa\s+com\s+fabricante", re.IGNORECASE), "tratativa fabricante"),
        (re.compile(r"(INV[-\s]\d+|Inversor\s+\d+)\s+normalizado[s]?", re.IGNORECASE), "normalizado"),
        (re.compile(r"(INV[-\s]\d+|Inversor\s+\d+)\s+em\s+(?:operação|funcionamento)", re.IGNORECASE), "normalizado"),
        (re.compile(r"(INV[-\s]\d+|Inversor\s+\d+)\s+em\s+garantia", re.IGNORECASE), "garantia"),
        (re.compile(r"(INV[-\s]\d+|Inversor\s+\d+)\s+em\s+tratativa\s+com\s+fabricante", re.IGNORECASE), "tratativa fabricante"),
    ]
    melhor = {}
    for padrao, status in padroes_ativo:
        for m in padrao.finditer(texto_acao):
            ativo_raw = m.group(1).strip()
            nums = re.findall(r"\d+", ativo_raw)
            tipo = re.search(r"(Tracker|INV|Inversor|TCU|Motor)", ativo_raw, re.IGNORECASE)
            tipo_str = tipo.group(1).capitalize() if tipo else "Tracker"
            if tipo_str.upper() == "INV":
                tipo_str = "Inversor"
            for num in nums:
                num_norm = normalizar_num(num)
                chave = f"{tipo_str.lower()}_{num_norm}"
                pri_nova = PRIORIDADE.get(status, 0)
                if chave not in melhor or pri_nova > PRIORIDADE.get(melhor[chave]["status_update"], 0):
                    melhor[chave] = {
                        "equipamento": f"{tipo_str} {num_norm}",
                        "status_update": status,
                        "normalizar": (status == "normalizado"),
                        "acao_resumida": {
                            "normalizado":          "Ocorrência normalizada em campo",
                            "garantia":             "Aguardando garantia com fabricante",
                            "tratativa fabricante": "Em tratativa com fabricante",
                        }.get(status, status.capitalize()),
                    }
    return list(melhor.values())

def equipamento_match(equip_planilha, equip_busca):
    def norm(s):
        s = s.lower().strip()
        s = re.sub(r"motor\s*", "tracker ", s)
        s = re.sub(r"tcu\s*tracker\s*", "tracker ", s)
        s = re.sub(r"inv-", "inversor ", s)
        nums = re.findall(r"\d+", s)
        tipo = re.search(r"(tracker|inversor|motor|tcu|inv)", s)
        tipo_str = tipo.group(1) if tipo else ""
        if tipo_str == "inv": tipo_str = "inversor"
        return tipo_str, [normalizar_num(n) for n in nums]
    tipo1, nums1 = norm(equip_planilha)
    tipo2, nums2 = norm(equip_busca)
    if not nums1 or not nums2: return False
    tipos_ok = tipo1 == tipo2 or not tipo1 or not tipo2
    nums_ok  = bool(set(nums1) & set(nums2))
    return tipos_ok and nums_ok

def separar_blocos(texto):
    if eh_formato_cos_grid(texto):
        partes = re.split(r"(?=(?:^|\n)Usina:)", texto, flags=re.MULTILINE | re.IGNORECASE)
        blocos = [p.strip() for p in partes if p.strip() and len(p.strip()) > 20]
        return blocos if blocos else [texto]

    partes = re.split(r"(?=(?:^|\n)[ \t]*(?:🔴|🟡|🟢|🟠|✅|⏸️))", texto, flags=re.MULTILINE)
    blocos = [p.strip() for p in partes if p.strip() and len(p.strip()) > 30]

    if len(blocos) <= 1:
        partes = re.split(r"(?=(?:^|\n)[ \t]*(?:🔴|🟡|🟢|🟠|✅|⏸️|🔧)?[ \t]*(?:DESVIO:?\s*)?(?:Usina|UFV):)", texto, flags=re.MULTILINE | re.IGNORECASE)
        blocos = [p.strip() for p in partes if p.strip() and len(p.strip()) > 30]

    return blocos if blocos else [texto]

def parse_bloco(bloco):
    if eh_formato_cos_grid(bloco):
        return parse_bloco_cos_grid(bloco)

    c = {k: extrair(bloco, p) for k, p in PADROES.items()}

    if not c["usina"] or len(c["usina"]) > 60:
        primeira = bloco.split('\n')[0].strip()
        m_desvio = re.search(r'(?:🔴|🟡|🟢|🟠|✅|⏸️)?\s*(?:DESVIO:\s*|Usina:\s*)?(?:UFV\s+)?(.+?)[\s:*]*$', primeira, re.IGNORECASE)
        if m_desvio:
            candidato = m_desvio.group(1).strip().rstrip(':*').strip()
            if candidato and len(candidato) < 60:
                c["usina"] = candidato

    if not c["usina"]:
        return None

    # Canoniza usando o catálogo oficial — resolve qualquer variação de nome
    usina_canonical = canonizar_usina(c["usina"])
    if not usina_canonical:
        log.info(f"Usina não reconhecida (formato original): {c['usina']!r}")
        return None
    usina = usina_canonical

    eh_formato_tracker = not vazio(c["identificacao"]) or not vazio(c["equip_problema"])

    if eh_formato_tracker:
        id_raw  = c["identificacao"] if not vazio(c["identificacao"]) else ""
        id_fmt  = re.sub(r"Tck\s*", "Tracker ", id_raw, flags=re.IGNORECASE).strip()
        equip   = id_fmt if id_fmt else inferir_equipamento(problema=c["problema"], descricao=c["descricao"], acao=c["acao"], impacto=c.get("impacto",""))
        equip_prob = c["equip_problema"] if not vazio(c["equip_problema"]) else ""
        m_acao  = re.search(r"(.+?)\.\s*(acionado.+)$", equip_prob, re.IGNORECASE | re.DOTALL)
        if m_acao:
            causa        = m_acao.group(1).strip()
            acao_tracker = m_acao.group(2).strip().capitalize()
        else:
            causa        = equip_prob
            acao_tracker = ""
        partes_acao = []
        if acao_tracker:
            partes_acao.append(acao_tracker)
        elif not vazio(c["acao"]):
            partes_acao.append(c["acao"])
        else:
            partes_acao.append("Inspeção em campo")
    else:
        equip = c["equipamento"] if not vazio(c["equipamento"]) else \
                inferir_equipamento(problema=c["problema"], descricao=c["descricao"], identificacao=c["identificacao"], equip_problema=c["equip_problema"], acao=c["acao"], impacto=c.get("impacto",""))
        causa        = c["causa"] if not vazio(c["causa"]) else ""
        acao_tracker = ""
        partes_acao  = []
        if not vazio(c["acao"]):
            partes_acao.append(c["acao"])
        else:
            partes_acao.append("Inspeção em campo")

    tec = extrair_tecnico(c["equipe"]) if not vazio(c["equipe"]) else ""
    if not vazio(tec): partes_acao.append(f"Técnico: {tec}")
    sup = re.sub(r"^[Ss]im[,\s]*", "", c["supervisor"]).strip() if not vazio(c["supervisor"]) else ""
    sup = re.sub(r"@", "", sup).strip()
    if not vazio(sup): partes_acao.append(f"Supervisor: {sup}")

    os_num = ""
    if not vazio(c["os"]):
        m_os = re.search(r"[\d/]+", c["os"])
        os_num = m_os.group() if m_os else ""

    status_emoji = detectar_status_emoji(bloco)
    normalizar   = (status_emoji == "normalizado")

    hoje       = agora_br().strftime("%d/%m")
    hist       = []
    data_inicio = extrair_data_fmt(c["inicio"], hoje)
    if normalizar:
        data_fim = extrair_data_fmt(c["fim"], hoje)
        hist.append(f"{data_inicio} - Registro inicial")
        hist.append(f"{data_fim} - Ocorrência normalizada")
    else:
        hist.append(f"{data_inicio} - Registro inicial")
        acao_hist = acao_tracker if eh_formato_tracker and not vazio(acao_tracker) else c["acao"]
        if not vazio(acao_hist):
            hist.append(f"{hoje} - {acao_hist}")

    return {
        "usina":       usina,
        "cliente":     inferir_cliente(usina),
        "equipamento": equip,
        "falha":       (c["problema"] or c["descricao"] or c["tipo_manut"] or (f"Tracker parado - {causa}" if eh_formato_tracker else "") or ""),
        "causa":       causa,
        "equip_impact":equip,
        "acao":        " | ".join(partes_acao),
        "status":      "Concluído" if normalizar else "Em Aberto",
        "historico":   "\n".join(hist),
        "os":          os_num,
        "normalizar":  normalizar,
        "acao_texto":  c["acao"],
        "mensagem_bruta": bloco.strip(),
    }


# ── Google Sheets ─────────────────────────────────────────────────────────────

def get_sheet():
    gc = get_gc()
    return gc.open_by_key(SHEET_ID).worksheet(SHEET_NAME)

ZELADORIA_GID = 987654321

def get_zeladoria_sheet():
    gc = get_gc()
    return gc.open_by_key(SHEET_ID).get_worksheet_by_id(ZELADORIA_GID)


ATIVIDADES_SHEET_NAME = "Painel de Atividades"
ATIVIDADES_HEADERS = ["ID", "Cliente", "Usina", "Equipamento", "Descricao", "Responsavel", "Prazo",
                       "Prioridade", "Status", "DataCriacao", "DataConclusao", "Historico", "Editor",
                       "NumeroOS"]

DESLIGAMENTO_MANUAL_SHEET_NAME = "_DesligamentoManual"
DESLIGAMENTO_MANUAL_HEADERS = ["origem", "id", "valor", "editor", "atualizadoEm"]

CHAMADOS_FABRICANTE_SHEET_NAME = "ChamadosFabricante"
CHAMADOS_FABRICANTE_HEADERS = [
    "Ativo", "UFV", "Cliente", "Equipe", "Supervisor", "OS de Abertura", "Ticket/RMA",
    "Fabricante", "Identificação Supervisório", "Identificação do Equipamento", "Serial Number",
    "Data da ocorrência", "Data da abertura do chamado", "Data da Última Atualização",
    "Motivo da abertura do chamado", "Causa da Falha", "Dias corridos", "Data de finalização",
    "Status", "Título do E-mail", "Observações", "N° da Solicitação de OS", "Supervisor Antigo",
    "Status OS", "Resolução",
]

def get_chamados_fabricante_sheet():
    gc = get_gc()
    ss = gc.open_by_key(SHEET_ID)
    try:
        ws = ss.worksheet(CHAMADOS_FABRICANTE_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=CHAMADOS_FABRICANTE_SHEET_NAME, rows=500,
                               cols=len(CHAMADOS_FABRICANTE_HEADERS))
        ws.append_row(CHAMADOS_FABRICANTE_HEADERS)
    return ws


def get_desligamento_manual_sheet():
    gc = get_gc()
    ss = gc.open_by_key(SHEET_ID)
    try:
        ws = ss.worksheet(DESLIGAMENTO_MANUAL_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=DESLIGAMENTO_MANUAL_SHEET_NAME, rows=200, cols=len(DESLIGAMENTO_MANUAL_HEADERS))
        ws.append_row(DESLIGAMENTO_MANUAL_HEADERS)
    return ws


SUPERVISAO_TEMP_SHEET_NAME = "_SupervisaoTemporaria"
SUPERVISAO_TEMP_HEADERS = ["cliente", "usina", "cluster", "responsavelOriginal", "adicionadoEm"]


def get_supervisao_temp_sheet():
    gc = get_gc()
    ss = gc.open_by_key(SHEET_ID)
    try:
        ws = ss.worksheet(SUPERVISAO_TEMP_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=SUPERVISAO_TEMP_SHEET_NAME, rows=200, cols=len(SUPERVISAO_TEMP_HEADERS))
        ws.append_row(SUPERVISAO_TEMP_HEADERS)
    return ws


LOCALIZACOES_SHEET_NAME = "Localizacoes"
LOCALIZACOES_HEADERS = ["Cliente", "Usina", "Endereco", "MapsLink", "Lat", "Lng"]

# Seed inicial (planilha "Localizações e Endereços" enviada pelo Fred,
# versão corrigida de 24/07/2026 — Matão I renomeado, Embu Guaçu removido,
# Sal Energia com coordenadas já convertidas de DMS pra decimal). Só é usado
# na primeira vez que a aba "Localizacoes" é criada; depois disso o Fred
# edita direto na planilha ou via /localizacoes-atualizar-coords.
LOCALIZACOES_SEED = [
    ('Renogrid', 'Nova Xavantina I', 'Rua Sete de Setembro, Nova Xavantina, MT, 78690-000', 'https://maps.app.goo.gl/ZrLXCvacAEfdmLG49', "", ""),
    ('Renogrid', 'Nova Xavantina II', 'Projeto Xavantina, Setor Nova Brasília, Gleba B, Nova Xavantina – MT, CEP: 78690-000', 'https://maps.app.goo.gl/qvLsNF1xkcoJHk1d7', "", ""),
    ('Renogrid', 'Colíder I', 'Estrada de Santa Luzia, Colíder - MT, 78500-000', 'https://maps.app.goo.gl/d1N8znU5FBrBh1FZ7', "", ""),
    ('Renogrid', 'Colíder II', 'Sítio Nossa Senhora Salete, Estrada Vicinal R, Colíder – MT, CEP: 78.500-000', 'https://maps.app.goo.gl/4CjanaMerTJYCDBs5', "", ""),
    ('Renogrid', 'Nobres', 'Fazenda Lavrinha, S/N, Nobres - MT, 78470-000', 'https://maps.app.goo.gl/ndrKoLEk3aDyroLJ9', "", ""),
    ('Renogrid', 'Elias Fausto', 'Sítio Santa Izabel, Gleba B, Elias Fausto – SP, CEP: 13.358-899', 'https://maps.app.goo.gl/uy57XM69H3ejTENJ6', "", ""),
    ('Renogrid', 'Crateús', 'Fazenda São Luiz, Crateús - CE, CEP: 63.709-899', 'https://maps.app.goo.gl/W9VS4Jdcd57rqCgC9', "", ""),
    ('Thopen', 'Boa Esperança do Sul I', 'Rua Sete de Setembro, Boa Esperança do Sul - SP, 14930-000', 'https://maps.app.goo.gl/T74pDLBrsd3yGDPP6', "", ""),
    ('Thopen', 'Boa Esperança do Sul II', 'Rua Sete de Setembro, Boa Esperança do Sul - SP, 14930-000', 'https://maps.app.goo.gl/T74pDLBrsd3yGDPP6', "", ""),
    ('Thopen', 'Ibaté I', 'R. Júlio Gonzaga, Ibaté - SP, 14815-000', 'https://maps.app.goo.gl/8fPjE3dNwtgZucqF8', "", ""),
    ('Thopen', 'Ibaté II', 'BR-267, 292-1040 - Jardim Nosso Teto, Ibaté - SP, 14815-000', 'https://maps.app.goo.gl/2m6UqXNHHhtr4zsb9', "", ""),
    ('Thopen', 'Matão I', 'Via Luís Gonzaga da Silva Leite - Pedreira, Matão - SP', 'https://maps.app.goo.gl/CFsFFwniYJffJH9w9', "", ""),
    ('Thopen', 'Matão II - Topázio', 'Via Carl Fisher, 5600 - Matão, SP, 15995-054', 'https://maps.app.goo.gl/MJnvbaMm3Ar2nUxF7', "", ""),
    ('Thopen', 'Sítio Bonfim', 'Sítio Morros, Limoeiro do Norte - CE, 62930-000', 'https://maps.app.goo.gl/7bc56bH8vXEAWAu18', "", ""),
    ('Thopen', 'Poconé', 'Assentamento Beija Flor, Lote 04, Zona Rural, Poconé - MT, 78175-000', 'https://maps.app.goo.gl/y3weGpiKFnVAGMHz8', "", ""),
    ('Thopen', 'Canarana I', 'MT-110, Canarana - MT, 78640-000', 'https://maps.app.goo.gl/2WeCCv27u1KSQXjs9', "", ""),
    ('Thopen', 'Canarana II', 'FPPJ+996 Canarana, MT', 'https://maps.app.goo.gl/rTLEK4FVzrRz4UNA7', "", ""),
    ('Thopen', 'Ribeirão Cascalheira', '35HX+8PR Ribeirão Cascalheira, MT', 'https://maps.app.goo.gl/BL8xuoi7hy5w4yev6', "", ""),
    ('2C Energia', 'Araputanga', 'Rodovia MT 175, Araputanga - MT', 'https://maps.app.goo.gl/AzpfpHjkqxyqBwAV7', "", ""),
    ('2C Energia', 'Sete Lagoas', 'Zona Rural, Sete Lagoas - MG, 35702-087', 'https://maps.app.goo.gl/mCoLpWAobk6K2qLG7', "", ""),
    ('GD Energy', 'Guajirú', 'Sítio Ilha, s/n - Vassouras, Trairi - CE, 62690-000', 'https://maps.app.goo.gl/rv2uhbD7o37ANCdSA', "", ""),
    ('GD Energy', 'Sol do Norte I', 'Sítio Ilha, s/n - Vassouras, Trairi - CE, 62690-000', 'https://maps.app.goo.gl/rv2uhbD7o37ANCdSA', "", ""),
    ('GD Energy', 'Sol do Norte II', 'Sítio Ilha, s/n - Vassouras, Trairi - CE, 62690-000', 'https://maps.app.goo.gl/rv2uhbD7o37ANCdSA', "", ""),
    ('Alves Lima', 'ABC Morada Nova', 'Estr. Tigre A Dourado, 2 - Pedras, Morada Nova - CE, 62940-000', 'https://maps.app.goo.gl/GPmb3Ea9JMGnHXBB6', "", ""),
    ('Sal Energia', 'SunPower (Cascavel)', 'Zona Rural, s/n – Cascavel - CE', 'https://maps.app.goo.gl/gogYXrktX8D3TfFPA', -4.105639, -38.319028),
    ('Sal Energia', 'Hortina (Quixadá I)', 'Zona Rural, s/n – Quixadá - CE', 'https://maps.app.goo.gl/oajX7AZNQfBdJNwF9', -4.997056, -38.991528),
    ('Sal Energia', 'Vitesse (Quixadá II)', 'Zona Rural, s/n – Quixadá - CE', 'https://maps.app.goo.gl/oajX7AZNQfBdJNwF9', -4.997056, -38.991528),
    ('Sal Energia', 'Salvales (Aquiraz I)', 'Zona Rural, s/n – Aquiraz - CE', 'https://maps.app.goo.gl/XXn34jb8jE1NRvfE9', -4.059083, -38.495917),
    ('Sal Energia', 'Carosa (Aquiraz II)', 'Zona Rural, s/n – Aquiraz - CE', 'https://maps.app.goo.gl/XXn34jb8jE1NRvfE9', -4.059083, -38.495917),
]


def get_localizacoes_sheet():
    gc = get_gc()
    ss = gc.open_by_key(SHEET_ID)
    try:
        ws = ss.worksheet(LOCALIZACOES_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=LOCALIZACOES_SHEET_NAME, rows=200, cols=len(LOCALIZACOES_HEADERS))
        ws.append_row(LOCALIZACOES_HEADERS)
        ws.append_rows([list(r) for r in LOCALIZACOES_SEED])
    return ws


def get_atividades_sheet():
    gc = get_gc()
    ss = gc.open_by_key(SHEET_ID)
    try:
        ws = ss.worksheet(ATIVIDADES_SHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=ATIVIDADES_SHEET_NAME, rows=1000, cols=len(ATIVIDADES_HEADERS))
        ws.append_row(ATIVIDADES_HEADERS)
        return ws
    # migração incremental: garante que colunas novas (ex: Equipamento, NumeroOS) existam
    header = ws.row_values(1)
    if len(header) < len(ATIVIDADES_HEADERS):
        if ws.col_count < len(ATIVIDADES_HEADERS):
            ws.add_cols(len(ATIVIDADES_HEADERS) - ws.col_count)  # expande a grade antes de escrever
        for i in range(len(header), len(ATIVIDADES_HEADERS)):
            ws.update_cell(1, i + 1, ATIVIDADES_HEADERS[i])
    return ws

def carregar_planilha(ws):
    """Lê uma aba inteira com retry automático em caso de 429 (cota de
    leitura excedida) — helper central usado por ~14 pontos do backend,
    incluindo /gerar-relatorio-semanal, que confirmadamente estourava a
    cota (erro 429 real em produção, 02/09/2026)."""
    return _gspread_retry(lambda: ws.get_all_values())

def proximo_id(todos):
    maior = 0
    for row in todos[1:]:
        if row and row[0]:
            try:
                maior = max(maior, int(row[0]))
            except ValueError:
                pass
    return maior + 1

# ── Fingerprint de deduplicação ───────────────────────────────────────────────

import unicodedata as _ud

def _norm(s):
    """Normaliza string para comparação: sem acento, minúsculo, só alfanum."""
    s = _ud.normalize("NFKD", (s or "").lower())
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return " ".join(s.split())

def fingerprint_ocorrencia(usina, equipamento, falha):
    """
    Chave de identidade única de uma ocorrência.
    Formato: usina | tipo_equip | num_equip | palavras_falha
    Exemplos:
      "boa esperanca do sul 1 | tracker | 6 | geracao inversores perda"
      "ibate ii | inversor | 4 | funcionamento parcial strings"
    """
    usina_n = _norm(usina)
    equip_n = _norm(equipamento)

    # Números do equipamento ("Tracker 06" → "6")
    nums    = [str(int(n)) for n in re.findall(r"\d+", equip_n)]
    num_str = "_".join(nums) if nums else ""

    # Tipo do equipamento
    tipo_m  = re.search(
        r"(tracker|inversor|motor|tcu|nobreak|camera|exaustor|piranometro|"
        r"fieldlogger|smartlogger|ep\d+|igate|rele|switch|transformador|"
        r"bateria|stringbox|anemometro|otimizador|seccionadora|combiner|ncu|gcu|etm|nvr)",
        equip_n
    )
    tipo_str = tipo_m.group(1) if tipo_m else equip_n[:12]

    # Top-5 palavras significativas da falha
    stop = {"para", "com", "que", "dos", "das", "nos", "nas", "pelo", "pela",
            "esse", "esta", "este", "uma", "uns", "umas", "nao", "sem"}
    palavras = sorted(set(
        p for p in _norm(falha).split()
        if len(p) > 3 and p not in stop
    ))[:5]

    return f"{usina_n}|{tipo_str}|{num_str}|{'_'.join(palavras)}"


def _norm_equip_key(equip):
    """
    Gera chave de comparação de equipamento:
    extrai tipo + números normalizados.
    Ex: "INV-03" → ("inversor", ["3"])
        "Tracker 08" → ("tracker", ["8"])
        "Motor Tracker 5" → ("tracker", ["5"])
    """
    s = _norm(equip)
    s = re.sub(r"motor\s+tracker", "tracker", s)
    s = re.sub(r"tcu\s+tracker", "tracker", s)
    s = re.sub(r"inv-?", "inversor ", s)
    tipo_m = re.search(
        r"(inversor|tracker|motor|tcu|nobreak|camera|exaustor|piranometro|"
        r"fieldlogger|smartlogger|igate|rele|switch|transformador|"
        r"bateria|stringbox|otimizador|seccionadora|combiner|ep\d+)",
        s
    )
    tipo = tipo_m.group(1) if tipo_m else s[:10]
    nums = [str(int(n)) for n in re.findall(r"\d+", s)]
    return tipo, nums


def equipamentos_sao_iguais(equip1, equip2):
    """
    Compara dois equipamentos de forma tolerante.
    Considera iguais se tipo E pelo menos um número coincidem.
    """
    if not equip1 or not equip2: return False
    tipo1, nums1 = _norm_equip_key(equip1)
    tipo2, nums2 = _norm_equip_key(equip2)
    if not nums1 or not nums2: return False
    tipos_ok = tipo1 == tipo2 or not tipo1 or not tipo2
    nums_ok  = bool(set(nums1) & set(nums2))
    return tipos_ok and nums_ok


def usinas_sao_iguais(usina1, usina2):
    """Compara usinas usando o catálogo canônico."""
    c1 = canonizar_usina(usina1) or _norm(usina1)
    c2 = canonizar_usina(usina2) or _norm(usina2)
    return c1 == c2


def buscar_por_fingerprint(todos, usina, equipamento, falha, os_num=""):
    """
    Busca ocorrência existente EM ABERTO usando hierarquia de critérios:

    NÍVEL 1 (mais forte) — OS + usina + equipamento:
      Se a mensagem tem número de OS, busca por OS+usina+equip.
      Isso garante que atualizações de um chamado específico sempre
      encontrem a ocorrência certa, independente da descrição da falha.

    NÍVEL 2 — usina + equipamento (tipo + número):
      Compara usina (via catálogo canônico) + tipo e número do equipamento.
      Ex: INV-03 e "Inversor 3" são o mesmo; Tracker 8 e Motor 08 também.

    NÍVEL 3 (fallback) — fingerprint de palavras:
      Só usa se os níveis anteriores não encontrarem nada.

    Retorna (num_linha, row) ou None.
    """
    candidatos = []

    candidatos_concluidos = []  # para reabrir recentemente concluídas

    for i, row in enumerate(todos[1:], start=2):
        if len(row) < 9: continue
        status = row[8].strip().lower()
        eh_concluido = "conclu" in status or "resolv" in status or "fechad" in status

        usina_plan = row[2].strip()
        equip_plan = row[3].strip()
        os_plan    = (row[10] if len(row) > 10 else "").strip()

        # Usinas devem ser a mesma (obrigatório em todos os níveis)
        if not usinas_sao_iguais(usina, usina_plan):
            continue

        # NÍVEL 1a: OS + usina (mais forte — mesma OS = mesma ocorrência)
        # Normaliza: remove prefixos "OS", "#", zeros à esquerda, espaços
        def _norm_os(s):
            s = s.strip()
            m = re.match(r"(?i)^(?:os|n[oº°]?|#)\s*(\d+)", s)
            if m: return str(int(m.group(1)))
            try: return str(int(s))
            except: return s.lower().strip()
        os_n  = _norm_os(os_num)
        os_p  = _norm_os(os_plan)
        invalidos = {"", "n/a", "na", "-", "s/n", "sn", "0"}
        if os_n and os_p and os_n not in invalidos and os_p not in invalidos and os_n == os_p:
            log.info(f"🎯 Match NÍVEL 1 (OS+usina): linha {i} | OS={os_num} | {equip_plan}")
            return (i, row)

        if eh_concluido:
            # Verifica se foi concluída recentemente (≤ 7 dias) pelo histórico
            hist_txt = row[11] if len(row) > 11 else ""
            hoje = agora_br()
            datas = re.findall(r"(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", hist_txt)
            reabrir = False
            for d_match in datas:
                try:
                    dia, mes = int(d_match[0]), int(d_match[1])
                    ano = int(d_match[2]) if d_match[2] else hoje.year
                    dt = datetime(ano, mes, dia)
                    if (hoje - dt).days <= 7:
                        reabrir = True
                        break
                except:
                    pass
            if reabrir and equipamentos_sao_iguais(equipamento, equip_plan):
                candidatos_concluidos.append((i, row, "reabrir"))
            continue  # não adiciona concluídas nos candidatos normais

        # NÍVEL 2: usina + equipamento
        if equipamentos_sao_iguais(equipamento, equip_plan):
            candidatos.append((i, row, "equip"))
            continue

        # NÍVEL 3: fingerprint de palavras (fallback)
        fp_novo   = fingerprint_ocorrencia(usina, equipamento, falha)
        fp_plan   = fingerprint_ocorrencia(usina_plan, equip_plan, row[4])
        if fp_novo == fp_plan:
            candidatos.append((i, row, "fingerprint"))

    if not candidatos:
        # Tenta reabrir ocorrência recentemente concluída (reincidência < 7 dias)
        if candidatos_concluidos:
            i, row, _ = candidatos_concluidos[0]
            log.info(f"🔄 Reincidência detectada — reabrindo linha {i} | {row[3]} (concluída há ≤ 7 dias)")
            return (i, row)
        return None

    # Prioriza match por equipamento sobre fingerprint
    por_equip = [c for c in candidatos if c[2] == "equip"]
    if por_equip:
        i, row, _ = por_equip[0]
        log.info(f"🎯 Match NÍVEL 2 (usina+equip): linha {i} | {row[3]}")
        return (i, row)

    i, row, _ = candidatos[0]
    log.info(f"🎯 Match NÍVEL 3 (fingerprint): linha {i} | {row[3]}")
    return (i, row)


def acao_mudou(row, acao_nova):
    """
    Retorna True se a ação nova contém informação não presente no campo Ação
    atual nem no Histórico cronológico da planilha.
    """
    if vazio(acao_nova):
        return False
    acao_atual = _norm(row[7] if len(row) > 7 else "")
    historico  = _norm(row[11] if len(row) > 11 else "")
    acao_norm  = _norm(acao_nova)
    # Considera mudança se pelo menos 60% das palavras novas não estão no conteúdo atual
    palavras_novas = [p for p in acao_norm.split() if len(p) > 3]
    if not palavras_novas:
        return False
    ja_conhecidas = sum(1 for p in palavras_novas if p in acao_atual or p in historico)
    return (ja_conhecidas / len(palavras_novas)) < 0.6


def status_mudou(row, novo_status):
    """Retorna True se o status da planilha é diferente do novo."""
    atual = (row[8] if len(row) > 8 else "").strip().lower()
    novo  = (novo_status or "").strip().lower()
    return atual != novo and not vazio(novo_status)


# ── Operações na planilha ─────────────────────────────────────────────────────

def detectar_aguardando_fabricante(texto):
    """
    Retorna True quando o texto indica:
      A) Número de chamado com fabricante (Case #XXXXX, Chamado Nº XXXXX, etc.)
      B) Normalidade em campo (normal em campo, 100% normal, etc.)
    Nessa combinação → status deve ser 'Aguardando Fabricante'.
    """
    if not texto:
        return False
    t = texto.lower()
    import re as _re
    tem_chamado = bool(_re.search(
        r'(?:case|chamado|ticket|n°)\s*[#°]?\s*\d{5,}',
        t
    ))
    tem_normal_campo = any(p in t for p in [
        "normal em campo", "normalizado em campo", "normalidade em campo",
        "em campo está normal", "campo está normal",
        "em campo esta normal", "campo esta normal",
        "100% normal", "100 % normal", "normalizado no campo",
        "apresenta normalidade em campo",
    ])
    return tem_chamado and tem_normal_campo


def extrair_ticket_fabricante(texto):
    """
    Extrai o número/código do chamado/ticket do fabricante a partir do texto
    da mensagem, cobrindo os formatos reais usados no dia a dia:
      - Prefixados: "SOL-10596", "SOL - 12634", "RMA 25814"
      - Explícitos: "Chamado 25065", "Ticket 6843263", "Case #45231"
      - Contextuais: "Caso deferido 6843263", "Acionamento Fabricante 15817311"
    Retorna a string do ticket (ex: "SOL-10596") ou "" se não encontrar.
    Não confundir com 'os_num' (Número da OS interna), que é capturado
    separadamente por outro padrão.
    """
    if not texto:
        return ""

    # Formato prefixado: SOL-12345, SOL 12345, RMA-25814 (aceita espaços/
    # hífen variáveis ao redor do prefixo). Exclui prefixos que na prática
    # são outra coisa (código de rastreio, OS interna, etc.)
    m = re.search(r'\b([A-Z]{2,5})\s*[-\s]\s*(\d{4,})\b', texto, re.IGNORECASE)
    if m:
        prefixo = m.group(1).upper()
        numero  = m.group(2)
        if prefixo not in ("OS", "PV", "EP", "ID", "QN", "AD", "OY", "BR"):
            return f"{prefixo}-{numero}"

    # Formato explícito: "chamado/case/ticket/n°/rma + número longo"
    m = re.search(
        r'(?:case|chamado|ticket|n°|rma)\s*[#°]?\s*(\d{5,})',
        texto, re.IGNORECASE
    )
    if m:
        return m.group(1)

    # Contexto de fabricante/garantia próximo a um número isolado de 6-8
    # dígitos (cobre "Caso deferido 6843263", "Acionamento Fabricante 15817311")
    m = re.search(
        r'(?:fabricante|deferid[oa]|garantia)\D{0,30}\b(\d{6,8})\b',
        texto, re.IGNORECASE
    )
    if m:
        return m.group(1)

    return ""


# ── Detecção de gatilhos T1 / T2 / T3 (Tempo Ativo O&M) ────────────────────
#
# Estas funções alimentam as colunas M/N/O (Data 1ª Ação, Data Encaminhamento,
# Data Retorno Externo) da planilha. A regra de ouro é: o app.py SÓ grava
# nessas colunas se a célula estiver VAZIA — nunca sobrescreve uma correção
# manual feita por Fred na planilha. Essa checagem é feita em
# gravar_nova_ocorrencia() e atualizar_ocorrencia(), não aqui — estas funções
# apenas respondem True/False para o texto analisado.

def detectar_primeira_acao(texto):
    """
    Retorna True quando o texto da mensagem (de abertura ou de uma atualização)
    já indica que a equipe Grid começou a atuar na ocorrência — ex: técnico
    acionado, equipe a caminho, já verificando, etc.

    Cobre tanto o caso em que a 1ª ação vem embutida na própria mensagem de
    abertura (T1 ~ 0) quanto uma atualização posterior que só agora informa
    que a equipe agiu.
    """
    if not texto:
        return False
    t = texto.lower()
    termos = [
        "técnico acionado", "tecnico acionado",
        "equipe acionada", "equipe foi acionada",
        "técnico a caminho", "tecnico a caminho",
        "equipe a caminho", "em deslocamento",
        "já estamos verificando", "ja estamos verificando",
        "já está verificando", "ja esta verificando",
        "estamos verificando", "equipe verificando",
        "verificando em campo", "verificando a ocorrência",
        "técnico em campo", "tecnico em campo",
        "equipe em campo", "equipe em atendimento",
        "atendimento iniciado", "iniciado atendimento",
        "já em atendimento", "ja em atendimento",
        "técnico foi enviado", "tecnico foi enviado",
        "enviamos técnico", "enviamos tecnico",
        "deslocando equipe", "deslocando técnico", "deslocando tecnico",
        "iniciada a verificação", "iniciada a verificacao",
        "já estamos atuando", "ja estamos atuando",
        "equipe já está no local", "equipe ja esta no local",
        "técnico já está no local", "tecnico ja esta no local",
        "já iniciamos o atendimento", "ja iniciamos o atendimento",
    ]
    return any(p in t for p in termos)


def detectar_encaminhamento(texto):
    """
    Retorna True quando o texto indica que a ocorrência foi encaminhada para
    fora do controle direto da equipe Grid — fabricante, cliente ou
    fornecedor de equipamento. Reaproveita a varredura de
    detectar_aguardando_fabricante() e expande para Aguardando Cliente /
    Aguardando Equipamento.

    Marca o fim do T2 e o início do T3 (espera externa).
    """
    if not texto:
        return False
    t = texto.lower()

    # Reaproveita a lógica de chamado fabricante + campo normal
    if detectar_aguardando_fabricante(texto):
        return True

    termos = [
        "aguardando fabricante", "aguardando o fabricante",
        "aguardando cliente", "aguardando o cliente",
        "aguardando equipamento", "aguardando peça", "aguardando peca",
        "aguardando material", "aguardando envio",
        "encaminhado ao fabricante", "encaminhado para o fabricante",
        "encaminhado ao cliente", "encaminhado para o cliente",
        "chamado aberto no fabricante", "chamado aberto com o fabricante",
        "chamado aberto com fabricante", "chamado aberto fabricante",
        "os aberta no fabricante", "os aberta com o fabricante",
        "solicitado ao fabricante", "solicitamos ao fabricante",
        "acionamos o fabricante", "acionado o fabricante",
        "aguardando retorno do fabricante", "aguardando retorno fabricante",
        "aguardando posição do fabricante", "aguardando posicao do fabricante",
        "aguardando garantia", "em garantia",
        "peça solicitada", "peca solicitada",
        "material solicitado",
    ]
    return any(p in t for p in termos)


def detectar_retorno_externo(texto):
    """
    Retorna True quando o texto indica que o fabricante/cliente respondeu
    ou que o material/equipamento chegou — fim da espera externa (T3),
    início da execução final pela equipe Grid (T4).
    """
    if not texto:
        return False
    t = texto.lower()
    termos = [
        "fabricante retornou", "fabricante respondeu",
        "cliente retornou", "cliente respondeu",
        "peça chegou", "peca chegou",
        "material chegou", "equipamento chegou",
        "peça recebida", "peca recebida",
        "material recebido", "equipamento recebido",
        "peça entregue", "peca entregue",
        "material entregue",
        "retorno do fabricante", "retorno do cliente",
        "posição do fabricante", "posicao do fabricante",
        "fabricante enviou", "fabricante autorizou",
        "garantia aprovada", "garantia autorizada",
        "liberado pelo fabricante", "liberado pelo cliente",
        "chegou a peça", "chegou a peca", "chegou o material",
        "chegou o equipamento",
        "já estamos com a peça", "ja estamos com a peca",
        "já estamos com o material", "ja estamos com o material",
    ]
    return any(p in t for p in termos)


def atualizar_ocorrencia(ws, num_linha, row, dados, origem="qualquer"):
    """
    Atualiza uma ocorrência existente.

    REGRAS:
    - Status de ocorrência existente NUNCA é alterado por esta função,
      EXCETO quando detecta chamado fabricante + campo normal
      → define 'Aguardando Fabricante'.
    - Status só é definido na CRIAÇÃO (gravar_nova_ocorrencia) ou
      na NORMALIZAÇÃO (normalizar_ocorrencia → Concluído).
    - Histórico e Ação: sempre acrescenta (nunca substitui).
    """
    hoje = agora_br().strftime("%d/%m")

    # Ação — acrescenta (não sobrescreve)
    acao_nova = (dados.get("acao_texto") or "").strip()
    if not vazio(acao_nova):
        acao_atual = (row[7] if len(row) > 7 else "").strip()
        if acao_nova not in acao_atual:
            nova_acao = (acao_atual + "\n" + acao_nova).strip() if acao_atual else acao_nova
            ws.update_cell(num_linha, 8, nova_acao)

    # Histórico — sempre acrescenta entrada nova
    hist_atual = (row[11] if len(row) > 11 else "").strip()
    # Monta entrada do histórico com o texto mais informativo disponível
    if not vazio(acao_nova):
        entrada_hist = f"{hoje} - {acao_nova}"
    else:
        novo_status = dados.get("status", "")
        if not vazio(novo_status):
            entrada_hist = f"{hoje} - Status: {novo_status}"
        else:
            entrada_hist = f"{hoje} - Atualização"
    if entrada_hist not in hist_atual:
        novo_hist = (hist_atual + "\n" + entrada_hist).strip() if hist_atual else entrada_hist
        ws.update_cell(num_linha, 12, novo_hist)

    # Status — NUNCA é alterado ao atualizar ocorrência existente.
    # O status só muda em dois casos:
    #   1. Criação de nova ocorrência (gravar_nova_ocorrencia) — usa o status do parse
    #   2. Normalização (normalizar_ocorrencia) — define Concluído
    #   3. Detecção explícita de "Aguardando Fabricante" (chamado + campo normal)
    #
    # Mensagens de grupo (webhook ou ronda) NÃO alteram o status de ocorrências abertas.
    status_atual = (row[8] if len(row) > 8 else "").strip().lower()
    ja_concluido = any(x in status_atual for x in ["conclu", "resolv", "fechad"])

    if not ja_concluido:
        # Única exceção permitida: detecção de chamado fabricante + campo normal
        texto_analise = " ".join(filter(None, [
            acao_nova,
            dados.get("causa", ""),
            dados.get("falha", ""),
        ]))
        if detectar_aguardando_fabricante(texto_analise):
            if status_atual not in ["aguardando fabricante"]:
                ws.update_cell(num_linha, 9, "Aguardando Fabricante")
                log.info(f"   → Status → Aguardando Fabricante (chamado+campo normal): linha {num_linha}")
        # NENHUMA outra condição altera o status de uma ocorrência existente

    # OS — preenche se estava vazio
    os_num = dados.get("os", "")
    if not vazio(os_num):
        os_atual = (row[10] if len(row) > 10 else "").strip()
        if vazio(os_atual):
            ws.update_cell(num_linha, 11, os_num)

    # Ticket Fabricante (J) — extrai do texto e preenche SOMENTE se vazio
    # (preserva qualquer ticket já preenchido manualmente por Fred).
    try:
        ticket_atual = (row[9] if len(row) > 9 else "").strip()
        if vazio(ticket_atual):
            texto_para_ticket = " ".join(filter(None, [
                acao_nova, dados.get("falha", ""), dados.get("causa", ""),
            ]))
            ticket_novo = extrair_ticket_fabricante(texto_para_ticket)
            if ticket_novo:
                ws.update_cell(num_linha, 10, ticket_novo)
                log.info(f"   → Ticket Fabricante detectado e gravado: linha {num_linha} = {ticket_novo}")
    except Exception as e:
        log.error(f"[Ticket] Erro ao extrair/gravar ticket fabricante: {e}")

    # ── Tempo Ativo O&M (T1-T3) — colunas M/N/O ─────────────────────────────
    # Cada gatilho só grava se a célula correspondente ainda estiver vazia
    # (preserva qualquer correção manual feita por Fred direto na planilha).
    try:
        texto_gatilho = " ".join(filter(None, [
            acao_nova,
            dados.get("causa", ""),
            dados.get("falha", ""),
        ]))

        # N — Data 1ª Ação (T1): equipe começou a atuar
        if detectar_primeira_acao(texto_gatilho):
            gravar_data_se_vazia(ws, num_linha, 13, row, label="Data 1ª Ação")

        # O — Data Encaminhamento (T2 → T3): foi pra fabricante/cliente/equipamento
        if detectar_encaminhamento(texto_gatilho):
            gravar_data_se_vazia(ws, num_linha, 14, row, label="Data Encaminhamento")

        # P — Data Retorno Externo (T3 → T4): fabricante/cliente retornou
        if detectar_retorno_externo(texto_gatilho):
            gravar_data_se_vazia(ws, num_linha, 15, row, label="Data Retorno Externo")
    except Exception as e:
        log.error(f"[T1-T4] Erro ao avaliar gatilhos de tempo ativo O&M: {e}")

    # ── Mensagem Original (V) — anexa o texto bruto desta atualização ─────────
    try:
        msg_bruta = dados.get("mensagem_bruta", "")
        if not vazio(msg_bruta):
            anexar_mensagem_original(ws, num_linha, 21, row, msg_bruta)
    except Exception as e:
        log.error(f"[Mensagens] Erro ao anexar mensagem original: {e}")

    log.info(f"🔄 Atualizado linha {num_linha} | {dados['usina']} / {dados.get('equipamento','')} [{origem}]")



def normalizar_ocorrencia(ws, num_linha, row, dados):
    """Fecha uma ocorrência: status → Concluído + entrada no histórico."""
    hoje = agora_br().strftime("%d/%m")
    ws.update_cell(num_linha, 9, "Concluído")

    if not vazio(dados.get("os", "")):
        ws.update_cell(num_linha, 11, dados["os"])

    hist_atual   = (row[11] if len(row) > 11 else "").strip()
    nova_entrada = f"{hoje} - Ocorrência normalizada"
    acao_txt = dados.get("acao_texto", "")
    if not vazio(acao_txt):
        nova_entrada += f"\n{hoje} - {acao_txt}"
    novo_hist = (hist_atual + "\n" + nova_entrada).strip() if hist_atual else nova_entrada
    ws.update_cell(num_linha, 12, novo_hist)

    # ── Data de Fechamento (U) — gravada na normalização, SOMENTE se vazia ──
    # Mesma regra de prioridade das demais datas (M/N/O/P): preenchimento
    # manual sempre tem prioridade sobre o automático. Se Fred já corrigiu
    # essa célula manualmente, o robô nunca sobrescreve.
    try:
        gravar_data_se_vazia(ws, num_linha, 20, row, label="Data de Fechamento")
    except Exception as e:
        log.error(f"[T1-T4] Erro ao gravar Data de Fechamento: {e}")

    # ── Mensagem Original (V) — anexa a mensagem de normalização ──────────────
    try:
        msg_bruta = dados.get("mensagem_bruta", "")
        if not vazio(msg_bruta):
            anexar_mensagem_original(ws, num_linha, 21, row, msg_bruta)
    except Exception as e:
        log.error(f"[Mensagens] Erro ao anexar mensagem original na normalização: {e}")

    log.info(f"✅ Normalizado linha {num_linha} | {dados['usina']}")


def primeira_linha_vazia(todos):
    ultima_com_dado = 1
    for i, row in enumerate(todos[1:], start=2):
        if row and row[0] and str(row[0]).strip():
            ultima_com_dado = i
    return ultima_com_dado + 1


def gravar_nova_ocorrencia(ws, todos, dados):
    novo_id       = proximo_id(todos)
    proxima_linha = primeira_linha_vazia(todos)
    agora_str = agora_br().strftime("%d/%m/%Y %H:%M:%S")

    # Tenta extrair o número do chamado/ticket do fabricante já na abertura
    texto_para_ticket = " ".join(filter(None, [
        dados.get("acao", ""), dados.get("acao_texto", ""),
        dados.get("falha", ""), dados.get("causa", ""), dados.get("historico", ""),
    ]))
    ticket = extrair_ticket_fabricante(texto_para_ticket)

    linha = [
        novo_id,
        dados["cliente"],
        dados["usina"],
        dados["equipamento"],
        dados["falha"],
        dados["causa"],
        dados["equip_impact"],
        dados["acao"],
        dados["status"],
        ticket,
        dados["os"],
        dados["historico"],
    ]
    ws.update(f"A{proxima_linha}:L{proxima_linha}", [linha])
    log.info(f"➕ Nova ocorrência ID={novo_id} | {dados['usina']} — {dados['equipamento']} | linha {proxima_linha}")

    # ── Data de Abertura (M) — sempre gravada na criação ───────────────────
    try:
        ws.update_cell(proxima_linha, 13, agora_str)
    except Exception as e:
        log.error(f"[T1-T4] Erro ao gravar Data de Abertura: {e}")

    # ── Data 1ª Ação (N) — se a própria mensagem de abertura já indicar que
    #    a equipe começou a atuar (técnico acionado, equipe a caminho, etc.)
    try:
        texto_analise = " ".join(filter(None, [
            dados.get("acao", ""), dados.get("falha", ""), dados.get("historico", ""),
        ]))
        if detectar_primeira_acao(texto_analise):
            ws.update_cell(proxima_linha, 14, agora_str)
            log.info(f"   → Data 1ª Ação gravada na abertura (gatilho na própria mensagem): linha {proxima_linha}")
    except Exception as e:
        log.error(f"[T1-T4] Erro ao gravar Data 1ª Ação na abertura: {e}")

    # ── Mensagem Original (V) — guarda o texto bruto que originou o registro ──
    try:
        msg_bruta = dados.get("mensagem_bruta", "")
        if not vazio(msg_bruta):
            ws.update_cell(proxima_linha, 22, f"{agora_br().strftime('%d/%m %H:%M')} - {msg_bruta.strip()}")
    except Exception as e:
        log.error(f"[Mensagens] Erro ao gravar mensagem original na abertura: {e}")


    # Notificação push — nova ocorrência
    try:
        usina_nome = dados.get("usina", "")
        equip_nome = dados.get("equipamento", "")
        falha_txt  = dados.get("falha", "")
        cliente    = dados.get("cliente", "")

        # Detecta desligamento
        fc = (falha_txt + " " + dados.get("causa", "")).lower()
        eh_deslig = bool(re.search(
            r"usina\s+desligad|ufv\s+desligad|desligamento\s+da\s+usina|usina\s+parad", fc
        ))

        if eh_deslig:
            enviar_push(
                titulo=f"⚡ USINA DESLIGADA — {usina_nome}",
                corpo=f"{falha_txt or 'Usina sem geração'} · {cliente}",
                tipo="desligamento",
                url=f"https://fred-alexandrino.github.io/PAINELDEFALHAS/?ocorrencia={novo_id}",
            )
        else:
            enviar_push(
                titulo=f"🔴 Nova falha — {usina_nome}",
                corpo=f"{equip_nome}: {falha_txt[:80] if falha_txt else 'Nova ocorrência registrada'} · {cliente}",
                tipo="nova_ocorrencia",
                url=f"https://fred-alexandrino.github.io/PAINELDEFALHAS/?ocorrencia={novo_id}",
            )
    except Exception as e:
        log.error(f"[Push] Erro ao notificar nova ocorrência: {e}")

    return novo_id


# ── Processamento principal ───────────────────────────────────────────────────
#
# LÓGICA POR BLOCO (mesma para tempo real e botão Verificar Rondas):
#
#  1. Parseia o bloco → extrai usina, equipamento, falha, ação, status, OS
#  2. Busca na planilha por fingerprint (usina + tipo_equip + num + palavras_falha)
#
#  CASO A — NÃO encontrou na planilha:
#    → CRIA nova linha
#
#  CASO B — Encontrou, é normalização (✅ NORMALIZADO):
#    → FECHA a ocorrência (status = Concluído, histórico atualizado)
#
#  CASO C — Encontrou, ação NÃO mudou e status NÃO mudou:
#    → IGNORA (mensagem repetida de ronda sem informação nova)
#
#  CASO D — Encontrou, ação OU status mudou:
#    → ATUALIZA (acrescenta ação + entrada no histórico + status se diferente)

def processar_texto(texto, origem="webhook"):
    ws     = get_sheet()
    todos  = carregar_planilha(ws)
    blocos = separar_blocos(texto)
    resultado = {"novos": [], "atualizados": [], "normalizados": [], "ignorados": 0}

    for bloco in blocos:
        dados = parse_bloco(bloco)
        if not dados:
            resultado["ignorados"] += 1
            continue

        usina     = dados.get("usina", "")
        equip     = dados.get("equipamento", "")
        falha     = dados.get("falha", "")
        normalizar = dados.get("normalizar", False)

        # ── Caso especial: formato com atualizações individuais por ativo ──
        # Ex: "Tracker 3 normalizado, Tracker 5 em garantia"
        atualizacoes_individuais = extrair_atualizacoes_por_ativo(dados.get("acao_texto", ""))

        if atualizacoes_individuais:
            alguma_acao = False
            for upd in atualizacoes_individuais:
                existente = buscar_por_fingerprint(todos, usina, upd["equipamento"], falha, dados.get("os",""))
                if existente:
                    num_linha, row = existente
                    if upd["normalizar"]:
                        normalizar_ocorrencia(ws, num_linha, row, {
                            **dados,
                            "acao_texto": upd["acao_resumida"],
                            "os": dados.get("os", ""),
                        })
                        resultado["normalizados"].append(f"{usina} - {upd['equipamento']}")
                    else:
                        atualizar_ocorrencia(ws, num_linha, row, {
                            **dados,
                            "acao_texto": upd["acao_resumida"],
                        }, origem=origem)
                        resultado["atualizados"].append(f"{usina} - {upd['equipamento']}")
                    alguma_acao = True
                    todos = carregar_planilha(ws)
            if not alguma_acao:
                # Nenhum ativo encontrado → cria novo
                novo_id = gravar_nova_ocorrencia(ws, todos, dados)
                resultado["novos"].append({"id": novo_id, "usina": usina})
                todos = carregar_planilha(ws)
            continue

        # ── Múltiplos inversores numa mesma mensagem ──────────────────────
        # Ex: "Inversores 6 e 7" → cria/atualiza INV-06 e INV-07 separadamente
        multi_inv = extrair_inversores_multiplos(bloco, dados)
        if multi_inv:
            for dados_inv in multi_inv:
                existente_inv = buscar_por_fingerprint(todos, dados_inv["usina"], dados_inv["equipamento"], dados_inv["falha"], dados_inv.get("os",""))
                if not existente_inv:
                    novo_id = gravar_nova_ocorrencia(ws, todos, dados_inv)
                    resultado["novos"].append({"id": novo_id, "usina": dados_inv["usina"]})
                    todos = carregar_planilha(ws)
                elif dados_inv.get("normalizar"):
                    num_linha, row = existente_inv
                    normalizar_ocorrencia(ws, num_linha, row, dados_inv)
                    resultado["normalizados"].append(f"{dados_inv['usina']} - {dados_inv['equipamento']}")
                    todos = carregar_planilha(ws)
                else:
                    num_linha, row = existente_inv
                    if acao_mudou(row, dados_inv.get("acao_texto","")):
                        atualizar_ocorrencia(ws, num_linha, row, dados_inv, origem="ronda")
                        resultado["atualizados"].append(f"{dados_inv['usina']} - {dados_inv['equipamento']}")
                        todos = carregar_planilha(ws)
                    else:
                        resultado["ignorados"] += 1
            continue  # pula o fluxo principal — já foi tratado acima

        # ── Normaliza nomenclatura de inversores na falha ──────────────────
        dados["falha"]        = normalizar_inversores(dados.get("falha", ""))
        dados["equipamento"]  = _limpar_equipamento(dados.get("equipamento", ""))
        dados["equip_impact"] = dados["equipamento"]
        equip = dados["equipamento"]
        falha = dados["falha"]

        # ── Fluxo principal ────────────────────────────────────────────────
        existente = buscar_por_fingerprint(todos, usina, equip, falha, dados.get("os",""))

        # Se não encontrou aberta e é normalização, busca também nas concluídas
        # (para não criar linha nova quando a ocorrência já estava concluída em outro grupo)
        if not existente and normalizar:
            for i2, row2 in enumerate(todos[1:], start=2):
                if len(row2) < 4: continue
                if not usinas_sao_iguais(usina, row2[2].strip()): continue
                if equipamentos_sao_iguais(equip, row2[3].strip()):
                    existente = (i2, row2)
                    log.info(f"[Normaliz] Encontrada ocorrência (incl. concluídas) para {usina} / {equip}: linha {i2}")
                    break

        if not existente and normalizar:
            # Normalização sem ocorrência existente — ignora, não cria linha nova
            log.info(f"[Normaliz] Sem ocorrência para normalizar — ignorando: {usina} / {equip}")
            resultado["ignorados"] += 1

        elif not existente:
            # CASO A — nova ocorrência
            novo_id = gravar_nova_ocorrencia(ws, todos, dados)
            resultado["novos"].append({"id": novo_id, "usina": usina})
            todos = carregar_planilha(ws)

        elif normalizar:
            # CASO B — normalização / conclusão
            num_linha, row = existente
            status_atual = row[8].strip().lower() if len(row) > 8 else ""
            if "conclu" in status_atual or "resolv" in status_atual:
                log.info(f"[Normaliz] Já concluída — ignorando duplicata: {usina} / {equip}")
                resultado["ignorados"] += 1
            else:
                normalizar_ocorrencia(ws, num_linha, row, dados)
                resultado["normalizados"].append(usina)
                todos = carregar_planilha(ws)

        else:
            num_linha, row = existente
            acao_nova = dados.get("acao_texto", "")

            # CASO D: atualiza apenas se houver ação nova OU chamado+campo normal
            # Status NUNCA é alterado ao atualizar ocorrência existente
            tem_info_nova = acao_mudou(row, acao_nova)
            tem_aguardando = detectar_aguardando_fabricante(
                " ".join(filter(None, [acao_nova, dados.get("causa",""), dados.get("falha","")]))
            )

            if tem_info_nova or tem_aguardando:
                atualizar_ocorrencia(ws, num_linha, row, dados, origem="ronda")
                resultado["atualizados"].append(usina)
                todos = carregar_planilha(ws)
            else:
                # CASO C — nenhuma informação nova → ignora
                log.info(f"⏭️  Sem novidade: {usina} / {equip} — ignorado")
                resultado["ignorados"] += 1

    return resultado


def eh_ronda_status_ok(texto):
    """
    Retorna True quando a mensagem é uma ronda de status informando que
    tudo está OK na usina — sem falhas, sem ocorrências.

    Detecta combinações como:
    - "RONDA DIÁRIA" + "Sem Ocorrência" (em Ocorrências durante o turno E pendentes)
    - "<usina> OK." sem emoji de falha (🔴🟡🟠)
    - "Status Atual:" + "<usina> OK" + "Sem Ocorrência"
    """
    t = texto.lower()

    # Presença de emoji de falha = há problema real → não ignorar
    tem_falha_emoji = bool(re.search(r"🔴|🟡|🟠|⏸️", texto))
    if tem_falha_emoji:
        return False

    # Padrão 1: RONDA DIÁRIA / RONDA / Status do dia com "Sem Ocorrência" explícito
    eh_ronda = bool(re.search(
        r"(?:ronda\s+di[aá]ria|status\s+do\s+dia|status\s+operacional|cos\s+[-–]\s*grid|ronda\s+de\s+campo)",
        t
    ))
    sem_ocorrencia = bool(re.search(
        r"sem\s+ocorr[eê]ncia|sem\s+ocorr[eê]ncias|sem\s+ocorr[eê]nci[ao]",
        t
    ))

    if eh_ronda and sem_ocorrencia:
        return True

    # Padrão 2: "<usina nome> OK." sem qualquer desvio (formato COS Grid OK)
    # Ex: "ABC Morada Nova OK." ou "Araputanga OK."
    tem_usina_ok = bool(re.search(r"\w[\w\s]+\s+ok\.", t))
    tem_desvio = bool(re.search(
        r"desvio|falha|problema|ocorr[eê]ncia|parado|desligad|comunica[cç]",
        t
    ))
    if tem_usina_ok and not tem_desvio and sem_ocorrencia:
        return True

    return False


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.route("/webhook", methods=["POST"])
def webhook():
    """
    Recebe mensagens em tempo real do server.js.
    Chamado automaticamente pelo monitoramento — não depende do botão.
    """
    try:
        payload = request.get_json(force=True)
        if not payload:
            return jsonify({"status": "ignored", "reason": "empty payload"}), 200

        if WEBHOOK_SECRET:
            secret = request.headers.get("X-Webhook-Secret", "")
            if secret != WEBHOOK_SECRET:
                return jsonify({"error": "unauthorized"}), 401

        evento = payload.get("event", "")
        if evento not in ("messages.upsert", "MESSAGES_UPSERT"):
            return jsonify({"status": "ignored", "event": evento}), 200

        data    = payload.get("data", {})
        msg_obj = data if "message" in data else payload

        if msg_obj.get("key", {}).get("fromMe"):
            return jsonify({"status": "ignored", "reason": "own message"}), 200

        message = msg_obj.get("message", {})
        texto   = (
            message.get("conversation") or
            message.get("extendedTextMessage", {}).get("text") or ""
        )

        if not texto:
            return jsonify({"status": "ignored", "reason": "no text"}), 200

        remote_jid = msg_obj.get("key", {}).get("remoteJid", "")
        if "@g.us" not in remote_jid:
            return jsonify({"status": "ignored", "reason": "not a group"}), 200

        if GRUPOS_FILTRO and GRUPOS_FILTRO[0]:
            if not any(g.strip() in remote_jid for g in GRUPOS_FILTRO):
                return jsonify({"status": "ignored", "reason": "group not in filter"}), 200

        if eh_atualizacao_atividade(texto):
            grupo_nome = remote_jid.split("@")[0]
            gravar_log_mensagem(remote_jid, grupo_nome, texto)
            resultado_ativ = processar_atualizacao_atividade(texto, editor=f"tecnico:{grupo_nome}")
            log.info(f"[Atividades WhatsApp] grupo={grupo_nome} resultado={resultado_ativ}")
            return jsonify({"status": "ok", "tipo": "atividade", **resultado_ativ}), 200

        tem_usina  = bool(re.search(r"Usina:", texto, re.IGNORECASE))
        tem_emoji  = bool(re.search(r"🔴|🟡|🟢|🟠|✅|⏸️", texto))
        tem_bullet = eh_formato_cos_grid(texto)

        if not tem_usina and not tem_emoji and not tem_bullet:
            return jsonify({"status": "ignored", "reason": "no failure content"}), 200

        # Ignora mensagens de ronda diária que informam tudo OK / sem ocorrência
        if eh_ronda_status_ok(texto):
            gravar_log_mensagem(remote_jid, remote_jid.split("@")[0], texto)
            return jsonify({"status": "ignored", "reason": "ronda_diaria_ok"}), 200

        # Grava no log antes de processar (para histórico de varredura)
        grupo_nome = remote_jid.split("@")[0]
        gravar_log_mensagem(remote_jid, grupo_nome, texto)

        resultado = processar_texto(texto)

        total = len(resultado["novos"]) + len(resultado["atualizados"]) + len(resultado["normalizados"])
        if total > 0:
            log.info(f"✅ [Tempo real] {len(resultado['novos'])} novos, {len(resultado['atualizados'])} atualizados, {len(resultado['normalizados'])} normalizados")

        # Limpeza automática: remove linhas do log com mais de 5 dias
        try:
            removidas = limpar_log_antigo()
            if removidas > 0:
                log.info(f"🧹 [Rondas] Log limpo: {removidas} linha(s) com mais de 5 dias removidas")
        except Exception as e_clean:
            log.warning(f"[Rondas] Limpeza do log falhou (não crítico): {e_clean}")

        return jsonify({"status": "ok", **resultado}), 200

    except Exception as e:
        log.error(f"❌ Erro no webhook: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/processar-texto-manual", methods=["POST"])
def processar_texto_manual():
    """Ferramenta de recuperação: processa manualmente o texto de uma
    mensagem de ocorrência (útil quando uma mensagem real chegou num
    grupo do WhatsApp mas não foi capturada — ex.: sessão do WhatsApp
    caiu no momento). Usa o mesmo parser do webhook normal, então o
    resultado fica idêntico ao que teria acontecido automaticamente."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    body = request.get_json(force=True, silent=True) or {}
    texto = body.get("texto", "")
    resultado = processar_texto(texto)
    return jsonify({"ok": True, **resultado}), 200


@app.route("/rondas", methods=["POST"])
def verificar_rondas():
    """
    Botão "Verificar Rondas" do dashboard.

    Busca as mensagens das últimas 6 horas em cada grupo configurado
    via GET /api/messages/:grupoId no server.js, e processa as relevantes.

    O monitoramento em tempo real NÃO é afetado por este endpoint.

    NOTA: Este endpoint é chamado diretamente pelo dashboard (GitHub Pages)
    via fetch(). Por isso NÃO exige WEBHOOK_SECRET — a autenticação é feita
    pelo login do próprio dashboard. O WEBHOOK_SECRET é usado apenas na
    comunicação interna entre server.js → /webhook.

    Body (opcional):
      { "horas": 6 }
    """
    try:
        payload = request.get_json(force=True) or {}
        horas   = int(payload.get("horas", 6))

        if not WPP_SERVER_URL:
            return jsonify({
                "ok":    False,
                "error": "WPP_SERVER_URL não configurado",
                "hint":  "Adicione a variável de ambiente WPP_SERVER_URL com a URL do servidor Baileys (server.js)",
            }), 400

        log.info(f"[Rondas] Iniciando varredura no log | últimas {horas}h")

        resultado_total = {
            "novos":        [],
            "atualizados":  [],
            "normalizados": [],
            "ignorados":    0,
            "mensagens_lidas": 0,
            "mensagens_processadas": 0,
        }

        # Lê mensagens do log das últimas N horas
        mensagens = ler_log_mensagens(horas)
        resultado_total["mensagens_lidas"] = len(mensagens)

        # Marca todas as mensagens como processadas em lote após processar
        ws_log = get_log_sheet()
        rows_log = ws_log.get_all_values()
        linhas_para_marcar = []

        for i, msg in enumerate(mensagens):
            texto = msg.get("texto", "")
            if not texto:
                continue

            # Filtra apenas mensagens de ronda/ocorrência
            tem_usina  = bool(re.search(r"Usina:", texto, re.IGNORECASE))
            tem_emoji  = bool(re.search(r"🔴|🟡|🟢|🟠|✅|⏸️", texto))
            tem_desvio = bool(re.search(r"DESVIO:", texto, re.IGNORECASE))
            tem_bullet = eh_formato_cos_grid(texto)

            relevante = tem_usina or tem_emoji or tem_desvio or tem_bullet

            # Mensagem de ronda diária informando tudo OK → não cria ocorrências
            if relevante and eh_ronda_status_ok(texto):
                relevante = False
                log.info(f"[Rondas] Ronda diária OK ignorada: {texto[:60]!r}")

            if relevante:
                try:
                    res = processar_texto(texto, origem="ronda")
                    resultado_total["novos"]        += res.get("novos", [])
                    resultado_total["atualizados"]  += res.get("atualizados", [])
                    resultado_total["normalizados"] += res.get("normalizados", [])
                    resultado_total["ignorados"]    += res.get("ignorados", 0)
                    resultado_total["mensagens_processadas"] += 1
                except Exception as e:
                    log.error(f"[Rondas] Erro ao processar mensagem: {e}")

            # Marca como processada (relevante ou não) para não reprocessar
            linhas_para_marcar.append(msg.get("linha_idx"))

        # Marca em lote no Sheets — uma única requisição para todas as linhas
        try:
            if linhas_para_marcar:
                idxs_validos = [idx for idx in linhas_para_marcar if idx]
                if idxs_validos:
                    # batch_update: uma única chamada à API
                    ws_log.batch_update([{
                        'range': f'E{idx}',
                        'values': [['✅']]
                    } for idx in idxs_validos])
                    log.info(f"[Rondas] {len(idxs_validos)} mensagens marcadas como processadas")
        except Exception as e:
            log.warning(f"[Rondas] Erro ao marcar processadas: {e}")

        total = (len(resultado_total["novos"]) +
                 len(resultado_total["atualizados"]) +
                 len(resultado_total["normalizados"]))

        log.info(f"[Rondas] Concluído: {total} ação(ões) | {resultado_total['mensagens_lidas']} msgs lidas do log")
        return jsonify({"ok": True, "horas_verificadas": horas, **resultado_total}), 200

    except Exception as e:
        log.error(f"[Rondas] Erro geral: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Notificações Push ────────────────────────────────────────────────────────

def get_push_sheet():
    """Retorna a aba 'Push Subscriptions', criando-a com cabeçalho se não existir."""
    gc = get_gc()
    sh = gc.open_by_key(SHEET_ID)
    try:
        return sh.worksheet(PUSH_SHEET_NAME)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=PUSH_SHEET_NAME, rows=200, cols=3)
        ws.update("A1:C1", [["Endpoint", "Subscription", "DataCriacao"]])
        return ws

def carregar_push_subscriptions():
    """
    Carrega as subscriptions salvas na planilha para dentro de _push_subscriptions
    (em memória). Chamado uma vez na inicialização do processo, para que os
    dispositivos cadastrados sobrevivam a reinícios do Render.
    """
    try:
        ws = get_push_sheet()
        rows = ws.get_all_values()[1:]  # pula cabeçalho
        carregadas = 0
        for row in rows:
            if len(row) < 2 or not row[0] or not row[1]:
                continue
            try:
                _push_subscriptions[row[0]] = json.loads(row[1])
                carregadas += 1
            except (json.JSONDecodeError, TypeError):
                continue
        log.info(f"[Push] {carregadas} subscription(ões) carregada(s) da planilha")
    except Exception as e:
        log.error(f"[Push] Erro ao carregar subscriptions da planilha: {e}")

def salvar_push_subscription(endpoint, sub):
    """Persiste (ou atualiza) uma subscription na planilha. Retorna True/False."""
    try:
        ws = get_push_sheet()
        cell = ws.find(endpoint, in_column=1)
        linha = [endpoint, json.dumps(sub), agora_br().strftime("%d/%m/%Y %H:%M:%S")]
        if cell:
            ws.update(f"A{cell.row}:C{cell.row}", [linha])
        else:
            ws.append_row(linha)
        return True
    except Exception as e:
        log.error(f"[Push] Erro ao salvar subscription na planilha: {e}", exc_info=True)
        return False

def remover_push_subscription(endpoint):
    """Remove uma subscription da planilha (expirada ou desativada pelo usuário)."""
    try:
        ws = get_push_sheet()
        cell = ws.find(endpoint, in_column=1)
        if cell:
            ws.delete_rows(cell.row)
    except Exception as e:
        log.error(f"[Push] Erro ao remover subscription da planilha: {e}")

def enviar_push(titulo, corpo, tipo="geral", url="https://fred-alexandrino.github.io/PAINELDEFALHAS/"):
    """
    Envia notificação push para todos os dispositivos registrados.
    tipo: "desligamento" | "nova_ocorrencia" | "geral"
    """
    if not PUSH_ENABLED:
        log.warning("[Push] pywebpush não disponível")
        return 0
    if not VAPID_PRIVATE_KEY:
        log.warning("[Push] VAPID_PRIVATE_KEY não configurada")
        return 0
    if not _push_subscriptions:
        log.info("[Push] Nenhum dispositivo registrado")
        return 0

    payload = json.dumps({
        "title": titulo,
        "body":  corpo,
        "tipo":  tipo,
        "url":   url,
        "tag":   f"painel-{tipo}",
        "icon":  "https://fred-alexandrino.github.io/PAINELDEFALHAS/icon-192.png",
    })

    enviados = 0
    expirados = []
    for endpoint, sub in list(_push_subscriptions.items()):
        try:
            webpush(
                subscription_info=sub,
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims=VAPID_CLAIMS,
                headers={"Urgency": "high"},
                ttl=86400,
            )
            enviados += 1
            log.info(f"[Push] Enviado para {endpoint[:40]}...")
        except WebPushException as e:
            if "410" in str(e) or "404" in str(e):
                # Subscription expirada — remove
                expirados.append(endpoint)
                log.info(f"[Push] Subscription expirada removida: {endpoint[:40]}")
            else:
                log.error(f"[Push] Erro ao enviar: {e}")
        except Exception as e:
            log.error(f"[Push] Erro inesperado: {e}")

    for ep in expirados:
        _push_subscriptions.pop(ep, None)
        remover_push_subscription(ep)

    log.info(f"[Push] {enviados} notificação(ões) enviada(s)")
    return enviados


@app.route("/push/subscribe", methods=["POST", "OPTIONS"])
def push_subscribe():
    """
    Registra subscription de notificação push de um dispositivo.
    Chamado pelo dashboard ao clicar em "Ativar Notificações".
    """
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200

    try:
        payload = request.get_json(force=True) or {}
        sub = payload.get("subscription")
        if not sub or not sub.get("endpoint"):
            return jsonify({"error": "subscription inválida"}), 400

        endpoint = sub["endpoint"]
        ja_existia = endpoint in _push_subscriptions
        _push_subscriptions[endpoint] = sub

        salvo = salvar_push_subscription(endpoint, sub)
        if not salvo:
            # Reverte o registro em memória — melhor reportar erro real do que
            # fingir sucesso e perder essa subscription num futuro restart.
            _push_subscriptions.pop(endpoint, None)
            log.error(f"[Push] FALHA ao persistir subscription (endpoint não salvo na planilha): {endpoint[:60]}...")
            return jsonify({"ok": False, "error": "Falha ao salvar a inscrição no servidor. Tente novamente em instantes."}), 500

        log.info(f"[Push] Subscription registrada ({'já existia' if ja_existia else 'nova'}): {endpoint[:60]}...")
        log.info(f"[Push] Total de dispositivos: {len(_push_subscriptions)}")

        # Envia notificação de boas-vindas apenas para dispositivos realmente novos
        if not ja_existia:
            try:
                webpush(
                    subscription_info=sub,
                    data=json.dumps({
                        "title": "🔔 Painel O&M — Notificações ativas!",
                        "body":  "Você receberá alertas de desligamentos e novas ocorrências.",
                        "tipo":  "geral",
                        "url":   "https://fred-alexandrino.github.io/PAINELDEFALHAS/",
                    }),
                    vapid_private_key=VAPID_PRIVATE_KEY,
                    vapid_claims=VAPID_CLAIMS,
                    headers={"Urgency": "high"},
                    ttl=86400,
                ) if PUSH_ENABLED and VAPID_PRIVATE_KEY else None
            except Exception as e:
                log.warning(f"[Push] Erro na notificação de boas-vindas: {e}")

        return jsonify({"ok": True, "total": len(_push_subscriptions)}), 200

    except Exception as e:
        log.error(f"[Push] Erro ao registrar subscription: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/push/unsubscribe", methods=["POST", "OPTIONS"])
def push_unsubscribe():
    """
    Remove a subscription de notificação push de um dispositivo (do backend
    e da planilha). Chamado pelo dashboard ao clicar em "Desativar Notificações".
    """
    if request.method == "OPTIONS":
        return jsonify({"ok": True}), 200

    try:
        payload = request.get_json(force=True) or {}
        endpoint = payload.get("endpoint") or (payload.get("subscription") or {}).get("endpoint")
        if not endpoint:
            return jsonify({"error": "endpoint não informado"}), 400

        _push_subscriptions.pop(endpoint, None)
        remover_push_subscription(endpoint)
        log.info(f"[Push] Subscription removida: {endpoint[:60]}...")
        return jsonify({"ok": True, "total": len(_push_subscriptions)}), 200

    except Exception as e:
        log.error(f"[Push] Erro ao remover subscription: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/push/test", methods=["POST"])
def push_test():
    """Envia notificação de teste para todos os dispositivos registrados."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"error": "unauthorized"}), 401
    n = enviar_push(
        titulo="🧪 Teste — Painel O&M",
        corpo="Se você está vendo isso, as notificações estão funcionando!",
        tipo="geral",
    )
    return jsonify({"ok": True, "enviados": n}), 200


@app.route("/notificar-edicao-planilha", methods=["POST"])
def notificar_edicao_planilha():
    """
    Recebido do Apps Script (gatilho onEdit) sempre que alguém edita
    manualmente qualquer célula da planilha pela interface do Google Sheets.
    Edições feitas por script/API (bot WhatsApp, dashboard, gspread) NÃO
    disparam o onEdit do Google — só edição humana direta na UI.
    """
    try:
        if SHEET_EDIT_SECRET:
            secret = request.headers.get("X-Sheet-Secret", "")
            if secret != SHEET_EDIT_SECRET:
                return jsonify({"error": "unauthorized"}), 401

        body = request.get_json(force=True) or {}
        aba          = body.get("aba", "planilha")
        linha        = body.get("linha", "")
        cabecalho    = body.get("cabecalho") or f"coluna {body.get('coluna', '?')}"
        valor_antigo = body.get("valorAntigo", "")
        valor_novo   = body.get("valorNovo", "")
        usuario      = body.get("usuario", "desconhecido")
        id_registro  = str(body.get("idValor", "")).strip()

        titulo = f"✏️ Edição manual — {aba}"
        corpo = (f"Linha {linha} · {cabecalho}: "
                 f"\"{valor_antigo or '—'}\" → \"{valor_novo or '—'}\" (por {usuario})")

        # Se for edição no Painel de Atividades, busca o número da OS e o
        # tema (Ação/Tarefa) da linha editada — sem isso a notificação só
        # dizia "linha X mudou", impossível saber do que se tratava sem
        # abrir a planilha (corrigido 31/07/2026; reaplicado 05/08/2026
        # depois de constatado que outra sessão tinha revertido pra versão
        # anterior — ver nota de vigilância em memória sobre correções
        # sendo sobrescritas sem aviso).
        if aba == ATIVIDADES_SHEET_NAME and id_registro:
            try:
                ws_ativ = get_atividades_sheet()
                todos_ativ = ws_ativ.get_all_values()
                encontrada = buscar_atividade_por_id_ou_os(todos_ativ, id_registro)
                if encontrada:
                    _, linha_ativ = encontrada
                    numero_os_ativ = linha_ativ[13].strip() if len(linha_ativ) > 13 else ""
                    descricao_ativ = linha_ativ[4].strip() if len(linha_ativ) > 4 else ""
                    usina_ativ = linha_ativ[2].strip() if len(linha_ativ) > 2 else ""
                    tema_ativ = descricao_ativ or "Descrição não informada"
                    titulo = f"✏️ Edição manual" + (f" — OS {numero_os_ativ}" if numero_os_ativ else "") + (f" — {usina_ativ}" if usina_ativ else "")
                    corpo = f"{tema_ativ}\n{cabecalho}: \"{valor_antigo or '—'}\" → \"{valor_novo or '—'}\" (por {usuario})"
            except Exception as e:
                log.error(f"[EdicaoPlanilha] Falha ao enriquecer com tema da OS: {e}")

        url = "https://fred-alexandrino.github.io/PAINELDEFALHAS/"
        if id_registro:
            if aba == ATIVIDADES_SHEET_NAME:
                url += f"?atividade={id_registro}"
            elif aba == SHEET_NAME:
                url += f"?ocorrencia={id_registro}"

        n = enviar_push(titulo=titulo, corpo=corpo, tipo="edicao_manual", url=url)
        return jsonify({"ok": True, "enviados": n}), 200
    except Exception as e:
        log.error(f"[EdicaoPlanilha] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/rondas/grupos", methods=["POST"])
def rondas_por_grupo():
    """
    Retorna as últimas mensagens de CADA grupo monitorado — somente leitura.
    Usa ler_log_historico() que inclui mensagens já processadas.
    """
    try:
        payload = request.get_json(force=True) or {}
        horas   = int(payload.get("horas", 24))
        # Usa histórico completo (inclui processadas) — somente para visualização
        mensagens = ler_log_historico(horas)

        # Agrupa por grupo_id
        grupos_map = {}
        for msg in mensagens:
            gid = msg.get("grupo_id", "")
            if gid not in grupos_map:
                grupos_map[gid] = []
            grupos_map[gid].append({
                "texto":      msg.get("texto", ""),
                "timestamp":  msg.get("timestamp", ""),
                "processado": msg.get("processado", False),
            })

        grupos = []
        for gid, msgs in grupos_map.items():
            grupos.append({
                "id":        gid,
                "total":     len(msgs),
                "mensagens": msgs[-5:],  # últimas 5 por grupo
            })

        # Garante que todos os grupos configurados aparecem (mesmo sem msgs)
        ids_com_msgs = {g["id"] for g in grupos}
        for gid in GRUPOS_FILTRO:
            gid = gid.strip()
            if gid and gid not in ids_com_msgs:
                grupos.append({"id": gid, "total": 0, "mensagens": []})

        return jsonify({"ok": True, "horas": horas, "grupos": grupos}), 200

    except Exception as e:
        log.error(f"[Grupos] Erro: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


APP_VERSION = "2026-07-01-fix-get_sheet"

@app.route("/health", methods=["GET"])
def health():
    # Reaproveita o ping de 5min que o monitor UptimeRobot já faz aqui pra
    # também rodar o ciclo do FV Energias (ronda + alertas). Roda em
    # try/except isolado: se a Solplanet estiver fora do ar, o /health
    # continua respondendo normalmente (não derruba o monitor de uptime).
    try:
        _fv_energias_processar_ciclo()
    except Exception as e:
        log.error(f"[FV Energias] Erro no ciclo automático via /health: {e}")

    return jsonify({
        "status":     "ok",
        "version":    APP_VERSION,
        "timestamp":  agora_br().isoformat(),
        "wpp_server": WPP_SERVER_URL or "não configurado",
    }), 200


# ── Mapeamento de campo (nome JS) → coluna na planilha (1-based) ──────────
CAMPO_COL = {
    "falha":               5,
    "causa":               6,
    "impactados":          7,
    "acao":                8,
    "status":              9,
    "ticketFabricante":    10,
    "numeroOS":            11,
    "historico":           12,
    "dataAbertura":        13,
    "dataPrimeiraAcao":    14,
    "dataEncaminhamento":  15,
    "dataRetornoExterno":  16,
    "dataFechamento":      21,
}

@app.route("/atualizar-campo", methods=["POST", "OPTIONS"])
def atualizar_campo():
    """
    Endpoint chamado pelo dashboard para salvar alterações de campo individual.
    Body JSON: { id, field, value, editor, append? }
    """
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido"}), 400

    ocorrencia_id = str(body.get("id", "")).strip()
    field         = str(body.get("field", "")).strip()
    value         = str(body.get("value", "")).strip()
    append        = body.get("append", False)
    editor        = str(body.get("editor", "dashboard")).strip()

    if not ocorrencia_id or not field:
        return jsonify({"ok": False, "error": "id e field são obrigatórios"}), 400

    col = CAMPO_COL.get(field)
    if col is None:
        return jsonify({"ok": False, "error": f"Campo '{field}' não mapeado"}), 400

    try:
        ws   = get_sheet()
        rows = ws.get_all_values()
    except Exception as e:
        log.error(f"[atualizar-campo] Erro ao abrir planilha: {e}")
        return jsonify({"ok": False, "error": f"Erro ao acessar planilha: {str(e)}"}), 500

    # Busca por ID + Equipamento + OS (chave composta para evitar colisão de IDs duplicados)
    # Body pode trazer campos extras: equipamento, numeroOS
    equip_busca = str(body.get("equipamento", "")).strip().upper()
    os_busca    = str(body.get("numeroOS", body.get("os", ""))).strip()

    def _norm_id(v):
        v = str(v).strip()
        try: v = str(int(float(v)))
        except: pass
        return v

    ocorrencia_id_norm = _norm_id(ocorrencia_id)
    candidatos = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or len(row) < 1:
            continue
        if _norm_id(row[0]) == ocorrencia_id_norm:
            candidatos.append((i, row))

    num_linha = None
    if len(candidatos) == 1:
        # ID único — usa direto
        num_linha = candidatos[0][0]
    elif len(candidatos) > 1:
        # ID duplicado — refina por Equipamento (col D = índice 3) e OS (col K = índice 10)
        for (i, row) in candidatos:
            row_equip = row[3].strip().upper() if len(row) > 3 else ""
            row_os    = row[10].strip() if len(row) > 10 else ""
            equip_match = (not equip_busca) or (equip_busca in row_equip) or (row_equip in equip_busca)
            os_match    = (not os_busca) or (os_busca == row_os)
            if equip_match and os_match:
                num_linha = i
                break
        # Se não achou com os dois critérios, tenta só por equipamento
        if num_linha is None and equip_busca:
            for (i, row) in candidatos:
                row_equip = row[3].strip().upper() if len(row) > 3 else ""
                if equip_busca in row_equip or row_equip in equip_busca:
                    num_linha = i
                    break
        # Último recurso: primeiro candidato
        if num_linha is None:
            num_linha = candidatos[0][0]
            log.warning(f"[atualizar-campo] ID {ocorrencia_id} duplicado, sem match por equip/OS — usando linha {num_linha}")

    if num_linha is None:
        ids_existentes = [_norm_id(r[0]) for r in rows[1:5] if r]
        log.warning(f"[atualizar-campo] ID {ocorrencia_id!r} não encontrado. Primeiros IDs: {ids_existentes}")
        return jsonify({"ok": False, "error": f"Ocorrência {ocorrencia_id} não encontrada"}), 404

    log.info(f"[atualizar-campo] ID={ocorrencia_id} → linha={num_linha} (candidatos={len(candidatos)}, equip={equip_busca!r})")

    try:
        if field == "historico" and append:
            # Acrescenta ao histórico existente sem sobrescrever
            hist_atual = (rows[num_linha - 1][11] if len(rows[num_linha - 1]) > 11 else "").strip()
            hoje = agora_br().strftime("%d/%m")
            nova_entrada = f"{hoje} - {value}"
            # Deduplicação: não adiciona se já existir
            if nova_entrada in hist_atual or value in hist_atual.split("\n")[-1]:
                return jsonify({"ok": True, "dedup": True}), 200
            novo_hist = (hist_atual + "\n" + nova_entrada).strip() if hist_atual else nova_entrada
            ws.update_cell(num_linha, col, novo_hist)
        else:
            ws.update_cell(num_linha, col, value)
        log.info(f"[atualizar-campo] ✅ GRAVADO ID={ocorrencia_id} linha={num_linha} campo={field} valor={value[:40]!r}")
        return jsonify({"ok": True, "linha": num_linha}), 200
    except Exception as e:
        log.error(f"[atualizar-campo] Erro ao gravar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/nova-ocorrencia", methods=["POST", "OPTIONS"])
def nova_ocorrencia_dashboard():
    """
    Registra uma nova ocorrência criada manualmente pelo dashboard.
    Body JSON: { cliente, usina, equipamento, falha, causa, acao, status, historico, editor }
    """
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido"}), 400

    cliente    = body.get("cliente", "").strip()
    usina      = body.get("usina", "").strip()
    equipamento= body.get("equipamento", "").strip()
    falha      = body.get("falha", "").strip()
    causa      = body.get("causa", "").strip()
    acao       = body.get("acao", "").strip()
    status     = body.get("status", "Em Aberto").strip()
    historico  = body.get("historico", "").strip()
    numero_os  = body.get("numeroOS", "").strip()
    editor     = body.get("editor", "dashboard").strip()

    if not equipamento or not falha:
        return jsonify({"ok": False, "error": "equipamento e falha são obrigatórios"}), 400

    try:
        ws   = get_sheet()
        todos = ws.get_all_values()
    except Exception as e:
        log.error(f"[nova-ocorrencia] Erro ao abrir planilha: {e}")
        return jsonify({"ok": False, "error": f"Erro ao acessar planilha: {str(e)}"}), 500

    try:
        dados = {
            "cliente":      cliente,
            "usina":        usina,
            "equipamento":  equipamento,
            "falha":        falha,
            "causa":        causa,
            "equip_impact": equipamento,
            "acao":         acao,
            "status":       status,
            "os":           numero_os,
            "historico":    historico or f"{agora_br().strftime('%d/%m')} - Registro inicial via dashboard.",
        }
        gravar_nova_ocorrencia(ws, todos, dados)
        log.info(f"[nova-ocorrencia] {usina} — {equipamento} | editor={editor}")
        return jsonify({"ok": True}), 200
    except Exception as e:
        log.error(f"[nova-ocorrencia] Erro ao gravar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

def _is_concluido_atividade(status):
    s = (status or "").lower()
    return any(x in s for x in ["concluído", "concluido", "resolvido", "fechado"])


def _proximo_id_atividade(todos):
    ids = []
    for row in todos[1:]:
        if row and str(row[0]).strip().isdigit():
            ids.append(int(str(row[0]).strip()))
    return str(max(ids) + 1) if ids else "1"


ATIV_HEADERS_JSON = ["id", "cliente", "usina", "equipamento", "descricao", "responsavel", "prazo",
                      "prioridade", "status", "dataCriacao", "dataConclusao", "historico", "editor",
                      "numeroOS", "statusOS", "observacoesOS", "linkOS", "statusTarefaOS", "etiquetasOS",
                      "anotacoesPessoais", "percentualOS", "statusGeralOS", "detalhesEquipamentosOS",
                      "ultimaVerificacaoOS", "visualizado"]

ATIV_CAMPO_COL = {
    "cliente": 2, "usina": 3, "equipamento": 4, "descricao": 5, "responsavel": 6,
    "prazo": 7, "prioridade": 8, "status": 9, "dataConclusao": 11, "historico": 12, "numeroOS": 14,
    "statusOS": 15, "observacoesOS": 16, "linkOS": 17, "statusTarefaOS": 18, "etiquetasOS": 19,
    "anotacoesPessoais": 20, "percentualOS": 21, "statusGeralOS": 22, "detalhesEquipamentosOS": 23,
    "ultimaVerificacaoOS": 24, "visualizado": 25,
}

ATIV_TOTAL_COLUNAS = 25

_ativ_headers_ensured = {"done": False}


def _garantir_headers_atividades(ws):
    """
    Garante que a aba Painel de Atividades tenha colunas suficientes na
    grade (a grade do Sheets tem um limite físico de colunas, separado do
    cabeçalho) e que o cabeçalho (linha 1) tenha as colunas novas.

    A expansão de colunas é tentada em TODA chamada (é uma checagem barata
    e idempotente — só chama a API se realmente precisar crescer), pra não
    ficar travada pra sempre caso uma tentativa anterior tenha falhado
    silenciosamente. Só o conteúdo do cabeçalho (linha 1) é cacheado por
    processo, já que isso sim é mais caro de checar toda hora.
    """
    try:
        if ws.col_count < ATIV_TOTAL_COLUNAS:
            ws.resize(cols=ATIV_TOTAL_COLUNAS)
            log.info(f"[Atividades] Grade expandida para {ATIV_TOTAL_COLUNAS} colunas")
    except Exception as e:
        log.error(f"[Atividades] Erro ao expandir colunas da grade: {e}")

    if _ativ_headers_ensured["done"]:
        return
    try:
        header = ws.row_values(1)
        extras = {15: "statusOS", 16: "observacoesOS", 17: "linkOS", 18: "statusTarefaOS",
                  19: "etiquetasOS", 20: "anotacoesPessoais", 21: "percentualOS",
                  22: "statusGeralOS", 23: "detalhesEquipamentosOS", 24: "ultimaVerificacaoOS",
                  25: "visualizado"}
        precisa = False
        visualizado_e_novo = (len(header) < 25 or header[24].strip() != "visualizado")
        for col, nome in extras.items():
            atual = header[col - 1] if len(header) >= col else ""
            if atual.strip() != nome:
                precisa = True
                break
        if precisa:
            novo_header = header + [""] * max(0, ATIV_TOTAL_COLUNAS - len(header))
            novo_header = novo_header[:ATIV_TOTAL_COLUNAS]
            for col, nome in extras.items():
                novo_header[col - 1] = nome
            ws.update(f"A1:{chr(64 + ATIV_TOTAL_COLUNAS)}1", [novo_header])
            log.info("[Atividades] Header estendido com todos os campos Fracttal")
        if visualizado_e_novo:
            # backfill: atividades JÁ existentes não devem aparecer como
            # "não lidas" quando essa funcionalidade é ligada pela primeira
            # vez — só atividades genuinamente novas (criadas depois disso)
            # devem nascer sem o marcador de "visualizado".
            try:
                total_linhas = len(ws.get_all_values())
                if total_linhas > 1:
                    coluna_letra = chr(64 + ATIV_CAMPO_COL["visualizado"])
                    valores = [["sim"]] * (total_linhas - 1)
                    ws.update(f"{coluna_letra}2:{coluna_letra}{total_linhas}", valores)
                    log.info(f"[Atividades] Backfill: {total_linhas - 1} atividades existentes marcadas como já visualizadas")
            except Exception as e:
                log.error(f"[Atividades] Erro no backfill de visualizado: {e}")
        _ativ_headers_ensured["done"] = True
    except Exception as e:
        log.error(f"[Atividades] Erro ao garantir conteúdo do header estendido: {e}")


@app.route("/disparar-comunicado-cluster", methods=["POST"])
def disparar_comunicado_cluster():
    """Dispara manualmente (via botão no dashboard) o comunicado de um
    cluster específico pro grupo de WhatsApp correspondente. O grupo é
    derivado das usinas do cluster (reaproveita o mapeamento grupo_usina
    já existente — normalmente todas as usinas de um mesmo cluster caem
    no mesmo grupo, já que representam a mesma equipe de campo).

    Sem exigência de WEBHOOK_SECRET aqui de propósito: é chamado direto do
    navegador pelo botão no dashboard (não tem como o frontend guardar o
    secret com segurança), então a única proteção é o próprio login no
    painel (role manager)."""
    if not WPP_SERVER_URL:
        return jsonify({"ok": False, "error": "WPP_SERVER_URL não configurado"}), 400

    dados = request.get_json(force=True, silent=True) or {}
    cluster = (dados.get("cluster") or "").strip()
    texto = (dados.get("texto") or "").strip()
    if not cluster or not texto:
        return jsonify({"ok": False, "error": "cluster e texto são obrigatórios"}), 400

    mapa_cluster_usina = _mapa_cluster_usina()
    mapa_grupo_usina = _mapa_grupo_usina()
    usinas_do_cluster = [u for u, c in mapa_cluster_usina.items() if c == cluster]
    grupo_id = next((mapa_grupo_usina[u] for u in usinas_do_cluster if u in mapa_grupo_usina), None)
    if not grupo_id:
        return jsonify({"ok": False, "error": (f"Nenhum grupo de WhatsApp configurado pras usinas do cluster "
                        f"\"{cluster}\". Configure em _Sistema: \"grupo_usina:<Usina>\" = \"<id>@g.us\".")}), 400

    try:
        r = requests.post(
            f"{WPP_SERVER_URL}/api/enviar-mensagem",
            json={"grupoId": grupo_id, "texto": texto},
            headers={"X-Webhook-Secret": WEBHOOK_SECRET} if WEBHOOK_SECRET else {},
            timeout=20,
        )
        if r.ok and r.json().get("ok"):
            return jsonify({"ok": True, "grupo": grupo_id}), 200
        return jsonify({"ok": False, "error": r.text[:300]}), 502
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _montar_texto_comunicado_zeladoria(cluster, usinas, semana):
    """Monta o texto do comunicado quinzenal pedindo as fotos de zeladoria
    (vegetação e sujidade) das usinas de um cluster."""
    lista_usinas = "\n".join(f"• {u}" for u in usinas)
    return (
        f"🌿 *Comunicado de Zeladoria — Semana {semana}*\n\n"
        f"Como de costume, precisamos das fotos de vegetação e sujidade dos módulos das usinas abaixo:\n\n"
        f"{lista_usinas}\n\n"
        f"📸 Padrão das fotos (por usina):\n"
        f"• 5 fotos de vegetação: próxima aos módulos, cabine primária, inversores e sala de O&M\n"
        f"• 3 fotos de sujidade: face do painel mostrando claramente a sujidade, com uma parte limpa ao lado pra comparação (usar pano com água)\n\n"
        f"Por favor, enviem aqui no grupo o quanto antes. Qualquer dúvida, me chamem."
    )


@app.route("/gerar-comunicado-zeladoria", methods=["GET"])
def gerar_comunicado_zeladoria():
    """Monta, por cluster/equipe, a lista de usinas e o texto do comunicado
    quinzenal pedindo as fotos de zeladoria (vegetação e sujidade).
    Reaproveita o mesmo mapeamento cluster/grupo usado no comunicado de
    Atividades (aba _Sistema: "cluster_usina:<Usina>" e "grupo_usina:
    <Usina>"), então os comunicados de Zeladoria saem pros mesmos grupos
    de WhatsApp das equipes."""
    mapa_cluster = _mapa_cluster_usina()  # usina -> cluster
    semana = agora_br().isocalendar()[1]

    por_cluster = {}
    for usina, cluster in mapa_cluster.items():
        por_cluster.setdefault(cluster, []).append(usina)

    resultado = []
    for cluster, usinas in sorted(por_cluster.items(), key=lambda kv: -len(kv[1])):
        usinas_ordenadas = sorted(usinas)
        texto = _montar_texto_comunicado_zeladoria(cluster, usinas_ordenadas, semana)
        resultado.append({"cluster": cluster, "usinas": usinas_ordenadas, "texto": texto})

    return jsonify({"ok": True, "semana": semana, "clusters": resultado}), 200



# ── Comunicados de Sobreaviso ────────────────────────────────────────────
# Fred sobe manualmente o arquivo Escala_Sobreaviso_GridCo_RXX.html (o mesmo
# artifact interativo de 3.Em Operação) toda vez que uma revisão nova sai.
# Guardamos o "estado" embutido no arquivo (blocos + grupos + contatos) em
# disco local na VM1 — sobrevive a deploy porque o deploy só faz
# `git reset --hard` (não `git clean`), o mesmo padrão já usado por
# zeladoria_fotos/ e mensagens_grupos.db.
#
# A fonte da escala real é o array 'grupos' (clusters FUNDIDOS por equipe de
# cobertura, com pool combinado e "dupla" quando 2 plantonistas cobrem o
# bloco junto) — NÃO o array 'campo' (rotação individual por cluster, que
# ignora as fusões e não bate com o que o dashboard de fato exibe; ver
# conversa de 04/09/2026 onde isso foi corrigido).
SOBREAVISO_ESCALA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sobreaviso_escala.json")

_MESES_PT_ABREV = {1: "jan", 2: "fev", 3: "mar", 4: "abr", 5: "mai", 6: "jun",
                    7: "jul", 8: "ago", 9: "set", 10: "out", 11: "nov", 12: "dez"}


def _sobreaviso_extrair_estado_do_html(html_text):
    """Extrai o objeto JSON embutido em <script id="estado" type="application/
    json"> do arquivo de escala. Usa JSONDecoder().raw_decode a partir do '>'
    de abertura da tag em vez de procurar o '</script>' de fechamento por
    regex: o próprio JSON carrega uma cópia auto-referente do HTML inteiro
    (função docFonte() do artifact, usada pro botão "Baixar escala
    atualizada"), com um '<\\/script>' ESCAPADO no meio do valor — um regex
    simples de fechamento acaba pegando a tag errada e quebra o parse."""
    marca = 'id="estado"'
    pos = html_text.find(marca)
    if pos == -1:
        raise ValueError("Tag <script id=\"estado\"> não encontrada — esse arquivo não parece ser a Escala de Sobreaviso.")
    inicio = html_text.find(">", pos) + 1
    decoder = json.JSONDecoder()
    estado, _ = decoder.raw_decode(html_text, inicio)
    return estado


def _sobreaviso_salvar_estado(estado, nome_arquivo=""):
    payload = {"estado": estado, "nome_arquivo": nome_arquivo, "carregado_em": agora_br().isoformat()}
    with open(SOBREAVISO_ESCALA_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)


def _sobreaviso_carregar_estado():
    if not os.path.exists(SOBREAVISO_ESCALA_PATH):
        return None
    try:
        with open(SOBREAVISO_ESCALA_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.error(f"[Sobreaviso] Erro ao ler estado salvo: {e}")
        return None


def _sobreaviso_indice_bloco_sugerido(blocos):
    """Primeiro bloco cujo fim ainda não passou (ou seja, o bloco vigente
    hoje, ou o próximo se hoje estiver fora de qualquer bloco). Se todos já
    passaram, cai no último bloco disponível."""
    hoje = agora_br().date().isoformat()
    for i, b in enumerate(blocos):
        if b["fim"] >= hoje:
            return i
    return max(0, len(blocos) - 1)


def _sobreaviso_fmt_bloco(b):
    ini_y, ini_m, ini_d = (int(x) for x in b["inicio"].split("-"))
    fim_y, fim_m, fim_d = (int(x) for x in b["fim"].split("-"))
    if b["inicio"] == b["fim"]:
        return f"{ini_d}/{_MESES_PT_ABREV[ini_m]}"
    if ini_m == fim_m:
        return f"{ini_d} a {fim_d}/{_MESES_PT_ABREV[fim_m]}"
    return f"{ini_d}/{_MESES_PT_ABREV[ini_m]} a {fim_d}/{_MESES_PT_ABREV[fim_m]}"


def _sobreaviso_montar_texto(grupo, bloco, pessoas, contatos):
    clusters_txt = " + ".join(grupo["clusters"])
    dupla_txt = " (dupla de plantão)" if grupo.get("por_bloco") == 2 else ""
    partes = []
    for p in pessoas:
        tel = contatos.get(p, "")
        partes.append(p + (f" — {tel}" if tel else ""))
    linha_pessoas = " + ".join(partes) if partes else "sem cobertura definida"
    return (
        f"📋 Escala de Sobreaviso — {clusters_txt}\n"
        f"Período: {_sobreaviso_fmt_bloco(bloco)}{dupla_txt}\n\n"
        f"De plantão: {linha_pessoas}\n\n"
        f"Qualquer dúvida, chamar o supervisor."
    )


@app.route("/sobreaviso-upload", methods=["POST"])
def sobreaviso_upload():
    """Recebe (multipart/form-data, campo 'arquivo') o HTML da Escala de
    Sobreaviso vigente e extrai/guarda o estado. Chamado pelo botão de
    upload da aba Sobreavisos, em Comunicados. Sem exigência de
    WEBHOOK_SECRET de propósito — mesmo padrão de outros endpoints de
    escrita chamados direto do dashboard (role manager)."""
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"ok": False, "error": "Envie o arquivo em 'arquivo' (multipart/form-data)."}), 400
    try:
        html_text = arquivo.read().decode("utf-8", errors="replace")
        estado = _sobreaviso_extrair_estado_do_html(html_text)
        if "blocos" not in estado or "grupos" not in estado:
            raise ValueError("O estado extraído não tem 'blocos'/'grupos' — arquivo inesperado.")
        _sobreaviso_salvar_estado(estado, nome_arquivo=arquivo.filename or "")
    except Exception as e:
        log.error(f"[Sobreaviso] Erro ao processar upload: {e}")
        return jsonify({"ok": False, "error": f"Erro ao processar o arquivo: {e}"}), 400

    return jsonify({
        "ok": True,
        "periodo": estado.get("periodo"),
        "total_blocos": len(estado.get("blocos", [])),
        "total_grupos": len(estado.get("grupos", [])),
        "nome_arquivo": arquivo.filename or "",
    }), 200


@app.route("/sobreaviso-blocos", methods=["GET"])
def sobreaviso_blocos():
    """Lista os blocos (períodos de sobreaviso) do arquivo carregado por
    último, com o índice do bloco sugerido (vigente/próximo) já calculado."""
    payload = _sobreaviso_carregar_estado()
    if not payload:
        return jsonify({"ok": False, "error": "Nenhuma escala de sobreaviso carregada ainda. Envie o arquivo primeiro."}), 404

    estado = payload["estado"]
    blocos = estado.get("blocos", [])
    return jsonify({
        "ok": True,
        "nome_arquivo": payload.get("nome_arquivo", ""),
        "carregado_em": payload.get("carregado_em"),
        "periodo": estado.get("periodo"),
        "blocos": [{"idx": i, "inicio": b["inicio"], "fim": b["fim"], "tipo": b["tipo"],
                    "feriado": b.get("feriado", False), "label": _sobreaviso_fmt_bloco(b)}
                   for i, b in enumerate(blocos)],
        "bloco_sugerido": _sobreaviso_indice_bloco_sugerido(blocos),
    }), 200


def _sobreaviso_montar_usinas_por_cluster(usinas):
    """cluster -> lista de usinas {usina, equipe, cliente, uf, supervisor}.
    'equipe' é o código de cluster usado no resto do dashboard (ex.: "SP
    Centro 01"), diferente do nome de cluster da escala (ex.: "Boa Esperança
    / Ibaté I e II") — é o que casa com o filtro lateral "Cluster"."""
    idx = {}
    for u in usinas:
        cl = u.get("cluster")
        if not cl:
            continue
        idx.setdefault(cl, []).append({
            "usina": u.get("usina"),
            "equipe": u.get("equipe"),
            "cliente": u.get("cliente"),
            "uf": u.get("uf"),
            "supervisor": u.get("resp_dash") or u.get("supervisor"),  # resp_dash = titularidade real do cluster; supervisor = quem assina a escala (pode ser um stand-in geografico, ex.: Cedro)
        })
    return idx


@app.route("/gerar-comunicado-sobreaviso", methods=["GET"])
def gerar_comunicado_sobreaviso():
    """Monta o texto do comunicado de sobreaviso pra cada grupo de cobertura
    (array 'grupos' — clusters fundidos que compartilham pool de técnicos)
    de um bloco/período específico. ?bloco=N escolhe o índice; se omitido,
    usa o bloco sugerido (vigente/próximo). Cada grupo devolvido carrega a
    lista de usinas cobertas (com cliente/equipe/supervisor) pra permitir
    filtrar no frontend pelos mesmos filtros laterais (cliente, usina,
    cluster) e por "só minhas usinas"."""
    payload = _sobreaviso_carregar_estado()
    if not payload:
        return jsonify({"ok": False, "error": "Nenhuma escala de sobreaviso carregada ainda. Envie o arquivo primeiro."}), 404

    estado = payload["estado"]
    blocos = estado.get("blocos", [])
    grupos = estado.get("grupos", [])
    contatos = (estado.get("contatos") or {}).get("pessoas", {})
    campo = estado.get("campo", [])
    usinas_todas = estado.get("usinas", [])
    meta_cluster = {c["cluster"]: {"cliente": c.get("cliente"), "uf": c.get("uf"),
                                    "supervisor": c.get("supervisor")} for c in campo}
    usinas_por_cluster = _sobreaviso_montar_usinas_por_cluster(usinas_todas)

    if not blocos:
        return jsonify({"ok": False, "error": "A escala carregada não tem blocos."}), 400

    bloco_idx = request.args.get("bloco", type=int)
    if bloco_idx is None:
        bloco_idx = _sobreaviso_indice_bloco_sugerido(blocos)
    if bloco_idx < 0 or bloco_idx >= len(blocos):
        return jsonify({"ok": False, "error": f"bloco {bloco_idx} fora do intervalo (0 a {len(blocos) - 1})"}), 400

    bloco = blocos[bloco_idx]
    resultado = []
    for g in grupos:
        escala_g = g.get("escala", [])
        pessoas = escala_g[bloco_idx] if bloco_idx < len(escala_g) else []
        texto = _sobreaviso_montar_texto(g, bloco, pessoas, contatos)
        clientes = []
        for cl in g.get("clusters", []):
            info = meta_cluster.get(cl)
            if info and info["cliente"] and info["cliente"] not in clientes:
                clientes.append(info["cliente"])

        usinas_grupo, vistas = [], set()
        for cl in g.get("clusters", []):
            for u in usinas_por_cluster.get(cl, []):
                if u["usina"] in vistas:
                    continue
                vistas.add(u["usina"])
                usinas_grupo.append(u)

        resultado.append({
            "nome": g.get("nome"),
            "clusters": g.get("clusters", []),
            "supervisores": g.get("supervisores", []),
            "clientes": clientes,
            "usinas": usinas_grupo,
            "dupla": g.get("por_bloco") == 2,
            "pessoas": pessoas,
            "texto": texto,
        })

    return jsonify({
        "ok": True,
        "bloco": {"idx": bloco_idx, "inicio": bloco["inicio"], "fim": bloco["fim"],
                   "tipo": bloco["tipo"], "label": _sobreaviso_fmt_bloco(bloco)},
        "grupos": resultado,
    }), 200


def _sobreaviso_montar_indice_grupo_por_cluster(grupos):
    """cluster -> grupo (dict), pra achar rápido qual grupo de cobertura
    fundido inclui cada cluster."""
    idx = {}
    for g in grupos:
        for cl in g.get("clusters", []):
            idx[cl] = g
    return idx


@app.route("/conferencia-sobreaviso", methods=["GET"])
def conferencia_sobreaviso():
    """Confere, usina por usina — a partir do catálogo 'usinas' embutido no
    próprio arquivo de escala (a lista mais completa: as ~100 usinas de
    todos os supervisores, sempre em sincronia com o que foi carregado) —
    se há alguém de sobreaviso cobrindo cada uma no bloco selecionado.
    Devolve 3 categorias de problema, pra garantir que nenhuma usina fique
    desassistida:
      - sem_cobertura: o cluster da usina não tem gente suficiente escalada
        nesse bloco (pool vazio ou menos gente do que o 'por_bloco' exige) —
        CRÍTICO, ninguém responde por essa usina nesse período.
      - fora_sla: o cluster TEM gente escalada, mas nenhum candidato do pool
        alcança o SLA de deslocamento contratual (fora_sla=true, herdado do
        próprio cálculo de cobertura do arquivo) — tem plantonista, mas fora
        do prazo combinado com o cliente.
      - sem_cluster: a usina não aparece em NENHUM grupo de cobertura — gap
        estrutural no cadastro da escala, independe do bloco escolhido.

    ?bloco=N escolhe o período (padrão: o sugerido/vigente).
    ?supervisor=Nome filtra só as usinas daquele supervisor (padrão: todos)."""
    payload = _sobreaviso_carregar_estado()
    if not payload:
        return jsonify({"ok": False, "error": "Nenhuma escala de sobreaviso carregada ainda. Envie o arquivo primeiro."}), 404

    estado = payload["estado"]
    blocos = estado.get("blocos", [])
    grupos = estado.get("grupos", [])
    usinas = estado.get("usinas", [])
    if not blocos or not usinas:
        return jsonify({"ok": False, "error": "A escala carregada não tem 'blocos' ou 'usinas'."}), 400

    bloco_idx = request.args.get("bloco", type=int)
    if bloco_idx is None:
        bloco_idx = _sobreaviso_indice_bloco_sugerido(blocos)
    if bloco_idx < 0 or bloco_idx >= len(blocos):
        return jsonify({"ok": False, "error": f"bloco {bloco_idx} fora do intervalo (0 a {len(blocos) - 1})"}), 400
    bloco = blocos[bloco_idx]

    supervisor_filtro = (request.args.get("supervisor") or "").strip()
    idx_grupo_por_cluster = _sobreaviso_montar_indice_grupo_por_cluster(grupos)

    sem_cobertura, fora_sla_lista, sem_cluster, ok_lista = [], [], [], []
    total_consideradas = 0

    for u in usinas:
        nome_usina = u.get("usina")
        cluster = u.get("cluster")
        supervisor = u.get("resp_dash") or u.get("supervisor")  # resp_dash = titularidade real do cluster; supervisor = quem assina a escala (pode ser um stand-in geografico, ex.: Cedro)
        cliente = u.get("cliente")
        equipe = u.get("equipe")
        if supervisor_filtro and supervisor != supervisor_filtro:
            continue
        total_consideradas += 1

        grupo = idx_grupo_por_cluster.get(cluster) if cluster else None
        if not grupo:
            sem_cluster.append({"usina": nome_usina, "cluster": cluster, "cliente": cliente,
                                 "supervisor": supervisor, "equipe": equipe})
            continue

        escala_g = grupo.get("escala", [])
        pessoas = escala_g[bloco_idx] if bloco_idx < len(escala_g) else []
        por_bloco = grupo.get("por_bloco", 1)
        if not pessoas or len(pessoas) < por_bloco:
            sem_cobertura.append({
                "usina": nome_usina, "cluster": cluster, "cliente": cliente, "supervisor": supervisor,
                "equipe": equipe, "pessoas": pessoas, "esperado": por_bloco,
            })
            continue

        if grupo.get("fora_sla"):
            fora_sla_lista.append({
                "usina": nome_usina, "cluster": cluster, "cliente": cliente, "supervisor": supervisor,
                "equipe": equipe, "pessoas": pessoas, "grupo_nome": grupo.get("nome"),
            })
        else:
            ok_lista.append({"usina": nome_usina, "cluster": cluster, "cliente": cliente,
                              "supervisor": supervisor, "equipe": equipe})

    return jsonify({
        "ok": True,
        "bloco": {"idx": bloco_idx, "inicio": bloco["inicio"], "fim": bloco["fim"],
                   "tipo": bloco["tipo"], "label": _sobreaviso_fmt_bloco(bloco)},
        "total_usinas": total_consideradas,
        "total_ok": total_consideradas - len(sem_cobertura) - len(sem_cluster),
        "sem_cobertura": sem_cobertura,
        "sem_cluster": sem_cluster,
        "fora_sla": fora_sla_lista,
        "ok_usinas": ok_lista,
    }), 200


_ZEL_GRUPOS = ["Roçada", "Poda Química", "Lavagem dos Módulos", "Controle de Pragas"]
_ZEL_SUBCOLS = ["Última Data", "Próxima Data", "Fornecedor", "Status"]  # 4 subcolunas por grupo


@app.route("/zeladoria-reestruturar-fornecedor", methods=["POST"])
def zeladoria_reestruturar_fornecedor():
    """Uso único: renomeia a 3ª subcoluna de cada grupo (Roçada, Poda
    Química, Lavagem dos Módulos, Controle de Pragas) de 'Quantidade' pra
    'Fornecedor' na linha 2 de cabeçalho da aba Zeladoria — mais útil pra
    rastrear qual empresa terceirizada está em cada frente de serviço."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    ws = get_zeladoria_sheet()
    row2 = ws.row_values(2)
    alterados = []
    for i, valor in enumerate(row2):
        if valor.strip().lower() == "quantidade":
            col = i + 1  # gspread é 1-indexed
            ws.update_cell(2, col, "Fornecedor")
            alterados.append(col)
    return jsonify({"ok": True, "colunas_renomeadas": alterados}), 200


_ZEL_GRUPO_ALIASES = {
    "rocada": "Roçada", "rocagem": "Roçada", "supressao": "Roçada",
    "supressao vegetal": "Roçada", "capina": "Roçada", "vegetal": "Roçada",
    "poda quimica": "Poda Química", "poda": "Poda Química", "herbicida": "Poda Química",
    "lavagem dos modulos": "Lavagem dos Módulos", "lavagem": "Lavagem dos Módulos",
    "limpeza": "Lavagem dos Módulos", "limpeza dos modulos": "Lavagem dos Módulos",
    "limpeza de modulos": "Lavagem dos Módulos", "modulos": "Lavagem dos Módulos",
    "controle de pragas": "Controle de Pragas", "pragas": "Controle de Pragas",
}


def _zel_resolver_grupo(grupo, indice_cols):
    """Acha o grupo real da planilha a partir do que a IA (ou o chamador)
    mandou, tolerando variações de escrita (ex: 'Limpeza' -> 'Lavagem dos
    Módulos'). Retorna None se não conseguir resolver."""
    grupo_norm = _normalizar_tema_comunicado(grupo)
    if grupo in indice_cols:
        return grupo
    alvo = _ZEL_GRUPO_ALIASES.get(grupo_norm)
    if alvo and alvo in indice_cols:
        return alvo
    for chave_real in indice_cols:
        if _normalizar_tema_comunicado(chave_real) == grupo_norm:
            return chave_real
    return None


def _zel_montar_indice_colunas(ws):
    """Lê as linhas 1 e 2 de cabeçalho da aba Zeladoria e monta um índice
    {grupo: {subcol_normalizada: coluna_1indexed}} pra permitir localizar
    a célula certa de cada usina/grupo/subcoluna sem hardcode de posição
    (a planilha pode ganhar colunas novas no futuro)."""
    row1 = ws.row_values(1)
    row2 = ws.row_values(2)
    largura = max(len(row1), len(row2))
    grupo_atual = None
    indice = {}
    for i in range(largura):
        v1 = row1[i].strip() if i < len(row1) else ""
        v2 = row2[i].strip() if i < len(row2) else ""
        if v1:
            grupo_atual = v1
        if grupo_atual and v2:
            indice.setdefault(grupo_atual, {})[_normalizar_tema_comunicado(v2)] = i + 1
    return indice


@app.route("/zeladoria-atualizar-lote", methods=["POST"])
def zeladoria_atualizar_lote():
    """Recebe uma lista de atualizações {usina, grupo, proximaData,
    fornecedor, status} e grava direto nas células correspondentes da
    aba Zeladoria, localizando a linha pela usina (coluna B) e as
    colunas pelo cabeçalho (grupo + subcoluna). Usina não encontrada na
    planilha é reportada em 'nao_encontradas', sem interromper o
    restante do lote. Chamado tanto pelo botão 'Enviar Print/Observação'
    do frontend (sem auth — mesmo padrão de outros endpoints de escrita
    chamados direto do dashboard) quanto por scripts administrativos."""
    body = request.get_json(force=True, silent=True) or {}
    itens = body.get("itens") or []
    if not itens:
        return jsonify({"ok": False, "error": "informe 'itens'"}), 400

    ws = get_zeladoria_sheet()
    todos = ws.get_all_values()
    indice_cols = _zel_montar_indice_colunas(ws)

    mapa_linha_usina = {}
    for i, row in enumerate(todos[2:], start=3):  # dados começam na linha 3
        if len(row) > 1 and row[1].strip():
            mapa_linha_usina[_normalizar_tema_comunicado(row[1])] = i

    aplicadas, nao_encontradas = [], []
    updates = []  # lista de (linha, coluna, valor) pra um único batch_update
    for item in itens:
        usina = (item.get("usina") or "").strip()
        grupo_bruto = (item.get("grupo") or "").strip()
        linha = mapa_linha_usina.get(_normalizar_tema_comunicado(usina))
        grupo = _zel_resolver_grupo(grupo_bruto, indice_cols)
        cols_grupo = indice_cols.get(grupo) if grupo else None
        if not linha or not cols_grupo:
            nao_encontradas.append({"usina": usina, "grupo": grupo_bruto})
            continue
        for campo, valor in [
            ("ultima data", item.get("ultimaData")),
            ("proxima data", item.get("proximaData")),
            ("fornecedor", item.get("fornecedor")),
            ("status", item.get("status")),
        ]:
            col = cols_grupo.get(campo)
            if col and valor is not None:
                updates.append({
                    "range": gspread.utils.rowcol_to_a1(linha, col),
                    "values": [[valor]],
                })
        aplicadas.append({"usina": usina, "grupo": grupo, "linha": linha})

    if updates:
        ws.batch_update(updates)

    return jsonify({
        "ok": True,
        "aplicadas": aplicadas,
        "nao_encontradas": nao_encontradas,
    }), 200


def _montar_prompt_extrair_zeladoria(texto_observacoes, usinas_validas):
    lista_usinas = "\n".join(f"- {u}" for u in usinas_validas)
    hoje = agora_br()
    data_hoje_fmt = hoje.strftime("%d/%m/%Y")
    ano_atual = hoje.year
    return f"""Você é um assistente que ajuda a extrair informações de controle de zeladoria (roçada/supressão vegetal, poda química, lavagem/limpeza de módulos fotovoltaicos, controle de pragas) de usinas solares, a partir de prints de conversas, cronogramas de fornecedores terceirizados ou anotações de reunião.

CONTEXTO DE DATA — a data de hoje é {data_hoje_fmt}. Datas mencionadas no texto/imagem SEM ano explícito (ex: "28/07", "dia 30/08") são sempre do ano corrente ({ano_atual}) ou, se o dia/mês já passou este ano, do próximo ano civil que fizer sentido — NUNCA de anos passados como 2024 ou 2025. Se o texto/imagem já trouxer o ano explicitamente, use o que foi informado.

USINAS VÁLIDAS (use exatamente esse nome no campo "usina" quando reconhecer a usina — se a informação for de uma usina que NÃO está nessa lista, ainda inclua no resultado, copie o nome como veio no texto/imagem, e marque "usina_reconhecida": false):
{lista_usinas}

GRUPOS VÁLIDOS (campo "grupo", escolha o que melhor descreve o serviço):
- "Roçada" (roçagem, supressão vegetal, capina)
- "Poda Química" (herbicida, poda química)
- "Lavagem dos Módulos" (limpeza/lavagem de módulos fotovoltaicos)
- "Controle de Pragas"

STATUS VÁLIDOS (campo "status", escolha o mais adequado ao que foi informado):
- "Programado" (data e fornecedor confirmados, aguardando execução)
- "Aguardando assinatura" (contrato ainda não assinado, previsão sujeita a mudança)
- "Em cotação" (negociação em andamento com fornecedor já identificado)
- "Buscando cotação" (ainda procurando fornecedor)
- "Em andamento" (serviço sendo executado agora)
- "Concluído" (serviço já finalizado)
- "Sem informações" (citado mas sem detalhe suficiente pra decidir)

Extraia do texto/imagem abaixo todas as atualizações de zeladoria que conseguir identificar. Pra cada usina+grupo mencionado, gere um item com:
- "usina": nome exato (da lista acima, se reconhecida) ou como veio no original
- "usina_reconhecida": true/false
- "grupo": um dos grupos válidos acima
- "ultimaData": data em que o serviço JÁ FOI EXECUTADO/CONCLUÍDO (data passada), formato DD/MM/AAAA, ou "" se não houver
- "proximaData": data AGENDADA/PROGRAMADA pra acontecer (data futura, ainda não executada), formato DD/MM/AAAA, ou "" se não houver
- "fornecedor": nome da empresa/fornecedor responsável, ou "" se não informado
- "status": um dos status válidos acima
- "observacao": nota curta livre só se houver algo relevante que não caiba nos campos acima (ex.: "previsão sujeita a confirmação pós-assinatura"), ou "" caso contrário

REGRA CRÍTICA sobre qual campo de data usar — isso é o erro mais comum, preste atenção:
- Se o texto diz que o serviço JÁ ACONTECEU ("foi concluída em X", "concluído dia X", "realizado em X", "feito em X", "executado em X"), essa data vai em "ultimaData" e o status correto é "Concluído" (ou "Em andamento" se foi parcial). NUNCA coloque essa data em "proximaData" — "próxima data" significa data futura ainda não executada, e colocar uma data de conclusão ali contradiz o status "Concluído".
- Se o texto diz que o serviço está AGENDADO/PREVISTO pra acontecer ("programado para X", "vai ser feito dia X", "agendado para X", "previsão de X"), essa data vai em "proximaData", com status "Programado" (ou "Aguardando assinatura"/"Em cotação"/"Buscando cotação" conforme o caso).
- Um item pode ter as duas datas preenchidas ao mesmo tempo (ex.: "concluímos a roçada dia 27/07, a próxima já tá programada pra 15/09" → ultimaData=27/07, proximaData=15/09), mas nunca a MESMA data nos dois campos.

REGRA CRÍTICA — não invente informação que não está no texto/imagem. Se uma usina for citada mas sem detalhes suficientes pra decidir o status, use "Sem informações" e deixe ultimaData/proximaData/fornecedor vazios em vez de supor. Isso vale especialmente pro ANO da data — siga a regra de contexto de data acima, nunca invente um ano aleatório.

Texto/observações fornecidas pelo usuário: {texto_observacoes or "(nenhuma observação em texto — considere só a imagem)"}

FORMATO DE SAÍDA (OBRIGATÓRIO): responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes ou depois), no formato:
{{"itens": [{{"usina": "...", "usina_reconhecida": true, "grupo": "...", "ultimaData": "...", "proximaData": "...", "fornecedor": "...", "status": "...", "observacao": "..."}}]}}"""


@app.route("/zeladoria-extrair-print", methods=["POST", "OPTIONS"])
def zeladoria_extrair_print():
    """Recebe um print (imagem em base64) e/ou texto de observações
    (repassados de reuniões, fornecedores etc.) e usa IA (Gemini, visão +
    texto) pra extrair uma lista estruturada de atualizações de zeladoria
    (usina, grupo, data, fornecedor, status), pronta pra revisão no
    frontend antes de aplicar via /zeladoria-atualizar-lote. Este
    endpoint NUNCA grava nada sozinho — só extrai e devolve pra revisão."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    imagem_b64 = body.get("imagemBase64") or ""
    imagem_mime = body.get("imagemMimeType") or "image/png"
    texto = (body.get("texto") or "").strip()
    if not imagem_b64 and not texto:
        return jsonify({"ok": False, "error": "envie uma imagem e/ou texto de observações"}), 400

    ws = get_zeladoria_sheet()
    todos = ws.get_all_values()
    usinas_validas = sorted({row[1].strip() for row in todos[2:] if len(row) > 1 and row[1].strip()})

    prompt = _montar_prompt_extrair_zeladoria(texto, usinas_validas)
    parts = [{"text": prompt}]
    if imagem_b64:
        parts.append({"inline_data": {"mime_type": imagem_mime, "data": imagem_b64}})

    diagnostico = request.args.get("diagnostico", "").lower() == "true"
    try:
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": parts}],
                "generationConfig": {
                    "temperature": 0.2,
                    "maxOutputTokens": 8192,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=45,
            usar_chave_teste=diagnostico,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        texto_bruto = candidato["content"]["parts"][0]["text"].strip()
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
        finish_reason = candidato.get("finishReason", "")
        try:
            parsed = json.loads(texto_limpo)
        except json.JSONDecodeError:
            if finish_reason == "MAX_TOKENS":
                # resposta cortada pelo limite de tokens antes de fechar o JSON —
                # tenta recuperar os itens já completos em vez de descartar tudo
                itens_parciais = re.findall(r"\{[^{}]*\"usina\"[^{}]*\}", texto_limpo)
                itens_ok = []
                for bruto in itens_parciais:
                    try:
                        itens_ok.append(json.loads(bruto))
                    except json.JSONDecodeError:
                        continue
                if itens_ok:
                    log.warning(f"[zeladoria-extrair-print] resposta truncada (MAX_TOKENS), recuperados {len(itens_ok)} item(ns) parcial(is)")
                    return jsonify({"ok": True, "itens": itens_ok, "usinasValidas": usinas_validas, "truncado": True})
            raise
        itens = parsed.get("itens") or []
        return jsonify({"ok": True, "itens": itens, "usinasValidas": usinas_validas})
    except Exception as e:
        log.error(f"[zeladoria-extrair-print] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 502


@app.route("/resolver-duplicata-8866", methods=["POST"])
def resolver_duplicata_8866():
    """Uso único: resolve a duplicidade da OS 8866 (atividades #24 e #35
    apontando pra mesma OS). Mantém #24 (vinculada pelo fluxo oficial
    Solicitar OS), cancela #35 (criada manualmente à parte), preservando
    o histórico das duas com uma nota cruzada explicando o motivo."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    agora = agora_br().strftime('%d/%m/%Y %H:%M')
    resultado = {}
    for row_idx, row in enumerate(todos[1:], start=2):
        if not row or not row[0].strip():
            continue
        id_ativ = row[0].strip()
        if id_ativ == "24":
            nota = f"{agora} - Atividade #35 (duplicata manual da mesma OS) foi cancelada e mesclada aqui."
            hist_atual = row[11] if len(row) > 11 else ""
            ws.update_cell(row_idx, 12, f"{hist_atual}\n{nota}".strip())
            resultado["24"] = "nota adicionada"
        elif id_ativ == "35":
            nota = f"{agora} - Cancelada: duplicata da atividade #24 pra mesma OS (8866). Mantida #24, vinculada pelo fluxo oficial Solicitar OS."
            hist_atual = row[11] if len(row) > 11 else ""
            ws.update_cell(row_idx, 12, f"{hist_atual}\n{nota}".strip())
            ws.update_cell(row_idx, 9, "Cancelado")
            resultado["35"] = "cancelada"

    return jsonify({"ok": True, "resultado": resultado}), 200


def _fracttal_verificar_e_atualizar_uma_os(ws, i, row, numero_os, enviar_notificacao=True):
    """Consulta a Fracttal AO VIVO pra uma única OS (linha i da planilha) e
    atualiza todos os campos derivados (statusOS, percentualOS,
    statusGeralOS, statusTarefaOS, detalhesEquipamentosOS) + aplica a
    correção de status interno via _status_interno_esperado(). Usada tanto
    pelo rodízio automático quanto pela auditoria completa — única função
    que efetivamente fala com a Fracttal pra revalidar uma OS, pra nunca
    ter dois lugares checando/decidindo isso de formas diferentes.

    enviar_notificacao=False quando quem chama está processando um lote
    (rodízio de várias OSs de uma vez) — nesse caso, quem chama deve
    mandar um único push resumido no final, em vez de um por item (evita
    disparar muitas notificações em sequência rápida, o que já fez o
    Chrome marcar o site como "possível spam" — relatado 14/07/2026).

    Retorna um dict com o resumo do que mudou (ou None em caso de erro,
    já logado)."""
    status_interno_atual = row[8].strip()
    status_os_atual = row[14].strip()
    percentual_atual = row[20].strip()
    status_geral_atual = row[21].strip()
    responsavel_atual = row[ATIV_CAMPO_COL["responsavel"] - 1].strip() if len(row) >= ATIV_CAMPO_COL["responsavel"] else ""
    cliente_atual = row[ATIV_CAMPO_COL["cliente"] - 1].strip() if len(row) >= ATIV_CAMPO_COL["cliente"] else ""
    usina_atual = row[ATIV_CAMPO_COL["usina"] - 1].strip() if len(row) >= ATIV_CAMPO_COL["usina"] else ""
    agora_iso = agora_br().strftime("%Y-%m-%dT%H:%M:%S")
    try:
        token = _fracttal_get_token()
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders/{numero_os}", headers=headers, timeout=15)
        resp.raise_for_status()
        tasks = (resp.json().get("data") or [])
    except Exception as e:
        log.error(f"[Fracttal] Erro ao checar/atualizar OS {numero_os}: {e}")
        # marca como verificada MESMO em erro — senão essa OS quebrada
        # (ex.: número inválido, removida da Fracttal, timeout) fica
        # sempre "a mais antiga" e monopoliza a fila de prioridade pra
        # sempre, nunca deixando outras OSs saudáveis serem revisitadas.
        try:
            ws.update_cell(i, ATIV_CAMPO_COL["ultimaVerificacaoOS"], agora_iso)
        except Exception:
            pass
        return None

    try:
        ws.update_cell(i, ATIV_CAMPO_COL["ultimaVerificacaoOS"], agora_iso)
        if not tasks:
            return {"numeroOS": numero_os, "mudou": False, "motivo": "OS sem tarefas na Fracttal"}

        status_novo_raw = str(tasks[0].get("id_status_work_order", "")).strip()
        status_novo = _FRACTTAL_STATUS_OS_MAP.get(status_novo_raw, "")
        percentual_novo = str(_fracttal_percentual_conclusao(tasks))
        status_geral_novo = _fracttal_status_geral(tasks)
        status_tarefa_novo = _fracttal_status_tarefa_agregado(tasks)
        detalhes_novo = _fracttal_detalhes_equipamentos(tasks)
        # Reatribuição de técnico na Fracttal depois da OS já criada aqui
        # (ex.: OS 12666 nasceu com Adriano Silva e foi reatribuída na
        # Fracttal pra Cláudio Ferreira) — sem isso o campo "responsavel"
        # fica gravado só no momento da criação e nunca mais é revisto,
        # mesmo passando por centenas de revalidações (bug identificado
        # 03/09/2026). Fracttal é fonte única de verdade pro técnico
        # também, igual já é pro status.
        tecnico_novo = (tasks[0].get("personnel_description") or tasks[0].get("responsible")
                        or tasks[0].get("created_by") or "").strip()

        # Revalidação de usina/cliente (adicionado 03/09/2026, mesma causa
        # raiz do fix de responsável): usina/cliente também só eram
        # gravados na CRIAÇÃO e nunca revistos — uma classificação errada
        # feita antes de algum fix de catálogo/cruzamento entrar no ar
        # (ex.: OS 12923/12925, criadas antes do fix cliente x usina)
        # ficava presa errada pra sempre, mesmo com o bug já corrigido no
        # código. Só sobrescreve quando o resolvedor CONSEGUE decidir uma
        # usina (nunca troca uma classificação existente por "revisão
        # manual" só porque a Fracttal mudou uma palavra num campo texto).
        _res_usina = _fracttal_resolver_usina_tecnico(tasks[0])
        usina_novo = _res_usina["usina"] or ""
        cliente_novo = _res_usina["cliente"] or ""

        mudou = False
        if status_novo and status_novo != status_os_atual:
            ws.update_cell(i, ATIV_CAMPO_COL["statusOS"], status_novo)
            mudou = True
        if tecnico_novo and tecnico_novo != responsavel_atual:
            ws.update_cell(i, ATIV_CAMPO_COL["responsavel"], tecnico_novo)
            mudou = True
        if usina_novo and (usina_novo != usina_atual or cliente_novo != cliente_atual):
            ws.update_cell(i, ATIV_CAMPO_COL["usina"], usina_novo)
            ws.update_cell(i, ATIV_CAMPO_COL["cliente"], cliente_novo)
            mudou = True
        if (percentual_novo != percentual_atual) or (status_geral_novo != status_geral_atual):
            ws.update_cell(i, ATIV_CAMPO_COL["statusTarefaOS"], status_tarefa_novo)
            ws.update(f"U{i}:W{i}", [[percentual_novo, status_geral_novo, detalhes_novo]])
            mudou = True

        hist_atual = row[ATIV_COL_HISTORICO - 1] if len(row) >= ATIV_COL_HISTORICO else ""
        mudanca_resumo = ""  # versão curta do que mudou, pra caber em notificação (push/central) — 29/07/2026
        if mudou:
            # Mensagem reescrita (17/07/2026): a versão anterior sempre
            # mostrava "status X → X, 0% → 0%" mesmo quando SÓ a situação
            # geral da tarefa tinha mudado — confuso, parecia que nada
            # tinha acontecido de verdade. Agora só entra na frase o que
            # de fato mudou, cada coisa em sua própria oração.
            partes = []
            partes_curtas = []
            if status_novo and status_novo != status_os_atual:
                partes.append(f"status na Fracttal mudou de \"{status_os_atual or '—'}\" para \"{status_novo}\"")
                partes_curtas.append(f"{status_os_atual or '—'} → {status_novo}")
            if tecnico_novo and tecnico_novo != responsavel_atual:
                partes.append(f"responsável mudou de \"{responsavel_atual or '—'}\" para \"{tecnico_novo}\" (reatribuição na Fracttal)")
                partes_curtas.append(f"Responsável: {responsavel_atual or '—'} → {tecnico_novo}")
            if usina_novo and (usina_novo != usina_atual or cliente_novo != cliente_atual):
                partes.append(f"usina/cliente corrigidos de \"{usina_atual or '—'}\"/\"{cliente_atual or '—'}\" "
                               f"para \"{usina_novo}\"/\"{cliente_novo}\" (reclassificação automática)")
                partes_curtas.append(f"Usina: {usina_atual or '—'} → {usina_novo}")
            if percentual_novo != percentual_atual:
                partes.append(f"progresso da tarefa foi de {percentual_atual or '0'}% para {percentual_novo}%")
                partes_curtas.append(f"{percentual_atual or '0'}% → {percentual_novo}%")
            if status_geral_novo != status_geral_atual:
                partes.append(f"situação geral da tarefa mudou de \"{status_geral_atual or '—'}\" para \"{status_geral_novo}\"")
                partes_curtas.append(f"{status_geral_novo}")
            if partes:
                entry = f"{agora_br().strftime('%d/%m/%Y %H:%M')} - " + "; ".join(partes) + "."
                ws.update_cell(i, ATIV_COL_HISTORICO, f"{hist_atual}\n{entry}".strip() if hist_atual else entry)
                hist_atual = f"{hist_atual}\n{entry}".strip() if hist_atual else entry
                mudanca_resumo = "; ".join(partes_curtas)

        # correção de status interno — roda SEMPRE, independente de "mudou"
        # (bug estrutural identificado e corrigido em 12/07/2026: se só
        # rodasse quando outro campo mudasse, um status já errado nunca
        # seria corrigido enquanto a Fracttal não mudasse de novo).
        status_efetivo = status_novo or status_os_atual
        novo_status_interno = _status_interno_esperado(status_efetivo, status_interno_atual)
        if novo_status_interno:
            _gravar_status_interno(ws, i, novo_status_interno)
            if novo_status_interno == "Em Aberto" and status_interno_atual in ("Concluído", "Cancelado"):
                correcao = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - ⚠️ OS reaberta automaticamente: "
                            f"estava marcada como \"{status_interno_atual}\", mas a Fracttal mostra estado "
                            f"\"{status_efetivo or '—'}\" (voltou pra Em Processo/Em Revisão — provavelmente "
                            f"reprovada ou reaberta).")
            else:
                correcao = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - ✅ Status interno corrigido pra "
                            f"\"{novo_status_interno}\" (estado na Fracttal: \"{status_efetivo or '—'}\").")
            ws.update_cell(i, ATIV_COL_HISTORICO, f"{hist_atual}\n{correcao}".strip() if hist_atual else correcao)

        if mudou and enviar_notificacao:
            try:
                usina_row = row[ATIV_CAMPO_COL["usina"] - 1] if len(row) >= ATIV_CAMPO_COL["usina"] else ""
                descricao_row = row[ATIV_CAMPO_COL["descricao"] - 1] if len(row) >= ATIV_CAMPO_COL["descricao"] else ""
                equipamento_row = row[ATIV_CAMPO_COL["equipamento"] - 1] if len(row) >= ATIV_CAMPO_COL["equipamento"] else ""
                id_atividade = row[0] if row else ""
                # Prioriza a Ação/Tarefa (o "tema" real da OS) em vez do
                # código do Ativo — mesmo ativo pode ter tarefas bem
                # diferentes entre OSs, então mostrar só o equipamento não
                # diz o que de fato precisa ser feito (corrigido 29/07/2026).
                tema_row = descricao_row or equipamento_row or "Descrição não informada"
                # Corpo agora mostra o que a OS É (tema/tarefa) e o que
                # de fato MUDOU (ex. "Em Processo → Em Revisão"), não só
                # o estado atual — sem isso não dava pra saber se a OS
                # tinha sido concluída, reaberta, etc (corrigido 29/07/2026).
                enviar_push(
                    titulo=f"🔄 OS {numero_os} — {usina_row or 'Usina não informada'}",
                    corpo=f"{tema_row}\n{mudanca_resumo or f'{status_geral_novo} — {percentual_novo}% concluído'}",
                    tipo="fracttal_status",
                    url=f"https://fred-alexandrino.github.io/PAINELDEFALHAS/?atividade={id_atividade}",
                )
            except Exception as e:
                log.error(f"[sync-fracttal] Falha ao enviar push de mudança de status {numero_os}: {e}")

        return {"numeroOS": numero_os, "id": row[0] if row else "", "mudou": mudou,
                "statusOS": status_novo or status_os_atual,
                "percentualOS": percentual_novo, "statusGeralOS": status_geral_novo,
                "statusInternoCorrigido": novo_status_interno,
                "responsavel": tecnico_novo or responsavel_atual,
                "usina": usina_novo or usina_atual, "cliente": cliente_novo or cliente_atual,
                "usina": row[ATIV_CAMPO_COL["usina"] - 1] if len(row) >= ATIV_CAMPO_COL["usina"] else "",
                "equipamento": row[ATIV_CAMPO_COL["equipamento"] - 1] if len(row) >= ATIV_CAMPO_COL["equipamento"] else "",
                "descricao": row[ATIV_CAMPO_COL["descricao"] - 1] if len(row) >= ATIV_CAMPO_COL["descricao"] else "",
                "mudancaResumo": mudanca_resumo}
    except Exception as e:
        log.error(f"[Fracttal] Erro ao checar/atualizar OS {numero_os}: {e}")
        return None


def _auditoria_consistencia_os_core(aplicar=True, limite_atraso_minutos=0, limite_recheck_ao_vivo=35, origem="automática"):
    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    divergencias = []
    desatualizadas = []
    agora = agora_br()
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        numero_os = row[13].strip()
        if not numero_os:
            continue  # só audita quem está vinculado a uma OS da Fracttal
        # Corrigido 20/08/2026, relatado pelo Fred: OS de uma usina que
        # já foi devolvida (removida da Supervisão Temporária) continuava
        # sendo revalidada ao vivo na Fracttal e gerando push de "OS
        # atualizada" indefinidamente — porque essa auditoria varre TODA
        # linha com numeroOS, sem checar se a usina ainda é reconhecida
        # agora (a linha antiga fica pra sempre na planilha, só o
        # cadastro da usina que sai do catálogo temporário). Mesmo
        # princípio da correção de 31/07/2026 no /atividades e de
        # 18/08/2026 no resumo/ronda — aplicado aqui pra fechar o último
        # caminho que ainda vazava usina devolvida.
        if canonizar_usina(row[2].strip()) is None:
            continue
        status_interno_atual = row[8].strip()
        status_os_atual = row[14].strip()

        # ── Parte 1: a OS está sendo verificada com a frequência que
        # deveria? Isso pega o caso mais grave — uma OS que por algum bug
        # ficou fora do rodízio e nunca mais é revisitada, então nem tem
        # como a consistência interna (parte 2) detectar problema nela,
        # porque o statusOS gravado pode estar simplesmente desatualizado
        # há muito tempo, sem ninguém perceber.
        if status_os_atual not in ("Finalizada", "Cancelada"):
            ultima_verificacao = row[23].strip()
            se_atrasada = True
            if ultima_verificacao:
                try:
                    try:
                        dt_verif = datetime.strptime(ultima_verificacao, "%Y-%m-%dT%H:%M:%S")
                    except ValueError:
                        # o Google Sheets reformata a data ao salvar/ler,
                        # trocando o "T" por espaço — aceita os dois formatos.
                        dt_verif = datetime.strptime(ultima_verificacao, "%Y-%m-%d %H:%M:%S")
                    if agora.tzinfo:
                        dt_verif = dt_verif.replace(tzinfo=agora.tzinfo)
                    minutos_desde = (agora - dt_verif).total_seconds() / 60
                    se_atrasada = minutos_desde > limite_atraso_minutos
                except Exception:
                    se_atrasada = True
            if se_atrasada:
                # BUG CRÍTICO identificado em 16/07/2026: o campo usado pra
                # ordenar (abaixo) usava "ultima_verificacao or 'nunca'" —
                # ou seja, OSs NUNCA verificadas recebiam o texto "nunca"
                # em vez de string vazia. Só que a string "nunca" começa
                # com 'n', que em ordenação alfabética vem DEPOIS de
                # qualquer timestamp (que começa com dígito) — o oposto do
                # pretendido pelo comentário original ("nunca vazio
                # primeiro"). Resultado: toda OS nunca verificada era
                # empurrada pro FIM da fila, e como só as N primeiras (35)
                # são de fato rechecadas ao vivo por rodada, uma OS nova
                # (ex.: 9513) ficava starved indefinidamente sempre que
                # havia 35+ outras OSs com QUALQUER timestamp anterior,
                # por mais antigo que fosse — nunca chegava a vez dela.
                # Corrigido usando uma chave de ordenação separada com o
                # valor cru (string vazia ordena primeiro de verdade),
                # mantendo "nunca" só como texto de exibição.
                desatualizadas.append({"id": row[0], "numeroOS": numero_os,
                                        "ultimaVerificacao": ultima_verificacao or "nunca",
                                        "_sortKey": ultima_verificacao, "linha": i, "row": row})

        # ── Parte 2: o status interno bate com o estado já gravado?
        if not status_os_atual:
            continue  # ainda sem estado conhecido — nada a auditar aqui

        esperado = _status_interno_esperado(status_os_atual, status_interno_atual)
        if esperado:
            divergencias.append({"linha": i, "id": row[0], "numeroOS": numero_os,
                                  "de": status_interno_atual, "para": esperado, "estadoFracttal": status_os_atual})
            if aplicar:
                _gravar_status_interno(ws, i, esperado)
                nota = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - 🔧 Auditoria {origem}: status interno "
                        f"corrigido de \"{status_interno_atual or '—'}\" pra \"{esperado}\" "
                        f"(estado na Fracttal: \"{status_os_atual}\").")
                hist_atual = row[ATIV_COL_HISTORICO - 1] if len(row) >= ATIV_COL_HISTORICO else ""
                ws.update_cell(i, ATIV_COL_HISTORICO, f"{hist_atual}\n{nota}".strip() if hist_atual else nota)

    # ── Parte 3: recheca AO VIVO na Fracttal as OSs mais desatualizadas —
    # isso é o que torna a auditoria de verdade "confiável", não só uma
    # conferência de campos que já podem estar todos errados juntos.
    #
    # MUDANÇA (17/07/2026, pedido do Fred): antes processava só um lote
    # fixo (limite_recheck_ao_vivo, 35) por chamada — significava que,
    # com fila grande (ex.: 114 OSs), uma OS específica podia esperar
    # vários ciclos de 5min pra ser rechecada de novo. Agora processa a
    # fila TODA em sequência dentro da mesma chamada (nº de "rodadas"
    # necessárias pra cobrir tudo, calculado a partir do total ÷ 35 —
    # só que aqui, em vez de rodadas HTTP separadas como o botão manual
    # faz, é o mesmo loop contínuo, sem reabrir conexão a cada 35).
    # Protegido por um orçamento de tempo (não um teto de contagem) pra
    # nunca estourar o timeout do gunicorn (120s) — se a fila for grande
    # demais pra caber no orçamento, processa o que der e para; o resto
    # continua com timestamp antigo, então cai automaticamente no topo
    # da fila (mais antigo primeiro) na PRÓXIMA chamada de 5min, sem
    # precisar de nenhuma lógica extra pra "lembrar onde parou".
    revalidadas_ao_vivo = []
    parou_por_orcamento = False
    if aplicar and desatualizadas:
        desatualizadas.sort(key=lambda d: d["_sortKey"])  # string vazia (nunca verificada) primeiro de verdade
        ORCAMENTO_SEGUNDOS = 60  # reduzido de 90 pra 60 (17/07/2026) — com só 1 worker
                                 # no gunicorn, cada segundo aqui é 1 segundo em que o
                                 # backend inteiro fica sem responder mais nada (frontend
                                 # trava em "Erro ao carregar atividades"). O fix real é
                                 # rodar com 2+ workers (systemd, fora do código) — isso
                                 # aqui é só uma margem extra de segurança complementar.
        inicio_recheck = time.time()
        for d in desatualizadas:
            if time.time() - inicio_recheck > ORCAMENTO_SEGUNDOS:
                parou_por_orcamento = True
                log.warning(f"[Auditoria] Orçamento de {ORCAMENTO_SEGUNDOS}s esgotado — "
                            f"{len(revalidadas_ao_vivo)}/{len(desatualizadas)} revalidadas nesta rodada, "
                            f"restante fica pro próximo ciclo automático (5min).")
                break
            resultado = _fracttal_verificar_e_atualizar_uma_os(ws, d["linha"], d["row"], d["numeroOS"],
                                                                enviar_notificacao=False)
            if resultado:
                revalidadas_ao_vivo.append(resultado)
            time.sleep(0.35)

    # um único push resumido pra tudo que mudou nessa rodada, em vez de um
    # por OS — o rodízio pode processar até 40 de uma vez, e um push por
    # item deixava a notificação "spammy" (o Chrome chegou a marcar o
    # site como "possível spam" por causa disso — relatado 14/07/2026).
    mudaram = [r for r in revalidadas_ao_vivo if r.get("mudou")]
    if mudaram:
        try:
            if len(mudaram) == 1:
                r = mudaram[0]
                usina_r = r.get("usina") or "Usina não informada"
                equip_r = r.get("descricao") or r.get("equipamento") or "Descrição não informada"
                enviar_push(
                    titulo=f"🔄 OS {r['numeroOS']} — {usina_r}",
                    corpo=f"{equip_r} · {r.get('statusGeralOS','')} — {r.get('percentualOS','0')}% concluído",
                    tipo="fracttal_status",
                    url=f"https://fred-alexandrino.github.io/PAINELDEFALHAS/?atividade={r.get('id','')}",
                )
            else:
                # Cada linha agora traz usina + tema da OS (Ação/Tarefa,
                # truncado) + o que mudou — antes faltava a usina, então
                # não dava pra saber de qual planta se tratava sem abrir
                # o painel (corrigido 05/08/2026).
                def _linha_resumo(r):
                    usina = (r.get("usina") or "Usina não informada").strip()
                    tema = (r.get("descricao") or r.get("equipamento") or "sem descrição").strip()
                    if len(tema) > 35:
                        tema = tema[:35].rstrip() + "…"
                    mudanca = r.get("mudancaResumo") or r.get("statusGeralOS") or ""
                    base = f"{r['numeroOS']} · {usina} — {tema}"
                    return f"{base} ({mudanca})" if mudanca else base
                linhas = "\n".join(_linha_resumo(r) for r in mudaram[:8])
                enviar_push(
                    titulo=f"🔄 {len(mudaram)} OSs atualizadas",
                    corpo=f"{linhas}{chr(10) + '...' if len(mudaram) > 8 else ''}",
                    tipo="fracttal_status",
                )
        except Exception as e:
            log.error(f"[Auditoria] Falha ao enviar push resumido de status atualizado: {e}")

    for d in desatualizadas:
        d.pop("linha", None)
        d.pop("row", None)
        d.pop("_sortKey", None)

    if divergencias:
        log.warning(f"[Auditoria] {len(divergencias)} divergência(s) de status encontrada(s) "
                    f"(aplicado={aplicar}): {[d['numeroOS'] for d in divergencias]}")
    if desatualizadas:
        log.warning(f"[Auditoria] {len(desatualizadas)} OS(s) sem verificação recente na Fracttal "
                    f"(>{limite_atraso_minutos}min), {len(revalidadas_ao_vivo)} revalidada(s) ao vivo agora: "
                    f"{[d['numeroOS'] for d in desatualizadas]}")

    return {"aplicado": aplicar, "total_divergencias": len(divergencias), "divergencias": divergencias,
            "total_desatualizadas": len(desatualizadas), "desatualizadas": desatualizadas,
            "revalidadas_ao_vivo": revalidadas_ao_vivo, "limite_recheck_ao_vivo": limite_recheck_ao_vivo,
            "parou_por_orcamento_tempo": parou_por_orcamento}


def _extrair_data_fallback_historico(historico, palavras_chave=None):
    """Varre o histórico (texto multi-linha) procurando a última data/hora
    (dd/mm/aaaa hh:mm) associada a uma transição de conclusão. Se
    palavras_chave for dado, prioriza linhas que contenham alguma delas
    (ex.: 'finalizada', 'concluíd', 'normalizad', 'cancelad'); senão usa a
    última data encontrada em qualquer linha. Retorna string
    'dd/mm/aaaa hh:mm:ss' pronta pra gravar, ou None se não achar nada."""
    linhas = (historico or "").strip().split("\n")
    padrao_data = re.compile(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2})")
    candidatas = []
    for linha in linhas:
        m = padrao_data.search(linha)
        if not m:
            continue
        prioridade = 0
        if palavras_chave and any(p in linha.lower() for p in palavras_chave):
            prioridade = 1
        candidatas.append((prioridade, linha, m.group(1)))
    if not candidatas:
        return None
    candidatas.sort(key=lambda t: t[0])  # prioridade 1 por último (a gente pega o último da lista com maior prioridade)
    melhor = [c for c in candidatas if c[0] == 1] or candidatas
    return f"{melhor[-1][2]}:00"


def _validar_integridade_relatorios_core(aplicar=True):
    """AUTOMAÇÃO DE VALIDAÇÃO — roda pra TODOS os clientes de uma vez (não
    é específica de nenhum caso pontual). Varre tanto o Painel de Falhas
    quanto o Painel de Atividades procurando o padrão de bug que fazia
    ocorrências sumirem dos relatórios semanais (13/07/2026): status
    marcado como concluído/cancelado mas sem a data de fechamento
    correspondente gravada — o relatório usa exatamente esse campo pra
    decidir se algo entra ou não.

    Corrige automaticamente usando a última data relevante do histórico
    de cada item (não usa "agora" como data, pra não distorcer em qual
    semana o item realmente foi concluído). Roda 3x/dia junto com a
    auditoria completa — funciona como um "aprovado/reprovado" contínuo
    da integridade dos dados que alimentam todos os relatórios, sem
    precisar esperar alguém notar um relatório com buraco."""
    problemas = {"falhas": [], "atividades": []}

    # ── Painel de Falhas ─────────────────────────────────────────────────
    try:
        ws_falhas = get_sheet()
        todos_falhas = ws_falhas.get_all_values()
        COL_CLIENTE, COL_STATUS, COL_HISTORICO = 1, 8, 11
        COL_DATA_FECHAMENTO = 20
        palavras_falha = ("normalizad", "resolvid", "concluíd", "concluid", "cancelad", "encerrad")
        for i, row in enumerate(todos_falhas[1:], start=2):
            if len(row) <= COL_DATA_FECHAMENTO:
                row = row + [""] * (COL_DATA_FECHAMENTO + 1 - len(row))
            status = row[COL_STATUS].strip().lower()
            concluida = any(x in status for x in ("resolvid", "concluíd", "concluid", "normalizad", "cancelad"))
            data_fechamento = row[COL_DATA_FECHAMENTO].strip()
            if concluida and not data_fechamento:
                fallback = _extrair_data_fallback_historico(row[COL_HISTORICO], palavras_falha)
                item = {"linha": i, "cliente": row[COL_CLIENTE].strip(), "corrigivel": bool(fallback)}
                problemas["falhas"].append(item)
                if aplicar and fallback:
                    ws_falhas.update_cell(i, COL_DATA_FECHAMENTO + 1, fallback)
    except Exception as e:
        log.error(f"[ValidacaoRelatorios] Erro ao varrer Painel de Falhas: {e}")

    # ── Painel de Atividades ─────────────────────────────────────────────
    try:
        ws_ativ = get_atividades_sheet()
        todos_ativ = ws_ativ.get_all_values()
        palavras_ativ = ("finalizada", "concluíd", "concluid", "cancelad")
        for i, row in enumerate(todos_ativ[1:], start=2):
            if len(row) < ATIV_TOTAL_COLUNAS:
                row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
            status = row[8].strip()
            data_conclusao = row[10].strip()
            if status in ("Concluído", "Cancelado") and not data_conclusao:
                fallback = _extrair_data_fallback_historico(row[ATIV_COL_HISTORICO - 1], palavras_ativ)
                item = {"linha": i, "id": row[0].strip(), "cliente": row[1].strip(),
                        "numeroOS": row[13].strip(), "corrigivel": bool(fallback)}
                problemas["atividades"].append(item)
                if aplicar and fallback:
                    ws_ativ.update_cell(i, ATIV_CAMPO_COL["dataConclusao"], fallback)
    except Exception as e:
        log.error(f"[ValidacaoRelatorios] Erro ao varrer Painel de Atividades: {e}")

    total = len(problemas["falhas"]) + len(problemas["atividades"])
    if total > 0:
        clientes_afetados = sorted(set(
            [p["cliente"] for p in problemas["falhas"] if p.get("cliente")] +
            [p["cliente"] for p in problemas["atividades"] if p.get("cliente")]
        ))
        log.warning(f"[ValidacaoRelatorios] {total} problema(s) de integridade encontrado(s) "
                    f"(clientes: {clientes_afetados}, aplicado={aplicar})")
        try:
            enviar_push(
                titulo=f"🔧 Validação de relatórios: {total} corrigido(s)" if aplicar else f"⚠️ Validação de relatórios: {total} problema(s)",
                corpo=f"Clientes afetados: {', '.join(clientes_afetados) or '—'}",
                tipo="validacao_relatorios",
            )
        except Exception as e:
            log.error(f"[ValidacaoRelatorios] Falha ao enviar push: {e}")

    return {"aplicado": aplicar, "total_problemas": total, "detalhes": problemas}


def _auditoria_completa_core(desde_horas_descoberta=24, limite_recheck_ao_vivo=40, origem="automática"):
    """AUDITORIA COMPLETA — varredura de verdade nas usinas/equipes do
    Fred, cobrindo tudo que uma auditoria de verdade precisa cobrir:
      1. DESCOBERTA: busca na Fracttal por OTs novas dentro da janela
         (padrão 24h) que ainda não estão no dashboard — pega OS nova
         que a descoberta rápida de rotina (2h) porventura tenha perdido.
      2. VARREDURA DE STATUS/ESTADO: revalida ao vivo na Fracttal um lote
         das OSs já existentes — detecta não só mudança de percentual,
         mas também cancelamentos e conclusões que tenham escapado.
      3. VALIDAÇÃO DE INTEGRIDADE DE RELATÓRIOS: varre Painel de Falhas +
         Painel de Atividades (todos os clientes) procurando o padrão que
         faz ocorrências sumirem dos relatórios semanais, corrigindo
         automaticamente.
    Roda automaticamente 3x/dia (7h/12h/16h) e sob demanda no botão
    "Auditoria". Mais pesada que a checagem de rotina (frequente, 5 em
    5 min) de propósito — por isso não roda toda hora, só nesses horários."""
    resultado_descoberta, _ = _sync_fracttal_core(desde_horas=desde_horas_descoberta)
    resultado_consistencia = _auditoria_consistencia_os_core(aplicar=True, limite_atraso_minutos=0,
                                                              limite_recheck_ao_vivo=limite_recheck_ao_vivo,
                                                              origem=origem)
    resultado_validacao_relatorios = _validar_integridade_relatorios_core(aplicar=True)
    return {"descoberta": resultado_descoberta, "consistencia": resultado_consistencia,
            "validacao_relatorios": resultado_validacao_relatorios}


def _verificar_e_disparar_auditoria_completa_se_necessario():
    """Só dispara a auditoria completa de verdade se estiver dentro de uma
    das 3 janelas do dia (07:00-07:09, 12:00-12:09, 16:00-16:09, horário
    de Brasília) e ainda não tiver rodado nessa janela hoje — mesmo
    padrão de trava usado pros comunicados, adaptado pra 3 horários."""
    try:
        agora = agora_br()
        janela_atual = None
        for h in (7, 12, 16):
            if agora.hour == h and agora.minute < 10:
                janela_atual = h
                break
        if janela_atual is None:
            return {"disparado": False, "motivo": f"fora das janelas 7h/12h/16h (agora {agora.strftime('%H:%M')})"}

        chave_trava = f"auditoria_completa_em_{janela_atual}h"
        hoje_str = agora.strftime("%Y-%m-%d")
        ja_rodou = _ler_trava(chave_trava)
        if ja_rodou == hoje_str:
            return {"disparado": False, "motivo": f"já rodou hoje na janela das {janela_atual}h"}

        _gravar_trava(chave_trava, hoje_str)
        resultado = _auditoria_completa_core()
        return {"disparado": True, "janela": f"{janela_atual}h", "resultado": resultado}
    except Exception as e:
        log.error(f"[AuditoriaCompleta] Erro na verificação/disparo: {e}")
        return {"disparado": False, "erro": str(e)}


def _verificar_e_disparar_descoberta_rapida_se_necessario(intervalo_minutos=30):
    """DESCOBERTA RÁPIDA — roda automaticamente a cada 30 min via piggyback
    no /sync-fracttal (mesmo gatilho confiável dos 5 min já usado pra
    atualização de status). Sem botão manual — existe só pra reduzir o
    gap de latência entre uma OS nova nascer na Fracttal e ela aparecer
    no dashboard, que antes podia chegar a ~9h (pior caso: OS criada logo
    depois da janela das 16h só entraria às 7h do dia seguinte).

    Deliberadamente LEVE, ao contrário da auditoria completa: só chama
    _sync_fracttal_core (descoberta pura, sem recheck de OSs existentes)
    com janela curta (2h) — não faz a varredura ampla nem a validação de
    integridade de relatórios que a auditoria completa faz. Isso evita
    reintroduzir o risco de 502 que já vimos quando descoberta ampla e
    recheck pesado rodaram juntos no mesmo request.

    Trava por timestamp (não por dia, como as outras) porque precisa
    rodar várias vezes ao dia, não uma vez só por janela."""
    try:
        agora = agora_br()
        chave_trava = "descoberta_rapida_ultima_em"
        ultima_str = _ler_trava(chave_trava)
        if ultima_str:
            try:
                ultima = datetime.strptime(ultima_str, "%Y-%m-%d %H:%M:%S")
                minutos_desde = (agora.replace(tzinfo=None) - ultima).total_seconds() / 60
                if minutos_desde < intervalo_minutos:
                    return {"disparado": False,
                            "motivo": f"rodou há {minutos_desde:.1f}min (< {intervalo_minutos}min); última em {ultima_str}"}
            except ValueError:
                pass  # trava com valor inválido/corrompido — trata como se nunca tivesse rodado

        _gravar_trava(chave_trava, agora.strftime("%Y-%m-%d %H:%M:%S"))
        resultado, status_http = _sync_fracttal_core(desde_horas=2)
        if status_http != 200:
            return {"disparado": True, "erro": resultado}
        return {"disparado": True, "resultado": resultado}
    except Exception as e:
        log.error(f"[DescobertaRapida] Erro na verificação/disparo: {e}")
        return {"disparado": False, "erro": str(e)}


@app.route("/auditoria-consistencia-os", methods=["POST", "GET"])
def auditoria_consistencia_os():
    """Rede de segurança definitiva: varre TODAS as atividades vinculadas
    a uma OS da Fracttal e confere se o status interno bate com o que
    _status_interno_esperado() diz que deveria ser, dado o estado
    (statusOS) atual já registrado — sem precisar chamar a API da
    Fracttal de novo (usa o que já está gravado, então é rápido e barato
    de rodar com frequência). Corrige qualquer divergência encontrada.

    Roda automaticamente a cada 5 min via piggyback no sync-fracttal
    (gatilho confiável), então qualquer inconsistência que escape da
    checagem normal (por bug futuro, edição manual, etc.) se autocorrige
    sozinha em poucos minutos, sem precisar de intervenção manual."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    aplicar = request.args.get("apply", "true").lower() != "false"
    resultado = _auditoria_consistencia_os_core(aplicar, origem="manual (diagnóstico)")
    return jsonify({"ok": True, **resultado}), 200


@app.route("/resumo", methods=["GET"])
def resumo_widget():
    """
    Endpoint leve pra consumo por widgets externos (ex.: apps de widget
    Android tipo KWGT/HTTP Request Widget, configurados pelo Fred na tela
    inicial do celular) — só os números-chave, sem os dados completos de
    cada atividade/chamado, pra ser rápido e simples de exibir.

    GET simples, sem secret (só números agregados, nada sensível de cada
    registro individual é exposto aqui).
    """
    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        status_excluidos = {"concluído", "concluido", "cancelado",
                             "convertida em ocorrência", "convertida em ocorrencia"}
        total = abertas = atrasadas = altas_abertas = concluidas_7d = 0
        hoje = agora_br().date()
        limite7 = hoje - timedelta(days=7)

        for row in todos[1:]:
            if len(row) < len(ATIV_HEADERS_JSON):
                row = row + [""] * (len(ATIV_HEADERS_JSON) - len(row))
            item = dict(zip(ATIV_HEADERS_JSON, row[:len(ATIV_HEADERS_JSON)]))
            if not item.get("id"):
                continue
            total += 1
            concluida = (item.get("status") or "").strip().lower() in status_excluidos
            if not concluida:
                abertas += 1
                prazo_str = (item.get("prazo") or "").strip()
                if prazo_str:
                    try:
                        if datetime.strptime(prazo_str, "%d/%m/%Y").date() < hoje:
                            atrasadas += 1
                    except Exception:
                        pass
                if (item.get("prioridade") or "").strip().lower() == "alta":
                    altas_abertas += 1
            else:
                dataconc = (item.get("dataConclusao") or "").strip().split(" ")[0]
                if dataconc:
                    try:
                        if datetime.strptime(dataconc, "%d/%m/%Y").date() >= limite7:
                            concluidas_7d += 1
                    except Exception:
                        pass

        chamados_total = chamados_abertos = 0
        try:
            ws_ch = get_chamados_fabricante_sheet()
            todos_ch = ws_ch.get_all_values()
            idx_supervisor = CHAMADOS_FABRICANTE_HEADERS.index("Supervisor")
            idx_status = CHAMADOS_FABRICANTE_HEADERS.index("Status")
            for row in todos_ch[1:]:
                if len(row) < len(CHAMADOS_FABRICANTE_HEADERS):
                    row = row + [""] * (len(CHAMADOS_FABRICANTE_HEADERS) - len(row))
                if row[idx_supervisor].strip().lower() != "fred alexandrino":
                    continue
                chamados_total += 1
                st = row[idx_status].strip().lower()
                if not any(k in st for k in ("conclu", "resolv", "fechad", "finaliz")):
                    chamados_abertos += 1
        except Exception:
            pass  # chamados é "bônus" no resumo, não deve derrubar o endpoint todo

        return jsonify({
            "ok": True,
            "atualizado_em": agora_br().strftime("%d/%m/%Y %H:%M"),
            "atividades": {
                "total": total,
                "em_aberto": abertas,
                "atrasadas": atrasadas,
                "prioridade_alta_abertas": altas_abertas,
                "concluidas_7d": concluidas_7d,
            },
            "chamados_fred": {
                "total": chamados_total,
                "em_aberto": chamados_abertos,
            },
        }), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/atividades", methods=["GET"])
def listar_atividades():
    try:
        ws = get_atividades_sheet()
        todos = _gspread_retry(lambda: ws.get_all_values())
        mapa_cluster = _mapa_cluster_usina()
        out = []
        for row in todos[1:]:
            if len(row) < len(ATIV_HEADERS_JSON):
                row = row + [""] * (len(ATIV_HEADERS_JSON) - len(row))
            if not row[0].strip():
                continue
            item = dict(zip(ATIV_HEADERS_JSON, row[:len(ATIV_HEADERS_JSON)]))
            # Correção 31/07/2026: se a atividade foi criada enquanto a
            # usina estava sob supervisão temporária e o Fred já devolveu
            # essa usina (removeu de /supervisao-temporaria), ela não deve
            # mais aparecer aqui — senão a devolução não devolve de fato,
            # só esconde a usina do catálogo mas deixa o histórico de
            # atividades grudado no painel do Fred pra sempre.
            if not usina_permitida(item.get("usina", "")):
                continue
            item["cluster"] = mapa_cluster.get(item.get("usina", "").strip(), "")
            out.append(item)
        return jsonify({"ok": True, "atividades": out})
    except Exception as e:
        log.error(f"[Atividades] Erro ao listar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _criar_atividade_interna(cliente, usina="", equipamento="", descricao="", responsavel="",
                              prazo="", prioridade="Média", status="Em Aberto", numeroOS="",
                              editor="dashboard", statusOS="", observacoesOS="", linkOS="",
                              statusTarefaOS="", etiquetasOS="", anotacoesPessoais="",
                              percentualOS="", statusGeralOS="", detalhesEquipamentosOS="",
                              ws=None, todos=None, enviar_notificacao=True):
    """
    Cria uma linha na aba Painel de Atividades. Usada tanto pelo endpoint
    HTTP /nova-atividade quanto pelo sync automático do Fracttal
    (/sync-fracttal, /backfill-fracttal), para evitar duplicar a lógica de
    escrita na planilha.

    Se `ws`/`todos` forem passados (leitura já feita por quem chamou, ex.
    sync em lote), evita reler a planilha inteira a cada chamada.

    enviar_notificacao=False quando quem chama vai criar várias atividades
    de uma vez (ex.: descoberta da Fracttal encontrando N OTs novas na
    mesma rodada) — nesse caso, quem chama deve mandar um único push
    resumido no final, em vez de um por item (evita disparar muitas
    notificações em sequência rápida — o Chrome já marcou o site como
    "possível spam" por causa disso antes, 14/07/2026. RESTAURADO em
    15/07/2026 depois de ter sido perdido numa edição de outra sessão que
    reconstruiu esta função a partir de uma versão mais antiga do arquivo).
    """
    cliente = (cliente or "").strip()
    descricao = (descricao or "").strip()
    if not cliente or not descricao:
        raise ValueError("cliente e descricao são obrigatórios")

    if ws is None:
        ws = get_atividades_sheet()
    if todos is None:
        todos = ws.get_all_values()

    numeroOS = (numeroOS or "").strip()
    if numeroOS:
        for row in todos[1:]:
            if len(row) < 14:
                continue
            numero_os_existente = row[13].strip()
            status_existente = row[8].strip()
            if numero_os_existente == numeroOS and not _is_concluido_atividade(status_existente):
                raise ValueError(f"Já existe uma atividade em aberto (id {row[0]}) pra essa OS ({numeroOS}). "
                                  f"Abra e edite a atividade existente em vez de criar uma nova.")

    _garantir_headers_atividades(ws)

    novo_id = _proximo_id_atividade(todos)
    agora = agora_br().strftime('%d/%m/%Y %H:%M:%S')
    historico_inicial = f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Atividade criada por {_editor_legivel(editor)}."
    data_conclusao_inicial = agora if status in ("Concluído", "Cancelado") else ""

    linha = [novo_id, cliente, usina, equipamento, descricao, responsavel, prazo,
             prioridade, status, agora, data_conclusao_inicial, historico_inicial, editor, numeroOS,
             statusOS, observacoesOS, linkOS, statusTarefaOS, etiquetasOS, anotacoesPessoais,
             percentualOS, statusGeralOS, detalhesEquipamentosOS, "", ""]
    ws.append_row(linha)
    # mantém `todos` coerente para quem estiver criando várias atividades em sequência
    todos.append(linha)
    log.info(f"[atividade] #{novo_id} {cliente}/{usina} — {descricao[:60]} | editor={editor}")

    if enviar_notificacao:
        try:
            enviar_push(
                titulo=f"🆕 Nova atividade" + (f" — OS {numeroOS}" if numeroOS else "") + f" — {usina or cliente}",
                corpo=(f"{equipamento} · " if equipamento else "") +
                      (f"{descricao[:80]}" if descricao else "Atividade criada"),
                tipo="nova_atividade",
                url=f"https://fred-alexandrino.github.io/PAINELDEFALHAS/?atividade={novo_id}",
            )
        except Exception as e:
            log.error(f"[Push] Erro ao notificar nova atividade: {e}")

    return novo_id


@app.route("/nova-atividade", methods=["POST", "OPTIONS"])
def nova_atividade():
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido"}), 400

    try:
        novo_id = _criar_atividade_interna(
            cliente=body.get("cliente", ""),
            usina=body.get("usina", ""),
            equipamento=body.get("equipamento", ""),
            descricao=body.get("descricao", ""),
            responsavel=body.get("responsavel", ""),
            prazo=body.get("prazo", ""),
            prioridade=body.get("prioridade", "Média").strip() or "Média",
            status=body.get("status", "Em Aberto").strip() or "Em Aberto",
            numeroOS=body.get("numeroOS", ""),
            editor=body.get("editor", "dashboard").strip() or "dashboard",
            statusOS=body.get("statusOS", ""),
            observacoesOS=body.get("observacoesOS", ""),
            linkOS=body.get("linkOS", ""),
            statusTarefaOS=body.get("statusTarefaOS", ""),
            etiquetasOS=body.get("etiquetasOS", ""),
            anotacoesPessoais=body.get("anotacoesPessoais", ""),
            percentualOS=body.get("percentualOS", ""),
            statusGeralOS=body.get("statusGeralOS", ""),
            detalhesEquipamentosOS=body.get("detalhesEquipamentosOS", ""),
        )
        return jsonify({"ok": True, "id": novo_id})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        log.error(f"[Atividades] Erro ao criar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _montar_prompt_extrair_atividade_imagem():
    lista_usinas = ", ".join(sorted(CATALOGO_USINAS.keys()))
    return f"""Aja como um Supervisor de O&M da Grid Co. lendo o print de uma ou mais Ordens de Serviço (OS)/tarefas
para cadastrar atividades no painel de gestão. A imagem pode ser: um único card da Fracttal, uma mensagem de
WhatsApp de um técnico, uma anotação de campo — OU uma tabela/planilha de repasse com VÁRIAS usinas, cada uma
com uma ou mais observações (ex.: colunas Contratante / UFV / Observações Gerais).

Extraia da imagem uma LISTA de atividades e responda APENAS com um JSON válido (sem markdown, sem crase,
sem texto antes ou depois), no formato:
{{
  "atividades": [
    {{
      "cliente": "",
      "usina": "",
      "equipamento": "",
      "descricao": "",
      "responsavel": "",
      "prazo": "",
      "prioridade": "",
      "numeroOS": "",
      "status": ""
    }}
  ]
}}

Regras gerais:
- Se a imagem mostrar uma tabela/lista com várias usinas, gere UMA atividade para CADA observação/bullet distinto
  e acionável dentro de cada usina — não junte várias observações diferentes numa única "descricao"; cada uma vira
  um item separado da lista, mesmo que a usina se repita várias vezes na lista.
- Se a imagem mostrar um único card/print de OS, retorne a lista com apenas 1 item.
- Cada "descricao" deve ser objetiva e fiel ao que está escrito — pode reescrever de forma mais clara e concisa,
  mas sem inventar informação nem misturar observações de usinas diferentes num mesmo item.

Regras por campo (aplicam-se a cada item da lista):
- "usina": nome da usina/planta como aparece na imagem. Usinas conhecidas no catálogo (tente casar com uma
  destas se fizer sentido, mas não force — se a imagem mostrar outra usina não listada, transcreva como
  está escrito mesmo): {lista_usinas}
- "cliente": só preencha se estiver explícito na imagem OU se você tiver certeza pela usina identificada;
  senão deixe vazio (o sistema tenta inferir pelo catálogo depois).
- "descricao": a AÇÃO/TAREFA real a ser feita (o que precisa ser executado), NUNCA apenas o código do
  ativo/equipamento. Ex.: se o ativo é "THPN-TPZ100-SSEG1-CMRA" mas a ação é "Recomposição de câmera de
  CFTV", "descricao" deve ser "Recomposição de câmera de CFTV", não o código do ativo.
- "equipamento": o ativo/equipamento em si (código ou nome), separado da ação. Deixe vazio se não for claro.
- "prazo": data no formato DD/MM/AAAA se houver uma data-limite visível; senão vazio.
- "prioridade": "Baixa", "Média" ou "Alta" — só preencha se houver indicação clara na imagem (palavras
  como urgente/crítico = Alta); senão deixe "Média" como neutro.
- "numeroOS": número da OS/OT se visível (só os dígitos, sem prefixo "OS"); senão vazio.
- "status": um destes valores, o mais coerente com o que a imagem mostra: "Em Aberto", "Em Andamento",
  "Aguardando Fabricante", "Aguardando Cliente", "Abrir chamado". Se não houver indicação clara, use
  "Em Aberto".
- "responsavel": nome da pessoa/técnico responsável, se citado; senão vazio.

REGRA CRÍTICA: se algum campo estiver ilegível, cortado, ambíguo ou simplesmente não aparecer na imagem,
deixe esse campo como string vazia "" — NUNCA presuma ou invente conteúdo. É melhor deixar em branco pra
o supervisor preencher manualmente do que registrar informação errada."""


@app.route("/extrair-atividade-de-imagem", methods=["POST", "OPTIONS"])
def extrair_atividade_de_imagem():
    """Lê o print de uma OS/tarefa (Fracttal, WhatsApp, anotação de campo)
    via Gemini (visão) e devolve os campos extraídos para pré-preencher o
    formulário de 'Nova Atividade' no dashboard — o supervisor revisa e
    confirma antes de efetivamente criar a atividade (POST /nova-atividade
    continua sendo um passo separado, feito pelo frontend depois que o
    usuário confere/ajusta os campos)."""
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500

    body = request.get_json(force=True, silent=True) or {}
    imagem_b64 = body.get("imagemBase64") or ""
    imagem_mime = body.get("imagemMimeType") or "image/png"
    if not imagem_b64:
        return jsonify({"ok": False, "error": "anexe uma imagem (print) da OS/atividade"}), 400

    prompt = _montar_prompt_extrair_atividade_imagem()
    parts = [{"text": prompt}, {"inline_data": {"mime_type": imagem_mime, "data": imagem_b64}}]

    diagnostico = request.args.get("diagnostico", "").lower() == "true"
    try:
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": parts}],
                "generationConfig": {
                    "temperature": 0.2,
                    "maxOutputTokens": 8192,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=55,
            usar_chave_teste=diagnostico,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        texto_bruto = candidato["content"]["parts"][0]["text"].strip()
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
        parsed = json.loads(texto_limpo)
    except Exception as e:
        log.error(f"[extrair-atividade-de-imagem] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 502

    # Aceita tanto {"atividades": [...]} (formato novo) quanto um objeto único
    # ou uma lista solta, pra ser resiliente a variações de resposta da IA.
    if isinstance(parsed, list):
        lista_bruta = parsed
    elif isinstance(parsed, dict) and isinstance(parsed.get("atividades"), list):
        lista_bruta = parsed["atividades"]
    elif isinstance(parsed, dict):
        lista_bruta = [parsed]
    else:
        return jsonify({"ok": False, "error": "formato inesperado retornado pela IA"}), 502

    if not lista_bruta:
        return jsonify({"ok": False, "error": "nenhuma atividade identificada na imagem"}), 502

    atividades = []
    for campos in lista_bruta:
        if not isinstance(campos, dict):
            continue
        usina_bruta = (campos.get("usina") or "").strip()
        usina_canonica = canonizar_usina(usina_bruta) if usina_bruta else None
        usina_reconhecida = usina_canonica is not None
        if usina_canonica:
            campos["usina"] = usina_canonica
            cliente_inferido = inferir_cliente(usina_canonica)
            if cliente_inferido:
                campos["cliente"] = cliente_inferido

        campos.setdefault("prioridade", "Média")
        if not campos.get("prioridade"):
            campos["prioridade"] = "Média"
        campos.setdefault("status", "Em Aberto")
        if not campos.get("status"):
            campos["status"] = "Em Aberto"
        campos["usinaReconhecida"] = usina_reconhecida
        atividades.append(campos)

    if not atividades:
        return jsonify({"ok": False, "error": "nenhuma atividade válida identificada na imagem"}), 502

    resposta = {"ok": True, "atividades": atividades}
    # Compat: quando só há 1 item, mantém também "campos"/"usinaReconhecida" no
    # nível raiz (formato antigo), pra não quebrar nenhum cliente que ainda
    # espere só um objeto único.
    if len(atividades) == 1:
        resposta["campos"] = atividades[0]
        resposta["usinaReconhecida"] = atividades[0]["usinaReconhecida"]

    return jsonify(resposta), 200


# ── Integração Fracttal (sync automático de OTs → Painel de Atividades) ───
FRACTTAL_CLIENT_KEY    = os.environ.get("FRACTTAL_CLIENT_KEY", "")
FRACTTAL_CLIENT_SECRET = os.environ.get("FRACTTAL_CLIENT_SECRET", "")
FRACTTAL_TOKEN_URL     = "https://one.fracttal.com/oauth/token"
FRACTTAL_API_BASE      = "https://app.fracttal.com/api"

_fracttal_token_cache = {"access_token": None, "expires_at": 0}


def _fracttal_get_token():
    """Obtém (com cache em memória) um access_token OAuth2 client_credentials da Fracttal."""
    agora = time.time()
    if _fracttal_token_cache["access_token"] and _fracttal_token_cache["expires_at"] > agora + 60:
        return _fracttal_token_cache["access_token"]

    if not FRACTTAL_CLIENT_KEY or not FRACTTAL_CLIENT_SECRET:
        raise RuntimeError("FRACTTAL_CLIENT_KEY / FRACTTAL_CLIENT_SECRET não configurados no Render")

    resp = requests.post(
        FRACTTAL_TOKEN_URL,
        auth=(FRACTTAL_CLIENT_KEY, FRACTTAL_CLIENT_SECRET),
        data={"grant_type": "client_credentials"},
        timeout=20,
    )
    resp.raise_for_status()
    data = resp.json()
    _fracttal_token_cache["access_token"] = data["access_token"]
    _fracttal_token_cache["expires_at"] = agora + int(data.get("expires_in", 7200))
    return _fracttal_token_cache["access_token"]


def _fracttal_listar_pagina(since=None, until=None, ot_status=None, start=0, limit=100):
    """
    Consulta uma página de work_orders na Fracttal usando os parâmetros
    OFICIAIS confirmados na documentação (api.fracttal.com/reference):
      since / until   — formato 'YYYY-MM-DDTHH:MM:SS-00:00', filtra por creation_date
      ot_status       — 1: Processo, 2: Revisão, 3: Finalizada, 4: Cancelada
      start / limit   — paginação (limit máximo 100, é o teto da própria Fracttal)

    Retorna (lista_de_ots, total_geral).
    """
    token = _fracttal_get_token()
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    params = {"start": start, "limit": min(limit, 100)}
    if since:
        params["since"] = since
    if until:
        params["until"] = until
    if ot_status:
        params["ot_status"] = ot_status

    resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders", headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    body = resp.json()
    return body.get("data", []) or [], body.get("total", 0)


def _fracttal_listar_ots_recentes(desde_horas=3):
    """Consulta OTs criadas/atualizadas nas últimas `desde_horas` horas (usado pelo cron de 2h em 2h)."""
    since = (datetime.utcnow() - timedelta(hours=desde_horas)).strftime("%Y-%m-%dT%H:%M:%S-00:00")
    ots, _total = _fracttal_listar_pagina(since=since, start=0, limit=100)
    return ots


_FRACTTAL_PRIORIDADE_MAP = {
    "HIGH": "Alta", "ALTA": "Alta",
    "MEDIUM": "Média", "MEDIA": "Média", "MÉDIA": "Média",
    "LOW": "Baixa", "BAIXA": "Baixa",
}

# 1: Processo, 2: Revisão, 3: Finalizada, 4: Cancelada (confirmado na doc oficial da Fracttal)
_FRACTTAL_STATUS_OS_MAP = {
    "1": "Em Processo", "2": "Em Revisão", "3": "Finalizada", "4": "Cancelada",
}


def _status_interno_esperado(estado_fracttal, status_interno_atual):
    """FONTE ÚNICA DE VERDADE pra decidir o status interno (coluna
    "status") a partir do ESTADO real da OS na Fracttal (coluna "statusOS"
    — Em Processo/Em Revisão/Finalizada/Cancelada). Usada tanto na criação
    de uma atividade nova quanto em TODA checagem de rotina subsequente —
    nunca duplicar essa lógica em outro lugar do código.

    Regras (only touches these two transitions, nunca mexe em status
    manuais tipo "Pausado"/"Aguardando Cliente" etc. definidos por
    técnico via WhatsApp):
      1. Estado = Finalizada  → status interno deve ser "Concluído"
         (a menos que já esteja "Cancelado", que é uma conclusão também).
      2. Estado = Cancelada   → status interno deve ser "Cancelado".
      3. Se o status interno atual é "Concluído"/"Cancelado" mas o estado
         NÃO é mais Finalizada/Cancelada (reaberta/reprovada na Fracttal)
         → volta pra "Em Aberto".
      4. Qualquer outro caso (estado ainda em Processo/Revisão e status
         interno não é Concluído/Cancelado) → não mexe, devolve None.

    Retorna o novo valor se precisar corrigir, ou None se já está certo.
    IMPORTANTE: essa função deve rodar em TODA checagem, independente de
    mais alguma coisa ter mudado na mesma passada — é isso que garante
    que o sistema se autocorrige, em vez de só corrigir "de carona" numa
    mudança de outro campo (bug estrutural corrigido em 12/07/2026)."""
    if estado_fracttal == "Finalizada":
        alvo = "Concluído"
    elif estado_fracttal == "Cancelada":
        alvo = "Cancelado"
    elif estado_fracttal in ("Em Processo", "Em Revisão") and status_interno_atual in ("Concluído", "Cancelado"):
        alvo = "Em Aberto"
    else:
        return None
    return alvo if alvo != status_interno_atual else None


def _gravar_status_interno(ws, i, novo_status):
    """FONTE ÚNICA que grava o status interno na planilha — usada nos 3
    pontos do código que podem mudar esse campo (checagem individual,
    auditoria, reabertura). Sempre que o novo status é uma conclusão
    (Concluído/Cancelado), também grava a Data de Conclusão — campo que
    ficava sempre vazio antes (bug identificado em 13/07/2026: relatórios
    semanais usam esse campo pra saber se algo fechou dentro do período,
    e como nunca era preenchido, OSs concluídas sumiam dos relatórios).
    Ao reabrir (volta pra "Em Aberto"), limpa a Data de Conclusão de novo.

    Quando vira "Cancelado", também sincroniza o statusGeralOS pra
    "Cancelada" — senão esse campo (progresso da tarefa) fica com o valor
    de antes do cancelamento pra sempre (a Fracttal não manda mais dado
    novo pra uma OS cancelada), fazendo o badge mostrar algo tipo "Não
    Iniciada" em vez de refletir que foi cancelada (bug identificado em
    14/07/2026)."""
    ws.update_cell(i, ATIV_CAMPO_COL["status"], novo_status)
    if novo_status in ("Concluído", "Cancelado"):
        ws.update_cell(i, ATIV_CAMPO_COL["dataConclusao"], agora_br().strftime("%d/%m/%Y %H:%M:%S"))
    elif novo_status == "Em Aberto":
        ws.update_cell(i, ATIV_CAMPO_COL["dataConclusao"], "")
    if novo_status == "Cancelado":
        ws.update_cell(i, ATIV_CAMPO_COL["statusGeralOS"], "Cancelada")

# A Fracttal tem o add-on "Share TOs" habilitado nesta conta, que gera uma
# URL pública específica por OT via /work_orders_shared_url/{folio}
# (confirmado com um teste real — ver histórico). Isso abre a OT direto,
# sem precisar buscar manualmente na lista. Se a chamada falhar por
# qualquer motivo, cai pro fallback da tela de OTs (fluxo antigo).
FRACTTAL_WEB_BASE = "https://app.fracttal.com/tasks/wo"


def _fracttal_montar_link(ot):
    folio = (ot.get("wo_folio") or "").strip()
    if not folio:
        return ""
    try:
        token = _fracttal_get_token()
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders_shared_url/{folio}",
                             headers=headers, timeout=15)
        resp.raise_for_status()
        dados = (resp.json().get("data") or [])
        if dados and dados[0].get("shared_wo_url"):
            return dados[0]["shared_wo_url"]
    except Exception as e:
        log.error(f"[fracttal] Erro ao buscar shared_wo_url da OT {folio}: {e}")
    return FRACTTAL_WEB_BASE


def _fracttal_formatar_data_br(iso_str):
    """Extrai apenas AAAA-MM-DD do timestamp ISO da Fracttal e devolve dd/mm/aaaa."""
    if not iso_str:
        return ""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(iso_str))
    if not m:
        return ""
    ano, mes, dia = m.groups()
    return f"{dia}/{mes}/{ano}"

# ── Cruzamento técnico responsável → usina(s) atendida(s) ─────────────────
# Usado para VALIDAR o match feito pelo nome do ativo (cross-check) e,
# quando o nome do ativo não bate com o catálogo, como fallback — mas só
# quando o técnico atende uma única usina do catálogo (senão é ambíguo e
# a OT vai para revisão manual em vez de arriscar um chute).
TECNICO_USINAS = {
    "rodolfo oliveira":  ["Boa Esperança do Sul I", "Boa Esperança do Sul II", "Ibaté I", "Ibaté II"],
    "andrick gouveia":   ["Boa Esperança do Sul I", "Boa Esperança do Sul II", "Ibaté I", "Ibaté II"],
    "equipe piracicaba": ["Elias Fausto"],
    "deivity saugo":     ["Colíder I", "Colíder II"],
    "deivity jhon cunha saugo": ["Colíder I", "Colíder II"],
    "railson gomes":     ["Crateús"],
    "valmir junior":     ["Nobres"],
    "lucas lima":        ["Nobres"],
    "gabriel oliveira":  ["Nova Xavantina I", "Nova Xavantina II"],
    "eduardo souza":     ["Matão I", "Matão II - Topázio"],
    "aniel rocha":       ["Araputanga", "Poconé"],
    "adriano moraes":    ["Araputanga", "Poconé"],
    "claudio ferreira":  ["Sítio Bonfim", "ABC Morada Nova", "Sol do Norte I", "Sol do Norte II", "Guajirú", "Hortina (Quixadá I)", "Vitesse (Quixadá II)"],
    "cláudio ferreira":  ["Sítio Bonfim", "ABC Morada Nova", "Sol do Norte I", "Sol do Norte II", "Guajirú", "Hortina (Quixadá I)", "Vitesse (Quixadá II)"],
    "isake costa":       ["Sítio Bonfim", "ABC Morada Nova", "Sol do Norte I", "Sol do Norte II", "Guajirú", "Hortina (Quixadá I)", "Vitesse (Quixadá II)"],
    "daniel de paula":   ["Sete Lagoas"],
    "adriano silva":     ["Solier (Cascavel)"],
}


def _normalizar_tecnico(nome):
    return _norm_usina(nome)  # mesma normalização (sem acento, minúsculo) já usada pra usina


def _extrair_nome_usina_fracttal(texto):
    """
    A Fracttal nomeia o campo groups_1_description no formato
    "Cliente - Nome da Usina - UF" (ex: "Thopen - Boa Esperança do Sul 1 - SP").
    Extrai só a parte do meio (nome da usina) pra comparar com o catálogo,
    removendo o prefixo de cliente e o sufixo de UF quando presentes.
    """
    if not texto:
        return ""
    partes = [p.strip() for p in texto.split(" - ") if p.strip()]
    if len(partes) >= 3:
        return " - ".join(partes[1:-1])
    if len(partes) == 2:
        return partes[1]
    return texto


_STATUS_TAREFA_MAP = {
    "NO_STARTED": "Não Iniciada",
    "IN_PROGRESS": "Em Progresso",
    "PAUSED": "Pausada",
    "DONE": "Concluída",
}


def _fracttal_status_tarefa_label(task_status_raw):
    return _STATUS_TAREFA_MAP.get((task_status_raw or "").strip().upper(), (task_status_raw or "").strip())


def _fracttal_agrupar_por_wo(ots):
    """Agrupa uma lista de linhas (uma por tarefa) da Fracttal pelo wo_folio (a mesma OS)."""
    grupos, ordem = {}, []
    for ot in ots:
        folio = (ot.get("wo_folio") or "").strip()
        if not folio:
            continue
        if folio not in grupos:
            grupos[folio] = []
            ordem.append(folio)
        grupos[folio].append(ot)
    return [(folio, grupos[folio]) for folio in ordem]


_PREVENTIVA_PERIODICIDADE_MAP = {
    "semestral": ("PREVENTIVA SEMESTRAL", "Múltiplos equipamentos (Preventiva Semestral)"),
    "anual": ("PREVENTIVA ANUAL", "Múltiplos equipamentos (Preventiva Anual)"),
    "mensal": ("PREVENTIVA MENSAL", "Múltiplos equipamentos (Preventiva Mensal)"),
}


def _fracttal_detectar_preventiva(tasks, texto_grupo_ativo=""):
    """Detecta se uma OS com múltiplas tarefas é uma manutenção preventiva
    periódica (MPM/MPS/MPA) — usa nomenclatura padronizada ("PREVENTIVA
    MENSAL/SEMESTRAL/ANUAL" e "Múltiplos equipamentos (Preventiva X)") em
    vez de listar cada tarefa individualmente. Detecção por palavra-chave
    OU pela sigla (MPM/MPS/MPA) nas descrições das tarefas — a Fracttal às
    vezes usa só a sigla (ex.: "[Grid Co.] - MPA") sem escrever "preventiva
    anual" por extenso. "Semestral"/"anual" são checados antes de "mensal"
    pra evitar falso-positivo (ex.: um texto que cite os dois por algum
    motivo).

    Retorna (titulo, equipamento) ou (None, None) se não for preventiva
    periódica reconhecida.
    """
    textos = [(t.get("description") or "") for t in tasks]
    textos.append(texto_grupo_ativo or "")
    junto = " ".join(textos).lower()

    if re.search(r"\bmps\b", junto):
        return _PREVENTIVA_PERIODICIDADE_MAP["semestral"]
    if re.search(r"\bmpa\b", junto):
        return _PREVENTIVA_PERIODICIDADE_MAP["anual"]
    if re.search(r"\bmpm\b", junto):
        return _PREVENTIVA_PERIODICIDADE_MAP["mensal"]

    if "preventiv" not in junto:
        return None, None
    if "semestral" in junto:
        return _PREVENTIVA_PERIODICIDADE_MAP["semestral"]
    if "anual" in junto:
        return _PREVENTIVA_PERIODICIDADE_MAP["anual"]
    if "mensal" in junto:
        return _PREVENTIVA_PERIODICIDADE_MAP["mensal"]
    return None, None


def _fracttal_eh_preventiva_mensal(tasks, texto_grupo_ativo=""):
    """Mantido por compatibilidade: só a variante mensal."""
    titulo, _ = _fracttal_detectar_preventiva(tasks, texto_grupo_ativo)
    return titulo == "PREVENTIVA MENSAL"


def _fracttal_descricao_agregada(tasks):
    descs = [(t.get("description") or "").strip() for t in tasks if (t.get("description") or "").strip()]
    if not descs:
        return ""
    if len(tasks) == 1:
        return descs[0]
    m = re.match(r"^(\[[^\]]*\]\s*-\s*[^-]+)\s*-", descs[0])
    if m:
        return m.group(1).strip()
    return descs[0]


def _fracttal_prazo_agregado(tasks):
    """Prazo mais próximo entre todas as tarefas da OS (a que vence primeiro)."""
    datas = []
    for t in tasks:
        bruta = t.get("final_date") or t.get("date_maintenance") or t.get("cal_date_maintenance") or t.get("initial_date")
        if bruta:
            datas.append(str(bruta))
    if not datas:
        return ""
    datas.sort()
    return _fracttal_formatar_data_br(datas[0])


def _fracttal_status_tarefa_agregado(tasks):
    contagem = {}
    for t in tasks:
        label = _fracttal_status_tarefa_label(t.get("task_status"))
        if label:
            contagem[label] = contagem.get(label, 0) + 1
    if not contagem:
        return ""
    if len(tasks) == 1:
        return next(iter(contagem))
    return " | ".join(f"{qtd} {label}" for label, qtd in contagem.items())


def _fracttal_etiquetas_agregadas(tasks):
    vistas, nomes = set(), []
    for t in tasks:
        for lbl in (t.get("labels") or []):
            desc = (lbl.get("description") or "").strip()
            if desc and desc not in vistas:
                vistas.add(desc)
                nomes.append(desc)
    return ", ".join(nomes)


def _fracttal_observacoes_agregadas(tasks):
    vistas, notas = set(), []
    for t in tasks:
        nota = (t.get("task_note") or t.get("note") or "").strip()
        if nota and nota not in vistas:
            vistas.add(nota)
            notas.append(nota)
    return "\n---\n".join(notas)


def _fracttal_historico_detalhe(tasks):
    """Detalhamento por equipamento — só gerado quando a OS tem mais de uma tarefa."""
    if len(tasks) <= 1:
        return ""
    linhas = [f"⚙️ OS com {len(tasks)} itens/equipamentos — detalhamento:"]
    for t in tasks:
        eq = (t.get("items_log_description") or t.get("code") or "?").split("{")[0].strip()
        status = _fracttal_status_tarefa_label(t.get("task_status"))
        prazo = _fracttal_formatar_data_br(t.get("final_date") or t.get("date_maintenance") or "")
        linha = f"• {eq} — {status}"
        if prazo:
            linha += f" (prazo {prazo})"
        linhas.append(linha)
    return "\n".join(linhas)


def _fracttal_percentual_conclusao(tasks):
    total = len(tasks)
    if not total:
        return 0
    valores = []
    for t in tasks:
        cp = t.get("completed_percentage")
        if cp is not None:
            try:
                valores.append(float(cp))
                continue
            except (TypeError, ValueError):
                pass
        valores.append(100.0 if (t.get("task_status") or "").strip().upper() == "DONE" else 0.0)
    return round(sum(valores) / total)


def _fracttal_status_geral(tasks):
    """
    Status agregado da OS inteira em uma das 4 categorias que a Fracttal usa
    na Vista Kanban: Não Iniciada, Em Progresso, Pausada, Concluída.
    """
    total = len(tasks)
    if not total:
        return ""
    concluidas = sum(1 for t in tasks if (t.get("task_status") or "").strip().upper() == "DONE")
    em_progresso = sum(1 for t in tasks if (t.get("task_status") or "").strip().upper() == "IN_PROGRESS")
    pausadas = sum(1 for t in tasks if (t.get("task_status") or "").strip().upper() == "PAUSED")
    if concluidas == total:
        return "Concluída"
    if em_progresso > 0 or concluidas > 0:
        return "Em Progresso"
    if pausadas > 0:
        return "Pausada"
    return "Não Iniciada"


def _fracttal_detalhes_equipamentos(tasks):
    """
    Lista estruturada (JSON) de cada equipamento/tarefa da OS — usada pelo
    drawer/card pra montar uma tabela organizada em vez de só um texto no
    histórico. Cada item: {equipamento, status, prazo}.
    """
    itens = []
    for t in tasks:
        eq = (t.get("items_log_description") or t.get("code") or "?").split("{")[0].strip()
        status = _fracttal_status_tarefa_label(t.get("task_status"))
        prazo = _fracttal_formatar_data_br(t.get("final_date") or t.get("date_maintenance") or "")
        itens.append({"equipamento": eq, "status": status, "prazo": prazo})
    return json.dumps(itens, ensure_ascii=False)


def _fracttal_resolver_usina_tecnico(representante):
    """
    Resolve usina + cliente + alertas a partir de UMA tarefa representante
    da Fracttal (groups_1_description, ativo, técnico). Extraído de
    _fracttal_mapear_grupo em 03/09/2026 pra ser reutilizável também na
    revalidação periódica (_fracttal_verificar_e_atualizar_uma_os) — antes
    disso, usina/cliente só eram decididos na CRIAÇÃO da atividade e nunca
    mais revistos, então uma classificação errada na hora da criação (ex.:
    OS 12923/12925, criadas antes do fix de cruzamento cliente x usina
    entrar no ar) ficava presa errada pra sempre, mesmo com o bug já
    corrigido — mesma classe de problema já resolvida pro campo
    "responsavel".

      1. Nome do ativo bate com o catálogo E técnico é esperado nessa usina
         → segue normal, sem alerta.
      2. Nome do ativo bate, mas o técnico não é dos que atendem essa usina
         → usa mesmo assim (nome do ativo é a fonte mais confiável), mas
           com alerta.
      3. Nome do ativo NÃO bate, mas o técnico atende só 1 usina do catálogo
         → usa a usina do técnico como fallback, com alerta.
      4. Nome do ativo não bate e o técnico atende mais de uma usina (ou é
         desconhecido) → não dá pra decidir sozinho (usina=None, motivo
         preenchido).

    Retorna: {"usina": str|None, "cliente": str|None, "alerta": str|None,
              "motivo_revisao": str|None, "texto_usado": str, "tecnico_raw": str}
    """
    texto_grupo = _extrair_nome_usina_fracttal(representante.get("groups_1_description") or "")
    texto_ativo = representante.get("items_log_description") or representante.get("parent_description") or representante.get("item_code") or ""

    usina_por_ativo = canonizar_usina(texto_grupo) or canonizar_usina(texto_ativo)
    texto_usado = texto_grupo or texto_ativo

    # Cruzamento cliente x usina (adicionado 03/09/2026, causa raiz da OS
    # 12908): a Fracttal manda groups_1_description no formato "Cliente -
    # Nome da Usina - UF", mas até aqui só a parte do meio era usada —
    # o cliente informado era jogado fora. Isso permite falso-positivo
    # quando duas usinas de CLIENTES DIFERENTES têm o mesmo nome de
    # cidade (ex.: "Sal Energia - Cascavel - CE" é a SunPower real, mas
    # "Solier - Cascavel - CE", do cliente Qair — usina emprestada
    # temporariamente do supervisor Iago —, bateu no alias "cascavel" e
    # foi classificada como SunPower por engano). Se o cliente que a
    # Fracttal informou não bate (nem por substring) com o cliente
    # cadastrado pra usina que deu match por nome, tenta resolver contra
    # o catálogo de usinas emprestadas (_SupervisaoTemporaria) antes de
    # desistir — só então cai no fallback por técnico ou revisão manual.
    _motivo_conflito_cliente = None
    if usina_por_ativo:
        _partes_grupo_raw = [p.strip() for p in (representante.get("groups_1_description") or "").split(" - ") if p.strip()]
        _cliente_fracttal_raw = _partes_grupo_raw[0] if len(_partes_grupo_raw) >= 2 else ""
        if _cliente_fracttal_raw:
            _cliente_esperado = inferir_cliente(usina_por_ativo)
            _cf_norm = _norm_usina(_cliente_fracttal_raw)
            _ce_norm = _norm_usina(_cliente_esperado)
            if _ce_norm and _cf_norm not in _ce_norm and _ce_norm not in _cf_norm:
                # Antes de desistir, tenta resolver contra usinas emprestadas
                # temporariamente (_SupervisaoTemporaria). Padrão atípico visto
                # no Qair/Solier: a Fracttal desse cliente NÃO manda "Qair" na
                # posição de cliente do groups_1_description — manda o
                # codinome da própria usina ("Solier - Cascavel - CE"). Por
                # isso aqui comparamos o texto contra o nome_oficial INTEIRO
                # da usina emprestada (que pode conter tanto o codinome quanto
                # a cidade), não só contra o campo "cliente" cadastrado.
                _resolvido_temp = None
                for _item_temp in _usinas_temporarias():
                    _nome_temp_norm = _norm_usina(_item_temp["usina"])
                    _bateu_nome = ((_norm_usina(texto_grupo) and _norm_usina(texto_grupo) in _nome_temp_norm) or
                                   (_norm_usina(texto_ativo) and _norm_usina(texto_ativo) in _nome_temp_norm))
                    _bateu_cliente = (_cf_norm in _nome_temp_norm or _cf_norm in _norm_usina(_item_temp["cliente"]))
                    if _bateu_nome and _bateu_cliente:
                        _resolvido_temp = _item_temp["usina"]
                        break
                if _resolvido_temp:
                    usina_por_ativo = _resolvido_temp
                else:
                    _motivo_conflito_cliente = (f"nome de usina bateu com \"{usina_por_ativo}\" mas o cliente informado "
                                                 f"pela Fracttal (\"{_cliente_fracttal_raw}\") não é \"{_cliente_esperado}\" "
                                                 f"— provável cliente novo/fora do catálogo com usina de nome parecido "
                                                 f"ou mesma cidade")
                    usina_por_ativo = None

    tecnico_raw = (representante.get("personnel_description") or representante.get("responsible") or representante.get("created_by") or "").strip()
    tecnico_norm = _normalizar_tecnico(tecnico_raw)
    usinas_do_tecnico = TECNICO_USINAS.get(tecnico_norm, [])

    usina = None
    alerta = None
    motivo_revisao = None

    if usina_por_ativo:
        usina = usina_por_ativo
        if usinas_do_tecnico and usina not in usinas_do_tecnico:
            alerta = (f"⚠️ Cruzamento: técnico \"{tecnico_raw}\" não está mapeado para {usina} "
                      f"(usinas esperadas dele: {', '.join(usinas_do_tecnico)}). Confira se a usina está certa.")
    elif len(usinas_do_tecnico) == 1 and not texto_usado.strip():
        # Só confia no fallback "técnico só atende 1 usina do catálogo" quando
        # a Fracttal NÃO informou nome de grupo/ativo nenhum (texto_usado
        # vazio). Se veio um nome e ele só não bateu com o catálogo (ex:
        # "Castelo do Piauí 1" pro técnico Railson Gomes, que também atende
        # a GreenYellow fora do catálogo do Fred), é sinal forte de outro
        # site/cliente — não é o mesmo caso de "Fracttal não informou nada".
        # Isso já causou um erro grave: OS da GreenYellow/Castelo do Piauí
        # entrando no painel como Crateús/Renogrid (24/07/2026). Vai pra
        # revisão manual em vez de assumir.
        usina = usinas_do_tecnico[0]
        alerta = (f"⚠️ Usina inferida pelo técnico responsável (\"{tecnico_raw}\"), pois a Fracttal não "
                  f"informou grupo nem ativo. Confira se está correto.")
    else:
        if _motivo_conflito_cliente:
            motivo_revisao = f"{_motivo_conflito_cliente} (técnico: \"{tecnico_raw or '—'}\")."
        elif usinas_do_tecnico:
            motivo_revisao = (f"Grupo/ativo (\"{texto_usado}\") não reconhecido e técnico \"{tecnico_raw}\" atende mais de "
                               f"uma usina ({', '.join(usinas_do_tecnico)}) — não dá pra decidir sozinho.")
        elif tecnico_raw:
            motivo_revisao = f"Grupo/ativo (\"{texto_usado}\") não reconhecido e técnico \"{tecnico_raw}\" não está no mapa de usinas."
        else:
            motivo_revisao = f"Grupo/ativo (\"{texto_usado}\") não reconhecido e OT sem técnico responsável informado."

    cliente = inferir_cliente(usina) if usina else None
    return {"usina": usina, "cliente": cliente, "alerta": alerta, "motivo_revisao": motivo_revisao,
            "texto_usado": texto_usado, "tecnico_raw": tecnico_raw, "texto_ativo": texto_ativo}


def _fracttal_mapear_grupo(tasks):
    """
    Converte um GRUPO de tarefas (todas da mesma OS, mesmo wo_folio) para
    os campos do Painel de Atividades — uma OS vira UMA atividade, mesmo
    quando tem várias tarefas/equipamentos (ex: preventivas mensais/anuais
    com dezenas de itens). Detalhamento por equipamento vai pro Histórico.

    O cruzamento usina x técnico responsável usa a primeira tarefa como
    representante (grupo/usina e técnico geralmente são os mesmos pra
    todas as tarefas de uma mesma OS). A resolução em si mora em
    _fracttal_resolver_usina_tecnico (compartilhada com a revalidação).
    """
    representante = tasks[0]
    _res = _fracttal_resolver_usina_tecnico(representante)
    usina, cliente, alerta = _res["usina"], _res["cliente"], _res["alerta"]
    texto_usado, tecnico_raw, texto_ativo = _res["texto_usado"], _res["tecnico_raw"], _res["texto_ativo"]

    if not usina:
        return {"_revisao_manual": True, "motivo": _res["motivo_revisao"], "wo_folio": representante.get("wo_folio", "?")}

    prioridade_raw = (representante.get("priorities_description") or "").strip().upper()
    prioridade = _FRACTTAL_PRIORIDADE_MAP.get(prioridade_raw, "Média")

    status_os_raw = str(representante.get("id_status_work_order", "")).strip()
    status_os = _FRACTTAL_STATUS_OS_MAP.get(status_os_raw, "")

    multiplos = len(tasks) > 1
    _titulo_prev, _equip_prev = _fracttal_detectar_preventiva(tasks, texto_usado) if multiplos else (None, None)
    if _equip_prev:
        equipamento = _equip_prev
    else:
        equipamento = "Múltiplas atividades" if multiplos else (representante.get("code") or texto_ativo or "Múltiplas atividades").strip()

    detalhe_hist = _fracttal_historico_detalhe(tasks)
    if alerta and detalhe_hist:
        alerta = f"{alerta}\n{detalhe_hist}"
    elif detalhe_hist:
        alerta = detalhe_hist

    return {
        "cliente": cliente,
        "usina": usina,
        "equipamento": equipamento,
        "descricao": (_titulo_prev if _titulo_prev
                       else (_fracttal_descricao_agregada(tasks) or f"OT {representante.get('wo_folio', '')} (Fracttal)")),
        "responsavel": tecnico_raw,
        "prazo": _fracttal_prazo_agregado(tasks),
        "prioridade": prioridade,
        # se a OS já nasce Finalizada/Cancelada na Fracttal (ex.: criada e
        # cancelada pelo operador antes do nosso sync sequer vê-la), o
        # status interno tem que refletir isso já na criação — senão essa
        # atividade nunca mais é revisitada pelo rodízio (que pula quem já
        # está Finalizada/Cancelada) e fica presa em "Em Aberto" pra sempre.
        "status": (_status_interno_esperado(status_os, "Em Aberto") or "Em Aberto"),
        "numeroOS": (representante.get("wo_folio") or "").strip(),
        "editor": "fracttal-sync",
        "statusOS": status_os,
        "observacoesOS": _fracttal_observacoes_agregadas(tasks),
        "linkOS": _fracttal_montar_link(representante),
        "statusTarefaOS": _fracttal_status_tarefa_agregado(tasks),
        "etiquetasOS": _fracttal_etiquetas_agregadas(tasks),
        "percentualOS": str(_fracttal_percentual_conclusao(tasks)),
        "statusGeralOS": _fracttal_status_geral(tasks),
        "detalhesEquipamentosOS": _fracttal_detalhes_equipamentos(tasks),
        "_alerta": alerta,
    }


@app.route("/fracttal-raw", methods=["GET"])
def fracttal_raw():
    """
    Endpoint de DIAGNÓSTICO — repassa query params direto pro /work_orders
    da Fracttal e devolve a resposta crua (sem mapear). Usado só pra
    confirmar nomes de parâmetros (status, paginação) antes de rodar
    sincronizações em lote. Protegido pelo mesmo WEBHOOK_SECRET.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    try:
        token = _fracttal_get_token()
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        endpoint = request.args.get("endpoint", "").strip()
        folio = request.args.get("folio", "").strip()
        if endpoint:
            url = f"{FRACTTAL_API_BASE}/{endpoint.lstrip('/')}"
            params = {k: v for k, v in request.args.items() if k not in ("secret", "endpoint")}
        elif folio:
            url = f"{FRACTTAL_API_BASE}/work_orders/{folio}"
            params = {}
        else:
            url = f"{FRACTTAL_API_BASE}/work_orders"
            params = {k: v for k, v in request.args.items() if k != "secret"}
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        try:
            body = resp.json()
        except Exception:
            body = resp.text[:3000]
        return jsonify({"ok": True, "status_code": resp.status_code, "url_chamada": resp.url, "body": body})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _fracttal_datahora_br(iso_str):
    """Converte um timestamp ISO da Fracttal (UTC) pra dd/mm/aaaa HH:MM no fuso de Brasília."""
    if not iso_str:
        return None
    try:
        s = str(iso_str).strip()
        # normaliza variações tipo "...81+00:00" / "...Z" pro formato aceito pelo fromisoformat
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt_br = dt.astimezone(_TZ_BR)
        return dt_br.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return None


@app.route("/os-historico-fracttal/<numero_os>", methods=["GET"])
def os_historico_fracttal(numero_os):
    """
    Monta uma linha do tempo da OS usando dados REAIS e ao vivo da Fracttal
    (não o histórico interno construído pelo nosso próprio sync). A API
    pública da Fracttal não expõe o log de eventos completo que aparece na
    tela "Histórico" dentro do app deles (anexos, cada troca de status) —
    isso é um recurso interno da UI, sem endpoint documentado. O que a API
    entrega, por tarefa/equipamento da OS, são os timestamps reais de
    criação, início e conclusão — o suficiente pra responder com precisão
    "quando essa OS foi iniciada", que é o que interessa aqui.
    """
    numero_os = (numero_os or "").strip()
    if not numero_os:
        return jsonify({"ok": False, "error": "numero_os vazio"}), 400
    try:
        token = _fracttal_get_token()
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders/{numero_os}", headers=headers, timeout=20)
        resp.raise_for_status()
        body = resp.json()
        tasks = body.get("data") or []
        if isinstance(tasks, dict):
            tasks = [tasks]

        eventos = []
        for t in tasks:
            equip = (t.get("items_log_description") or t.get("parent_description") or "").strip() or "—"
            desc_tarefa = (t.get("description") or "").strip() or "—"
            tecnico = (t.get("personnel_description") or "").strip() or "—"
            criado_por = (t.get("created_by") or "").strip() or "—"

            dt_criacao = _fracttal_datahora_br(t.get("creation_date"))
            if dt_criacao:
                eventos.append({
                    "data": t.get("creation_date"), "dataFormatada": dt_criacao,
                    "titulo": "Ordem de trabalho criada",
                    "detalhe": f"{desc_tarefa} — {equip}", "autor": criado_por,
                })

            dt_inicio = _fracttal_datahora_br(t.get("initial_date"))
            if dt_inicio:
                eventos.append({
                    "data": t.get("initial_date"), "dataFormatada": dt_inicio,
                    "titulo": "Tarefa iniciada",
                    "detalhe": f"{desc_tarefa} — {equip}", "autor": tecnico,
                })

            dt_fim = _fracttal_datahora_br(t.get("final_date"))
            if dt_fim:
                eventos.append({
                    "data": t.get("final_date"), "dataFormatada": dt_fim,
                    "titulo": "Tarefa concluída" if t.get("done") else "Tarefa finalizada",
                    "detalhe": f"{desc_tarefa} — {equip}", "autor": tecnico,
                })

        eventos.sort(key=lambda e: e["data"] or "")
        # mais recente primeiro, igual ao padrão de histórico já usado no dashboard
        eventos.reverse()

        return jsonify({
            "ok": True,
            "numeroOS": numero_os,
            "totalTarefas": len(tasks),
            "eventos": eventos,
        })
    except requests.exceptions.HTTPError as e:
        status = e.response.status_code if e.response is not None else 502
        return jsonify({"ok": False, "error": f"Fracttal respondeu {status} pra OS {numero_os}"}), 502
    except Exception as e:
        log.error(f"[Historico Fracttal] Erro na OS {numero_os}: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/backfill-fracttal", methods=["POST", "GET"])
def backfill_fracttal():
    """
    Backfill histórico de OTs da Fracttal pro Painel de Atividades.
    Processa UMA PÁGINA por chamada (start/limit) — pensado pra ser chamado
    repetidas vezes pelo workflow do GitHub Actions, avançando o `start`,
    pra não estourar o timeout de 60s do Render numa carga grande.

    Query params:
      since      (default 2026-03-01T00:00:00-00:00)
      until      (opcional)
      ot_status  (default 1 = Em Processo)
      start      (default 0)
      limit      (default 100, teto de 100 — limite da própria Fracttal)
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    since = request.args.get("since", "2026-03-01T00:00:00-00:00")
    until = request.args.get("until", "") or None
    ot_status = request.args.get("ot_status", "1")
    start = int(request.args.get("start", 0))
    limit = min(int(request.args.get("limit", 100)), 100)

    try:
        ots, total = _fracttal_listar_pagina(since=since, until=until, ot_status=ot_status,
                                              start=start, limit=limit)
    except Exception as e:
        log.error(f"[backfill-fracttal] Erro ao consultar Fracttal (start={start}): {e}")
        return jsonify({"ok": False, "error": str(e), "start": start}), 502

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        os_existentes = {row[13].strip() for row in todos[1:] if len(row) > 13 and row[13].strip()}
    except Exception as e:
        log.error(f"[backfill-fracttal] Erro ao ler Painel de Atividades: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

    criadas, revisao_manual, erros = [], [], []
    revisao_folios_vistos = set()

    for folio, tasks in _fracttal_agrupar_por_wo(ots):
        if folio in os_existentes:
            continue  # já registrada (nesta página ou em página anterior do mesmo backfill)
        mapeado = _fracttal_mapear_grupo(tasks)
        if not mapeado:
            continue
        if mapeado.get("_revisao_manual"):
            if folio not in revisao_folios_vistos:
                revisao_folios_vistos.add(folio)
                revisao_manual.append({"wo_folio": folio, "motivo": mapeado["motivo"]})
            continue

        alerta = mapeado.pop("_alerta", None)
        mapeado["editor"] = "fracttal-backfill"
        try:
            novo_id = _criar_atividade_interna(ws=ws, todos=todos, enviar_notificacao=False, **mapeado)
            if alerta:
                _aplicar_update_campo_atividade(ws, len(todos), todos[-1], "historico", alerta,
                                                 "fracttal-backfill", append=True)
            criadas.append({"numeroOS": mapeado["numeroOS"], "id": novo_id, "itens": len(tasks)})
            os_existentes.add(mapeado["numeroOS"])
        except Exception as e:
            log.error(f"[backfill-fracttal] Erro ao criar atividade para OT {mapeado.get('numeroOS')}: {e}")
            erros.append(mapeado.get("numeroOS", "?"))

    proximo_start = start + limit
    log.info(f"[backfill-fracttal] start={start} total_geral={total} criadas={len(criadas)} "
             f"revisao_manual={len(revisao_manual)} erros={len(erros)}")
    return jsonify({
        "ok": True,
        "total_geral": total,
        "start": start,
        "limit": limit,
        "processados_nesta_pagina": len(ots),
        "proximo_start": proximo_start,
        "tem_mais": proximo_start < total,
        "criadas": criadas,
        "revisao_manual": revisao_manual,
        "erros": erros,
    })


@app.route("/completar-fracttal-backfill", methods=["POST", "GET"])
def completar_fracttal_backfill():
    """
    Complementa atividades já criadas (tipicamente pelo /backfill-fracttal
    antes da introdução dos campos statusOS/observacoesOS/linkOS/prazo
    corrigido) buscando cada OT individualmente na Fracttal por wo_folio
    (GET /work_orders/{folio}) e preenchendo os campos que faltam.

    Só mexe em linhas com editor == "fracttal-backfill" (ou "fracttal-sync")
    E que ainda não têm statusOS preenchido — não sobrescreve nada que já
    foi completado ou editado manualmente.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    limit = int(request.args.get("limit", 8))

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        _garantir_headers_atividades(ws)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    token = _fracttal_get_token()
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    atualizadas, erros, puladas = [], [], []
    processadas = 0

    for i, row in enumerate(todos[1:], start=2):
        if processadas >= limit:
            break
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        editor = row[12].strip()
        numero_os = row[13].strip()
        status_os_atual = row[14].strip()

        if editor not in ("fracttal-backfill", "fracttal-sync") or not numero_os or status_os_atual:
            continue

        processadas += 1
        try:
            resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders/{numero_os}", headers=headers, timeout=20)
            resp.raise_for_status()
            body = resp.json()
            tasks = (body.get("data") or [])
            if not tasks:
                puladas.append({"numeroOS": numero_os, "motivo": "OT não encontrada na Fracttal"})
                continue

            representante = tasks[0]
            status_os_raw = str(representante.get("id_status_work_order", "")).strip()
            status_os = _FRACTTAL_STATUS_OS_MAP.get(status_os_raw, "")
            observacoes = _fracttal_observacoes_agregadas(tasks)
            link = _fracttal_montar_link(representante)
            prazo_novo = _fracttal_prazo_agregado(tasks)
            status_tarefa = _fracttal_status_tarefa_agregado(tasks)
            etiquetas = _fracttal_etiquetas_agregadas(tasks)

            # escreve os campos novos numa única chamada (evita estourar cota de escrita)
            ws.update(f"O{i}:S{i}", [[status_os, observacoes, link, status_tarefa, etiquetas]])
            prazo_atual = row[6].strip()
            if prazo_novo and prazo_novo != prazo_atual:
                ws.update_cell(i, 7, prazo_novo)
            if len(tasks) > 1:
                _, _equip_prev_check = _fracttal_detectar_preventiva(tasks)
                _equip_esperado = _equip_prev_check or "Múltiplas atividades"
            else:
                _equip_esperado = None
            if _equip_esperado and row[3].strip() != _equip_esperado:
                ws.update_cell(i, 4, _equip_esperado)
                detalhe = _fracttal_historico_detalhe(tasks)
                if detalhe:
                    hist_atual = row[ATIV_COL_HISTORICO - 1] if len(row) >= ATIV_COL_HISTORICO else ""
                    ws.update_cell(i, ATIV_COL_HISTORICO, f"{hist_atual}\n{detalhe}".strip() if hist_atual else detalhe)

            atualizadas.append({"linha": i, "numeroOS": numero_os, "statusOS": status_os, "itens": len(tasks)})
            time.sleep(1.2)  # respeita a cota de escrita por minuto do Google Sheets
        except Exception as e:
            log.error(f"[completar-fracttal-backfill] Erro na OT {numero_os}: {e}")
            erros.append({"numeroOS": numero_os, "erro": str(e)})

    log.info(f"[completar-fracttal-backfill] atualizadas={len(atualizadas)} puladas={len(puladas)} erros={len(erros)}")
    return jsonify({"ok": True, "atualizadas": atualizadas, "puladas": puladas, "erros": erros,
                     "processadas_nesta_chamada": processadas, "limit": limit})


@app.route("/corrigir-descricoes-multiplas", methods=["POST", "GET"])
def corrigir_descricoes_multiplas():
    """
    Corrige atividades multi-tarefa antigas (criadas antes da mudança de
    texto "Múltiplos equipamentos" -> "Múltiplas atividades" e do corte da
    descrição pro prefixo só): reconsulta a OS na Fracttal e reescreve
    equipamento + descrição no padrão atual.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    limit = int(request.args.get("limit", 8))

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    token = _fracttal_get_token()
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    corrigidas, erros = [], []
    processadas = 0

    for i, row in enumerate(todos[1:], start=2):
        if processadas >= limit:
            break
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        equipamento_atual = row[3].strip()
        descricao_atual = row[4].strip()
        numero_os = row[13].strip()

        precisa = (equipamento_atual == "Múltiplos equipamentos") or (
            equipamento_atual == "Múltiplas atividades" and re.search(r"Múltiplas atividades \(\d+ itens\)", descricao_atual)
        )
        if not precisa or not numero_os:
            continue

        processadas += 1
        try:
            resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders/{numero_os}", headers=headers, timeout=20)
            resp.raise_for_status()
            tasks = (resp.json().get("data") or [])
            if not tasks:
                continue
            nova_descricao = _fracttal_descricao_agregada(tasks)
            ws.update(f"D{i}:E{i}", [["Múltiplas atividades", nova_descricao]])
            corrigidas.append({"linha": i, "numeroOS": numero_os, "descricao": nova_descricao})
            time.sleep(1.0)
        except Exception as e:
            log.error(f"[corrigir-descricoes-multiplas] Erro na OT {numero_os}: {e}")
            erros.append({"numeroOS": numero_os, "erro": str(e)})

    log.info(f"[corrigir-descricoes-multiplas] corrigidas={len(corrigidas)} erros={len(erros)}")
    return jsonify({"ok": True, "corrigidas": corrigidas, "erros": erros,
                     "processadas_nesta_chamada": processadas, "limit": limit})


@app.route("/normalizar-usinas-clientes", methods=["POST", "GET"])
def normalizar_usinas_clientes():
    """
    Varre TODAS as atividades (independente de origem/editor) e corrige
    usina/cliente pra forma canônica do catálogo sempre que divergirem —
    resolve nomenclaturas duplicadas (ex: "Sete Lagoas 2" vs "Sete Lagoas")
    que escaparam da canonização por terem sido criadas manualmente antes
    da integração Fracttal existir.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    corrigidas = []
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        usina_atual = row[2].strip()
        cliente_atual = row[1].strip()
        if not usina_atual or usina_atual in ("Geral", "Administrativo"):
            continue

        canonica = canonizar_usina(usina_atual)
        if not canonica:
            continue  # não reconhecida — não mexe (pode ser usina legítima fora do catálogo atual)

        cliente_correto = inferir_cliente(canonica) or cliente_atual
        if canonica != usina_atual or cliente_correto != cliente_atual:
            ws.update(f"B{i}:C{i}", [[cliente_correto, canonica]])
            corrigidas.append({"linha": i, "usina_antes": usina_atual, "usina_depois": canonica,
                                "cliente_antes": cliente_atual, "cliente_depois": cliente_correto})
            time.sleep(0.8)

    log.info(f"[normalizar-usinas-clientes] corrigidas={len(corrigidas)}")
    return jsonify({"ok": True, "corrigidas": corrigidas})


@app.route("/completar-campos-v3-fracttal", methods=["POST", "GET"])
def completar_campos_v3_fracttal():
    """
    Preenche percentualOS/statusGeralOS/detalhesEquipamentosOS em atividades
    da Fracttal que já foram completadas pelas versões anteriores (já têm
    statusTarefaOS) mas ainda não têm esses 3 campos mais novos.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    limit = int(request.args.get("limit", 6))

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        _garantir_headers_atividades(ws)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    token = _fracttal_get_token()
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    atualizadas, erros = [], []
    processadas = 0

    for i, row in enumerate(todos[1:], start=2):
        if processadas >= limit:
            break
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        editor = row[12].strip()
        numero_os = row[13].strip()
        percentual_atual = row[20].strip()

        if editor not in ("fracttal-backfill", "fracttal-sync", "claude-chat") or not numero_os:
            continue
        if percentual_atual:
            continue

        processadas += 1
        try:
            resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders/{numero_os}", headers=headers, timeout=20)
            resp.raise_for_status()
            tasks = (resp.json().get("data") or [])
            if not tasks:
                continue

            percentual = str(_fracttal_percentual_conclusao(tasks))
            status_geral = _fracttal_status_geral(tasks)
            detalhes = _fracttal_detalhes_equipamentos(tasks)
            ws.update(f"U{i}:W{i}", [[percentual, status_geral, detalhes]])

            atualizadas.append({"linha": i, "numeroOS": numero_os, "percentualOS": percentual,
                                 "statusGeralOS": status_geral})
            time.sleep(1.2)
        except Exception as e:
            log.error(f"[completar-campos-v3-fracttal] Erro na OT {numero_os}: {e}")
            erros.append({"numeroOS": numero_os, "erro": str(e)})

    log.info(f"[completar-campos-v3-fracttal] atualizadas={len(atualizadas)} erros={len(erros)}")
    return jsonify({"ok": True, "atualizadas": atualizadas, "erros": erros,
                     "processadas_nesta_chamada": processadas, "limit": limit})


@app.route("/completar-campos-v2-fracttal", methods=["POST", "GET"])
def completar_campos_v2_fracttal():
    """
    Reprocessa atividades da Fracttal que já têm statusOS preenchido (então
    o /completar-fracttal-backfill as ignora) mas ainda não têm os campos
    mais novos: statusTarefaOS, etiquetasOS, e a correção de "Múltiplos
    equipamentos" quando a OS tem mais de uma tarefa. Existe só pra
    completar o que ficou faltando em atividades criadas antes desses
    campos existirem.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    limit = int(request.args.get("limit", 6))

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        _garantir_headers_atividades(ws)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    token = _fracttal_get_token()
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    atualizadas, erros = [], []
    processadas = 0

    for i, row in enumerate(todos[1:], start=2):
        if processadas >= limit:
            break
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        editor = row[12].strip()
        numero_os = row[13].strip()
        status_os_atual = row[14].strip()
        status_tarefa_atual = row[17].strip()

        if editor not in ("fracttal-backfill", "fracttal-sync", "claude-chat") or not numero_os:
            continue
        if not status_os_atual or status_tarefa_atual:
            continue  # ou ainda não foi completado pelo v1, ou já tem os campos novos

        processadas += 1
        try:
            resp = requests.get(f"{FRACTTAL_API_BASE}/work_orders/{numero_os}", headers=headers, timeout=20)
            resp.raise_for_status()
            tasks = (resp.json().get("data") or [])
            if not tasks:
                continue

            status_tarefa = _fracttal_status_tarefa_agregado(tasks)
            etiquetas = _fracttal_etiquetas_agregadas(tasks)
            ws.update(f"R{i}:S{i}", [[status_tarefa, etiquetas]])

            if len(tasks) > 1:
                _, _equip_prev_check = _fracttal_detectar_preventiva(tasks)
                _equip_esperado = _equip_prev_check or "Múltiplas atividades"
            else:
                _equip_esperado = None
            if _equip_esperado and row[3].strip() != _equip_esperado:
                ws.update_cell(i, 4, _equip_esperado)
                detalhe = _fracttal_historico_detalhe(tasks)
                if detalhe:
                    hist_atual = row[ATIV_COL_HISTORICO - 1] if len(row) >= ATIV_COL_HISTORICO else ""
                    ws.update_cell(i, ATIV_COL_HISTORICO, f"{hist_atual}\n{detalhe}".strip() if hist_atual else detalhe)

            atualizadas.append({"linha": i, "numeroOS": numero_os, "itens": len(tasks),
                                 "statusTarefaOS": status_tarefa})
            time.sleep(1.2)
        except Exception as e:
            log.error(f"[completar-campos-v2-fracttal] Erro na OT {numero_os}: {e}")
            erros.append({"numeroOS": numero_os, "erro": str(e)})

    log.info(f"[completar-campos-v2-fracttal] atualizadas={len(atualizadas)} erros={len(erros)}")
    return jsonify({"ok": True, "atualizadas": atualizadas, "erros": erros,
                     "processadas_nesta_chamada": processadas, "limit": limit})


@app.route("/atualizar-links-fracttal", methods=["POST", "GET"])
def atualizar_links_fracttal():
    """
    Reprocessa SÓ o campo linkOS de atividades vindas da Fracttal, usando
    a URL pública por OT (add-on Share TOs). Ignora se já tem statusOS
    preenchido (diferente do /completar-fracttal-backfill) — existe
    justamente pra corrigir links antigos que ficaram com o fallback
    genérico (tela de OTs) em vez do link direto.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    limit = int(request.args.get("limit", 8))

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        _garantir_headers_atividades(ws)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    atualizadas, erros = [], []
    processadas = 0

    for i, row in enumerate(todos[1:], start=2):
        if processadas >= limit:
            break
        if len(row) < 17:
            row = row + [""] * (17 - len(row))
        editor = row[12].strip()
        numero_os = row[13].strip()
        link_atual = row[16].strip()

        if editor not in ("fracttal-backfill", "fracttal-sync") or not numero_os:
            continue
        if link_atual and link_atual != FRACTTAL_WEB_BASE:
            continue  # já tem link direto, não mexe

        processadas += 1
        try:
            link_novo = _fracttal_montar_link({"wo_folio": numero_os})
            if link_novo and link_novo != link_atual:
                ws.update_cell(i, 17, link_novo)
                atualizadas.append({"linha": i, "numeroOS": numero_os, "link": link_novo})
            time.sleep(1.2)
        except Exception as e:
            log.error(f"[atualizar-links-fracttal] Erro na OT {numero_os}: {e}")
            erros.append({"numeroOS": numero_os, "erro": str(e)})

    log.info(f"[atualizar-links-fracttal] atualizadas={len(atualizadas)} erros={len(erros)}")
    return jsonify({"ok": True, "atualizadas": atualizadas, "erros": erros,
                     "processadas_nesta_chamada": processadas, "limit": limit})


def _sync_fracttal_core(desde_horas=8):
    """
    Núcleo da DESCOBERTA: busca OTs recentes na Fracttal e cria atividades
    novas pra qualquer uma que ainda não esteja no Painel de Atividades.
    Só isso — não recheca status de OSs já existentes (isso é
    responsabilidade exclusiva da auditoria, _auditoria_consistencia_os_core,
    que faz esse trabalho de forma mais completa e sem duplicar lógica).

    Separar descoberta de "manter status em dia" é proposital: descobrir
    OS nova não é urgente (uma janela de horas é tranquila), enquanto
    manter o status atualizado precisa rodar com frequência — juntar os
    dois na mesma chamada só deixava tudo mais pesado e lento sem
    necessidade.
    """
    try:
        ots = _fracttal_listar_ots_recentes(desde_horas=desde_horas)
    except Exception as e:
        log.error(f"[sync-fracttal] Erro ao consultar Fracttal: {e}")
        return {"ok": False, "error": str(e)}, 502

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        _garantir_headers_atividades(ws)
        os_existentes = {row[13].strip() for row in todos[1:] if len(row) > 13 and row[13].strip()}
    except Exception as e:
        log.error(f"[sync-fracttal] Erro ao ler Painel de Atividades: {e}")
        return {"ok": False, "error": str(e)}, 500

    criadas, revisao_manual, erros = [], [], []
    revisao_folios_vistos = set()

    for folio, tasks in _fracttal_agrupar_por_wo(ots):
        if folio in os_existentes:
            continue  # já registrada — evita duplicata (mesma OT, outra linha de tarefa/componente)
        mapeado = _fracttal_mapear_grupo(tasks)
        if not mapeado:
            continue  # OT de outro cliente/supervisor, totalmente fora de escopo
        if mapeado.get("_revisao_manual"):
            if folio not in revisao_folios_vistos:
                revisao_folios_vistos.add(folio)
                revisao_manual.append({"wo_folio": folio, "motivo": mapeado["motivo"]})
            continue

        alerta = mapeado.pop("_alerta", None)
        try:
            novo_id = _criar_atividade_interna(ws=ws, todos=todos, enviar_notificacao=False, **mapeado)
            if alerta:
                _aplicar_update_campo_atividade(ws, len(todos), todos[-1], "historico", alerta,
                                                 "fracttal-sync", append=True)
            criadas.append({"numeroOS": mapeado["numeroOS"], "id": novo_id, "itens": len(tasks), "alerta": alerta,
                             "usina": mapeado["usina"], "cliente": mapeado["cliente"],
                             "descricao": mapeado.get("descricao", "")})
            os_existentes.add(mapeado["numeroOS"])
        except Exception as e:
            log.error(f"[sync-fracttal] Erro ao criar atividade para OT {mapeado.get('numeroOS')}: {e}")
            erros.append(mapeado.get("numeroOS", "?"))

    # um único push resumido pra todas as OSs novas descobertas nessa
    # rodada — restaurado 15/07/2026 (tinha sido perdido numa reconstrução
    # desta função por outra sessão, reintroduzindo notificação em
    # duplicidade/spam por item que já tinha sido corrigido antes).
    # 05/08/2026: mesma função teve o tema/descrição perdido de novo por
    # outra sobrescrita — reaplicado, agora mostrando tema (Ação/Tarefa)
    # em vez de só cliente/usina, sem o quê não dava pra saber do que a
    # OS se tratava sem abrir o painel.
    if criadas:
        try:
            if len(criadas) == 1:
                c = criadas[0]
                tema = (c.get("descricao") or "Descrição não informada").strip()
                enviar_push(
                    titulo=f"🆕 Nova OS Fracttal — {c['numeroOS']} — {c['usina']}",
                    corpo=f"{tema}\n{c['cliente']}",
                    tipo="fracttal_nova_os",
                    url=f"https://fred-alexandrino.github.io/PAINELDEFALHAS/?atividade={c['id']}",
                )
            else:
                def _linha_nova_os(c):
                    usina = (c.get("usina") or "Usina não informada").strip()
                    tema = (c.get("descricao") or "sem descrição").strip()
                    if len(tema) > 35:
                        tema = tema[:35].rstrip() + "…"
                    return f"{c['numeroOS']} · {usina} — {tema}"
                linhas = "\n".join(_linha_nova_os(c) for c in criadas[:6])
                enviar_push(
                    titulo=f"🆕 {len(criadas)} novas OSs na Fracttal",
                    corpo=f"{linhas}{chr(10) + '...' if len(criadas) > 6 else ''}",
                    tipo="fracttal_nova_os",
                    url="https://fred-alexandrino.github.io/PAINELDEFALHAS/",
                )
        except Exception as e:
            log.error(f"[sync-fracttal] Falha ao enviar push resumido de novas OSs: {e}")

    log.info(f"[sync-fracttal] criadas={len(criadas)} revisao_manual={len(revisao_manual)} erros={len(erros)}")
    return {"ok": True, "criadas": criadas, "revisao_manual": revisao_manual, "erros": erros}, 200



# ── Correção retroativa de fuso horário (uso único) ─────────────────────
# Antes do deploy que introduziu agora_br(), todo timestamp gravado no
# Painel de Atividades usava datetime.now() puro do servidor (UTC), mas era
# exibido como se já fosse horário de Brasília (GMT-3) — ficando 3h
# adiantado. Este endpoint corrige retroativamente as colunas afetadas
# (dataCriacao, dataConclusao, historico, ultimaVerificacaoOS) só nas
# entradas anteriores ao corte (momento em que o deploy corrigido entrou
# no ar). Entradas iguais/depois do corte já estão corretas e são
# ignoradas. Protegido por WEBHOOK_SECRET; sempre roda em modo simulação
# (dry_run=true) por padrão — precisa de ?apply=true explícito pra gravar.
_HOJE_DEPLOY = datetime(2026, 7, 8).date()
# O deploy da correção foi ao ar por volta de 21:05-21:09 UTC (= 18:05-18:09
# em Brasília, mesmo instante real). Uma entrada ANTIGA (com bug) grava o
# relógio UTC bruto como se fosse local — então seu valor literal só pode
# ir até, no máximo, o instante em que o código antigo parou de rodar
# (~21:15 UTC). Uma entrada NOVA (já corrigida) grava o horário real de
# Brasília — então seu valor literal só pode começar a partir do instante
# em que o deploy entrou no ar (~18:05 BR). Como os dois usam o mesmo
# "relógio de parede" pra escrever a célula, a faixa [18:05, 21:15] no dia
# do deploy é onde as duas interpretações se sobrepõem e não dá pra
# decidir com segurança só pelo valor — por isso fica de fora da correção
# automática e é sinalizada pra revisão manual.
_JANELA_INICIO = datetime(2026, 7, 8, 18, 5, 0).time()
_JANELA_FIM = datetime(2026, 7, 8, 21, 15, 0).time()

_HIST_LINHA_RE = re.compile(r'^(\d{2}/\d{2}/\d{4}) (\d{2}:\d{2})(:\d{2})? - ')


def _classificar_ts_fuso(ts_str, fmt):
    """Classifica um timestamp gravado antes da correção de fuso horário.

    Regra extra: uma entrada NOVA (corrigida) nunca pode ter horário no
    futuro em relação a "agora" (horário real de Brasília no momento da
    checagem) — se tiver, só pode ser uma entrada ANTIGA (valor bruto de
    UTC, que naturalmente "parece" mais tarde). Isso desambiguiza a maior
    parte da janela conforme o tempo passa.

    Retorna ('antigo', novo_valor) | ('ambiguo', ts_str) | ('atual', ts_str) | ('invalido', ts_str)
    """
    try:
        dt = datetime.strptime(ts_str, fmt)
    except Exception:
        return "invalido", ts_str

    if dt.date() < _HOJE_DEPLOY:
        return "antigo", (dt - timedelta(hours=3)).strftime(fmt)

    if dt.date() == _HOJE_DEPLOY:
        if dt.time() < _JANELA_INICIO:
            return "antigo", (dt - timedelta(hours=3)).strftime(fmt)
        agora_time = agora_br().time()
        if dt.time() > agora_time:
            # horário "no futuro" em relação a agora só é possível se for
            # valor bruto de UTC (entrada antiga) — uma entrada nova jamais
            # gravaria um horário à frente do relógio real de Brasília.
            return "antigo", (dt - timedelta(hours=3)).strftime(fmt)
        if dt.time() <= _JANELA_FIM:
            return "ambiguo", ts_str
        return "atual", ts_str

    return "atual", ts_str


@app.route("/travar-fuso-retroativo", methods=["POST", "GET"])
def travar_fuso_retroativo():
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    _gravar_trava("fuso_retroativo_concluido", "true")
    return jsonify({"ok": True, "trava": "ativada"}), 200


@app.route("/fix-pontual-9173-9154", methods=["POST", "GET"])
def fix_pontual_9173_9154():
    """Endpoint de uso único: reescreve o historico limpo e correto das OSs
    9173 e 9154, que ficaram com uma linha de log duplicada/suja por causa
    de uma tentativa de correção manual que esbarrou num bug pré-existente
    do endpoint /atualizar-campo-atividade (ele reloga a alteração usando
    dados em cache antigos ao invés de sobrescrever de forma limpa)."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    correcoes = {
        "9173": "08/07/2026 15:55 - Atividade criada por fracttal-sync.",
        "9154": ("08/07/2026 15:55 - Atividade criada por fracttal-sync.\n"
                 "08/07/2026 17:53 - Status na OS (Fracttal) atualizado: "
                 "\"Em Processo\" → \"Em Revisão\", 0% → 100% (Concluída)."),
    }

    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    batch_updates = []
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < 14:
            continue
        numero_os = row[13].strip()
        if numero_os in correcoes:
            batch_updates.append({
                "range": gspread.utils.rowcol_to_a1(i, 10),
                "values": [["08/07/2026 15:55:00"]],
            })
            batch_updates.append({
                "range": gspread.utils.rowcol_to_a1(i, 12),
                "values": [[correcoes[numero_os]]],
            })

    if batch_updates:
        ws.batch_update(batch_updates, value_input_option="RAW")

    return jsonify({"ok": True, "celulas_corrigidas": len(batch_updates)}), 200


@app.route("/corrigir-fuso-retroativo", methods=["POST", "GET"])
def corrigir_fuso_retroativo():
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    aplicar = request.args.get("apply", "false").lower() == "true"
    forcar_raw = request.args.get("forcar", "").strip()
    forcar_set = {x.strip() for x in forcar_raw.split(",") if x.strip()} if forcar_raw else set()
    ignorar_trava = request.args.get("ignorar_trava", "false").lower() == "true"

    if aplicar and not ignorar_trava:
        try:
            if _ler_trava("fuso_retroativo_concluido") == "true":
                return jsonify({
                    "ok": False,
                    "error": ("Correção retroativa já foi concluída em 2026-07-08 e está travada "
                              "pra evitar redescontar horas já corrigidas. Use ?ignorar_trava=true "
                              "só se tiver certeza absoluta do que está fazendo.")
                }), 409
        except Exception:
            pass  # se a aba de config falhar por algum motivo, não bloqueia a leitura normal (dry-run)

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        _garantir_headers_atividades(ws)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    alteracoes = []
    ambiguos = []
    batch_updates = []
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        numero_os = row[13].strip()
        id_atividade = row[0].strip()
        updates = {}
        ambiguos_linha = []
        forcar_linha = bool(forcar_set) and (numero_os in forcar_set or id_atividade in forcar_set)

        def _checar(campo_nome, valor, fmt, col):
            if not valor:
                return
            estado, novo = _classificar_ts_fuso(valor, fmt)
            if estado == "ambiguo" and forcar_linha:
                dt_forcado = datetime.strptime(valor, fmt)
                estado, novo = "antigo", (dt_forcado - timedelta(hours=3)).strftime(fmt)
            if estado == "antigo":
                updates[col] = novo
            elif estado == "ambiguo":
                ambiguos_linha.append({"campo": campo_nome, "valor": valor})

        _checar("dataCriacao", row[9].strip(), '%d/%m/%Y %H:%M:%S', 10)
        _checar("dataConclusao", row[10].strip(), '%d/%m/%Y %H:%M:%S', 11)
        _checar("ultimaVerificacaoOS", row[23].strip(), '%Y-%m-%dT%H:%M:%S', 24)

        # historico (col 12) — multilinha, cada linha pode começar com timestamp
        hist = row[11]
        if hist:
            linhas_novas = []
            hist_mudou = False
            for linha_h in hist.split("\n"):
                m = _HIST_LINHA_RE.match(linha_h)
                if m:
                    data_str, hora_str, seg = m.group(1), m.group(2), m.group(3) or ""
                    ts_str = f"{data_str} {hora_str}{seg}"
                    fmt = '%d/%m/%Y %H:%M:%S' if seg else '%d/%m/%Y %H:%M'
                    estado, novo_ts = _classificar_ts_fuso(ts_str, fmt)
                    if estado == "ambiguo" and forcar_linha:
                        dt_forcado = datetime.strptime(ts_str, fmt)
                        estado, novo_ts = "antigo", (dt_forcado - timedelta(hours=3)).strftime(fmt)
                    if estado == "antigo":
                        linha_h = novo_ts + linha_h[len(ts_str):]
                        hist_mudou = True
                    elif estado == "ambiguo":
                        ambiguos_linha.append({"campo": "historico", "valor": ts_str, "linha_texto": linha_h[:80]})
                linhas_novas.append(linha_h)
            if hist_mudou:
                updates[12] = "\n".join(linhas_novas)

        if updates:
            alteracoes.append({"linha": i, "id": id_atividade, "numeroOS": numero_os,
                                "colunas_alteradas": list(updates.keys())})
            if aplicar:
                for col, novo_val in updates.items():
                    batch_updates.append({
                        "range": gspread.utils.rowcol_to_a1(i, col),
                        "values": [[novo_val]],
                    })
        if ambiguos_linha:
            ambiguos.append({"linha": i, "id": id_atividade, "numeroOS": numero_os, "campos": ambiguos_linha})

    if aplicar and batch_updates:
        # grava tudo em poucas chamadas (lotes de 200 células) em vez de uma
        # chamada por célula — evita estourar o timeout do Gunicorn (60s)
        TAMANHO_LOTE = 200
        for k in range(0, len(batch_updates), TAMANHO_LOTE):
            ws.batch_update(batch_updates[k:k + TAMANHO_LOTE], value_input_option="RAW")

    return jsonify({"ok": True, "aplicado": aplicar, "linhas_afetadas": len(alteracoes),
                     "detalhes": alteracoes, "ambiguos": ambiguos}), 200


def _get_config_sheet():
    """Aba minúscula usada só pra guardar flags de controle (ex.: travas de
    operações de uso único). Cria a aba se ainda não existir."""
    sh = get_atividades_sheet().spreadsheet
    try:
        return sh.worksheet("_Sistema")
    except gspread.exceptions.WorksheetNotFound:
        ws_cfg = sh.add_worksheet(title="_Sistema", rows=20, cols=4)
        ws_cfg.update("A1", [["chave", "valor"]])
        return ws_cfg


def _ler_trava(chave):
    ws_cfg = _get_config_sheet()
    valores = ws_cfg.get_all_values()
    for row in valores[1:]:
        if row and row[0].strip() == chave:
            return row[1].strip() if len(row) > 1 else ""
    return ""


def _gravar_trava(chave, valor):
    ws_cfg = _get_config_sheet()
    valores = ws_cfg.get_all_values()
    for i, row in enumerate(valores[1:], start=2):
        if row and row[0].strip() == chave:
            ws_cfg.update_cell(i, 2, valor)
            return
    ws_cfg.append_row([chave, valor])


def _ler_travas(chaves):
    """Lê várias chaves da aba _Sistema numa única leitura (1 chamada à
    API do Sheets), em vez de uma chamada por chave via _ler_trava — usado
    onde múltiplas travas são lidas juntas (ex.: status de sincronização
    de chamados) pra não somar latência desnecessária num endpoint que já
    lê uma planilha grande."""
    ws_cfg = _get_config_sheet()
    valores = ws_cfg.get_all_values()
    mapa = {}
    for row in valores[1:]:
        if row and row[0].strip() in chaves:
            mapa[row[0].strip()] = row[1].strip() if len(row) > 1 else ""
    return {chave: mapa.get(chave, "") for chave in chaves}


# ══════════════════════════════════════════════════════════════════════
# COMPROMISSOS (Boletim de Medição, Relatório de Performance, Relatório
# PCM) — checklist de prazos recorrentes por cliente/usina, com engine
# de cálculo de dia útil/dia fixo e alertas push automáticos.
# ══════════════════════════════════════════════════════════════════════

COMPROMISSO_ETAPAS = {
    "BM": ["Envio do BM", "Aprovação do Cliente", "Emissão da NF"],
    "RelatorioPerformance": ["Envio do Relatório de Performance"],
    "RelatorioPCM": ["Envio do Relatório de Manutenção (PCM)"],
}

COMPROMISSO_LABEL = {
    "BM": "Boletim de Medição",
    "RelatorioPerformance": "Relatório de Performance",
    "RelatorioPCM": "Relatório de Manutenção (PCM)",
}


def _feriados_nacionais_brasil(ano):
    """Feriados nacionais fixos + móveis (baseados na Páscoa, algoritmo
    de Gauss). Não cobre feriados estaduais/municipais — só o suficiente
    pra não antecipar prazo em cima de feriado nacional por engano."""
    a = ano
    # Páscoa (algoritmo de Meeus/Jones/Butcher)
    y = a
    g = y % 19
    c = y // 100
    h = (c - c // 4 - (8 * c + 13) // 25 + 19 * g + 15) % 30
    i = h - (h // 28) * (1 - (h // 28) * (29 // (h + 1)) * ((21 - g) // 11))
    j = (y + y // 4 + i + 2 - c + c // 4) % 7
    l = i - j
    mes = 3 + (l + 40) // 44
    dia = l + 28 - 31 * (mes // 4)
    pascoa = datetime(y, mes, dia)

    fixos = [
        datetime(a, 1, 1), datetime(a, 4, 21), datetime(a, 5, 1),
        datetime(a, 9, 7), datetime(a, 10, 12), datetime(a, 11, 2),
        datetime(a, 11, 15), datetime(a, 12, 25),
    ]
    moveis = [
        pascoa - timedelta(days=47),  # carnaval segunda
        pascoa - timedelta(days=46),  # carnaval terça
        pascoa - timedelta(days=2),   # sexta-feira santa
        pascoa + timedelta(days=60),  # corpus christi
    ]
    return {d.date() for d in (fixos + moveis)}


def _e_dia_util(dt):
    return dt.weekday() < 5 and dt.date() not in _feriados_nacionais_brasil(dt.year)


def _enesimo_dia_util(ano, mes, n):
    dt = datetime(ano, mes, 1)
    contados = 0
    while True:
        if _e_dia_util(dt):
            contados += 1
            if contados == n:
                return dt
        dt += timedelta(days=1)
        if dt.month != mes:  # segurança: não vaza pro mês seguinte
            return dt - timedelta(days=1)


def _ultimo_dia_do_mes(ano, mes):
    if mes == 12:
        prox = datetime(ano + 1, 1, 1)
    else:
        prox = datetime(ano, mes + 1, 1)
    return prox - timedelta(days=1)


def _ultimo_dia_util(ano, mes):
    dt = _ultimo_dia_do_mes(ano, mes)
    while not _e_dia_util(dt):
        dt -= timedelta(days=1)
    return dt


def _dia_fixo_com_antecipacao(ano, mes, dia):
    ultimo = _ultimo_dia_do_mes(ano, mes).day
    dt = datetime(ano, mes, min(dia, ultimo))
    while not _e_dia_util(dt):
        dt -= timedelta(days=1)
    return dt


def _calcular_prazo_compromisso(regra_tipo, regra_valor, ano, mes):
    regra_valor = int(regra_valor)
    if regra_tipo == "nDiaUtil":
        return _enesimo_dia_util(ano, mes, regra_valor)
    if regra_tipo == "diaFixo":
        return _dia_fixo_com_antecipacao(ano, mes, regra_valor)
    if regra_tipo == "diaAoUltimoUtil":
        return _ultimo_dia_util(ano, mes)
    raise ValueError(f"regra_tipo desconhecido: {regra_tipo}")


def _subtrair_dias_uteis(dt, n):
    """Volta N dias úteis a partir de dt (não conta o próprio dt)."""
    atual = dt
    contados = 0
    while contados < n:
        atual -= timedelta(days=1)
        if _e_dia_util(atual):
            contados += 1
    return atual


def _somar_dias_uteis(dt, n):
    """Avança N dias úteis a partir de dt (não conta o próprio dt)."""
    atual = dt
    contados = 0
    while contados < n:
        atual += timedelta(days=1)
        if _e_dia_util(atual):
            contados += 1
    return atual


# Regra padrão pro fluxo interno de BM (Boletim de Medição), levantada em
# 03/09/2026: nenhum dos contratos analisados (ABC/Alves Lima, GD Energy,
# Sal Energia) define prazo próprio de envio do BM nem de aprovação do
# cliente — só a data final de emissão da NF está no texto contratual.
# A pedido do Fred, na ausência de cláusula específica no contrato:
#   Envio do BM = DataLimite da NF menos 5 dias úteis
#   Aprovação do Cliente = Envio do BM mais 2 dias úteis
BM_ENVIO_DIAS_UTEIS_ANTES_NF = 5
BM_APROVACAO_DIAS_UTEIS_APOS_ENVIO = 2


def _calcular_subprazos_bm(data_limite_nf):
    """Calcula as datas de envio do BM e prazo de aprovação do cliente a
    partir da data-limite da NF, usando a regra padrão (sem cláusula
    contratual específica encontrada nos contratos analisados)."""
    envio_bm = _subtrair_dias_uteis(data_limite_nf, BM_ENVIO_DIAS_UTEIS_ANTES_NF)
    aprovacao = _somar_dias_uteis(envio_bm, BM_APROVACAO_DIAS_UTEIS_APOS_ENVIO)
    return envio_bm, aprovacao


def _garantir_colunas_bm_prazos(ws_comp):
    """Garante que a planilha Compromissos tenha as colunas M/N
    (DataLimiteEnvioBM / DataLimiteAprovacao). Sheet foi criada
    originalmente só até a coluna L (12) — expande sob demanda."""
    if ws_comp.col_count < 14:
        ws_comp.add_cols(14 - ws_comp.col_count)
    header = ws_comp.row_values(1)
    if len(header) < 13 or header[12] != "DataLimiteEnvioBM":
        ws_comp.update_cell(1, 13, "DataLimiteEnvioBM")
    if len(header) < 14 or header[13] != "DataLimiteAprovacao":
        ws_comp.update_cell(1, 14, "DataLimiteAprovacao")


def _get_compromissos_regras_sheet():
    sh = get_atividades_sheet().spreadsheet
    try:
        return sh.worksheet("_ComprometimentosRegras")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="_ComprometimentosRegras", rows=50, cols=8)
        ws.update("A1", [["ID", "Tipo", "Cliente", "Usina", "RegraTipo", "RegraValor", "Ativo"]])
        # Seed inicial — só os clientes já mapeados no painel, regra de BM
        # tirada do calendário de emissão enviado pelo Fred (13/07/2026).
        seed = [
            ["1", "BM", "Renogrid", "", "diaFixo", "25", "TRUE"],
            ["2", "BM", "Thopen", "", "diaFixo", "15", "TRUE"],
            ["3", "BM", "2C Energia", "", "nDiaUtil", "5", "TRUE"],
            ["4", "BM", "GD Energy", "", "nDiaUtil", "5", "TRUE"],
            ["5", "BM", "Alves Lima", "", "nDiaUtil", "5", "TRUE"],
        ]
        ws.append_rows(seed)
        return ws


def _get_compromissos_sheet():
    sh = get_atividades_sheet().spreadsheet
    try:
        return sh.worksheet("Compromissos")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Compromissos", rows=200, cols=14)
        ws.update("A1", [["ID", "Tipo", "Cliente", "Usina", "Competencia", "DataLimite",
                           "Etapas", "EtapasConcluidas", "Status", "DataCriacao",
                           "DataConclusao", "Historico", "DataLimiteEnvioBM", "DataLimiteAprovacao"]])
        return ws


def _proximo_id_compromisso(todos):
    maior = 0
    for row in todos[1:]:
        if row and row[0].strip().isdigit():
            maior = max(maior, int(row[0].strip()))
    return str(maior + 1)


def _status_compromisso(etapas_concluidas, data_limite, hoje):
    concluidas = [e for e in etapas_concluidas if e]
    if len(concluidas) == len(etapas_concluidas):
        return "Concluído"
    if hoje.date() > data_limite.date() and not etapas_concluidas[0]:
        return "Atrasado"
    if concluidas:
        return "Em Andamento"
    return "Pendente"


def _gerar_compromissos_mes_atual_se_necessario():
    """Versão econômica de _gerar_compromissos_mes_atual(): só faz a
    varredura pesada (2 leituras completas de planilha) 1x por competência,
    usando a trava em _Sistema (leitura pequena e barata) pra decidir se
    vale a pena. Sem isso, todo GET /compromissos gastava 2 leituras
    completas — e como o frontend recarregava a lista a cada clique de
    checkbox, isso estourou a cota de leitura do Google Sheets (429) e
    derrubou o resto do painel junto (13/07/2026)."""
    agora = agora_br()
    competencia = agora.strftime("%m/%Y")
    ja_gerado = _ler_trava("compromissos_gerados_em")
    if ja_gerado == competencia:
        return []
    criados = _gerar_compromissos_mes_atual()
    _gravar_trava("compromissos_gerados_em", competencia)
    return criados


def _gerar_compromissos_mes_atual():
    """Idempotente: cria o card do mês corrente pra cada regra ativa que
    ainda não tenha um card gerado nessa competência. Fecha sozinho o
    ciclo anterior (o card antigo simplesmente fica com seu status real —
    Concluído ou Atrasado — e um novo é aberto pra competência atual)."""
    ws_regras = _get_compromissos_regras_sheet()
    regras = ws_regras.get_all_values()[1:]

    ws_comp = _get_compromissos_sheet()
    _garantir_colunas_bm_prazos(ws_comp)
    todos = ws_comp.get_all_values()
    existentes = {(r[1], r[2], r[3], r[4]) for r in todos[1:] if len(r) >= 5}

    agora = agora_br()
    competencia = agora.strftime("%m/%Y")
    criados = []

    for r in regras:
        if len(r) < 7 or r[6].strip().upper() != "TRUE":
            continue
        _id, tipo, cliente, usina, regra_tipo, regra_valor = r[0], r[1], r[2], r[3], r[4], r[5]
        chave = (tipo, cliente, usina, competencia)
        if chave in existentes:
            continue
        try:
            prazo = _calcular_prazo_compromisso(regra_tipo, regra_valor, agora.year, agora.month)
        except Exception as e:
            log.error(f"[Compromissos] Erro ao calcular prazo pra regra {_id}: {e}")
            continue

        envio_bm_str, aprovacao_str = "", ""
        if tipo == "BM":
            envio_bm, aprovacao = _calcular_subprazos_bm(prazo)
            envio_bm_str, aprovacao_str = envio_bm.strftime("%d/%m/%Y"), aprovacao.strftime("%d/%m/%Y")

        etapas = COMPROMISSO_ETAPAS.get(tipo, ["Envio"])
        novo_id = _proximo_id_compromisso(todos)
        linha = [novo_id, tipo, cliente, usina, competencia, prazo.strftime("%d/%m/%Y"),
                  json.dumps(etapas, ensure_ascii=False), json.dumps([""] * len(etapas)),
                  "Pendente", agora.strftime("%d/%m/%Y %H:%M:%S"), "",
                  f"{agora.strftime('%d/%m/%Y %H:%M')} - Card criado automaticamente pra competência {competencia}.",
                  envio_bm_str, aprovacao_str]
        ws_comp.append_row(linha)
        todos.append(linha)
        existentes.add(chave)
        criados.append({"id": novo_id, "tipo": tipo, "cliente": cliente, "competencia": competencia,
                         "dataLimite": prazo.strftime("%d/%m/%Y"),
                         "dataLimiteEnvioBM": envio_bm_str, "dataLimiteAprovacao": aprovacao_str})

    return criados


def _listar_compromissos_core():
    _gerar_compromissos_mes_atual_se_necessario()
    ws = _get_compromissos_sheet()
    todos = ws.get_all_values()
    agora = agora_br()
    resultado = []
    for row in todos[1:]:
        if len(row) < 12 or not row[0].strip():
            continue
        try:
            data_limite = datetime.strptime(row[5].strip(), "%d/%m/%Y")
        except Exception:
            continue
        etapas = json.loads(row[6]) if row[6] else []
        etapas_concluidas = json.loads(row[7]) if row[7] else []
        status_calc = _status_compromisso(etapas_concluidas, data_limite, agora)
        dias_restantes = (data_limite.date() - agora.date()).days

        data_limite_envio_bm = row[12] if len(row) > 12 else ""
        data_limite_aprovacao = row[13] if len(row) > 13 else ""
        dias_restantes_envio_bm, dias_restantes_aprovacao = None, None
        try:
            if data_limite_envio_bm:
                dias_restantes_envio_bm = (datetime.strptime(data_limite_envio_bm, "%d/%m/%Y").date() - agora.date()).days
            if data_limite_aprovacao:
                dias_restantes_aprovacao = (datetime.strptime(data_limite_aprovacao, "%d/%m/%Y").date() - agora.date()).days
        except Exception:
            pass

        resultado.append({
            "id": row[0], "tipo": row[1], "tipoLabel": COMPROMISSO_LABEL.get(row[1], row[1]),
            "cliente": row[2], "usina": row[3], "competencia": row[4],
            "dataLimite": row[5], "diasRestantes": dias_restantes,
            "dataLimiteEnvioBM": data_limite_envio_bm, "diasRestantesEnvioBM": dias_restantes_envio_bm,
            "dataLimiteAprovacao": data_limite_aprovacao, "diasRestantesAprovacao": dias_restantes_aprovacao,
            "etapas": etapas, "etapasConcluidas": etapas_concluidas,
            "status": status_calc, "dataConclusao": row[10],
        })
    # Mais urgente primeiro: atrasado > vence antes > já concluído por último
    ordem_status = {"Atrasado": 0, "Pendente": 1, "Em Andamento": 1, "Concluído": 2}
    resultado.sort(key=lambda c: (ordem_status.get(c["status"], 1), c["diasRestantes"]))
    return resultado


@app.route("/compromissos", methods=["GET"])
def listar_compromissos():
    try:
        return jsonify({"ok": True, "compromissos": _listar_compromissos_core()}), 200
    except Exception as e:
        log.error(f"[Compromissos] Erro ao listar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/compromissos/regras", methods=["GET"])
def listar_regras_compromissos():
    try:
        ws_regras = _get_compromissos_regras_sheet()
        valores = ws_regras.get_all_values()
        cabecalho = valores[0] if valores else []
        regras = [dict(zip(cabecalho, row)) for row in valores[1:] if row and row[0].strip()]
        return jsonify({"ok": True, "regras": regras}), 200
    except Exception as e:
        log.error(f"[Compromissos] Erro ao listar regras: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/compromissos/regras/atualizar", methods=["POST"])
def atualizar_regra_compromisso():
    """Atualiza RegraTipo/RegraValor de uma regra em _ComprometimentosRegras
    e, se já existir um card gerado pra competência atual com essa regra
    (ainda não concluído), recalcula o DataLimite dele também — porque
    _gerar_compromissos_mes_atual é idempotente e não regeneraria um card
    já existente só porque a regra mudou."""
    try:
        body = request.get_json(force=True) or {}
        regra_id = str(body.get("id", "")).strip()
        novo_tipo = str(body.get("regraTipo", "")).strip()
        novo_valor = str(body.get("regraValor", "")).strip()
        editor = body.get("editor", "desconhecido")
        if not regra_id or not novo_tipo or not novo_valor:
            return jsonify({"ok": False, "error": "id, regraTipo e regraValor são obrigatórios"}), 400
        if novo_tipo not in ("nDiaUtil", "diaFixo", "diaAoUltimoUtil"):
            return jsonify({"ok": False, "error": f"regraTipo desconhecido: {novo_tipo}"}), 400

        ws_regras = _get_compromissos_regras_sheet()
        valores = ws_regras.get_all_values()
        linha_idx, linha = None, None
        for i, row in enumerate(valores[1:], start=2):
            if row and row[0].strip() == regra_id:
                linha_idx, linha = i, row
                break
        if not linha_idx:
            return jsonify({"ok": False, "error": f"regra {regra_id} não encontrada"}), 404

        tipo, cliente, usina = linha[1], linha[2], linha[3]
        ws_regras.update_cell(linha_idx, 5, novo_tipo)   # RegraTipo
        ws_regras.update_cell(linha_idx, 6, novo_valor)  # RegraValor

        agora = agora_br()
        competencia = agora.strftime("%m/%Y")
        card_atualizado = None
        try:
            novo_prazo = _calcular_prazo_compromisso(novo_tipo, novo_valor, agora.year, agora.month)
        except Exception as e:
            return jsonify({"ok": True, "regraAtualizada": True,
                             "aviso": f"regra salva, mas não foi possível recalcular o card do mês: {e}"}), 200

        ws_comp = _get_compromissos_sheet()
        _garantir_colunas_bm_prazos(ws_comp)
        todos = ws_comp.get_all_values()
        for i, row in enumerate(todos[1:], start=2):
            if len(row) < 12:
                continue
            if row[1] == tipo and row[2] == cliente and row[3] == usina and row[4] == competencia and row[8] != "Concluído":
                data_antiga = row[5]
                nova_data_str = novo_prazo.strftime("%d/%m/%Y")
                ws_comp.update_cell(i, 6, nova_data_str)  # DataLimite
                historico_novo = row[11] + f"\n{agora.strftime('%d/%m/%Y %H:%M')} - Prazo corrigido de {data_antiga} para {nova_data_str} (regra ajustada por {editor})."
                ws_comp.update_cell(i, 12, historico_novo)
                card_atualizado = {"id": row[0], "dataLimiteAnterior": data_antiga, "dataLimiteNova": nova_data_str}
                if tipo == "BM":
                    novo_envio_bm, nova_aprovacao = _calcular_subprazos_bm(novo_prazo)
                    ws_comp.update_cell(i, 13, novo_envio_bm.strftime("%d/%m/%Y"))
                    ws_comp.update_cell(i, 14, nova_aprovacao.strftime("%d/%m/%Y"))
                    card_atualizado["dataLimiteEnvioBM"] = novo_envio_bm.strftime("%d/%m/%Y")
                    card_atualizado["dataLimiteAprovacao"] = nova_aprovacao.strftime("%d/%m/%Y")
                break

        return jsonify({"ok": True, "regraAtualizada": True, "cardAtualizado": card_atualizado}), 200
    except Exception as e:
        log.error(f"[Compromissos] Erro ao atualizar regra: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/compromissos/testar-alerta-etapas-abertas", methods=["GET", "POST"])
def testar_alerta_etapas_abertas():
    """Dispara o alerta de etapas abertas na hora, ignorando a janela e a
    trava diária — só pra teste manual. Query param ?enviar=false devolve
    a lista sem mandar a mensagem pro WhatsApp."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    try:
        enviar = request.args.get("enviar", "true").lower() != "false"
        compromissos = _listar_compromissos_core()
        etapas_abertas = []
        for c in compromissos:
            if c["status"] == "Concluído":
                continue
            for idx, nome_etapa in enumerate(c["etapas"]):
                if c["etapasConcluidas"][idx]:
                    continue
                data_str, dias = _prazo_etapa_compromisso(c, idx)
                if not data_str or dias is None:
                    continue
                if dias <= 0:
                    etapas_abertas.append({
                        "cliente": c["cliente"], "tipoLabel": c["tipoLabel"],
                        "usina": c["usina"], "etapa": nome_etapa,
                        "dataLimite": data_str, "diasAtraso": -dias,
                    })
        etapas_abertas.sort(key=lambda e: -e["diasAtraso"])

        if not etapas_abertas:
            return jsonify({"ok": True, "etapasAbertas": 0, "mensagem": "nenhuma etapa em aberto agora"}), 200

        linhas = []
        for e in etapas_abertas:
            usina_txt = f" ({e['usina']})" if e["usina"] else ""
            prazo_txt = f"vence hoje ({e['dataLimite']})" if e["diasAtraso"] == 0 \
                else f"vencida há {e['diasAtraso']} dia(s) — prazo era {e['dataLimite']}"
            linhas.append(f"• {e['cliente']}{usina_txt} — {e['tipoLabel']} / {e['etapa']}: {prazo_txt}")
        texto = (
            f"⚠️ *Etapas em aberto — {agora_br().strftime('%d/%m/%Y')}*\n\n"
            + "\n".join(linhas)
            + "\n\nAbra o Painel Gerencial > Boletins de Medição pra marcar as etapas concluídas."
        )
        resultado_envio = _enviar_mensagem_grupo(GRUPO_GESTAO_OM_ID, texto) if enviar else None
        return jsonify({"ok": True, "etapasAbertas": len(etapas_abertas), "texto": texto, "envio": resultado_envio}), 200
    except Exception as e:
        log.error(f"[Compromissos] Erro no teste manual de alerta de etapas abertas: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/compromissos/backfill-subprazos-bm", methods=["POST"])
def backfill_subprazos_bm():
    """Preenche DataLimiteEnvioBM/DataLimiteAprovacao pra cards tipo=BM
    que já existiam antes dessas colunas serem adicionadas (03/09/2026).
    Idempotente: só escreve nas linhas onde as colunas ainda estão vazias."""
    try:
        ws_comp = _get_compromissos_sheet()
        _garantir_colunas_bm_prazos(ws_comp)
        todos = ws_comp.get_all_values()
        atualizados = []
        for i, row in enumerate(todos[1:], start=2):
            if len(row) < 6 or row[1] != "BM":
                continue
            ja_tem = len(row) > 13 and row[12].strip() and row[13].strip()
            if ja_tem:
                continue
            try:
                data_limite = datetime.strptime(row[5].strip(), "%d/%m/%Y")
            except Exception:
                continue
            envio_bm, aprovacao = _calcular_subprazos_bm(data_limite)
            envio_bm_str, aprovacao_str = envio_bm.strftime("%d/%m/%Y"), aprovacao.strftime("%d/%m/%Y")
            ws_comp.update_cell(i, 13, envio_bm_str)
            ws_comp.update_cell(i, 14, aprovacao_str)
            atualizados.append({"id": row[0], "cliente": row[2], "competencia": row[4],
                                 "dataLimiteEnvioBM": envio_bm_str, "dataLimiteAprovacao": aprovacao_str})
        return jsonify({"ok": True, "atualizados": atualizados}), 200
    except Exception as e:
        log.error(f"[Compromissos] Erro no backfill de subprazos BM: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/compromissos/regras/criar", methods=["POST"])
def criar_regra_compromisso():
    """Cria uma nova regra em _ComprometimentosRegras (novo cliente/tipo
    entrando no ciclo de compromissos). Não gera o card do mês na hora —
    isso continua a cargo de _gerar_compromissos_mes_atual_se_necessario()
    no próximo GET /compromissos, igual às regras do seed."""
    try:
        body = request.get_json(force=True) or {}
        tipo = str(body.get("tipo", "")).strip()
        cliente = str(body.get("cliente", "")).strip()
        usina = str(body.get("usina", "")).strip()
        regra_tipo = str(body.get("regraTipo", "")).strip()
        regra_valor = str(body.get("regraValor", "")).strip()
        ativo = body.get("ativo", True)
        if not tipo or not cliente or not regra_tipo or not regra_valor:
            return jsonify({"ok": False, "error": "tipo, cliente, regraTipo e regraValor são obrigatórios"}), 400
        if regra_tipo not in ("nDiaUtil", "diaFixo", "diaAoUltimoUtil"):
            return jsonify({"ok": False, "error": f"regraTipo desconhecido: {regra_tipo}"}), 400

        ws_regras = _get_compromissos_regras_sheet()
        todas = ws_regras.get_all_values()
        for row in todas[1:]:
            if len(row) >= 4 and row[1] == tipo and row[2] == cliente and row[3] == usina:
                return jsonify({"ok": False, "error": f"já existe uma regra {tipo} pra {cliente}"
                                 + (f" / {usina}" if usina else "")}), 409

        maior = 0
        for row in todas[1:]:
            if row and row[0].strip().isdigit():
                maior = max(maior, int(row[0].strip()))
        novo_id = str(maior + 1)

        linha = [novo_id, tipo, cliente, usina, regra_tipo, regra_valor, "TRUE" if ativo else "FALSE"]
        ws_regras.append_row(linha)

        return jsonify({"ok": True, "regraCriada": {
            "id": novo_id, "tipo": tipo, "cliente": cliente, "usina": usina,
            "regraTipo": regra_tipo, "regraValor": regra_valor, "ativo": bool(ativo),
        }}), 200
    except Exception as e:
        log.error(f"[Compromissos] Erro ao criar regra: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/compromissos/marcar-etapa", methods=["POST"])
def marcar_etapa_compromisso():
    try:
        body = request.get_json(force=True) or {}
        comp_id = str(body.get("id", "")).strip()
        etapa_index = int(body.get("etapaIndex", -1))
        concluida = bool(body.get("concluida", True))
        editor = body.get("editor", "desconhecido")
        if not comp_id or etapa_index < 0:
            return jsonify({"ok": False, "error": "id e etapaIndex são obrigatórios"}), 400

        ws = _get_compromissos_sheet()
        todos = ws.get_all_values()
        linha_idx, linha = None, None
        for i, row in enumerate(todos[1:], start=2):
            if row and row[0].strip() == comp_id:
                linha_idx, linha = i, row
                break
        if not linha_idx:
            return jsonify({"ok": False, "error": f"compromisso {comp_id} não encontrado"}), 404

        etapas = json.loads(linha[6]) if linha[6] else []
        etapas_concluidas = json.loads(linha[7]) if len(linha) > 7 and linha[7] else [""] * len(etapas)
        if etapa_index >= len(etapas):
            return jsonify({"ok": False, "error": "etapaIndex fora do intervalo"}), 400

        agora = agora_br()
        etapas_concluidas[etapa_index] = agora.strftime("%d/%m/%Y %H:%M") if concluida else ""

        try:
            data_limite = datetime.strptime(linha[5].strip(), "%d/%m/%Y")
        except Exception:
            data_limite = agora
        novo_status = _status_compromisso(etapas_concluidas, data_limite, agora)

        data_conclusao = agora.strftime("%d/%m/%Y %H:%M:%S") if novo_status == "Concluído" else ""

        nome_etapa = etapas[etapa_index]
        acao = "concluída" if concluida else "reaberta"
        entry = f"{agora.strftime('%d/%m/%Y %H:%M')} - Etapa \"{nome_etapa}\" {acao} por {editor}."
        hist_atual = linha[11] if len(linha) > 11 else ""
        novo_hist = f"{hist_atual}\n{entry}".strip() if hist_atual else entry

        ws.update(f"H{linha_idx}:L{linha_idx}", [[
            json.dumps(etapas_concluidas, ensure_ascii=False), novo_status,
            linha[9] if len(linha) > 9 else agora.strftime("%d/%m/%Y %H:%M:%S"),
            data_conclusao, novo_hist,
        ]])

        return jsonify({"ok": True, "id": comp_id, "status": novo_status,
                         "etapasConcluidas": etapas_concluidas}), 200
    except Exception as e:
        log.error(f"[Compromissos] Erro ao marcar etapa: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Sketchbook / Anotações ───────────────────────────────────────────────
# Bloco de notas rápidas do Fred: anotações soltas ou vinculadas a
# usina/cliente que depois alimentam o Chat-IA (ferramenta
# consultar_anotacoes), relatórios e comunicados como contexto extra.
# Entrada só pelo dashboard (campo de texto rápido) — decidido 03/09/2026.
ANOTACOES_SHEET_NAME = "Anotacoes"
ANOTACOES_HEADERS = ["ID", "Data", "Autor", "Categoria", "Texto", "Usina", "Cliente", "Status"]


def _get_anotacoes_sheet():
    sh = get_atividades_sheet().spreadsheet
    try:
        return _gspread_retry(lambda: sh.worksheet(ANOTACOES_SHEET_NAME))
    except gspread.exceptions.WorksheetNotFound:
        ws = _gspread_retry(lambda: sh.add_worksheet(title=ANOTACOES_SHEET_NAME, rows=500, cols=len(ANOTACOES_HEADERS)))
        _gspread_retry(lambda: ws.update("A1", [ANOTACOES_HEADERS]))
        return ws


def _proximo_id_anotacao(todos):
    maior = 0
    for row in todos[1:]:
        if row and row[0].strip().isdigit():
            maior = max(maior, int(row[0].strip()))
    return str(maior + 1)


def _ia_consultar_anotacoes(usina="", cliente="", categoria=""):
    """Handler da ferramenta consultar_anotacoes do Chat-IA — mesmo padrão
    de campos compactos + limite de itens das outras ferramentas (ver nota
    de 03/09/2026 sobre payload gigante estourando timeout da Gemini)."""
    ws = _get_anotacoes_sheet()
    todos = _gspread_retry(lambda: ws.get_all_values())
    usina_norm = canonizar_usina(usina) if usina else None
    out = []
    for row in todos[1:]:
        if len(row) < len(ANOTACOES_HEADERS):
            row = row + [""] * (len(ANOTACOES_HEADERS) - len(row))
        item = dict(zip(ANOTACOES_HEADERS, row[:len(ANOTACOES_HEADERS)]))
        if not item.get("ID", "").strip():
            continue
        if item.get("Status", "Ativa").strip() != "Ativa":
            continue
        item_usina = item.get("Usina", "").strip()
        if item_usina and not usina_permitida(item_usina):
            continue
        if usina_norm and canonizar_usina(item_usina) != usina_norm:
            continue
        if cliente and cliente.strip().lower() not in item.get("Cliente", "").strip().lower():
            continue
        if categoria and categoria.strip().lower() not in item.get("Categoria", "").strip().lower():
            continue
        out.append({
            "id": item.get("ID"), "data": item.get("Data"), "autor": item.get("Autor"),
            "categoria": item.get("Categoria"), "texto": _ia_trunc(item.get("Texto"), 180),
            "usina": item_usina, "cliente": item.get("Cliente"),
        })
    limitado = out[:25]
    return {"total_encontrado": len(out), "mostrando": len(limitado), "anotacoes": limitado}


@app.route("/anotacoes", methods=["GET"])
def listar_anotacoes():
    try:
        incluir_arquivadas = request.args.get("todas", "").strip().lower() in ("1", "true", "sim")
        usina_f = request.args.get("usina", "").strip()
        cliente_f = request.args.get("cliente", "").strip()
        categoria_f = request.args.get("categoria", "").strip()
        ws = _get_anotacoes_sheet()
        todos = _gspread_retry(lambda: ws.get_all_values())
        usina_norm = canonizar_usina(usina_f) if usina_f else None
        out = []
        for row in todos[1:]:
            if len(row) < len(ANOTACOES_HEADERS):
                row = row + [""] * (len(ANOTACOES_HEADERS) - len(row))
            item = dict(zip(ANOTACOES_HEADERS, row[:len(ANOTACOES_HEADERS)]))
            if not item.get("ID", "").strip():
                continue
            status = item.get("Status", "Ativa").strip() or "Ativa"
            if not incluir_arquivadas and status != "Ativa":
                continue
            item_usina = item.get("Usina", "").strip()
            if item_usina and not usina_permitida(item_usina):
                continue
            if usina_norm and canonizar_usina(item_usina) != usina_norm:
                continue
            if cliente_f and cliente_f.lower() not in item.get("Cliente", "").strip().lower():
                continue
            if categoria_f and categoria_f.lower() not in item.get("Categoria", "").strip().lower():
                continue
            item["Status"] = status
            out.append(item)
        out.sort(key=lambda i: i.get("ID", "0").zfill(10), reverse=True)
        return jsonify({"ok": True, "anotacoes": out}), 200
    except Exception as e:
        log.error(f"[Anotacoes] Erro ao listar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/anotacoes", methods=["POST"])
def criar_anotacao():
    try:
        body = request.get_json(force=True) or {}
        texto = str(body.get("texto", "")).strip()
        autor = str(body.get("autor", "")).strip() or "Fred Alexandrino"
        categoria = str(body.get("categoria", "")).strip()
        usina = str(body.get("usina", "")).strip()
        cliente = str(body.get("cliente", "")).strip()
        if not texto:
            return jsonify({"ok": False, "error": "texto é obrigatório"}), 400

        ws = _get_anotacoes_sheet()
        todos = _gspread_retry(lambda: ws.get_all_values())
        novo_id = _proximo_id_anotacao(todos)
        agora = agora_br().strftime("%d/%m/%Y %H:%M")
        linha = [novo_id, agora, autor, categoria, texto, usina, cliente, "Ativa"]
        _gspread_retry(lambda: ws.append_row(linha))
        return jsonify({"ok": True, "anotacao": dict(zip(ANOTACOES_HEADERS, linha))}), 200
    except Exception as e:
        log.error(f"[Anotacoes] Erro ao criar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/anotacoes/arquivar", methods=["POST"])
def arquivar_anotacao():
    try:
        body = request.get_json(force=True) or {}
        anot_id = str(body.get("id", "")).strip()
        if not anot_id:
            return jsonify({"ok": False, "error": "id é obrigatório"}), 400
        ws = _get_anotacoes_sheet()
        todos = _gspread_retry(lambda: ws.get_all_values())
        for i, row in enumerate(todos[1:], start=2):
            if row and row[0].strip() == anot_id:
                _gspread_retry(lambda: ws.update_cell(i, 8, "Arquivada"))
                return jsonify({"ok": True, "id": anot_id}), 200
        return jsonify({"ok": False, "error": f"anotação {anot_id} não encontrada"}), 404
    except Exception as e:
        log.error(f"[Anotacoes] Erro ao arquivar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _verificar_compromissos_se_necessario():
    """Piggyback no /sync-fracttal: roda 1x por dia na janela 07:00-08:30
    (mesma janela alargada dos comunicados, mesmo motivo — cold-start do
    Render podia consumir a janela de 9 minutos inteira em dias ruins).
    Gera os cards do mês corrente e dispara push pra compromissos
    vencendo em 3/1/0 dias ou já atrasados."""
    try:
        agora = agora_br()
        hoje_str = agora.strftime("%Y-%m-%d")
        if not (agora.hour == 7 or (agora.hour == 8 and agora.minute <= 30)):
            return {"disparado": False, "motivo": f"fora da janela (agora {agora.strftime('%H:%M')})"}
        ja_feito = _ler_trava("compromissos_verificados_em")
        if ja_feito == hoje_str:
            return {"disparado": False, "motivo": "já verificado hoje"}
        _gravar_trava("compromissos_verificados_em", hoje_str)

        criados = _gerar_compromissos_mes_atual()
        compromissos = _listar_compromissos_core()
        alertados = []
        for c in compromissos:
            if c["status"] == "Concluído":
                continue
            dias = c["diasRestantes"]
            if c["status"] == "Atrasado":
                enviar_push(
                    titulo=f"🔴 {c['tipoLabel']} atrasado — {c['cliente']}",
                    corpo=f"Competência {c['competencia']} — prazo era {c['dataLimite']}. Etapa pendente: {c['etapas'][0]}.",
                    tipo="compromisso_atrasado",
                    url="https://fred-alexandrino.github.io/PAINELDEFALHAS/",
                )
                alertados.append(c["id"])
            elif dias in (0, 1, 3):
                prazo_txt = "vence hoje" if dias == 0 else f"faltam {dias} dia(s)"
                enviar_push(
                    titulo=f"📋 {c['tipoLabel']} — {c['cliente']}",
                    corpo=f"Competência {c['competencia']}: {prazo_txt} ({c['dataLimite']}).",
                    tipo="compromisso_alerta",
                    url="https://fred-alexandrino.github.io/PAINELDEFALHAS/",
                )
                alertados.append(c["id"])
        return {"disparado": True, "cardsCriados": criados, "alertados": alertados}
    except Exception as e:
        log.error(f"[Compromissos] Erro na verificação diária: {e}")
        return {"disparado": False, "erro": str(e)}


def _prazo_etapa_compromisso(c, idx):
    """Devolve (data_str, dias_restantes) da etapa idx de um compromisso —
    já vindo de _listar_compromissos_core(). Pra tipo BM, cada etapa tem
    prazo próprio (Envio/Aprovação/NF); pros demais tipos (RelatorioPerformance,
    RelatorioPCM) só existe uma etapa, com prazo = dataLimite do card."""
    if c["tipo"] == "BM":
        if idx == 0:
            return c.get("dataLimiteEnvioBM") or "", c.get("diasRestantesEnvioBM")
        if idx == 1:
            return c.get("dataLimiteAprovacao") or "", c.get("diasRestantesAprovacao")
        return c["dataLimite"], c["diasRestantes"]
    return c["dataLimite"], c["diasRestantes"]


def _verificar_alertas_etapas_abertas_se_necessario():
    """Piggyback no /sync-fracttal: roda 1x por dia na mesma janela de
    _verificar_compromissos_se_necessario (07:00-08:30). Varre TODAS as
    etapas de TODOS os compromissos não concluídos (qualquer cliente,
    qualquer tipo — BM, RelatorioPerformance, RelatorioPCM) e manda um
    único alerta consolidado pro grupo Gestão O&M (mesmo grupo dos
    resumos) sempre que alguma etapa já passou do prazo próprio dela e
    ainda não foi marcada como concluída. Pedido do Fred em 03/09/2026:
    alerta não filtra por cliente nem por etapa — é genérico."""
    try:
        agora = agora_br()
        hoje_str = agora.strftime("%Y-%m-%d")
        if not (agora.hour == 7 or (agora.hour == 8 and agora.minute <= 30)):
            return {"disparado": False, "motivo": f"fora da janela (agora {agora.strftime('%H:%M')})"}
        ja_feito = _ler_trava("alertas_etapas_abertas_enviado_em")
        if ja_feito == hoje_str:
            return {"disparado": False, "motivo": "já verificado hoje"}
        _gravar_trava("alertas_etapas_abertas_enviado_em", hoje_str)

        compromissos = _listar_compromissos_core()
        etapas_abertas = []
        for c in compromissos:
            if c["status"] == "Concluído":
                continue
            for idx, nome_etapa in enumerate(c["etapas"]):
                if c["etapasConcluidas"][idx]:
                    continue
                data_str, dias = _prazo_etapa_compromisso(c, idx)
                if not data_str or dias is None:
                    continue
                if dias <= 0:
                    etapas_abertas.append({
                        "cliente": c["cliente"], "tipoLabel": c["tipoLabel"],
                        "usina": c["usina"], "etapa": nome_etapa,
                        "dataLimite": data_str, "diasAtraso": -dias,
                    })

        if not etapas_abertas:
            return {"disparado": True, "etapasAbertas": 0}

        etapas_abertas.sort(key=lambda e: -e["diasAtraso"])
        linhas = []
        for e in etapas_abertas:
            usina_txt = f" ({e['usina']})" if e["usina"] else ""
            if e["diasAtraso"] == 0:
                prazo_txt = f"vence hoje ({e['dataLimite']})"
            else:
                prazo_txt = f"vencida há {e['diasAtraso']} dia(s) — prazo era {e['dataLimite']}"
            linhas.append(f"• {e['cliente']}{usina_txt} — {e['tipoLabel']} / {e['etapa']}: {prazo_txt}")

        texto = (
            f"⚠️ *Etapas em aberto — {agora.strftime('%d/%m/%Y')}*\n\n"
            + "\n".join(linhas)
            + "\n\nAbra o Painel Gerencial > Boletins de Medição pra marcar as etapas concluídas."
        )
        resultado_envio = _enviar_mensagem_grupo(GRUPO_GESTAO_OM_ID, texto)
        return {"disparado": True, "etapasAbertas": len(etapas_abertas), "envio": resultado_envio}
    except Exception as e:
        log.error(f"[Compromissos] Erro no alerta de etapas abertas: {e}")
        return {"disparado": False, "erro": str(e)}


def _verificar_e_disparar_resumo_diario_se_necessario():
    """Piggyback no /sync-fracttal: gera e envia o resumo diário todo dia
    na janela 17:00-17:30 (Brasília).

    Migrado do cron do GitHub Actions em 03/08/2026: o workflow estava
    agendado pra 20h UTC (17h BRT) mas o GitHub Actions atrasava o
    disparo em 35 a 55min em TODOS os dias observados (fila de runners
    compartilhados, comum em horários cheios como hora exata) — a
    geração e o envio em si sempre levaram menos de 30s, o atraso era
    100% do agendamento do Actions, não da geração. Mesmo padrão de
    piggyback já usado pra comunicados/compromissos: o gatilho real é o
    ping do UptimeRobot a cada 5min no /sync-fracttal, que não depende
    de fila de agendamento nenhuma.

    Se a geração falhar (ex.: timeout do Gemini), a trava é liberada
    pra tentar de novo no próximo ping dentro da mesma janela (até 6
    tentativas, uma a cada 5min entre 17:00 e 17:30).

    ATENÇÃO (bug corrigido em 06/08/2026): antes a trava era gravada
    ANTES de rodar a geração, com reset no except em caso de falha —
    mas isso não protege contra o worker do gunicorn ser matado no meio
    do caminho por timeout (SIGABRT, não é uma exceção Python normal,
    pula o except). Foi exatamente o que aconteceu em 06/08: o worker
    estourou os 160s de timeout no meio da auditoria de OS (fila de
    revalidação pesada colidindo com a quota de escrita do Google
    Sheets), foi abortado pelo gunicorn, a trava ficou gravada como
    "enviado hoje" sem o resumo ter sido gerado, e não teve nova
    tentativa no resto da janela. Agora a trava só é gravada DEPOIS que
    a geração retorna com sucesso — o pior caso passa a ser, na pior
    das hipóteses, duas tentativas quase simultâneas (risco desprezível
    dado o intervalo de 5min entre pings), nunca mais um dia inteiro
    sem resumo."""
    try:
        agora = agora_br()
        hoje_str = agora.strftime("%Y-%m-%d")
        if not (agora.hour == 17 and agora.minute <= 30):
            return {"disparado": False, "motivo": f"fora da janela (agora {agora.strftime('%H:%M')})"}
        ja_feito = _ler_trava("resumo_diario_enviado_em")
        if ja_feito == hoje_str:
            return {"disparado": False, "motivo": "já enviado hoje"}
        resultado = _gerar_resumo_diario_core(data_str=hoje_str, enviar=True)
        _gravar_trava("resumo_diario_enviado_em", hoje_str)
        return {"disparado": True, "resultado": resultado}
    except Exception as e:
        log.error(f"[ResumoDiario] Erro no piggyback: {e}")
        return {"disparado": False, "erro": str(e)}


def _verificar_e_disparar_resumo_semanal_se_necessario():
    """Piggyback no /sync-fracttal: gera e envia o resumo semanal só às
    sextas-feiras, mesma janela 17:00-17:30 e mesmo motivo da migração
    do resumo diário (atraso do cron do GitHub Actions). Trava só
    gravada após sucesso — mesma correção de 06/08/2026 do resumo
    diário (ver docstring de _verificar_e_disparar_resumo_diario_se_necessario)."""
    try:
        agora = agora_br()
        if agora.weekday() != 4:  # 4 = sexta-feira
            return {"disparado": False, "motivo": "não é sexta-feira"}
        hoje_str = agora.strftime("%Y-%m-%d")
        if not (agora.hour == 17 and agora.minute <= 30):
            return {"disparado": False, "motivo": f"fora da janela (agora {agora.strftime('%H:%M')})"}
        ja_feito = _ler_trava("resumo_semanal_enviado_em")
        if ja_feito == hoje_str:
            return {"disparado": False, "motivo": "já enviado hoje"}
        resultado = _gerar_resumo_semanal_core(data_fim_str=hoje_str, enviar=True)
        _gravar_trava("resumo_semanal_enviado_em", hoje_str)
        return {"disparado": True, "resultado": resultado}
    except Exception as e:
        log.error(f"[ResumoSemanal] Erro no piggyback: {e}")
        return {"disparado": False, "erro": str(e)}


@app.route("/corrigir-nomenclatura-preventiva", methods=["POST", "GET"])
def corrigir_nomenclatura_preventiva():
    """Correção retroativa de uso único: atividades multi-tarefa criadas
    antes da padronização MPM/MPS/MPA ainda têm o texto bruto da Fracttal
    como descrição (ex.: "[Grid Co.] - MPM") em vez de "PREVENTIVA MENSAL"
    / "Múltiplos equipamentos (Preventiva Mensal)". Esse endpoint varre e
    corrige as já existentes; daqui pra frente isso já é automático."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    aplicar = request.args.get("apply", "false").lower() == "true"

    _MULTI_EQUIP_MARCADORES = ("Múltiplas atividades", "Múltiplos equipamentos")

    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    corrigidas = []
    batch_updates = []
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        numero_os = row[13].strip()
        equipamento_atual = row[3].strip()
        descricao_atual = row[4].strip()
        if not numero_os or not equipamento_atual.startswith(_MULTI_EQUIP_MARCADORES):
            continue
        titulo, equip_novo = _fracttal_detectar_preventiva([{"description": descricao_atual}])
        if not titulo:
            continue
        if descricao_atual == titulo and equipamento_atual == equip_novo:
            continue  # já está correto
        corrigidas.append({"linha": i, "numeroOS": numero_os,
                            "de": {"descricao": descricao_atual, "equipamento": equipamento_atual},
                            "para": {"descricao": titulo, "equipamento": equip_novo}})
        if aplicar:
            batch_updates.append({"range": gspread.utils.rowcol_to_a1(i, 4), "values": [[equip_novo]]})
            batch_updates.append({"range": gspread.utils.rowcol_to_a1(i, 5), "values": [[titulo]]})

    if aplicar and batch_updates:
        TAMANHO_LOTE = 200
        for k in range(0, len(batch_updates), TAMANHO_LOTE):
            ws.batch_update(batch_updates[k:k + TAMANHO_LOTE], value_input_option="RAW")

    return jsonify({"ok": True, "aplicado": aplicar, "total": len(corrigidas), "corrigidas": corrigidas}), 200


@app.route("/corrigir-estado-revisao", methods=["POST", "GET"])
def corrigir_estado_revisao():
    """Correção retroativa de uso único: reabre atividades marcadas como
    'Concluído' internamente cujo statusOS na Fracttal ainda é 'Em Revisão'
    (não 'Finalizada'). Bug anterior tratava Em Revisão como conclusão."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    aplicar = request.args.get("apply", "false").lower() == "true"

    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    corrigidas = []
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        numero_os = row[13].strip()
        status_interno = row[8].strip()
        status_os = row[14].strip()
        if numero_os and status_interno == "Concluído" and status_os == "Em Revisão":
            corrigidas.append({"linha": i, "numeroOS": numero_os})
            if aplicar:
                _gravar_status_interno(ws, i, "Em Aberto")
                entry = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - ⚠️ Correção retroativa: OS reaberta — "
                         f"estava marcada como concluída, mas o estado na Fracttal é \"Em Revisão\", "
                         f"não \"Finalizada\" (bug de interpretação de estado corrigido).")
                hist_atual = row[ATIV_COL_HISTORICO - 1] if len(row) >= ATIV_COL_HISTORICO else ""
                ws.update_cell(i, ATIV_COL_HISTORICO, f"{hist_atual}\n{entry}".strip() if hist_atual else entry)
                time.sleep(0.3)

    return jsonify({"ok": True, "aplicado": aplicar, "total": len(corrigidas), "corrigidas": corrigidas}), 200


@app.route("/config-ler", methods=["GET"])
def config_ler():
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    ws_cfg = _get_config_sheet()
    valores = ws_cfg.get_all_values()
    pares = {row[0]: (row[1] if len(row) > 1 else "") for row in valores[1:] if row and row[0].strip()}
    return jsonify({"ok": True, "pares": pares}), 200


@app.route("/config-limpar-cache", methods=["POST"])
def config_limpar_cache():
    """Zera o cache em memória de _mapa_grupo_usina/_mapa_cluster_usina —
    útil depois de editar a aba _Sistema (ex.: corrigir nome de usina)
    quando não dá pra esperar os 10 minutos do cache normal."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    _mapa_grupo_usina_cache["dados"] = None
    _mapa_grupo_usina_cache["expira_em"] = 0
    _mapa_cluster_usina_cache["dados"] = None
    _mapa_cluster_usina_cache["expira_em"] = 0
    return jsonify({"ok": True}), 200


@app.route("/config-remover", methods=["POST"])
def config_remover():
    """Remove uma ou mais chaves da aba _Sistema (linha inteira apagada).
    Usado pra corrigir duplicatas/erros de cadastro, ex.: uma usina
    registrada duas vezes com nomes ligeiramente diferentes. Compara
    normalizando acentuação (NFC) — strings com "ã" podem estar
    guardadas com codificação Unicode diferente (precomposta vs
    combinando caracteres) dependendo de como foram digitadas
    originalmente, e uma comparação direta de string falha nesse caso
    sem dar nenhum aviso."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    dados = request.get_json(force=True, silent=True) or {}
    chaves = dados.get("chaves", [])
    resultado = _config_remover_chaves_core(chaves)
    return jsonify({"ok": True, **resultado}), 200


def _config_remover_chaves_core(chaves):
    """Núcleo de /config-remover, reaproveitado internamente (ex.: por
    /supervisao-temporaria/remover) sem precisar de uma chamada HTTP
    própria. Remove da aba _Sistema todas as linhas cuja chave (coluna A)
    bata com alguma da lista, normalizando acentuação (NFC) pra evitar
    falso-negativo por codificação Unicode diferente."""
    if not chaves:
        return {"removidos": 0, "chavesEncontradas": []}

    import unicodedata
    chaves_norm = {unicodedata.normalize("NFC", c.strip()) for c in chaves}

    ws_cfg = _get_config_sheet()
    valores = ws_cfg.get_all_values()
    encontradas = []
    linhas_para_remover = []
    for i, row in enumerate(valores[1:], start=2):
        if row and unicodedata.normalize("NFC", row[0].strip()) in chaves_norm:
            linhas_para_remover.append(i)
            encontradas.append(row[0])
    linhas_para_remover.sort(reverse=True)  # de baixo pra cima, pra não bagunçar os índices ao deletar
    for idx in linhas_para_remover:
        ws_cfg.delete_rows(idx)

    return {"removidos": len(linhas_para_remover), "chavesEncontradas": encontradas}


@app.route("/config-set-lote", methods=["POST"])
def config_set_lote():
    """Grava múltiplos pares chave/valor na aba _Sistema de uma vez, numa
    única leitura + uma única escrita em lote (evita estourar a cota da
    API do Google Sheets, que é o que acontecia gravando um por um)."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    dados = request.get_json(force=True, silent=True) or {}
    pares = dados.get("pares", {})
    gravados = _config_set_lote_core(pares)
    return jsonify({"ok": True, "gravados": gravados}), 200


def _config_set_lote_core(pares):
    """Núcleo de /config-set-lote, reaproveitado internamente (ex.: por
    /supervisao-temporaria/adicionar) sem precisar de uma chamada HTTP
    própria. Grava múltiplos pares chave/valor na aba _Sistema numa
    única leitura + uma única escrita em lote."""
    if not pares:
        return []

    ws_cfg = _get_config_sheet()
    valores = ws_cfg.get_all_values()
    linha_existente = {row[0].strip(): i for i, row in enumerate(valores[1:], start=2) if row}

    batch_updates = []
    novas_linhas = []
    for chave, valor in pares.items():
        if chave in linha_existente:
            batch_updates.append({
                "range": gspread.utils.rowcol_to_a1(linha_existente[chave], 2),
                "values": [[valor]],
            })
        else:
            novas_linhas.append([chave, valor])

    if batch_updates:
        ws_cfg.batch_update(batch_updates, value_input_option="RAW")
    if novas_linhas:
        ws_cfg.append_rows(novas_linhas, value_input_option="RAW")

    return list(pares.keys())


# ── Comunicados diários automáticos (WhatsApp) ──────────────────────────
# Mapeamento usina → grupo do WhatsApp fica na aba "_Sistema", chaves no
# formato "grupo_usina:<Nome da Usina>" = "<id>@g.us". Fred edita essa
# aba diretamente na planilha pra adicionar/trocar grupos, sem precisar de
# um novo deploy. Usinas sem grupo configurado são simplesmente ignoradas
# (não dá erro, só não recebem comunicado).

_MAPA_CACHE_DISCO_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zeladoria_fotos")


def _mapa_cache_disco_salvar(nome, mapa):
    try:
        os.makedirs(_MAPA_CACHE_DISCO_DIR, exist_ok=True)
        with open(os.path.join(_MAPA_CACHE_DISCO_DIR, f"cache_{nome}.json"), "w", encoding="utf-8") as f:
            json.dump(mapa, f)
    except Exception as e:
        log.error(f"[_mapa_cache_disco_salvar] Falha ao salvar cache '{nome}' em disco: {e}")


def _mapa_cache_disco_carregar(nome):
    caminho = os.path.join(_MAPA_CACHE_DISCO_DIR, f"cache_{nome}.json")
    if not os.path.exists(caminho):
        return None
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.error(f"[_mapa_cache_disco_carregar] Falha ao ler cache '{nome}' do disco: {e}")
        return None


_mapa_grupo_usina_cache = {"dados": None, "expira_em": 0}


def _mapa_grupo_usina():
    """Cache de 10 min (era 30s) — essa função é chamada muitas vezes
    durante o processamento de fotos de zeladoria, e mesmo com cache
    curto ainda contribuía pra estourar a cota de leitura do Google
    Sheets, principalmente logo após um restart do servidor (cache
    vazio, todo mundo pedindo ao mesmo tempo). _Sistema muda raramente
    (só quando Fred edita manualmente), então 10 min de defasagem não é
    problema. Se a atualização falhar (ex.: cota momentaneamente
    esgotada), usa o cache em memória se tiver, senão um cache salvo em
    disco (sobrevive a restart do servidor) — só levanta erro de
    verdade se nenhum dos dois existir ainda."""
    agora_ts = time.time()
    if _mapa_grupo_usina_cache["dados"] is not None and agora_ts < _mapa_grupo_usina_cache["expira_em"]:
        return _mapa_grupo_usina_cache["dados"]
    try:
        ws_cfg = _get_config_sheet()
        valores = _gspread_retry(lambda: ws_cfg.get_all_values())
    except Exception as e:
        if _mapa_grupo_usina_cache["dados"] is not None:
            log.error(f"[_mapa_grupo_usina] Falha ao atualizar ({e}) — usando cache em memória")
            return _mapa_grupo_usina_cache["dados"]
        do_disco = _mapa_cache_disco_carregar("grupo_usina")
        if do_disco is not None:
            log.error(f"[_mapa_grupo_usina] Falha ao atualizar ({e}) — usando cache salvo em disco")
            _mapa_grupo_usina_cache["dados"] = do_disco
            _mapa_grupo_usina_cache["expira_em"] = agora_ts + 60  # tenta de novo em breve
            return do_disco
        raise
    mapa = {}
    for row in valores[1:]:
        if row and row[0].strip().startswith("grupo_usina:"):
            usina = row[0].strip()[len("grupo_usina:"):].strip()
            grupo_id = row[1].strip() if len(row) > 1 else ""
            if usina and grupo_id:
                mapa[usina] = grupo_id
    _mapa_grupo_usina_cache["dados"] = mapa
    _mapa_grupo_usina_cache["expira_em"] = agora_ts + 600
    _mapa_cache_disco_salvar("grupo_usina", mapa)
    return mapa


def _grupos_ids_ativos():
    """Retorna o conjunto de grupo_id do WhatsApp que correspondem a
    usinas ATUALMENTE reconhecidas (catálogo permanente OU sob
    Supervisão Temporária ativa agora), mais o grupo pessoal de gestão
    (que não é atrelado a nenhuma usina específica).

    Corrigido 18/08/2026, relatado pelo Fred: quando uma usina é
    removida da Supervisão Temporária, o /atividades já parava de
    mostrar as atividades dela (correção de 31/07/2026, usina_permitida
    aplicada linha a linha), mas resumo diário/semanal e ronda liam as
    mensagens capturadas do WhatsApp DIRETO do banco sqlite
    (_buscar_mensagens_periodo), sem checar se o grupo da mensagem ainda
    correspondia a uma usina válida — então o grupo da usina devolvida
    (ex.: "O&M - San. Bárb./Pirac. - SP LESTE 03") continuava sendo
    resumido normalmente na seção "RESUMO POR EQUIPE" e entrando no
    cruzamento de OS por número, trazendo a usina de volta pro relatório
    mesmo depois de removida. Esta função é o filtro que faltava — use
    sempre antes de agrupar mensagens pra IA num resumo/ronda."""
    grupos = {GRUPO_GESTAO_OM_ID}
    try:
        for usina, grupo_id in _mapa_grupo_usina().items():
            if grupo_id and usina_permitida(usina):
                grupos.add(grupo_id)
    except Exception as e:
        log.error(f"[GruposAtivos] Erro ao montar lista de grupos válidos: {e}")
    return grupos


def _nome_amigavel_grupo(grupo_id):
    """Resolve um ID bruto de grupo do WhatsApp (ex.: '120363...@g.us')
    pro nome da(s) usina(s) que esse grupo atende, usando o mapeamento já
    existente — em vez de expor o número do grupo cru no histórico."""
    try:
        mapa = _mapa_grupo_usina()  # usina -> grupo_id
        usinas = sorted({u for u, g in mapa.items() if g == grupo_id})
        if usinas:
            return " / ".join(usinas[:3]) + (" e outras" if len(usinas) > 3 else "")
    except Exception:
        pass
    return None


def _editor_legivel(editor):
    """Traduz identificadores internos de 'quem fez a alteração' pra texto
    apresentável (inclusive pro cliente ver) — nunca mostra ID de grupo do
    WhatsApp, nome de rotina interna (ex.: 'fracttal-sync'), etc. direto no
    histórico. Adicionado em 17/07/2026 a pedido do Fred, depois de reparar
    que o histórico mostrava coisas como 'tecnico:120363...' pro cliente."""
    editor = (editor or "").strip()
    if not editor:
        return "sistema"
    if editor.startswith("tecnico:"):
        grupo_id = editor[len("tecnico:"):]
        nome = _nome_amigavel_grupo(grupo_id)
        return f"técnico de campo ({nome})" if nome else "técnico de campo (via WhatsApp)"
    mapa = {
        "fracttal-sync": "sincronização automática com a Fracttal",
        "fracttal-backfill": "sincronização automática com a Fracttal",
        "claude-chat": "assistente (Claude)",
        "reprogramacao-ia": "reprogramação sugerida por IA",
    }
    return mapa.get(editor, editor)


_MIGRAR_HIST_PADRAO_CRIACAO = re.compile(
    r'^(?P<data>\d{2}/\d{2}/\d{4} \d{2}:\d{2}) - Atividade criada por (fracttal-sync|fracttal-backfill)\.$')
_MIGRAR_HIST_PADRAO_VISUALIZADO = re.compile(
    r'^(?P<data>\d{2}/\d{2}/\d{4} \d{2}:\d{2}) - visualizado alterado de "(?P<de>.*?)" para "(?P<para>.*?)" por (?P<editor>.*?)\.$')
_MIGRAR_HIST_PADRAO_TECNICO_DESC = re.compile(
    r'^(?P<data>\d{2}/\d{2}/\d{4} \d{2}:\d{2}) - tecnico:(?P<grupo>\d+): (?P<texto>.*)$')
_MIGRAR_HIST_PADRAO_TECNICO_STATUS = re.compile(
    r'^(?P<data>\d{2}/\d{2}/\d{4} \d{2}:\d{2}) - tecnico:(?P<grupo>\d+) reportou status "(?P<status>.*?)" '
    r'pelo WhatsApp — verificando direto na Fracttal \(o status real vem de lá, não da mensagem\)\.$')
_MIGRAR_HIST_PADRAO_STATUS_OS = re.compile(
    r'^(?P<data>\d{2}/\d{2}/\d{4} \d{2}:\d{2}) - Status na OS \(Fracttal\) atualizado: '
    r'"(?P<de>.*?)" → "(?P<para>.*?)", (?P<pde>\d+)% → (?P<ppara>\d+)% \((?P<geral>.*?)\)\.$')
_MIGRAR_HIST_PADRAO_GENERICO_TECNICO = re.compile(r'^(?P<prefixo>.*) por tecnico:(?P<grupo>\d+)\.$')


def _migrar_linha_historico(linha):
    """Reescreve uma única linha de histórico (formato antigo) pro formato
    novo, mais legível. Devolve a linha original sem alterar se nenhum dos
    padrões conhecidos bater — nunca inventa nem apaga informação."""
    m = _MIGRAR_HIST_PADRAO_CRIACAO.match(linha)
    if m:
        return f'{m["data"]} - Atividade criada por sincronização automática com a Fracttal.'

    m = _MIGRAR_HIST_PADRAO_VISUALIZADO.match(linha)
    if m and m["para"].strip().lower() == "sim":
        return f'{m["data"]} - Marcado como visualizado ({m["editor"]}).'

    m = _MIGRAR_HIST_PADRAO_TECNICO_DESC.match(linha)
    if m:
        nome = _nome_amigavel_grupo(m["grupo"]) or "via WhatsApp"
        return f'{m["data"]} - técnico de campo ({nome}): {m["texto"]}'

    m = _MIGRAR_HIST_PADRAO_TECNICO_STATUS.match(linha)
    if m:
        nome = _nome_amigavel_grupo(m["grupo"]) or "via WhatsApp"
        return (f'{m["data"]} - técnico de campo ({nome}) reportou status "{m["status"]}" '
                f'pelo WhatsApp (confirmado em seguida direto com a Fracttal).')

    m = _MIGRAR_HIST_PADRAO_STATUS_OS.match(linha)
    if m:
        partes = []
        if m["de"] != m["para"]:
            partes.append(f'status na Fracttal mudou de "{m["de"]}" para "{m["para"]}"')
        if m["pde"] != m["ppara"]:
            partes.append(f'progresso da tarefa foi de {m["pde"]}% para {m["ppara"]}%')
        if not partes:
            # nem status nem percentual mudaram no texto antigo — o único
            # jeito de "mudou" ter sido true na época é a situação geral da
            # tarefa ter mudado; não temos o valor "de" no texto antigo
            # (só foi gravado o "para"), então afirmamos só o que sabemos
            # de verdade, sem inventar uma transição que não temos como
            # confirmar.
            partes.append(f'situação geral da tarefa: "{m["geral"]}"')
        return f'{m["data"]} - ' + "; ".join(partes) + "."

    # Padrão genérico: PEGA QUALQUER linha que termine em "por tecnico:ID."
    # (ex.: "Responsável alterado de X para Y por tecnico:123."), não só os
    # formatos específicos já tratados acima — cobre qualquer campo editado
    # por um técnico via WhatsApp, sem precisar prever cada rótulo de campo.
    m = _MIGRAR_HIST_PADRAO_GENERICO_TECNICO.match(linha)
    if m:
        nome = _nome_amigavel_grupo(m["grupo"]) or "via WhatsApp"
        return f'{m["prefixo"]} por técnico de campo ({nome}).'

    return linha


@app.route("/migrar-historico-legivel", methods=["POST", "OPTIONS"])
def migrar_historico_legivel():
    """
    Reescreve retroativamente o texto já salvo no histórico de TODAS as
    atividades, aplicando os mesmos formatos mais legíveis usados a partir
    de 17/07/2026 (sem ID de grupo do WhatsApp cru, sem nome de rotina
    interna tipo 'fracttal-sync', sem 'X → X' quando nada mudou de verdade).

    Por padrão roda em modo TESTE (aplicar=false): não grava nada, só
    devolve quantas linhas/atividades seriam alteradas e uma amostra, pra
    conferir antes de aplicar de verdade.

    Corpo esperado (opcional): {"aplicar": true}
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    body = request.get_json(force=True, silent=True) or {}
    aplicar = bool(body.get("aplicar", False))

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()

        atualizacoes = []  # {"range": f"L{i}", "values": [[novo_historico]]}
        amostra = []
        atividades_afetadas = 0
        linhas_afetadas = 0

        for i, row in enumerate(todos[1:], start=2):
            hist_col = ATIV_COL_HISTORICO - 1
            if len(row) <= hist_col or not row[hist_col].strip():
                continue
            hist_original = row[hist_col]
            linhas_originais = hist_original.split("\n")
            linhas_novas = [_migrar_linha_historico(l) for l in linhas_originais]
            if linhas_novas == linhas_originais:
                continue

            atividades_afetadas += 1
            n_mudou = sum(1 for a, b in zip(linhas_originais, linhas_novas) if a != b)
            linhas_afetadas += n_mudou
            hist_novo = "\n".join(linhas_novas)

            if len(amostra) < 5:
                amostra.append({
                    "id": row[0] if row else "?",
                    "antes": [l for l, n in zip(linhas_originais, linhas_novas) if l != n][:3],
                    "depois": [n for l, n in zip(linhas_originais, linhas_novas) if l != n][:3],
                })

            if aplicar:
                col_letra = chr(64 + ATIV_COL_HISTORICO) if ATIV_COL_HISTORICO <= 26 else "AA"
                atualizacoes.append({"range": f"{col_letra}{i}", "values": [[hist_novo]]})

        if aplicar and atualizacoes:
            ws.batch_update(atualizacoes)

        return jsonify({
            "ok": True,
            "aplicado": aplicar,
            "atividades_afetadas": atividades_afetadas,
            "linhas_afetadas": linhas_afetadas,
            "amostra": amostra,
        }), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500



_mapa_cluster_usina_cache = {"dados": None, "expira_em": 0}

_usinas_temporarias_cache = {"dados": None, "expira_em": 0}
_indices_temporarios_cache = {"alias": {}, "cliente": {}, "expira_em": 0}


def _usinas_temporarias():
    """Lista as usinas atualmente sob supervisão temporária do Fred
    (emprestadas de outro supervisor, ex.: cobertura de férias) — lidas
    da aba _SupervisaoTemporaria. Cache de 60s (era 3 min, subiu pra 10
    min mais cedo em 31/07/2026 pra aliviar a cota de leitura do Sheets,
    depois voltou pra 60s no mesmo dia: com Gunicorn rodando 2 workers,
    cada um tem seu próprio cache em memória — não há estado
    compartilhado entre processos — então invalidar o cache num worker
    (ex.: ao remover uma usina) não afeta o outro, e esse painel filtra
    /atividades por essa lista (usina_permitida), então cache velho
    demais fazia usina já devolvida continuar aparecendo por até 10 min
    dependendo de qual worker atendesse a requisição. 60s é o meio-termo:
    ainda alivia bastante a frequência de leitura comparado ao original,
    mas limita a inconsistência entre workers a no máximo 1 min.
    Implementado em 30/07/2026 a pedido do Fred; restaurado em 31/07/2026
    depois que uma sessão paralela sobrescreveu o app.py sem essa
    funcionalidade (ver histórico de commits — commit 3b1d1e77c0)."""
    agora_ts = time.time()
    if _usinas_temporarias_cache["dados"] is not None and agora_ts < _usinas_temporarias_cache["expira_em"]:
        return _usinas_temporarias_cache["dados"]
    try:
        ws = get_supervisao_temp_sheet()
        valores = ws.get_all_values()
    except Exception as e:
        if _usinas_temporarias_cache["dados"] is not None:
            log.error(f"[_usinas_temporarias] Falha ao atualizar ({e}) — usando cache em memória")
            return _usinas_temporarias_cache["dados"]
        return []
    itens = []
    for row in valores[1:]:
        if len(row) >= 2 and row[1].strip():
            itens.append({
                "cliente": row[0].strip() if len(row) > 0 else "",
                "usina": row[1].strip(),
                "cluster": row[2].strip() if len(row) > 2 else "",
                "responsavelOriginal": row[3].strip() if len(row) > 3 else "",
                "adicionadoEm": row[4].strip() if len(row) > 4 else "",
                "grupoId": row[5].strip() if len(row) > 5 else "",
            })
    _usinas_temporarias_cache["dados"] = itens
    _usinas_temporarias_cache["expira_em"] = agora_ts + 60
    return itens


def _indices_temporarios():
    """Constrói (alias_index, cliente_index) a partir de _usinas_temporarias
    — mesma ideia do _ALIAS_INDEX/_CLIENTE_INDEX estáticos, só que
    recarregado periodicamente em vez de fixo na inicialização (pois essa
    lista muda em tempo real conforme o Fred adiciona/remove usinas)."""
    agora_ts = time.time()
    if agora_ts < _indices_temporarios_cache["expira_em"]:
        return _indices_temporarios_cache["alias"], _indices_temporarios_cache["cliente"]
    alias_temp, cliente_temp = {}, {}
    for item in _usinas_temporarias():
        nome_oficial = item["usina"]  # usa o nome exato do PCM como "oficial" pra essas emprestadas
        cliente_temp[nome_oficial] = item["cliente"]
        alias_temp[_norm_usina(nome_oficial)] = nome_oficial
        # adiciona também variações comuns: sem o sufixo "- UF", só a parte do meio
        m = re.match(r"^(.+?)\s*-\s*(.+?)\s*-\s*\w{2}$", nome_oficial)
        if m:
            alias_temp[_norm_usina(m.group(2))] = nome_oficial
            alias_temp[_norm_usina(f"{m.group(1)} - {m.group(2)}")] = nome_oficial
    _indices_temporarios_cache["alias"] = alias_temp
    _indices_temporarios_cache["cliente"] = cliente_temp
    _indices_temporarios_cache["expira_em"] = agora_ts + 60
    return alias_temp, cliente_temp


def _mapa_cluster_usina():
    """Mapeia usina -> código de cluster/equipe regional (ex.: 'SP Centro
    01'), configurado na aba _Sistema como 'cluster_usina:<Usina>'.
    Cache de 10 min com fallback pro cache em memória e depois pro cache
    salvo em disco (sobrevive a restart) em caso de falha — mesmo motivo
    do _mapa_grupo_usina."""
    agora_ts = time.time()
    if _mapa_cluster_usina_cache["dados"] is not None and agora_ts < _mapa_cluster_usina_cache["expira_em"]:
        return _mapa_cluster_usina_cache["dados"]
    try:
        ws_cfg = _get_config_sheet()
        valores = _gspread_retry(lambda: ws_cfg.get_all_values())
    except Exception as e:
        if _mapa_cluster_usina_cache["dados"] is not None:
            log.error(f"[_mapa_cluster_usina] Falha ao atualizar ({e}) — usando cache em memória")
            return _mapa_cluster_usina_cache["dados"]
        do_disco = _mapa_cache_disco_carregar("cluster_usina")
        if do_disco is not None:
            log.error(f"[_mapa_cluster_usina] Falha ao atualizar ({e}) — usando cache salvo em disco")
            _mapa_cluster_usina_cache["dados"] = do_disco
            _mapa_cluster_usina_cache["expira_em"] = agora_ts + 60
            return do_disco
        raise
    mapa = {}
    for row in valores[1:]:
        if row and row[0].strip().startswith("cluster_usina:"):
            usina = row[0].strip()[len("cluster_usina:"):].strip()
            cluster = row[1].strip() if len(row) > 1 else ""
            if usina and cluster:
                mapa[usina] = cluster
    _mapa_cluster_usina_cache["dados"] = mapa
    _mapa_cluster_usina_cache["expira_em"] = agora_ts + 600
    _mapa_cache_disco_salvar("cluster_usina", mapa)
    return mapa


_mapa_coordenador_cluster_cache = {"dados": None, "expira_em": 0}


def _mapa_coordenador_cluster():
    """Mapeia cluster -> coordenador/técnico responsável, configurado na
    aba _Sistema como 'coordenador_cluster:<Cluster>' (ex.: gravado via
    /config-set-lote). Levantado por vistoria em 26/08/2026, cruzando o
    campo 'responsavel' das atividades reais com o cluster de cada usina
    — substitui uma tabela fixa que existia hardcoded no prompt do
    chat-ia (ficava desatualizada sem ninguém perceber). Mesmo padrão de
    cache de _mapa_cluster_usina."""
    agora_ts = time.time()
    if _mapa_coordenador_cluster_cache["dados"] is not None and agora_ts < _mapa_coordenador_cluster_cache["expira_em"]:
        return _mapa_coordenador_cluster_cache["dados"]
    try:
        ws_cfg = _get_config_sheet()
        valores = _gspread_retry(lambda: ws_cfg.get_all_values())
    except Exception as e:
        if _mapa_coordenador_cluster_cache["dados"] is not None:
            log.error(f"[_mapa_coordenador_cluster] Falha ao atualizar ({e}) — usando cache em memória")
            return _mapa_coordenador_cluster_cache["dados"]
        do_disco = _mapa_cache_disco_carregar("coordenador_cluster")
        if do_disco is not None:
            log.error(f"[_mapa_coordenador_cluster] Falha ao atualizar ({e}) — usando cache salvo em disco")
            _mapa_coordenador_cluster_cache["dados"] = do_disco
            _mapa_coordenador_cluster_cache["expira_em"] = agora_ts + 60
            return do_disco
        raise
    mapa = {}
    for row in valores[1:]:
        if row and row[0].strip().startswith("coordenador_cluster:"):
            cluster = row[0].strip()[len("coordenador_cluster:"):].strip()
            nome = row[1].strip() if len(row) > 1 else ""
            if cluster and nome:
                mapa[cluster] = nome
    _mapa_coordenador_cluster_cache["dados"] = mapa
    _mapa_coordenador_cluster_cache["expira_em"] = agora_ts + 600
    _mapa_cache_disco_salvar("coordenador_cluster", mapa)
    return mapa


def _montar_texto_comunicado_usina(usina, atividades):
    def dias_atraso(prazo):
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", prazo or "")
        if not m:
            return None
        d, mth, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return (agora_br().date() - datetime(y, mth, d).date()).days

    atividades = sorted(atividades, key=lambda a: dias_atraso(a.get("prazo", "")) or -999, reverse=True)
    hoje_str = agora_br().strftime("%d/%m/%Y")
    txt = f"📋 *Atividades em aberto — {usina}*\n📅 {hoje_str}\n\n"
    txt += f"Existem *{len(atividades)} atividade{'s' if len(atividades) != 1 else ''}* pendente{'s' if len(atividades) != 1 else ''}:\n\n"
    for i, a in enumerate(atividades, start=1):
        dias = dias_atraso(a.get("prazo", ""))
        atrasada = dias is not None and dias > 0
        numero_os = a.get("numeroOS", "")
        txt += f"{i}. {'🔴' if atrasada else '🟢'} {'OS ' + numero_os + ' — ' if numero_os else ''}{a.get('equipamento', '')}\n"
        txt += f"   {a.get('descricao', '')}\n"
        if a.get("prazo"):
            txt += f"   📅 Prazo: {a['prazo']}" + (f" (atrasada há {dias} dia{'s' if dias > 1 else ''})" if atrasada else "") + "\n"
        txt += "\n"
    txt += "Por favor, atualizem o andamento dessas atividades o quanto antes. Qualquer dificuldade, me avisem."
    return txt


@app.route("/verificar-e-enviar-comunicados", methods=["POST", "GET"])
def _verificar_e_disparar_comunicados_se_necessario():
    """DESATIVADA a pedido do Fred em 15/07/2026 — ver detalhes no
    corpo da função. Mantida com esse nome (em vez de apagada) porque
    está diretamente exposta como rota pública (/verificar-e-enviar-
    comunicados) — se algum monitor externo (UptimeRobot ou outro) ainda
    estiver batendo nela numa agenda própria, precisa continuar
    respondendo 200 sem disparar nada, em vez de dar erro.

    Histórico do problema: o piggyback dentro de /sync-fracttal já tinha
    sido desligado (ligava aqui indiretamente), mas essa função também
    é chamada DIRETO por essa rota própria — que outra sessão/monitor
    pode estar acionando de forma independente, a cada poucos minutos,
    sem eu saber. Isso explicava comunicados continuando a sair sozinhos
    mesmo depois do primeiro desligamento. Agora a desativação está na
    fonte única (aqui dentro), então funciona não importa quem chame."""
    return {"disparado": False, "motivo": "disparo automático desativado — use o botão Comunicados no painel"}


def _verificar_e_disparar_comunicados_se_necessario_DESATIVADA_ORIGINAL():
    """Lógica compartilhada: só dispara o envio de verdade se for dia útil,
    estiver na janela 07:00-08:30 (BRT) e ainda não tiver sido enviado
    hoje. Retorna um dict com o resultado (nunca levanta exceção pro
    chamador, pra nunca quebrar quem estiver piggybackando nela).

    Janela alargada de 9 minutos pra 90 minutos em 15/07/2026: um dia em
    que o Render aparentemente estava com cold-start bem lento resultou
    em 502 em TODAS as tentativas dentro da janela original (confirmado
    manualmente: 1ª tentativa deu 502, 2ª — minutos depois — funcionou).
    Como o ciclo de 5 em 5 min só tinha ~2 chances dentro de 9 minutos,
    um dia ruim de Render bastava pra perder o comunicado inteiro. Com
    90 minutos de janela, sobra bastante margem pro serviço esquentar
    sozinho sem precisar de intervenção manual."""
    try:
        agora = agora_br()
        hoje_str = agora.strftime("%Y-%m-%d")

        if agora.weekday() >= 5:  # sábado=5, domingo=6
            return {"disparado": False, "motivo": "fim de semana"}
        if not (agora.hour == 7 or (agora.hour == 8 and agora.minute <= 30)):
            return {"disparado": False, "motivo": f"fora da janela (agora {agora.strftime('%H:%M')})"}

        ja_enviado = _ler_trava("comunicados_enviados_em")
        if ja_enviado == hoje_str:
            return {"disparado": False, "motivo": "já enviado hoje"}

        # a trava só é gravada DEPOIS de confirmar que o envio foi tentado
        # de verdade — antes ela era gravada logo antes de chamar a função
        # de envio, então uma falha (ex.: WhatsApp reconectando naquele
        # minuto) deixava o dia "marcado como enviado" sem nada ter saído,
        # bloqueando qualquer nova tentativa até o dia seguinte (bug
        # identificado em 14/07/2026, depois de um dia sem comunicado).
        resultado = _enviar_comunicados_diarios_core()
        if resultado.get("ok", True) is False and not resultado.get("enviados"):
            log.error(f"[ComunicadosDiarios] Envio falhou, trava NÃO gravada (tenta de novo no próximo ciclo): {resultado}")
            try:
                enviar_push(
                    titulo="⚠️ Comunicados de hoje não saíram",
                    corpo=f"Falha no envio às {agora.strftime('%H:%M')}: {resultado.get('error', 'motivo desconhecido')}. Tentará de novo em 5 min.",
                    tipo="comunicados_falha",
                )
            except Exception:
                pass
            return {"disparado": False, "motivo": "falha no envio, tentará de novo no próximo ciclo", "resultado": resultado}

        _gravar_trava("comunicados_enviados_em", hoje_str)
        return {"disparado": True, "resultado": resultado}
    except Exception as e:
        log.error(f"[ComunicadosDiarios] Erro na verificação/disparo: {e}")
        return {"disparado": False, "erro": str(e)}


def verificar_e_enviar_comunicados():
    """Ponto de entrada seguro pra ser chamado com frequência (ex.: a cada
    5 min via UptimeRobot) — só dispara o envio de verdade se:
      1. for dia útil (seg-sex) e estiver dentro da janela 07:00-07:09 (BRT)
      2. ainda não tiver sido enviado hoje (trava em _Sistema)
    Isso substitui o cron do GitHub Actions como gatilho principal, porque
    ele atrasa de forma imprevisível (chegou a disparar 7h30 depois do
    horário configurado). Fora da janela ou já enviado hoje, retorna sem
    fazer nada (barato, seguro de chamar repetidamente)."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    r = _verificar_e_disparar_comunicados_se_necessario()
    return jsonify({"ok": True, **r}), 200


@app.route("/enviar-comunicados-diarios", methods=["POST", "GET"])
def enviar_comunicados_diarios():
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    return jsonify(_enviar_comunicados_diarios_core()), 200


def _enviar_comunicados_diarios_core():
    if not WPP_SERVER_URL:
        return {"ok": False, "error": "WPP_SERVER_URL não configurado"}

    mapa_grupos = _mapa_grupo_usina()
    if not mapa_grupos:
        return {"ok": False, "error": ("Nenhum grupo configurado. Adicione linhas na aba "
                "_Sistema no formato \"grupo_usina:<Usina>\" = \"<id>@g.us\".")}

    ws = get_atividades_sheet()
    todos = ws.get_all_values()

    # ── Revalida AO VIVO na Fracttal cada candidata antes de decidir se
    # entra no comunicado — o dado gravado (statusOS) pode estar
    # ligeiramente desatualizado se essa OS específica ainda não tiver
    # caído no rodízio automático desde que o técnico a moveu pra
    # "Em Verificação" na Fracttal. Como o comunicado só roda 1x/dia,
    # vale o custo de checar direto na fonte pra garantir que nenhuma OS
    # já resolvida pelo técnico seja cobrada de novo (bug relatado pelos
    # técnicos em 14/07/2026 — "Em Verificação" sendo enviada mesmo
    # assim, por causa de dado desatualizado no momento do envio).
    candidatas_recheck = []
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        if not row[0].strip():
            continue
        status = row[8].strip()
        if _is_concluido_atividade(status):
            continue
        numero_os = row[13].strip()
        status_os = row[14].strip()
        if numero_os and status_os != "Em Processo":
            # regra simples e direta: só OS com estado "Em Processo" na
            # Fracttal entra no comunicado — qualquer outra coisa (Em
            # Revisão, Finalizada, Cancelada, vazio, ou qualquer estado
            # futuro que a Fracttal venha a usar) fica de fora por padrão,
            # em vez de tentar prever e listar cada estado que deveria
            # excluir (regra anterior, mais frágil — deixava passar coisa
            # nova que não estivesse na lista). Ajustado 15/07/2026.
            continue
        etiquetas = row[ATIV_CAMPO_COL["etiquetasOS"] - 1].strip().upper()
        if "PERFORMANCE" in etiquetas:
            # etiquetada na Fracttal como tarefa de análise de performance —
            # normalmente atribuída a um analista, não ao técnico de campo.
            # Mandar isso pro grupo da equipe só confunde quem recebe (não
            # é responsabilidade deles) — identificado 14/07/2026 com a OS
            # 8025 (Boa Esperança do Sul I), etiquetada PERFORMANCE e
            # atribuída a um analista, mas enviada ao grupo de campo.
            continue
        if numero_os:
            candidatas_recheck.append((i, row, numero_os))

    LIMITE_RECHECK_COMUNICADOS = 20  # trava de segurança de tempo — essa função
    # roda no mesmo ciclo que outras checagens (auditoria), então não pode
    # crescer sem limite. Prioriza as mais desatualizadas primeiro.
    candidatas_recheck.sort(key=lambda t: t[1][ATIV_CAMPO_COL["ultimaVerificacaoOS"] - 1] or "")
    candidatas_recheck = candidatas_recheck[:LIMITE_RECHECK_COMUNICADOS]

    for i, row, numero_os in candidatas_recheck:
        _fracttal_verificar_e_atualizar_uma_os(ws, i, row, numero_os, enviar_notificacao=False)
        time.sleep(0.35)

    # rebusca do zero — garante que a seleção final usa o dado que acabou
    # de ser gravado pela revalidação acima, sem depender de referências
    # de lista em memória (que podem se desconectar quando uma linha
    # precisa de padding).
    todos = ws.get_all_values()

    por_usina = {}
    for row in todos[1:]:
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        if not row[0].strip():
            continue
        status = row[8].strip()
        if _is_concluido_atividade(status):
            continue
        status_os = row[14].strip()
        numero_os = row[13].strip()
        if numero_os and status_os != "Em Processo":
            # mesma regra positiva do primeiro loop: só "Em Processo" entra.
            continue
        etiquetas = row[ATIV_CAMPO_COL["etiquetasOS"] - 1].strip().upper()
        if "PERFORMANCE" in etiquetas:
            continue
        usina = row[2].strip()
        if not usina:
            continue
        d = {
            "usina": usina,
            "equipamento": row[3].strip(),
            "descricao": row[4].strip(),
            "prazo": row[6].strip(),
            "numeroOS": numero_os,
        }
        por_usina.setdefault(usina, []).append(d)

    dry_run = request.args.get("dry_run", "false").lower() == "true"

    enviados, pulados, erros = [], [], []
    for usina, grupo_id in mapa_grupos.items():
        atividades = por_usina.get(usina, [])
        if not atividades:
            pulados.append({"usina": usina, "motivo": "sem atividades em aberto"})
            continue
        texto = _montar_texto_comunicado_usina(usina, atividades)

        if dry_run:
            enviados.append({"usina": usina, "grupo": grupo_id, "atividades": len(atividades), "texto": texto})
            continue

        try:
            r = requests.post(
                f"{WPP_SERVER_URL}/api/enviar-mensagem",
                json={"grupoId": grupo_id, "texto": texto},
                headers={"X-Webhook-Secret": WEBHOOK_SECRET} if WEBHOOK_SECRET else {},
                timeout=20,
            )
            if r.ok and r.json().get("ok"):
                enviados.append({"usina": usina, "grupo": grupo_id, "atividades": len(atividades)})
            else:
                erros.append({"usina": usina, "erro": r.text[:200]})
        except Exception as e:
            erros.append({"usina": usina, "erro": str(e)})

    log.info(f"[ComunicadosDiarios] dry_run={dry_run} enviados={len(enviados)} pulados={len(pulados)} erros={len(erros)}")
    return {"ok": True, "dry_run": dry_run, "enviados": enviados, "pulados": pulados, "erros": erros}


# ── Reversão de excesso (recuperação de uso único) ──────────────────────
# O endpoint acima foi rodado 3x por engano em 2026-07-08 antes de ter uma
# trava, e cada rodada redescontou -3h de linhas que já estavam corretas
# (bug de idempotência: a classificação não sabia diferenciar "ainda não
# corrigido" de "já corrigido"). Este endpoint soma de volta +6h nas linhas
# afetadas (2 descontos extras) pra restaurar o valor correto de um único
# desconto. As OSs em _EXCLUIR_REVERSAO tiveram um histórico de correção
# diferente (parcial/manual) e são tratadas à parte, não por aqui.
_EXCLUIR_REVERSAO = {"9173", "9154"}


@app.route("/reverter-excesso-fuso", methods=["POST", "GET"])
def reverter_excesso_fuso():
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    aplicar = request.args.get("apply", "false").lower() == "true"

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        _garantir_headers_atividades(ws)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    def _precisa_reverter(ts_str, fmt):
        try:
            dt = datetime.strptime(ts_str, fmt)
        except Exception:
            return None
        if dt.date() < _HOJE_DEPLOY:
            return (dt + timedelta(hours=6)).strftime(fmt)
        if dt.date() == _HOJE_DEPLOY and dt.time() < _JANELA_INICIO:
            return (dt + timedelta(hours=6)).strftime(fmt)
        return None

    alteracoes = []
    batch_updates = []
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        numero_os = row[13].strip()
        id_atividade = row[0].strip()
        if numero_os in _EXCLUIR_REVERSAO:
            continue
        updates = {}

        for campo_col, fmt in ((9, '%d/%m/%Y %H:%M:%S'), (10, '%d/%m/%Y %H:%M:%S')):
            val = row[campo_col].strip()
            if val:
                novo = _precisa_reverter(val, fmt)
                if novo:
                    updates[campo_col + 1] = novo

        hist = row[11]
        if hist:
            linhas_novas = []
            hist_mudou = False
            for linha_h in hist.split("\n"):
                m = _HIST_LINHA_RE.match(linha_h)
                if m:
                    data_str, hora_str, seg = m.group(1), m.group(2), m.group(3) or ""
                    ts_str = f"{data_str} {hora_str}{seg}"
                    fmt = '%d/%m/%Y %H:%M:%S' if seg else '%d/%m/%Y %H:%M'
                    novo_ts = _precisa_reverter(ts_str, fmt)
                    if novo_ts:
                        linha_h = novo_ts + linha_h[len(ts_str):]
                        hist_mudou = True
                linhas_novas.append(linha_h)
            if hist_mudou:
                updates[12] = "\n".join(linhas_novas)

        val = row[23].strip()
        if val:
            novo = _precisa_reverter(val, '%Y-%m-%dT%H:%M:%S')
            if novo:
                updates[24] = novo

        if updates:
            alteracoes.append({"linha": i, "id": id_atividade, "numeroOS": numero_os,
                                "colunas_alteradas": list(updates.keys())})
            if aplicar:
                for col, novo_val in updates.items():
                    batch_updates.append({
                        "range": gspread.utils.rowcol_to_a1(i, col),
                        "values": [[novo_val]],
                    })

    if aplicar and batch_updates:
        TAMANHO_LOTE = 200
        for k in range(0, len(batch_updates), TAMANHO_LOTE):
            ws.batch_update(batch_updates[k:k + TAMANHO_LOTE], value_input_option="RAW")

    return jsonify({"ok": True, "aplicado": aplicar, "linhas_afetadas": len(alteracoes),
                     "detalhes": alteracoes, "excluidas": list(_EXCLUIR_REVERSAO)}), 200


# Estado do último ciclo do piggyback /sync-fracttal, rodado em background
# (ver nota abaixo). Guardado só em memória do processo — serve pra
# depuração via /sync-fracttal-status, não é fonte de verdade de nada.
_sync_fracttal_lock = threading.Lock()
_sync_fracttal_last_result = {"em_andamento": False, "iniciado_em": None, "concluido_em": None, "body": None}


def _sync_fracttal_worker():
    """
    Corpo de fato do piggyback (descrito em detalhe no docstring de
    sync_fracttal). Roda em thread separada — ver nota na rota sobre o
    motivo de ter virado assíncrono em 15/08/2026.
    """
    body = {"ok": True}

    # ORDEM IMPORTA (corrigido em 16/08/2026): ronda e resumo diário/semanal
    # têm janela ESTREITA de 30min (08:00-08:30 e 17:00-17:30) e reavaliam
    # a hora atual só quando chegam a rodar — se rodassem depois da
    # varredura de status (atualizacao_status, abaixo, que sozinha já leva
    # 60-90s+ e pode ser seguida de auditoria completa/descoberta rápida,
    # ambas potencialmente demoradas), o acúmulo de tempo das etapas
    # anteriores podia empurrar a hora pra fora da janela antes da ronda
    # sequer ser tentada — foi exatamente o que aconteceu na janela de
    # ronda de 16/08, gerando um dia inteiro sem ronda mesmo com o piggyback
    # rodando normalmente a cada 5min. Daqui pra frente, tudo que depende de
    # janela estreita roda ANTES da varredura pesada de status.
    try:
        body["compromissos_check"] = _verificar_compromissos_se_necessario()
    except Exception as e:
        log.error(f"[Compromissos] Erro no piggyback: {e}")
        body["compromissos_check"] = {"erro": str(e)}

    try:
        body["alertas_etapas_abertas_check"] = _verificar_alertas_etapas_abertas_se_necessario()
    except Exception as e:
        log.error(f"[Compromissos] Erro no piggyback de alertas de etapas abertas: {e}")
        body["alertas_etapas_abertas_check"] = {"erro": str(e)}

    try:
        body["resumo_diario_check"] = _verificar_e_disparar_resumo_diario_se_necessario()
    except Exception as e:
        log.error(f"[ResumoDiario] Erro no piggyback: {e}")
        body["resumo_diario_check"] = {"erro": str(e)}

    try:
        body["ronda_check"] = _verificar_e_disparar_ronda_se_necessario()
    except Exception as e:
        log.error(f"[Ronda] Erro no piggyback: {e}")
        body["ronda_check"] = {"erro": str(e)}

    try:
        body["resumo_semanal_check"] = _verificar_e_disparar_resumo_semanal_se_necessario()
    except Exception as e:
        log.error(f"[ResumoSemanal] Erro no piggyback: {e}")
        body["resumo_semanal_check"] = {"erro": str(e)}

    # DESATIVADO a pedido do Fred em 15/07/2026: o disparo automático não
    # estava rodando de forma confiável às 7h (mesmo com a janela alargada
    # pra 90min) e, quando ele intervinha manualmente pra investigar, às
    # vezes resultava em envio em duplicidade (até 3x o mesmo comunicado
    # pras mesmas equipes) — prejudicando a credibilidade da ferramenta.
    # O botão "Comunicados" no painel continua funcionando normalmente,
    # sob demanda — só o gatilho automático (piggyback no /sync-fracttal)
    # foi desligado.
    body["comunicados_check"] = {"disparado": False, "motivo": "disparo automático desativado — use o botão Comunicados"}

    # A partir daqui, etapas mais lentas e/ou de janela larga (ou sem
    # janela) — não travam mais nada de horário estreito, já que as
    # checagens acima já rodaram primeiro.
    try:
        body["atualizacao_status"] = _auditoria_consistencia_os_core(aplicar=True)
    except Exception as e:
        log.error(f"[Atualizacao] Erro no piggyback: {e}")
        body["atualizacao_status"] = {"erro": str(e)}

    try:
        body["auditoria_completa_check"] = _verificar_e_disparar_auditoria_completa_se_necessario()
    except Exception as e:
        log.error(f"[AuditoriaCompleta] Erro no piggyback: {e}")
        body["auditoria_completa_check"] = {"erro": str(e)}

    try:
        body["descoberta_rapida_check"] = _verificar_e_disparar_descoberta_rapida_se_necessario()
    except Exception as e:
        log.error(f"[DescobertaRapida] Erro no piggyback: {e}")
        body["descoberta_rapida_check"] = {"erro": str(e)}

    _sync_fracttal_last_result["body"] = body
    _sync_fracttal_last_result["concluido_em"] = datetime.now(_TZ_BR).isoformat()
    _sync_fracttal_last_result["em_andamento"] = False
    log.info(f"[sync-fracttal] ciclo em background concluído às {_sync_fracttal_last_result['concluido_em']}")
    _sync_fracttal_lock.release()


@app.route("/sync-fracttal", methods=["POST", "GET"])
def sync_fracttal():
    """
    Gatilho automático confiável (chamado a cada 5 min via UptimeRobot).
    Faz quatro coisas com cadências diferentes, de propósito:
      1. VARREDURA DE STATUS/ESTADO das OSs já no dashboard — roda em
         TODA chamada (5 em 5 min), porque isso precisa ficar em dia com
         frequência (é o que o botão "Atualizar OS" também faz sob demanda).
      2. AUDITORIA COMPLETA (descoberta ampla de 24h + varredura ampla,
         incluindo detectar cancelamentos/conclusões, + validação de
         integridade de relatórios) — só roda de fato nas janelas das
         7h/12h/16h (throttle via _Sistema), porque é mais pesada e não
         precisa de frequência maior que isso.
      3. DESCOBERTA RÁPIDA (só descoberta, janela curta de 2h, sem recheck
         amplo) — roda a cada 30 min (throttle por timestamp via _Sistema),
         pra reduzir o gap de latência entre a criação de uma OS nova na
         Fracttal e ela aparecer no dashboard, sem esperar a próxima
         janela fixa de auditoria completa.
      4. Comunicados diários das 7h (piggyback, gatilho confiável).
      5. Resumo diário (17h) e resumo semanal (sexta 17h), migrados do
         cron do GitHub Actions em 03/08/2026 por atraso recorrente.

    ASSÍNCRONO desde 15/08/2026: o ciclo completo leva ~60-70s pra rodar
    (29 usinas), e o monitor do UptimeRobot só tolera timeout de 30s (teto
    do plano atual, não ajustável via API). Isso fazia o UptimeRobot
    marcar o endpoint como DOWN por "Connection Timeout" repetidamente,
    mesmo o servidor respondendo com sucesso — e como resumo diário/ronda
    dependem desse piggyback rodar até o fim, eles paravam de disparar.
    Agora a rota devolve 200 quase imediatamente (só dispara a thread) e o
    trabalho de fato roda em background, sem depender do cliente esperar.
    Ver /sync-fracttal-status pra inspecionar o resultado do último ciclo.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    # Evita sobrepor ciclos: se o anterior ainda está rodando (ex: pico
    # de latência da Fracttal), só confirma o recebimento sem empilhar
    # outra thread — o próximo ping em 5 min tenta de novo.
    #
    # PROTEÇÃO CONTRA LOCK TRAVADO (adicionada 17/08/2026): se uma thread
    # travar de vez (ex.: chamada de rede sem timeout que nunca retorna —
    # Sheets, Gemini, Fracttal), o lock nunca seria liberado e TODO o
    # piggyback ficaria "vivo" pro UptimeRobot (responde rápido, parece
    # saudável) mas sem processar mais nada — foi exatamente esse padrão
    # observado na manhã de 17/08: gap de quase 1h sem nenhum ciclo real
    # rodando, respostas idênticas e rápidas (só o "ciclo anterior ainda em
    # andamento"), e a ronda das 08h nunca chegou a ser tentada dentro da
    # janela. Se o lock estiver preso há mais de LOCK_TIMEOUT_SEGUNDOS,
    # tratamos como travado e liberamos à força pra não perder o dia
    # inteiro de novo.
    LOCK_TIMEOUT_SEGUNDOS = 300  # ciclo normal leva ~60-90s; auditoria completa pode passar de 4min — 5min de folga
    if _sync_fracttal_lock.locked():
        iniciado_str = _sync_fracttal_last_result.get("iniciado_em")
        travado_ha_muito = False
        if iniciado_str:
            try:
                iniciado_dt = datetime.fromisoformat(iniciado_str)
                travado_ha_muito = (datetime.now(_TZ_BR) - iniciado_dt).total_seconds() > LOCK_TIMEOUT_SEGUNDOS
            except Exception:
                pass
        if travado_ha_muito:
            log.error(f"[sync-fracttal] lock travado desde {iniciado_str} (> {LOCK_TIMEOUT_SEGUNDOS}s) — liberando à força")
            try:
                _sync_fracttal_lock.release()
            except RuntimeError:
                pass

    if not _sync_fracttal_lock.acquire(blocking=False):
        return jsonify({
            "ok": True,
            "background": True,
            "status": "ciclo_anterior_ainda_em_andamento",
            "iniciado_em": _sync_fracttal_last_result.get("iniciado_em"),
        }), 200

    _sync_fracttal_last_result["em_andamento"] = True
    _sync_fracttal_last_result["iniciado_em"] = datetime.now(_TZ_BR).isoformat()
    _sync_fracttal_last_result["concluido_em"] = None

    threading.Thread(target=_sync_fracttal_worker, daemon=True).start()

    return jsonify({
        "ok": True,
        "background": True,
        "status": "ciclo_iniciado",
        "iniciado_em": _sync_fracttal_last_result["iniciado_em"],
    }), 200


@app.route("/sync-fracttal-status", methods=["GET"])
def sync_fracttal_status():
    """Inspeciona o resultado do último ciclo do piggyback /sync-fracttal,
    já que a rota principal agora responde antes do trabalho terminar."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    return jsonify({"ok": True, **_sync_fracttal_last_result}), 200


# ══════════════════════════════════════════════════════════════════════
# PROGRAMAÇÃO PCM — espelha a MESMA fonte que alimenta a aba "Programação
# Semanal" do portal do PCM (Fillipe Figueiro): banco_dados.json,
# publicado no repo gridco-pcm-data (GitHub Pages), gerado por
# gerar_pcm_json.py. Fonte de verdade é esse JSON estático — este painel
# é somente leitura, nunca grava nada de volta lá.
#
# IMPORTANTE (corrigido em 28/07/2026): a primeira versão lia
# gestao_pcm.json (lista bruta de OS da Fracttal, campo dataProg), o que
# causava divergência com o portal — dataProg reflete a data bruta da OS
# na Fracttal, não o cronograma efetivamente PUBLICADO pelo PCM (que já
# passou por reprogramações manuais, tem horário definido, etc.).
# banco_dados.json é o que o portal de fato usa pra montar a Programação
# Semanal, então é a fonte certa pra espelhar aqui.
#
# banco_dados.json guarda as ~4 semanas mais recentes (semana_ativa +
# passadas/futuras), cada uma com "week" no formato ISO "AAAA-Www" e uma
# lista "rows" com um item por tarefa/dia/responsável, incluindo horário
# (h_ini/h_fim), status e dia da semana por extenso em português.
# ══════════════════════════════════════════════════════════════════════

_PCM_BANCO_URL = "https://raw.githubusercontent.com/fillipefigueiro-source/gridco-pcm-data/main/banco_dados.json"
_PCM_RESPONSAVEL = "Fred Alexandrino"
_PCM_CACHE_TTL_SEGUNDOS = 180  # banco_dados.json não tem hash publicado (~2.6MB) — cache curto por tempo em vez de checagem por hash

_cache_programacao_pcm = {"dados": None, "buscado_em": 0, "gerado_em": None}

_DIA_SEMANA_PT = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]

# banco_dados.json usa esses rótulos brutos de status; normalizamos pros
# mesmos rótulos que o resto do painel já usa (Finalizada/Pausada/etc.).
_PCM_STATUS_NORMALIZADO = {
    "Finalizados": "Finalizada",
    "Não Iniciada": "Não Iniciada",
    "pausado": "Pausada",
    "Em progresso": "Em progresso",
}


def _buscar_programacao_pcm_core(forcar=False):
    """Baixa (com cache curto por TTL, já que não há hash publicado pra
    esse arquivo) o banco_dados.json público do PCM — mesma fonte da
    Programação Semanal do portal."""
    global _cache_programacao_pcm
    agora = time.time()
    cache_valido = (
        _cache_programacao_pcm.get("dados") is not None
        and (agora - _cache_programacao_pcm.get("buscado_em", 0)) < _PCM_CACHE_TTL_SEGUNDOS
    )
    if not forcar and cache_valido:
        return _cache_programacao_pcm

    try:
        resp = requests.get(_PCM_BANCO_URL, timeout=60)
        resp.raise_for_status()
        dados = resp.json()
        _cache_programacao_pcm = {
            "dados": dados,
            "buscado_em": agora,
            "gerado_em": dados.get("geradoEm"),
        }
        log.info(f"[ProgramacaoPCM] banco_dados.json atualizado — {len(dados.get('semanas', []))} semana(s) publicadas.")
    except Exception as e:
        log.error(f"[ProgramacaoPCM] Erro ao baixar banco_dados.json: {e}")
        if _cache_programacao_pcm.get("dados") is None:
            raise

    return _cache_programacao_pcm


@app.route("/programacao-pcm", methods=["GET"])
def programacao_pcm():
    """
    Painel de Programações — espelha a Programação Semanal do PCM
    (banco_dados.json), filtrada pro responsável O&M Fred Alexandrino num
    dia específico. Somente leitura, puxado ao vivo (com cache curto).
    Parâmetros opcionais:
      ?data=YYYY-MM-DD  (default: hoje, America/Sao_Paulo)
      ?forcar=1          força rebaixar banco_dados.json ignorando o cache
    """
    forcar = request.args.get("forcar") in ("1", "true", "True")
    data_filtro = request.args.get("data", "").strip()
    if not data_filtro:
        data_filtro = datetime.now(_TZ_BR).strftime("%Y-%m-%d")

    try:
        dt = datetime.strptime(data_filtro, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"ok": False, "error": "parâmetro 'data' inválido, use YYYY-MM-DD"}), 400

    try:
        cache = _buscar_programacao_pcm_core(forcar=forcar)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Falha ao consultar PCM: {e}"}), 502

    dados = cache.get("dados") or {}
    semanas = dados.get("semanas", [])
    iso_year, iso_week, _ = dt.isocalendar()
    semana_alvo = f"{iso_year}-W{iso_week:02d}"
    semana = next((s for s in semanas if s.get("week") == semana_alvo), None)

    if semana is None:
        janela = ", ".join(s.get("week", "") for s in semanas)
        return jsonify({
            "ok": True, "data": data_filtro, "diaSemana": None, "total": 0, "atrasadas": 0,
            "resumoEstado": {}, "grupos": [],
            "aviso": f"Sem programação publicada pelo PCM pra semana {semana_alvo} (janela disponível: {janela}).",
            "fonte": {"geradoEmPCM": cache.get("gerado_em")},
        }), 200

    dia_pt = _DIA_SEMANA_PT[dt.weekday()]
    hoje_str = datetime.now(_TZ_BR).strftime("%Y-%m-%d")

    usinas_temp_nomes = {item["usina"] for item in _usinas_temporarias()}
    linhas_dia = [
        r for r in semana.get("rows", [])
        if (r.get("responsavel") == _PCM_RESPONSAVEL or r.get("usina") in usinas_temp_nomes) and r.get("dia") == dia_pt
    ]

    mapa_cluster_pcm = _mapa_cluster_usina()
    por_usina = {}
    for r in linhas_dia:
        usina = r.get("usina") or "(sem usina)"
        estado = _PCM_STATUS_NORMALIZADO.get(r.get("status"), r.get("status") or "")
        atrasado = data_filtro < hoje_str and estado != "Finalizada"
        # O campo "cluster" que vem no banco_dados.json é da fonte do PCM
        # (Power Automate do Fillipe) e pode ficar desatualizado depois de
        # uma reorganização de cluster feita só no nosso lado (ex.:
        # GD Energy migrou de "CE Leste 01" pra "CE Norte 01" em 07/08/2026
        # e a fonte do PCM continuou mandando "CE Leste 01" — isso fazia o
        # cluster "CE Norte 01" nunca aparecer nos Comunicados de
        # Programação, mesmo com atividades abertas). Corrigido 19/08/2026:
        # sempre que a usina for reconhecida no nosso catálogo, o cluster
        # usado é o do NOSSO mapeamento interno (_mapa_cluster_usina,
        # fonte de verdade); só cai no valor cru do PCM como fallback
        # quando a usina não é reconhecida (ex.: usina temporária de outro
        # supervisor sem cluster configurado ainda).
        usina_canonica = canonizar_usina(_extrair_nome_usina_fracttal(usina) or usina)
        cluster_correto = mapa_cluster_pcm.get(usina_canonica) if usina_canonica else None
        por_usina.setdefault(usina, []).append({
            "os": r.get("os_id"),
            "cliente": r.get("cliente"),
            "cluster": cluster_correto or r.get("cluster"),
            "tipo": r.get("tipo"),
            "tarefa": r.get("tarefa"),
            "estado": estado,
            "hIni": r.get("h_ini"),
            "hFim": r.get("h_fim"),
            "duracaoH": r.get("duracao"),
            "reprogramavel": (r.get("reprog") or "").strip().lower() == "sim",
            "atrasado": atrasado,
        })

    for itens in por_usina.values():
        itens.sort(key=lambda x: x.get("hIni") or "")

    grupos = [{"usina": usina, "itens": itens} for usina, itens in sorted(por_usina.items())]

    resumo_estado = {}
    for itens in por_usina.values():
        for it in itens:
            est = it["estado"] or "—"
            resumo_estado[est] = resumo_estado.get(est, 0) + 1

    total = sum(len(v) for v in por_usina.values())
    atrasadas = sum(1 for v in por_usina.values() for it in v if it["atrasado"])

    return jsonify({
        "ok": True,
        "data": data_filtro,
        "diaSemana": dia_pt,
        "total": total,
        "atrasadas": atrasadas,
        "resumoEstado": resumo_estado,
        "grupos": grupos,
        "fonte": {"geradoEmPCM": cache.get("gerado_em")},
    }), 200


@app.route("/alertar-wpp-status", methods=["POST"])
def alertar_wpp_status():
    """
    Chamado pela ponte do WhatsApp (server.js) sempre que a conexão cai ou
    precisa de novo QR code. Dispara push imediato pro celular do Fred —
    sem isso, uma queda de sessão só é percebida dias depois (mensagens
    de ocorrência chegam nos grupos mas não são capturadas enquanto a
    sessão estiver caída, e não ficam "na fila" esperando reconexão).
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    body = request.get_json(force=True, silent=True) or {}
    status = body.get("status", "desconhecido")
    detalhe = body.get("detalhe", "")
    try:
        if status == "aguardando_qr":
            enviar_push(
                titulo="⚠️ WhatsApp precisa de novo QR Code",
                corpo="A automação de ocorrências está sem conexão — mensagens de rondas não estão sendo capturadas até reconectar. Acesse /qr na ponte pra escanear.",
                tipo="wpp_desconectado",
            )
        elif status == "desconectado":
            enviar_push(
                titulo="⚠️ WhatsApp desconectado",
                corpo=f"Conexão caiu ({detalhe}). Tentando reconectar automaticamente.",
                tipo="wpp_desconectado",
            )
        elif status == "reconectado":
            enviar_push(
                titulo="✅ WhatsApp reconectado",
                corpo="A automação de ocorrências voltou a capturar mensagens normalmente.",
                tipo="wpp_reconectado",
            )
        elif status == "falha_encaminhamento":
            # A conexão com o WhatsApp em si pode estar normal, mas a
            # ponte não conseguiu repassar essa mensagem específica pro
            # backend depois de 3 tentativas — sem esse alerta, isso
            # ficava só no console, invisível, e a ocorrência real se
            # perdia por dias sem ninguém saber (identificado 13/07/2026).
            enviar_push(
                titulo="⚠️ Mensagem de ronda não foi registrada",
                corpo=f"Falha ao gravar após 3 tentativas: {detalhe[:150]}",
                tipo="wpp_falha_encaminhamento",
            )
        return jsonify({"ok": True}), 200
    except Exception as e:
        log.error(f"[AlertaWPP] Erro ao enviar push: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500



@app.route("/desligamento-manual", methods=["GET"])
def listar_desligamento_manual():
    """Lista todos os overrides manuais de classificação de desligamento
    (Falhas e Atividades juntos) — usado pelo frontend pra sobrepor a
    detecção automática por palavra-chave quando o Fred marcar manualmente
    que algo É ou NÃO É um desligamento de usina de verdade."""
    try:
        ws = get_desligamento_manual_sheet()
        todos = ws.get_all_values()
        itens = []
        for row in todos[1:]:
            if len(row) < 3 or not row[0].strip():
                continue
            itens.append({
                "origem": row[0].strip(), "id": row[1].strip(), "valor": row[2].strip(),
                "editor": row[3].strip() if len(row) > 3 else "",
                "atualizadoEm": row[4].strip() if len(row) > 4 else "",
            })
        return jsonify({"ok": True, "itens": itens}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/marcar-desligamento-manual", methods=["POST", "OPTIONS"])
def marcar_desligamento_manual():
    """Grava (ou limpa) a classificação manual de desligamento pra uma
    ocorrência/atividade específica. valor: "sim" (força tratar como
    desligamento), "nao" (força tratar como NÃO desligamento, mesmo que a
    detecção automática por palavra-chave tivesse batido), ou "" (remove o
    override, volta a valer só a detecção automática)."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    origem = str(body.get("origem", "")).strip()
    item_id = str(body.get("id", "")).strip()
    valor = str(body.get("valor", "")).strip().lower()
    editor = str(body.get("editor", "dashboard")).strip()
    if origem not in ("falha", "atividade") or not item_id:
        return jsonify({"ok": False, "error": "origem (falha|atividade) e id são obrigatórios"}), 400
    if valor not in ("sim", "nao", ""):
        return jsonify({"ok": False, "error": "valor deve ser 'sim', 'nao' ou vazio"}), 400

    try:
        ws = get_desligamento_manual_sheet()
        todos = ws.get_all_values()
        linha_existente = None
        for i, row in enumerate(todos[1:], start=2):
            if len(row) >= 2 and row[0].strip() == origem and row[1].strip() == item_id:
                linha_existente = i
                break
        agora = agora_br().strftime("%d/%m/%Y %H:%M:%S")
        if valor == "":
            if linha_existente:
                ws.delete_rows(linha_existente)
            return jsonify({"ok": True, "removido": bool(linha_existente)}), 200
        if linha_existente:
            ws.update(f"A{linha_existente}:E{linha_existente}", [[origem, item_id, valor, editor, agora]])
        else:
            ws.append_row([origem, item_id, valor, editor, agora])
        return jsonify({"ok": True}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/excluir-atividade", methods=["POST", "OPTIONS"])
def excluir_atividade():
    """Remove definitivamente uma atividade do Painel de Atividades pelo
    id interno (coluna A da planilha). Uso: correção manual de atividades
    criadas incorretamente (ex: OS de outro cliente/site que entrou por
    engano). Não existia endpoint de exclusão até 24/07/2026 — o painel só
    tinha edição de campos."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    ativ_id = str(body.get("id", "")).strip()
    if not ativ_id:
        return jsonify({"ok": False, "error": "id é obrigatório"}), 400
    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        linha = None
        for i, row in enumerate(todos[1:], start=2):
            if row and str(row[0]).strip() == ativ_id:
                linha = i
                break
        if not linha:
            return jsonify({"ok": False, "error": "atividade não encontrada"}), 404
        removida = todos[linha - 1]
        ws.delete_rows(linha)
        log.info(f"[excluir-atividade] id={ativ_id} removido. cliente={removida[1] if len(removida)>1 else '?'} "
                 f"usina={removida[2] if len(removida)>2 else '?'} numeroOS={removida[13] if len(removida)>13 else '?'}")
        return jsonify({"ok": True}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/localizacoes", methods=["GET"])
def listar_localizacoes():
    """Lista as localizações (endereço + link do Maps + lat/lng já
    geocodificados, quando disponíveis) de todas as usinas cadastradas na
    aba 'Localizacoes'. Usado pela seção 'Localização' do painel (Mapa +
    Lista). Lat/Lng em branco significa que o frontend ainda não
    geocodificou aquele endereço — nesse caso a usina some do Mapa até ser
    geocodificada (ver /localizacoes-atualizar-coords) mas continua
    aparecendo na Lista com o link do Google Maps."""
    try:
        ws = get_localizacoes_sheet()
        todos = ws.get_all_values()
        itens = []
        for row in todos[1:]:
            if len(row) < 2 or not row[1].strip():
                continue
            lat = row[4].strip() if len(row) > 4 else ""
            lng = row[5].strip() if len(row) > 5 else ""
            def _num(v):
                # a planilha está em locale pt-BR: números podem vir com
                # vírgula decimal (ex: "-4,105639") em vez de ponto
                if not v:
                    return None
                try:
                    return float(v.replace(".", "").replace(",", ".")) if "," in v else float(v)
                except ValueError:
                    return None
            itens.append({
                "cliente": row[0].strip() if len(row) > 0 else "",
                "usina": row[1].strip(),
                "endereco": row[2].strip() if len(row) > 2 else "",
                "mapsLink": row[3].strip() if len(row) > 3 else "",
                "lat": _num(lat),
                "lng": _num(lng),
            })
        return jsonify({"ok": True, "itens": itens}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/localizacoes-atualizar-coords", methods=["POST", "OPTIONS"])
def atualizar_coords_localizacao():
    """Grava lat/lng geocodificados no navegador (via Nominatim/OSM,
    client-side) de volta na planilha, pra não precisar geocodificar de
    novo a cada carregamento da página. Casamento por cliente+usina
    (case-insensitive, ignorando espaços nas pontas)."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    cliente = str(body.get("cliente", "")).strip()
    usina = str(body.get("usina", "")).strip()
    lat = body.get("lat")
    lng = body.get("lng")
    if not usina or lat is None or lng is None:
        return jsonify({"ok": False, "error": "usina, lat e lng são obrigatórios"}), 400
    try:
        ws = get_localizacoes_sheet()
        todos = ws.get_all_values()
        linha = None
        for i, row in enumerate(todos[1:], start=2):
            if len(row) >= 2 and row[1].strip().lower() == usina.lower() and (
                not cliente or row[0].strip().lower() == cliente.lower()
            ):
                linha = i
                break
        if not linha:
            return jsonify({"ok": False, "error": "usina não encontrada"}), 404
        ws.update(f"E{linha}:F{linha}", [[lat, lng]])
        return jsonify({"ok": True}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/localizacoes-adicionar", methods=["POST", "OPTIONS"])
def adicionar_localizacao():
    """Adiciona (ou atualiza, se já existir cliente+usina) uma linha na aba
    'Localizacoes'. Diferente de /localizacoes-atualizar-coords (que só
    grava lat/lng em usina já cadastrada), este endpoint cria o registro
    completo do zero — usado quando uma usina nova entra no mapeamento e
    ainda não tem nenhuma linha na planilha."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    cliente = str(body.get("cliente", "")).strip()
    usina = str(body.get("usina", "")).strip()
    endereco = str(body.get("endereco", "")).strip()
    maps_link = str(body.get("mapsLink", "")).strip()
    lat = body.get("lat")
    lng = body.get("lng")
    if not usina:
        return jsonify({"ok": False, "error": "usina é obrigatória"}), 400
    try:
        ws = get_localizacoes_sheet()
        todos = ws.get_all_values()
        linha = None
        for i, row in enumerate(todos[1:], start=2):
            if len(row) >= 2 and row[1].strip().lower() == usina.lower() and (
                not cliente or row[0].strip().lower() == cliente.lower()
            ):
                linha = i
                break
        nova_linha = [cliente, usina, endereco, maps_link,
                       "" if lat is None else lat, "" if lng is None else lng]
        if linha:
            ws.update(f"A{linha}:F{linha}", [nova_linha])
            return jsonify({"ok": True, "acao": "atualizada", "linha": linha}), 200
        else:
            ws.append_row(nova_linha)
            return jsonify({"ok": True, "acao": "criada"}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


_MAPS_PIN_REGEX = re.compile(r"!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)")


@app.route("/localizacoes-resolver-links-servidor", methods=["POST", "OPTIONS"])
def resolver_links_maps_servidor():
    """Pega a coordenada EXATA que o Fred já pinou no Google Maps,
    resolvendo o link curto (maps.app.goo.gl) que ele mandou pra cada
    usina — em vez de adivinhar a partir do texto do endereço (que falha
    pra fazenda/sítio rural não mapeado no OSM). Isso é definitivo: uma
    vez resolvido e gravado na planilha, nunca mais precisa rodar de novo
    pra essa usina, mesmo que o Nominatim nunca ache o endereço dela."""
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        ws = get_localizacoes_sheet()
        todos = ws.get_all_values()
        resolvidos, falhas = [], []
        for i, row in enumerate(todos[1:], start=2):
            if len(row) < 4 or not row[1].strip():
                continue
            lat_atual = row[4].strip() if len(row) > 4 else ""
            lng_atual = row[5].strip() if len(row) > 5 else ""
            if lat_atual and lng_atual:
                continue
            link = row[3].strip() if len(row) > 3 else ""
            usina = row[1].strip()
            if not link:
                falhas.append(usina + " (sem link do Maps)")
                continue
            try:
                resp = requests.get(link, allow_redirects=True, timeout=12,
                                     headers={"User-Agent": "Mozilla/5.0 (compatible; PainelOM-GridCo/1.0)"})
                m = _MAPS_PIN_REGEX.search(resp.url)
                if m:
                    lat, lng = float(m.group(1)), float(m.group(2))
                    ws.update(f"E{i}:F{i}", [[lat, lng]])
                    resolvidos.append(usina)
                else:
                    falhas.append(usina + " (não achou coordenada na URL resolvida)")
            except Exception as e:
                falhas.append(usina + f" (erro: {e})")
        return jsonify({"ok": True, "resolvidos": resolvidos, "falhas": falhas}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


NOMINATIM_USER_AGENT = "PainelOM-GridCo/1.0 (contato: fred@gridco.com.br)"


@app.route("/localizacoes-geocodificar-servidor", methods=["POST", "OPTIONS"])
def geocodificar_localizacoes_servidor():
    """Geocodifica no SERVIDOR (não no navegador do Fred) os endereços que
    ainda não têm lat/lng na aba 'Localizacoes', usando Nominatim/OSM.
    Mais confiável que geocodificar no navegador: o Nominatim é rígido com
    chamadas client-side sem um User-Agent de aplicação identificável
    (headers custom não são permitidos no fetch do browser), o que fazia
    boa parte das usinas ficarem travadas sem coordenada mesmo recarregando
    a página. Aqui a gente identifica a aplicação corretamente e respeita
    1 req/s, dentro do timeout de 160s do Gunicorn (dá pra ~140 endereços)."""
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        ws = get_localizacoes_sheet()
        todos = ws.get_all_values()
        atualizados, falhas = [], []
        for i, row in enumerate(todos[1:], start=2):
            if len(row) < 3 or not row[1].strip():
                continue
            lat_atual = row[4].strip() if len(row) > 4 else ""
            lng_atual = row[5].strip() if len(row) > 5 else ""
            if lat_atual and lng_atual:
                continue
            endereco = row[2].strip() if len(row) > 2 else ""
            usina = row[1].strip()
            if not endereco:
                falhas.append(usina + " (sem endereço)")
                continue
            try:
                resp = requests.get(
                    "https://nominatim.openstreetmap.org/search",
                    params={"q": endereco + ", Brasil", "format": "json", "limit": 1},
                    headers={"User-Agent": NOMINATIM_USER_AGENT, "Accept-Language": "pt-BR"},
                    timeout=10,
                )
                arr = resp.json() if resp.ok else []
                if arr:
                    lat, lng = float(arr[0]["lat"]), float(arr[0]["lon"])
                    ws.update(f"E{i}:F{i}", [[lat, lng]])
                    atualizados.append(usina)
                else:
                    falhas.append(usina + " (endereço não encontrado)")
            except Exception as e:
                falhas.append(usina + f" (erro: {e})")
            time.sleep(1.1)  # respeita 1 req/s da política de uso do Nominatim
        return jsonify({"ok": True, "atualizados": atualizados, "falhas": falhas}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/sincronizar-chamados", methods=["POST", "OPTIONS"])
def sincronizar_chamados():
    """
    Recebe dados da tabela de chamados de fabricante (hoje: exportação
    manual da planilha do SharePoint enviada pelo Fred; futuramente,
    talvez um fluxo automático). Faz upsert por uma CHAVE COMPOSTA
    (Ticket/RMA + Ativo + Identificação do Equipamento + Data da
    ocorrência), não só pelo Ticket/RMA sozinho.

    Por quê: na planilha real, valores de Ticket/RMA como "00" são
    usados como placeholder em dezenas de chamados completamente
    diferentes (usinas/equipamentos diferentes) até o número real ser
    aberto — e alguns tickets legítimos (ex.: "00574/26") cobrem
    várias peças de equipamento diferentes na mesma usina/data. Usar só
    o Ticket/RMA como chave colapsaria esses grupos numa linha só,
    apagando dados de verdade. A combinação dos 4 campos já foi
    validada como livre de colisão na importação inicial de 553
    registros reais (17/07/2026).

    Aceita tanto uma linha única (objeto) quanto várias de uma vez
    (lista de objetos).

    Escreve tudo em LOTE — no máximo duas chamadas à API do Google
    Sheets no total (uma pra criar todas as linhas novas, outra pra
    atualizar todas as existentes), não importa se são 5 ou 5.000
    linhas recebidas — evita estourar a cota de escrita da API do
    Google (~60/min).
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    body = request.get_json(force=True, silent=True)
    if body is None:
        return jsonify({"ok": False, "error": "corpo da requisição precisa ser JSON"}), 400
    linhas = body if isinstance(body, list) else [body]

    idx_ticket = CHAMADOS_FABRICANTE_HEADERS.index("Ticket/RMA")
    idx_ativo = CHAMADOS_FABRICANTE_HEADERS.index("Ativo")
    idx_equip = CHAMADOS_FABRICANTE_HEADERS.index("Identificação do Equipamento")
    idx_data_ocorrencia = CHAMADOS_FABRICANTE_HEADERS.index("Data da ocorrência")
    n_cols = len(CHAMADOS_FABRICANTE_HEADERS)
    colunas_letra_fim = chr(64 + n_cols) if n_cols <= 26 else "Z"

    def _chave(row_vals):
        return (row_vals[idx_ticket], row_vals[idx_ativo], row_vals[idx_equip], row_vals[idx_data_ocorrencia])

    try:
        ws = get_chamados_fabricante_sheet()
        todos = ws.get_all_values()

        por_chave = {}
        for i, row in enumerate(todos[1:], start=2):
            if len(row) < n_cols:
                row = row + [""] * (n_cols - len(row))
            por_chave[_chave(row)] = i

        atualizadas_map = {}
        criadas_map = {}
        erros = []

        for linha_recebida in linhas:
            if not isinstance(linha_recebida, dict):
                erros.append("item não é um objeto JSON válido")
                continue

            def _buscar_campo(nome_coluna):
                if nome_coluna in linha_recebida:
                    return str(linha_recebida[nome_coluna] or "").strip()
                alvo_norm = nome_coluna.lower().strip()
                for k, v in linha_recebida.items():
                    if k.lower().strip() == alvo_norm:
                        return str(v or "").strip()
                return ""

            nova_linha = [_buscar_campo(h) for h in CHAMADOS_FABRICANTE_HEADERS]
            chave = _chave(nova_linha)

            linha_existente = por_chave.get(chave)
            if linha_existente:
                atualizadas_map[linha_existente] = nova_linha
            else:
                criadas_map[chave] = nova_linha  # última ocorrência no batch vence, se repetir

        if atualizadas_map:
            ws.batch_update([
                {"range": f"A{linha}:{colunas_letra_fim}{linha}", "values": [valores]}
                for linha, valores in atualizadas_map.items()
            ])

        if criadas_map:
            ws.append_rows(list(criadas_map.values()))

        # registra quando essa sincronização aconteceu — usado pelo painel
        # pra mostrar "última sincronização" e o Fred conseguir verificar
        # se está em dia sem precisar comparar contagem manualmente
        # (pedido em 24/07/2026).
        try:
            _gravar_trava("chamados_ultima_sincronizacao", agora_br().strftime("%d/%m/%Y %H:%M:%S"))
        except Exception as e:
            log.error(f"[ChamadosFabricante] Falha ao gravar timestamp de sincronização: {e}")

        return jsonify({"ok": True, "criadas": len(criadas_map), "atualizadas": len(atualizadas_map),
                         "erros": erros}), 200
    except Exception as e:
        log.error(f"[ChamadosFabricante] Erro ao sincronizar: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _mapa_notas_chamados():
    """Notas que o Fred escreve no popup de detalhe do painel de Chamados
    — ficam só na aba _Sistema (chave 'nota_chamado:<ticket>|<ufv>|<equip>'),
    NUNCA na aba ChamadosFabricante, pra sobreviver a reimportações
    futuras da planilha do SharePoint sem serem apagadas."""
    ws_cfg = _get_config_sheet()
    valores = _gspread_retry(lambda: ws_cfg.get_all_values())
    mapa = {}
    for row in valores[1:]:
        if row and row[0].strip().startswith("nota_chamado:"):
            chave = row[0].strip()[len("nota_chamado:"):]
            mapa[chave] = row[1].strip() if len(row) > 1 else ""
    return mapa


@app.route("/atualizar-observacao-chamado", methods=["POST", "OPTIONS"])
def atualizar_observacao_chamado():
    """
    Salva a nota que o Fred escreve no popup de detalhe de um chamado —
    fica só na aba _Sistema (dashboard), NUNCA na aba ChamadosFabricante,
    pra não ser apagada quando a planilha do SharePoint for reimportada
    de novo no futuro.

    Endpoint PÚBLICO (sem secret) — chamado direto do navegador pelo
    popup de detalhe do chamado, igual /atualizar-campo-atividade.

    Corpo esperado: {"ticket": "...", "ufv": "...", "equipamento": "...",
    "novaObservacao": "..."}
    """
    if request.method == "OPTIONS":
        return ("", 204)

    body = request.get_json(force=True, silent=True) or {}
    ticket = (body.get("ticket") or "").strip()
    ufv = (body.get("ufv") or "").strip()
    equipamento = (body.get("equipamento") or "").strip()
    nova_obs = body.get("novaObservacao") or ""
    chave = f"nota_chamado:{ticket}|{ufv}|{equipamento}"

    try:
        ws_cfg = _get_config_sheet()
        valores = ws_cfg.get_all_values()
        linha_existente = None
        for i, row in enumerate(valores[1:], start=2):
            if row and row[0].strip() == chave:
                linha_existente = i
                break
        if linha_existente:
            ws_cfg.update(f"B{linha_existente}", [[nova_obs]])
        else:
            ws_cfg.append_row([chave, nova_obs])
        return jsonify({"ok": True}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/corrigir-tickets-empilhados", methods=["POST", "OPTIONS"])
def corrigir_tickets_empilhados():
    """
    Correção em lote (17/07/2026): na planilha original do SharePoint,
    várias linhas tinham MÚLTIPLOS tickets de fabricante empilhados
    numa célula só (ex.: "Ticket Novo: 15817311\\n15751652\\n15651216"),
    e vieram assim pro import inicial. Esse endpoint localiza cada linha
    pelo valor ANTIGO (bruto, empilhado) de Ticket/RMA + UFV, troca o
    Ticket/RMA pelo primeiro/mais recente já limpo (sem o rótulo tipo
    "Ticket Novo:"), e acrescenta os tickets mais antigos como nota no
    campo Observações (sem apagar o que já tinha lá).

    Corpo esperado: {"correcoes": [{"ticketAntigo": "...", "ufv": "...",
    "novoTicket": "...", "notaObservacao": "..."}, ...]}
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    body = request.get_json(force=True, silent=True) or {}
    correcoes = body.get("correcoes", [])
    if not correcoes:
        return jsonify({"ok": False, "error": "nenhuma correção informada"}), 400

    idx_ticket = CHAMADOS_FABRICANTE_HEADERS.index("Ticket/RMA")
    idx_ufv = CHAMADOS_FABRICANTE_HEADERS.index("UFV")
    idx_obs = CHAMADOS_FABRICANTE_HEADERS.index("Observações")
    n_cols = len(CHAMADOS_FABRICANTE_HEADERS)
    col_letra_ticket = chr(65 + idx_ticket)
    col_letra_obs = chr(65 + idx_obs)

    try:
        ws = get_chamados_fabricante_sheet()
        todos = ws.get_all_values()

        atualizacoes, nao_encontradas = [], []
        for c in correcoes:
            ticket_antigo = (c.get("ticketAntigo") or "").strip()
            ufv = (c.get("ufv") or "").strip()
            novo_ticket = c.get("novoTicket") or ""
            nota = c.get("notaObservacao") or ""
            achou = False
            for i, row in enumerate(todos[1:], start=2):
                if len(row) < n_cols:
                    row = row + [""] * (n_cols - len(row))
                if row[idx_ticket].strip() == ticket_antigo and row[idx_ufv].strip() == ufv:
                    obs_atual = row[idx_obs].strip()
                    obs_nova = f"{obs_atual}\n{nota}".strip() if obs_atual and nota else (nota or obs_atual)
                    atualizacoes.append({"range": f"{col_letra_ticket}{i}", "values": [[novo_ticket]]})
                    if nota:
                        atualizacoes.append({"range": f"{col_letra_obs}{i}", "values": [[obs_nova]]})
                    achou = True
                    break
            if not achou:
                nao_encontradas.append({"ticketAntigo": ticket_antigo, "ufv": ufv})

        if atualizacoes:
            ws.batch_update(atualizacoes)

        return jsonify({"ok": True, "corrigidas": len(correcoes) - len(nao_encontradas),
                         "nao_encontradas": nao_encontradas}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/corrigir-ativo-chamado", methods=["POST", "OPTIONS"])
def corrigir_ativo_chamado():
    """
    Correção pontual: atualiza SÓ a coluna "Ativo" de uma ou mais linhas
    já existentes na aba ChamadosFabricante, localizadas por Ticket/RMA
    + UFV (não usa a chave composta normal do /sincronizar-chamados,
    porque nesse caso o próprio "Ativo" é o campo que está sendo
    corrigido — geralmente linhas que vieram com Ativo vazio por causa
    de célula mesclada no Excel original).

    Corpo esperado: {"correcoes": [{"ticket": "...", "ufv": "...", "novoAtivo": "..."}, ...]}
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    body = request.get_json(force=True, silent=True) or {}
    correcoes = body.get("correcoes", [])
    if not correcoes:
        return jsonify({"ok": False, "error": "nenhuma correção informada"}), 400

    idx_ticket = CHAMADOS_FABRICANTE_HEADERS.index("Ticket/RMA")
    idx_ufv = CHAMADOS_FABRICANTE_HEADERS.index("UFV")
    n_cols = len(CHAMADOS_FABRICANTE_HEADERS)

    try:
        ws = get_chamados_fabricante_sheet()
        todos = ws.get_all_values()

        atualizacoes, nao_encontradas = [], []
        for c in correcoes:
            ticket, ufv, novo_ativo = (c.get("ticket") or "").strip(), (c.get("ufv") or "").strip(), c.get("novoAtivo") or ""
            achou = False
            for i, row in enumerate(todos[1:], start=2):
                if len(row) < n_cols:
                    row = row + [""] * (n_cols - len(row))
                if row[idx_ticket].strip() == ticket and row[idx_ufv].strip() == ufv:
                    atualizacoes.append({"range": f"A{i}", "values": [[novo_ativo]]})
                    achou = True
                    break
            if not achou:
                nao_encontradas.append({"ticket": ticket, "ufv": ufv})

        if atualizacoes:
            ws.batch_update(atualizacoes)

        return jsonify({"ok": True, "corrigidas": len(atualizacoes), "nao_encontradas": nao_encontradas}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _chamados_fabricante_itens():
    """Lê a aba ChamadosFabricante inteira + mescla as notas do dashboard
    (aba _Sistema). Reaproveitada tanto pelo endpoint GET /chamados-fabricante
    quanto pela geração do relatório semanal — uma única fonte de verdade.

    FILTRO POR USINA CONHECIDA (adicionado 29/07/2026): a planilha do
    SharePoint que alimenta essa aba tem chamados de OUTRAS empresas
    também (ex.: Santarém, Aparecida — nada a ver com o Fred), e essa
    função sincronizava tudo sem filtrar. Isso vazou no resumo diário
    (chamados de usinas que não são do Fred aparecendo como se fossem).
    Agora só entram linhas cujo campo UFV bate com uma usina do catálogo
    (mesma fonte usada pra reconhecer OS da Fracttal)."""
    ws = get_chamados_fabricante_sheet()
    todos = _gspread_retry(lambda: ws.get_all_values())
    notas = _mapa_notas_chamados()
    itens = []
    for row in todos[1:]:
        if len(row) < len(CHAMADOS_FABRICANTE_HEADERS):
            row = row + [""] * (len(CHAMADOS_FABRICANTE_HEADERS) - len(row))
        # linha em branco de verdade = TODAS as células vazias, não só a
        # primeira coluna (Ativo pode legitimamente vir vazio — ex.:
        # célula mesclada no Excel original — enquanto o resto da linha
        # tem dados reais; checar só row[0] descartava chamados válidos)
        if not any(cell.strip() for cell in row):
            continue
        item = dict(zip(CHAMADOS_FABRICANTE_HEADERS, row[:len(CHAMADOS_FABRICANTE_HEADERS)]))
        ufv_bruta = item.get("UFV", "")
        if ufv_bruta and canonizar_usina(ufv_bruta) is None:
            continue  # usina não reconhecida = não é do Fred, não deveria estar nessa planilha
        chave_nota = f"{item.get('Ticket/RMA','')}|{item.get('UFV','')}|{item.get('Identificação do Equipamento','')}"
        item["NotaDashboard"] = notas.get(chave_nota, "")
        itens.append(item)
    return itens


@app.route("/chamados-fabricante", methods=["GET"])
def listar_chamados_fabricante():
    """Devolve a tabela de chamados de fabricante inteira, pro frontend
    exibir no Painel de Chamados sem precisar de nenhuma cópia manual."""
    try:
        itens = _chamados_fabricante_itens()
        travas = _ler_travas(["chamados_ultima_sincronizacao", "chamados_ultima_tentativa",
                               "chamados_ultimo_status", "chamados_ultimo_erro"])
        return jsonify({"ok": True, "itens": itens, "total": len(itens),
                         "ultimaSincronizacao": travas["chamados_ultima_sincronizacao"],
                         "ultimaTentativa": travas["chamados_ultima_tentativa"],
                         "ultimoStatus": travas["chamados_ultimo_status"],
                         "ultimoErro": travas["chamados_ultimo_erro"]}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/reportar-status-sincronizacao-chamados", methods=["POST", "OPTIONS"])
def reportar_status_sincronizacao_chamados():
    """Recebe um relatório de status do script local (Agendador de Tarefas
    no PC do Fred), chamado no final de TODA execução — sucesso ou erro.

    Por quê existe separado do /sincronizar-chamados: quando o script
    falha ANTES de conseguir montar/enviar os dados (ex.: Excel não
    encontrado, OneDrive não sincronizado, erro de rede), o endpoint de
    sincronização de verdade nunca é chamado — então o painel ficaria
    "cego" pra esse tipo de falha, achando que só não rodou. Esse
    endpoint garante que TODA tentativa (com sucesso ou não) fica
    registrada, pra exibir um indicador de status no Painel de Chamados
    (pedido pelo Fred em 28/07/2026, após diagnosticar uma falha
    silenciosa do Agendador de Tarefas — Launch Failure por caminho de
    python.exe incorreto).
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    body = request.get_json(force=True, silent=True) or {}
    ok = bool(body.get("ok"))
    erro = str(body.get("erro") or "").strip()
    criadas = body.get("criadas")
    atualizadas = body.get("atualizadas")

    agora_str = agora_br().strftime("%d/%m/%Y %H:%M:%S")
    try:
        _gravar_trava("chamados_ultima_tentativa", agora_str)
        _gravar_trava("chamados_ultimo_status", "ok" if ok else "erro")
        _gravar_trava("chamados_ultimo_erro", erro if not ok else "")
        if ok:
            detalhe = f"criadas: {criadas}, atualizadas: {atualizadas}"
            _gravar_trava("chamados_ultimo_detalhe", detalhe)
    except Exception as e:
        log.error(f"[ChamadosFabricante] Falha ao gravar status de sincronização: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

    return jsonify({"ok": True}), 200


FOTOS_ATIVIDADES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fotos_atividades")
EXTENSOES_FOTO_PERMITIDAS = {"jpg", "jpeg", "png", "webp", "heic", "heif"}


def _extensao_permitida(nome_arquivo):
    return "." in nome_arquivo and nome_arquivo.rsplit(".", 1)[1].lower() in EXTENSOES_FOTO_PERMITIDAS


def _atividade_por_id(atividade_id):
    """Busca uma linha da aba Painel de Atividades pelo id. Retorna
    (linha_idx, row) ou (None, None) se não encontrar."""
    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    for i, row in enumerate(todos[1:], start=2):
        if row and row[0].strip() == str(atividade_id).strip():
            if len(row) < ATIV_TOTAL_COLUNAS:
                row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
            return i, row
    return None, None


@app.route("/atividade-anexar-foto", methods=["POST", "OPTIONS"])
def atividade_anexar_foto():
    """
    Anexa uma foto a uma atividade — SOMENTE atividades criadas
    manualmente (sem numeroOS, ou seja, não vinculadas à Fracttal).
    Fica guardada localmente no disco do servidor, ao lado de "Minhas
    anotações" no drawer — não sincroniza com a Fracttal, mesma lógica
    das anotações pessoais (implementado 24/07/2026, a pedido do Fred).
    """
    if request.method == "OPTIONS":
        return ("", 204)

    atividade_id = (request.form.get("id") or "").strip()
    if not atividade_id:
        return jsonify({"ok": False, "error": "id da atividade é obrigatório"}), 400
    if "foto" not in request.files:
        return jsonify({"ok": False, "error": "nenhuma foto enviada"}), 400

    arquivo = request.files["foto"]
    if not arquivo.filename or not _extensao_permitida(arquivo.filename):
        return jsonify({"ok": False, "error": "formato de arquivo não suportado (use jpg, png, webp ou heic)"}), 400

    _, row = _atividade_por_id(atividade_id)
    if row is None:
        return jsonify({"ok": False, "error": "atividade não encontrada"}), 404
    numero_os = row[ATIV_CAMPO_COL["numeroOS"] - 1].strip()
    if numero_os:
        return jsonify({"ok": False, "error": "essa atividade veio da Fracttal (tem OS vinculada) — anexo de fotos por enquanto só é permitido em atividades criadas manualmente"}), 403

    try:
        pasta = os.path.join(FOTOS_ATIVIDADES_DIR, secure_filename(atividade_id))
        os.makedirs(pasta, exist_ok=True)
        ext = arquivo.filename.rsplit(".", 1)[1].lower()
        nome_seguro = f"{uuid.uuid4().hex}.{ext}"
        caminho = os.path.join(pasta, nome_seguro)
        arquivo.save(caminho)
        return jsonify({"ok": True, "arquivo": nome_seguro}), 200
    except Exception as e:
        log.error(f"[Fotos Atividade] Erro ao salvar foto da atividade {atividade_id}: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/atividade-fotos/<atividade_id>", methods=["GET"])
def atividade_listar_fotos(atividade_id):
    """Lista as fotos já anexadas a uma atividade."""
    pasta = os.path.join(FOTOS_ATIVIDADES_DIR, secure_filename(atividade_id))
    if not os.path.isdir(pasta):
        return jsonify({"ok": True, "fotos": []}), 200
    try:
        arquivos = sorted(os.listdir(pasta))
        return jsonify({"ok": True, "fotos": arquivos}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/foto-atividade/<atividade_id>/<nome_arquivo>", methods=["GET"])
def atividade_servir_foto(atividade_id, nome_arquivo):
    """Serve o arquivo de imagem em si. secure_filename nos dois
    parâmetros previne path traversal (ex.: id="../../etc")."""
    pasta = os.path.join(FOTOS_ATIVIDADES_DIR, secure_filename(atividade_id))
    nome_seguro = secure_filename(nome_arquivo)
    caminho = os.path.join(pasta, nome_seguro)
    if not os.path.isfile(caminho):
        return jsonify({"ok": False, "error": "foto não encontrada"}), 404
    return send_file(caminho)


@app.route("/atividade-remover-foto", methods=["POST", "OPTIONS"])
def atividade_remover_foto():
    """Remove uma foto anexada a uma atividade."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    atividade_id = (body.get("id") or "").strip()
    nome_arquivo = (body.get("arquivo") or "").strip()
    if not atividade_id or not nome_arquivo:
        return jsonify({"ok": False, "error": "id e arquivo são obrigatórios"}), 400
    caminho = os.path.join(FOTOS_ATIVIDADES_DIR, secure_filename(atividade_id), secure_filename(nome_arquivo))
    try:
        if os.path.isfile(caminho):
            os.remove(caminho)
        return jsonify({"ok": True}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/outras-usinas-supervisores", methods=["GET"])
def outras_usinas_supervisores():
    """
    Lista as usinas de TODOS os outros supervisores (não o Fred), agrupadas
    por responsável, direto da fonte pública do PCM (banco_dados.json,
    semana ativa) — pro Fred escolher quais quer assumir temporariamente
    (ex.: cobertura de férias de outro supervisor).

    Implementado em 30/07/2026; restaurado em 31/07/2026.
    """
    try:
        resp = requests.get(_PCM_BANCO_URL, timeout=25)
        resp.raise_for_status()
        dados = resp.json()
    except Exception as e:
        return jsonify({"ok": False, "error": f"Falha ao buscar dados do PCM: {e}"}), 502

    semana_ativa = dados.get("semana_ativa")
    semana = next((s for s in dados.get("semanas", []) if s.get("week") == semana_ativa), None)
    if semana is None and dados.get("semanas"):
        semana = dados["semanas"][0]
    rows = semana.get("rows", []) if semana else []

    ja_adicionadas = {item["usina"] for item in _usinas_temporarias()}
    vistos = {}
    for r in rows:
        resp_nome = (r.get("responsavel") or "").strip()
        if not resp_nome or resp_nome == _PCM_RESPONSAVEL:
            continue
        usina = (r.get("usina") or "").strip()
        if not usina:
            continue
        chave = (resp_nome, usina)
        cluster = (r.get("cluster") or "").strip()
        if chave not in vistos or cluster.isupper():
            vistos[chave] = {
                "cliente": (r.get("cliente") or "").strip(),
                "usina": usina,
                "cluster": cluster,
                "responsavel": resp_nome,
                "jaAdicionada": usina in ja_adicionadas,
            }

    por_supervisor = {}
    for v in vistos.values():
        por_supervisor.setdefault(v["responsavel"], []).append(v)
    for lista in por_supervisor.values():
        lista.sort(key=lambda x: x["usina"])

    return jsonify({"ok": True, "porSupervisor": por_supervisor,
                     "semanaFonte": semana.get("week") if semana else None}), 200


@app.route("/supervisao-temporaria", methods=["GET"])
def listar_supervisao_temporaria():
    """Lista as usinas atualmente sob supervisão temporária do Fred.
    Correção 31/07/2026: não força mais releitura do Sheets a cada
    abertura da tela — isso somava com o resto do sistema (auditorias,
    sync-fracttal etc.) e estourava a cota de leitura do Google Sheets.
    O cache (10 min) já é invalidado nas ações de adicionar/remover, que
    é quando o estado realmente muda; abrir a tela sem mexer em nada
    pode reaproveitar o cache tranquilamente."""
    return jsonify({"ok": True, "itens": _usinas_temporarias()}), 200


@app.route("/supervisao-temporaria/adicionar", methods=["POST", "OPTIONS"])
def adicionar_supervisao_temporaria():
    """Adiciona uma usina de outro supervisor à supervisão temporária do
    Fred — a partir desse momento, ela passa a ser reconhecida em TODO o
    sistema (catálogo de usinas, chamados, comunicados, filtros,
    programação PCM) como se fosse dele, até ser removida."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    cliente = (body.get("cliente") or "").strip()
    usina = (body.get("usina") or "").strip()
    cluster = (body.get("cluster") or "").strip()
    responsavel_original = (body.get("responsavelOriginal") or "").strip()
    grupo_id = (body.get("grupoId") or "").strip()
    if not usina:
        return jsonify({"ok": False, "error": "usina é obrigatória"}), 400
    try:
        ws = get_supervisao_temp_sheet()
        valores = ws.get_all_values()
        linha_existente = next((i for i, row in enumerate(valores[1:], start=2) if len(row) > 1 and row[1].strip() == usina), None)
        if linha_existente:
            # Já existe: se um grupo foi informado agora, atualiza mesmo
            # assim (permite usar esta mesma tela pra corrigir/trocar o
            # grupo de uma usina já sob supervisão, sem precisar remover
            # e adicionar de novo) — tanto na coluna F da planilha quanto
            # no mapeamento grupo_usina/cluster_usina em _Sistema.
            if grupo_id:
                ws.update_cell(linha_existente, 6, grupo_id)
                _config_set_lote_core({
                    f"grupo_usina:{usina}": grupo_id,
                    **({f"cluster_usina:{usina}": cluster} if cluster else {}),
                })
                _usinas_temporarias_cache["expira_em"] = 0
                _mapa_grupo_usina_cache["expira_em"] = 0
                _mapa_cluster_usina_cache["expira_em"] = 0
            return jsonify({"ok": True, "jaExistia": True}), 200
        ws.append_row([cliente, usina, cluster, responsavel_original, agora_br().strftime("%d/%m/%Y %H:%M:%S"), grupo_id])
        _usinas_temporarias_cache["expira_em"] = 0
        _indices_temporarios_cache["expira_em"] = 0
        # Vincula automaticamente a usina ao grupo do WhatsApp informado
        # (grupo_usina) e ao cluster (cluster_usina) na aba _Sistema —
        # antes isso exigia um passo manual separado via /config-set-lote.
        if grupo_id:
            _config_set_lote_core({
                f"grupo_usina:{usina}": grupo_id,
                **({f"cluster_usina:{usina}": cluster} if cluster else {}),
            })
            _mapa_grupo_usina_cache["expira_em"] = 0
            _mapa_cluster_usina_cache["expira_em"] = 0
        return jsonify({"ok": True}), 200
    except Exception as e:
        log.error(f"[SupervisaoTemporaria] Erro ao adicionar {usina}: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/supervisao-temporaria/remover", methods=["POST", "OPTIONS"])
def remover_supervisao_temporaria():
    """Remove uma usina da supervisão temporária — volta ao normal
    (deixa de ser reconhecida como do Fred em todo o sistema)."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    usina = (body.get("usina") or "").strip()
    if not usina:
        return jsonify({"ok": False, "error": "usina é obrigatória"}), 400
    try:
        ws = get_supervisao_temp_sheet()
        valores = ws.get_all_values()
        for i, row in enumerate(valores[1:], start=2):
            if len(row) > 1 and row[1].strip() == usina:
                ws.delete_rows(i)
                break
        _usinas_temporarias_cache["expira_em"] = 0
        _indices_temporarios_cache["expira_em"] = 0
        # Limpa também o vínculo grupo_usina/cluster_usina dessa usina na
        # aba _Sistema, se existir — evita deixar mapeamento órfão
        # apontando pra uma usina que a Fred já devolveu.
        _config_remover_chaves_core([f"grupo_usina:{usina}", f"cluster_usina:{usina}"])
        _mapa_grupo_usina_cache["expira_em"] = 0
        _mapa_cluster_usina_cache["expira_em"] = 0
        return jsonify({"ok": True}), 200
    except Exception as e:
        log.error(f"[SupervisaoTemporaria] Erro ao remover {usina}: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/clientes-configurados", methods=["GET"])
def listar_clientes_configurados():
    """
    Lista os clientes cadastrados — derivada diretamente de
    CATALOGO_USINAS (a mesma fonte usada pra reconhecer OS da Fracttal),
    em vez de ser uma lista solta mantida à parte no frontend.

    Criado em 27/07/2026 depois de descobrir que a Sal Energia estava
    cadastrada na lista de clientes do frontend, mas ausente do
    catálogo de usinas do backend — duas fontes de verdade que
    precisavam ser lembradas separadamente, e uma ficou pra trás. Isso
    faz o frontend buscar a lista aqui, então cadastrar um cliente novo
    num lugar só (aqui) já reflete em tudo.

    Também inclui clientes de usinas sob supervisão temporária (ver
    /supervisao-temporaria), consistente com o resto do sistema.
    """
    _, cliente_temp = _indices_temporarios()
    clientes = sorted(set(_CLIENTE_INDEX.values()) | set(cliente_temp.values()))
    return jsonify({"ok": True, "clientes": clientes}), 200


# ══════════════════════════════════════════════════════════════════════
# CAPTURA DE MENSAGENS DOS GRUPOS — pro resumo diário/semanal (Gestão
# O&M). Usa SQLite local em vez de Google Sheets de propósito: volume
# potencialmente alto (16 grupos, todo dia) e mais frequente que
# qualquer outra escrita do sistema — colocar isso nas Sheets arriscaria
# estourar a cota de escrita/leitura que já vimos dar 429 antes
# (27/07/2026). Implementado a pedido do Fred pra alimentar o resumo
# diário/semanal com o que foi tratado em cada grupo, não só o que já
# era parseado (status de OS via palavras-chave).
# ══════════════════════════════════════════════════════════════════════
MENSAGENS_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mensagens_grupos.db")


def _get_mensagens_db():
    conn = sqlite3.connect(MENSAGENS_DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS mensagens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            grupo_id TEXT NOT NULL,
            nome_grupo TEXT NOT NULL,
            remetente TEXT,
            texto TEXT,
            data_hora TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mensagens_grupo_data ON mensagens(grupo_id, data_hora)")
    # Histórico dos resumos diário/semanal já gerados — pro Fred conseguir
    # ver no próprio painel, não só no WhatsApp (pedido em 29/07/2026).
    conn.execute("""
        CREATE TABLE IF NOT EXISTS resumos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            data_referencia TEXT NOT NULL,
            data_inicio TEXT,
            data_fim TEXT,
            texto TEXT NOT NULL,
            enviado_whatsapp INTEGER NOT NULL DEFAULT 0,
            criado_em TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_resumos_tipo_data ON resumos(tipo, data_referencia)")
    return conn


def _salvar_resumo(tipo, texto, data_referencia, data_inicio=None, data_fim=None, enviado=False):
    conn = _get_mensagens_db()
    conn.execute(
        "INSERT INTO resumos (tipo, data_referencia, data_inicio, data_fim, texto, enviado_whatsapp, criado_em) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (tipo, data_referencia, data_inicio, data_fim, texto, 1 if enviado else 0, agora_br().strftime("%Y-%m-%d %H:%M:%S")),
    )
    conn.commit()
    novo_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return novo_id


@app.route("/resumos", methods=["GET"])
def listar_resumos():
    """Lista o histórico de resumos diários/semanais já gerados, pro
    painel mostrar (não só o WhatsApp). Filtros opcionais: ?tipo=diario
    ou ?tipo=semanal, ?limit=N (default 30)."""
    tipo = request.args.get("tipo", "").strip()
    limit = min(int(request.args.get("limit", 30) or 30), 200)
    conn = _get_mensagens_db()
    conn.row_factory = sqlite3.Row
    query = "SELECT id, tipo, data_referencia, data_inicio, data_fim, texto, enviado_whatsapp, criado_em FROM resumos"
    params = []
    if tipo:
        query += " WHERE tipo = ?"
        params.append(tipo)
    query += " ORDER BY criado_em DESC LIMIT ?"
    params.append(limit)
    linhas = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify({"ok": True, "itens": [dict(r) for r in linhas]}), 200


@app.route("/capturar-mensagem-grupo", methods=["POST", "OPTIONS"])
def capturar_mensagem_grupo():
    """Arquiva uma mensagem de um dos 16 grupos mapeados pro resumo
    diário/semanal. Chamado pela ponte do WhatsApp (server.js) pra TODA
    mensagem desses grupos, sem filtro de conteúdo."""
    if request.method == "OPTIONS":
        return ("", 204)
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    body = request.get_json(force=True, silent=True) or {}
    grupo_id = (body.get("grupoId") or "").strip()
    nome_grupo = (body.get("nomeGrupo") or "").strip()
    remetente = (body.get("remetente") or "").strip()
    texto = (body.get("texto") or "").strip()
    if not grupo_id or not texto:
        return jsonify({"ok": False, "error": "grupoId e texto são obrigatórios"}), 400

    try:
        conn = _get_mensagens_db()
        conn.execute(
            "INSERT INTO mensagens (grupo_id, nome_grupo, remetente, texto, data_hora) VALUES (?, ?, ?, ?, ?)",
            (grupo_id, nome_grupo, remetente, texto, agora_br().strftime("%Y-%m-%d %H:%M:%S")),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True}), 200
    except Exception as e:
        log.error(f"[MensagensGrupos] Erro ao gravar mensagem de {nome_grupo}: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/mensagens-capturadas", methods=["GET"])
def mensagens_capturadas():
    """Consulta de diagnóstico: lista as mensagens realmente capturadas
    no banco local (via /capturar-mensagem-grupo, tempo real) num
    período/grupo, pra investigar buracos de captura — ex.: quando a
    ponte do WhatsApp (VM2) fica desconectada, nenhuma mensagem chega
    nesse período, independente da janela que a ronda/resumo consulte
    depois. Query params: ?data_inicio=YYYY-MM-DD (obrigatório),
    ?data_fim=YYYY-MM-DD (default = data_inicio), ?grupo_id=... (opcional)."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip() or data_inicio
    grupo_id = request.args.get("grupo_id", "").strip() or None
    if not data_inicio:
        return jsonify({"ok": False, "error": "data_inicio é obrigatório (YYYY-MM-DD)"}), 400
    try:
        msgs = _buscar_mensagens_periodo(data_inicio, data_fim, grupo_id=grupo_id)
        por_grupo = {}
        for m in msgs:
            por_grupo.setdefault(m["nome_grupo"], 0)
            por_grupo[m["nome_grupo"]] += 1
        return jsonify({"ok": True, "total": len(msgs), "por_grupo": por_grupo, "mensagens": msgs}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _buscar_mensagens_periodo(data_inicio, data_fim, grupo_id=None):
    """Busca mensagens capturadas entre data_inicio e data_fim (strings
    'YYYY-MM-DD', inclusive dos dois lados), opcionalmente filtrando por
    um grupo específico. Retorna lista de dicts."""
    conn = _get_mensagens_db()
    conn.row_factory = sqlite3.Row
    query = "SELECT * FROM mensagens WHERE date(data_hora) >= ? AND date(data_hora) <= ?"
    params = [data_inicio, data_fim]
    if grupo_id:
        query += " AND grupo_id = ?"
        params.append(grupo_id)
    query += " ORDER BY data_hora ASC"
    linhas = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in linhas]


GRUPO_GESTAO_OM_ID = "120363429317295622@g.us"


def _pcm_linhas_do_dia(data_str):
    """Retorna as linhas cruas da programação do PCM pro Fred, numa data
    específica (YYYY-MM-DD) — extraído do endpoint /programacao-pcm pra
    ser reaproveitado pelo resumo diário/semanal sem duplicar lógica."""
    try:
        dt = datetime.strptime(data_str, "%Y-%m-%d").date()
    except ValueError:
        return []
    try:
        cache = _buscar_programacao_pcm_core()
    except Exception:
        return []
    dados = cache.get("dados") or {}
    semanas = dados.get("semanas", [])
    iso_year, iso_week, _ = dt.isocalendar()
    semana_alvo = f"{iso_year}-W{iso_week:02d}"
    semana = next((s for s in semanas if s.get("week") == semana_alvo), None)
    if semana is None:
        return []
    dia_pt = _DIA_SEMANA_PT[dt.weekday()]
    usinas_temp_nomes = {item["usina"] for item in _usinas_temporarias()}
    return [r for r in semana.get("rows", [])
            if (r.get("responsavel") == _PCM_RESPONSAVEL or r.get("usina") in usinas_temp_nomes) and r.get("dia") == dia_pt]


FALHAS_SHEET_NAME_CANDIDATOS = ["Painel de Falhas - Fred Alexandrino", "Painel de Falhas"]


def _falhas_itens():
    """Lê a aba do Painel de Falhas direto via gspread (mesma
    credencial/conexão usada em todo o resto do sistema), com
    correspondência de coluna por palavra-chave no cabeçalho — igual à
    lógica já usada no frontend (fetchSheet), pra não depender de posição
    fixa de coluna. Criado em 27/07/2026 pro resumo diário/semanal."""
    gc = get_gc()
    ss = gc.open_by_key(SHEET_ID)
    ws = None
    for nome in FALHAS_SHEET_NAME_CANDIDATOS:
        try:
            ws = ss.worksheet(nome)
            break
        except gspread.WorksheetNotFound:
            continue
    if ws is None:
        ws = ss.get_worksheet(0)  # fallback: primeira aba (gid=0), mesma usada pelo frontend

    todos = ws.get_all_values()
    if not todos:
        return []
    header = [h.lower().strip() for h in todos[0]]

    def _achar_idx(*palavras_chave):
        for i, h in enumerate(header):
            if any(p in h for p in palavras_chave):
                return i
        return None

    idx = {
        "id": _achar_idx("id") or 0,
        "dataAbertura": _achar_idx("data_abertura", "data abertura"),
        "cliente": _achar_idx("cliente"),
        "usina": _achar_idx("usina"),
        "falha": _achar_idx("falha"),
        "status": _achar_idx("status"),
    }
    itens = []
    for row in todos[1:]:
        if not row or not (row[idx["id"]] if idx["id"] < len(row) else "").strip():
            continue
        def _get(campo, default_idx=None):
            i = idx.get(campo, default_idx)
            return row[i].strip() if i is not None and i < len(row) else ""
        itens.append({
            "id": _get("id"), "dataAbertura": _get("dataAbertura"),
            "cliente": _get("cliente"), "usina": _get("usina"),
            "falha": _get("falha"), "status": _get("status") or "Em Aberto",
        })
    return itens


def _coletar_dados_resumo_diario(data_str):
    """Junta tudo que o resumo diário precisa pra uma data específica
    (YYYY-MM-DD): programação do PCM cruzada com status real das OS,
    atividades concluídas no dia que NÃO estavam programadas, chamados
    de fabricante do dia, ocorrências novas do dia, desligamentos
    ativos, prazos vencendo amanhã, e as mensagens capturadas nos
    grupos mapeados."""
    resultado = {"data": data_str}

    # ── Programação do PCM x status real das OS ─────────────────────────
    linhas_pcm = _pcm_linhas_do_dia(data_str)
    try:
        ws_ativ = get_atividades_sheet()
        todos_ativ = ws_ativ.get_all_values()
    except Exception as e:
        log.error(f"[ResumoDiario] Erro ao ler Atividades: {e}")
        todos_ativ = []

    por_numero_os = {}
    for i, row in enumerate(todos_ativ[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        numero_os = row[ATIV_CAMPO_COL["numeroOS"] - 1].strip()
        if numero_os:
            por_numero_os[numero_os] = (i, row)

    # ── Revalidação AO VIVO na Fracttal antes de classificar cumprido/pendente
    # (corrigido em 17/08/2026, relatado pelo Fred): o resumo antes confiava
    # cegamente no statusOS já gravado na planilha — mas esse campo só é
    # tão fresco quanto a última vez que o rodízio de 5min passou por
    # aquela OS específica. Numa janela de recuperação de backlog grande
    # (ex.: dias de sync quebrado sendo processados de uma vez), dava pra
    # o resumo rodar e ler "Em Revisão"/"Finalizada" desatualizado
    # SEGUNDOS antes da auditoria corrigir aquele mesmo registro pra
    # "Em Processo" de verdade (ex.: 17/08/2026 — 3 OSs viradas como
    # concluídas às 17h, só que a Fracttal mostrava 12%-76% concluído,
    # porque a correção da auditoria pra essas 3 OSs específicas aconteceu
    # minutos DEPOIS do resumo já ter sido gerado e enviado). Como a
    # programação do dia é tipicamente uma lista curta (poucas dezenas de
    # itens), vale a pena confirmar CADA UMA ao vivo na hora de montar o
    # resumo, em vez de reaproveitar o cache — elimina esse tipo de corrida
    # de vez, não só nesse caso específico.
    ORCAMENTO_RECHECK_RESUMO_SEGUNDOS = 40
    inicio_recheck_resumo = time.time()
    parou_por_orcamento_resumo = False

    programado_cumprido, programado_pendente = [], []
    for r in linhas_pcm:
        os_id = str(r.get("os_id") or "").strip()
        item = {
            "usina": r.get("usina"), "cliente": r.get("cliente"), "tarefa": r.get("tarefa"),
            "os": os_id, "hIni": r.get("h_ini"), "tipo": r.get("tipo"),
        }
        entry = por_numero_os.get(os_id) if os_id else None
        if entry is not None:
            linha_idx, row_real = entry
            status_interno_cache = row_real[ATIV_CAMPO_COL["status"] - 1].strip()
            status_os_cache = row_real[ATIV_CAMPO_COL["statusOS"] - 1].strip()

            status_os_fresco = status_os_cache
            status_interno_fresco = status_interno_cache
            if not parou_por_orcamento_resumo and (time.time() - inicio_recheck_resumo) <= ORCAMENTO_RECHECK_RESUMO_SEGUNDOS:
                try:
                    resultado_live = _fracttal_verificar_e_atualizar_uma_os(
                        ws_ativ, linha_idx, row_real, os_id, enviar_notificacao=False)
                    if resultado_live:
                        status_os_fresco = resultado_live.get("statusOS") or status_os_cache
                        status_interno_fresco = _status_interno_esperado(status_os_fresco, status_interno_cache) or status_interno_cache
                    time.sleep(0.3)
                except Exception as e:
                    log.error(f"[ResumoDiario] Erro ao revalidar OS {os_id} ao vivo (usando cache): {e}")
            else:
                parou_por_orcamento_resumo = True

            item["statusReal"] = status_interno_fresco
            item["statusOSReal"] = status_os_fresco
            item["equipamento"] = row_real[ATIV_CAMPO_COL["equipamento"] - 1].strip()
            item["responsavel"] = row_real[ATIV_CAMPO_COL["responsavel"] - 1].strip() if len(row_real) > ATIV_CAMPO_COL["responsavel"] - 1 else ""
            if _is_concluido_atividade(status_interno_fresco) or status_os_fresco in ("Em Revisão", "Finalizada"):
                programado_cumprido.append(item)
            else:
                programado_pendente.append(item)
        else:
            item["statusReal"] = "sem OS correspondente no painel"
            programado_pendente.append(item)

    if parou_por_orcamento_resumo:
        log.warning(f"[ResumoDiario] Orçamento de {ORCAMENTO_RECHECK_RESUMO_SEGUNDOS}s pra revalidação ao vivo "
                    f"esgotado — parte da programação do dia usou o cache já salvo em vez de checar a Fracttal na hora.")

    resultado["programacao"] = {"cumprido": programado_cumprido, "pendente": programado_pendente}

    # ── Atividades concluídas no dia que NÃO estavam na programação ─────
    # Corrigido em 28/07/2026: só olhar dataConclusao não bastava — uma
    # auditoria automática pode marcar uma OS antiga como "Concluído" só
    # hoje (catch-up de backlog), sem que o trabalho de campo tenha
    # acontecido hoje de verdade. Agora exige também uma entrada no
    # histórico datada de hoje, como evidência real de atividade no dia
    # (mensagem de técnico, atualização da Fracttal, etc.) — não só uma
    # mudança de status administrativa.
    #
    # CORREÇÃO 28/07/2026 (relatado pelo Fred): dataConclusao reflete
    # quando a Fracttal finalizou ADMINISTRATIVAMENTE a OS (mudou pra
    # "Finalizada"), que pode ser dias depois do trabalho de campo real
    # (que termina quando o técnico manda pra "Em Revisão", 100%). Usar
    # dataConclusao sozinha fez o resumo dizer "fizemos religamento hoje"
    # pra OSs cujo trabalho de campo foi feito 4 dias antes — só a
    # Fracttal demorou pra confirmar. Agora busca especificamente a data
    # da transição de progresso/campo no histórico, não qualquer menção
    # à data de hoje (que aparecia de qualquer forma por causa da
    # finalização administrativa).
    _padrao_trabalho_campo = re.compile(
        r"(\d{2}/\d{2}/\d{4}) \d{2}:\d{2} - .*(?:progresso da tarefa foi de|"
        r'mudou de ".*?" para "Em Revisão"|situação geral da tarefa mudou.*?"Concluída")',
    )
    data_str_br = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    numeros_programados = {str(r.get("os_id") or "").strip() for r in linhas_pcm if r.get("os_id")}
    extras_nao_programadas = []
    # mesmo orçamento/lógica de revalidação ao vivo do bloco de programação
    # acima (corrigido 17/08/2026) — esse bloco também decide "concluída"
    # só olhando o HISTÓRICO salvo (padrão de texto batendo com a data de
    # hoje), sem nunca confirmar contra o estado atual da Fracttal. Mesmo
    # risco de corrida: se o histórico registrou "mudou pra Em Revisão"
    # hoje de manhã e a OS foi reaberta hoje à tarde (checklist reprovado,
    # trabalho incompleto), esse bloco reportava como concluída mesmo já
    # reaberta.
    inicio_recheck_extras = time.time()
    parou_por_orcamento_extras = False
    for i, row in enumerate(todos_ativ[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        historico = row[ATIV_CAMPO_COL["historico"] - 1] if len(row) > ATIV_CAMPO_COL["historico"] - 1 else ""
        datas_trabalho_campo = _padrao_trabalho_campo.findall(historico)
        if data_str_br not in datas_trabalho_campo:
            continue  # trabalho de campo não foi feito hoje (mesmo que a OS tenha sido finalizada administrativamente hoje)
        numero_os = row[ATIV_CAMPO_COL["numeroOS"] - 1].strip()
        if numero_os and numero_os in numeros_programados:
            continue  # já contabilizada como programação cumprida
        usina_bruta = row[ATIV_CAMPO_COL["usina"] - 1].strip()
        if canonizar_usina(usina_bruta) is None:
            continue  # usina de outro cliente/supervisor vazada no painel — não é do Fred, ignora

        status_os_fresco = row[ATIV_CAMPO_COL["statusOS"] - 1].strip()
        if numero_os and not parou_por_orcamento_extras and (time.time() - inicio_recheck_extras) <= 20:
            try:
                resultado_live = _fracttal_verificar_e_atualizar_uma_os(
                    ws_ativ, i, row, numero_os, enviar_notificacao=False)
                if resultado_live:
                    status_os_fresco = resultado_live.get("statusOS") or status_os_fresco
                    # reflete no registro em memória pra qualquer loop
                    # seguinte (ex.: progresso do dia, logo abaixo) já
                    # enxergar o dado atualizado, sem precisar reler a
                    # planilha inteira de novo.
                    row[ATIV_CAMPO_COL["statusOS"] - 1] = status_os_fresco
                    row[ATIV_CAMPO_COL["percentualOS"] - 1] = resultado_live.get("percentualOS") or row[ATIV_CAMPO_COL["percentualOS"] - 1]
                    row[ATIV_CAMPO_COL["statusGeralOS"] - 1] = resultado_live.get("statusGeralOS") or row[ATIV_CAMPO_COL["statusGeralOS"] - 1]
                    todos_ativ[i - 1] = row
                time.sleep(0.3)
            except Exception as e:
                log.error(f"[ResumoDiario] Erro ao revalidar OS {numero_os} (extras) ao vivo (usando cache): {e}")
        else:
            parou_por_orcamento_extras = True

        if status_os_fresco not in ("Em Revisão", "Finalizada"):
            continue  # reaberta/ainda incompleta — não é notícia de conclusão do dia, cai pro bloco de progresso abaixo

        extras_nao_programadas.append({
            "usina": usina_bruta,
            "cliente": row[ATIV_CAMPO_COL["cliente"] - 1].strip(),
            "equipamento": row[ATIV_CAMPO_COL["equipamento"] - 1].strip(),
            "descricao": row[ATIV_CAMPO_COL["descricao"] - 1].strip(),
            "numeroOS": numero_os,
        })
    resultado["extrasNaoProgramadas"] = extras_nao_programadas

    # ── Atividades com progresso hoje (mudança de % ou estado registrada
    #    no histórico), mesmo sem chegar a "Concluído" — pra não deixar
    #    de fora OS que avançaram mas ainda não fecharam. Não repete o
    #    que já apareceu em "cumprido" ou "extras". Adicionado em
    #    28/07/2026 a pedido do Fred.
    progresso_do_dia = []
    numeros_ja_contabilizados = numeros_programados | {e["numeroOS"] for e in extras_nao_programadas if e.get("numeroOS")}
    # mesmo cuidado do bloco de "extras": uma linha de histórico com a
    # data de hoje pode ser SÓ a finalização administrativa (Fracttal
    # mudou pra "Finalizada"), sem nenhum progresso de campo real hoje —
    # nesse caso não é notícia do dia, é catch-up atrasado. Só entra aqui
    # se tiver pelo menos uma linha de hoje que NÃO seja exclusivamente
    # essa transição administrativa.
    _padrao_so_finalizacao_administrativa = re.compile(
        r'^\d{2}/\d{2}/\d{4} \d{2}:\d{2} - status na Fracttal mudou de ".*?" para "Finalizada"\.?$'
        r'|^\d{2}/\d{2}/\d{4} \d{2}:\d{2} - .*Status interno corrigido pra "Conclu[ií]do".*$'
    )
    for row in todos_ativ[1:]:
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        numero_os = row[ATIV_CAMPO_COL["numeroOS"] - 1].strip()
        if not numero_os or numero_os in numeros_ja_contabilizados:
            continue
        historico = row[ATIV_CAMPO_COL["historico"] - 1] if len(row) > ATIV_CAMPO_COL["historico"] - 1 else ""
        linhas_de_hoje = [l for l in historico.split("\n") if l.strip().startswith(data_str_br)]
        linhas_relevantes = [l for l in linhas_de_hoje if not _padrao_so_finalizacao_administrativa.match(l.strip())]
        if not linhas_relevantes:
            continue  # só teve finalização administrativa hoje, sem progresso de campo real
        usina_bruta = row[ATIV_CAMPO_COL["usina"] - 1].strip()
        if canonizar_usina(usina_bruta) is None:
            continue  # usina de outro cliente/supervisor vazada no painel — não é do Fred, ignora
        progresso_do_dia.append({
            "usina": usina_bruta,
            "cliente": row[ATIV_CAMPO_COL["cliente"] - 1].strip(),
            "equipamento": row[ATIV_CAMPO_COL["equipamento"] - 1].strip(),
            "descricao": row[ATIV_CAMPO_COL["descricao"] - 1].strip(),
            "numeroOS": numero_os,
            "percentualAtual": row[ATIV_CAMPO_COL["percentualOS"] - 1].strip() if len(row) > ATIV_CAMPO_COL["percentualOS"] - 1 else "",
            "statusAtual": row[ATIV_CAMPO_COL["statusOS"] - 1].strip(),
        })
    resultado["progressoDoDia"] = progresso_do_dia

    # ── Desligamentos que ocorreram HOJE ─────────────────────────────────
    # Checa tanto Falhas (ocorrências novas de hoje) quanto Atividades
    # (extras concluídas + progresso do dia) — um desligamento pode ter
    # sido registrado em qualquer um dos dois painéis. Corrigido em
    # 28/07/2026: a versão anterior só olhava Falhas e caía num bug de
    # fallback que trazia desligamentos crônicos antigos (Araputanga,
    # Nova Xavantina II) em vez do dia real, além de não pegar
    # desligamentos registrados só como Atividade (ex.: Morada Nova).
    try:
        overrides = {}
        ws_desl = get_desligamento_manual_sheet()
        for row in ws_desl.get_all_values()[1:]:
            if len(row) >= 3 and row[0].strip():
                overrides[f"{row[0].strip()}:{row[1].strip()}"] = row[2].strip()

        padrao_desligamento = re.compile(
            r"(?:usina|ufv)\s+(?:\w+\s+){0,3}(?:desligad[ao]|parad[ao]|sem\s+energia|desenergizad[ao]|offline|sem\s+comunica[çc][ãa]o)"
            r"|(?:desligad[ao]|parad[ao]|offline)\s+(?:\w+\s+){0,3}(?:usina|ufv)"
            r"|desligamento\s+(?:total\s+)?(?:da|de)\s+(?:usina|ufv)",
            re.IGNORECASE,
        )
        desligamentos = []
        for f in resultado.get("ocorrenciasNovasDoDia", []):
            override = overrides.get(f"falha:{f.get('id')}")
            texto = (f.get("falha") or "")
            if override == "sim" or (override != "nao" and padrao_desligamento.search(texto)):
                desligamentos.append({"usina": f.get("usina"), "cliente": f.get("cliente"), "descricao": texto, "origem": "Falha"})
        for a in (extras_nao_programadas + progresso_do_dia):
            override = overrides.get(f"atividade:{a.get('numeroOS')}")
            texto = (a.get("descricao") or "")
            if override == "sim" or (override != "nao" and padrao_desligamento.search(texto)):
                desligamentos.append({"usina": a.get("usina"), "cliente": a.get("cliente"), "descricao": texto, "origem": f"Atividade OS {a.get('numeroOS')}"})
        resultado["desligamentosAtivos"] = desligamentos
    except Exception as e:
        log.error(f"[ResumoDiario] Erro ao checar desligamentos: {e}")
        resultado["desligamentosAtivos"] = []

    # ── Chamados de fabricante abertos/atualizados no dia ────────────────
    try:
        chamados = _chamados_fabricante_itens()
        chamados_do_dia = [
            c for c in chamados
            if (c.get("Data da abertura do chamado") or "").strip() == datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        ]
        resultado["chamadosDoDia"] = chamados_do_dia
    except Exception as e:
        log.error(f"[ResumoDiario] Erro ao ler Chamados: {e}")
        resultado["chamadosDoDia"] = []

    # ── Ocorrências novas do dia (Painel de Falhas) ──────────────────────
    try:
        falhas = _falhas_itens()
        data_br = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        resultado["ocorrenciasNovasDoDia"] = [
            f for f in falhas if str(f.get("dataAbertura") or "").strip().startswith(data_br)
            and canonizar_usina(f.get("usina")) is not None
        ]
    except Exception as e:
        log.error(f"[ResumoDiario] Erro ao ler Falhas: {e}")
        resultado["ocorrenciasNovasDoDia"] = []

    # ── OS de alta prioridade ainda em aberto ────────────────────────────
    altas_abertas = []
    for row in todos_ativ[1:]:
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        usina_bruta = row[ATIV_CAMPO_COL["usina"] - 1].strip()
        if (row[ATIV_CAMPO_COL["prioridade"] - 1].strip().lower() == "alta"
                and not _is_concluido_atividade(row[ATIV_CAMPO_COL["status"] - 1].strip())
                and canonizar_usina(usina_bruta) is not None):
            altas_abertas.append({
                "usina": usina_bruta,
                "cliente": row[ATIV_CAMPO_COL["cliente"] - 1].strip(),
                "descricao": row[ATIV_CAMPO_COL["descricao"] - 1].strip(),
                "prazo": row[ATIV_CAMPO_COL["prazo"] - 1].strip(),
            })
    resultado["altaPrioridadeAberta"] = altas_abertas

    # ── Mensagens capturadas nos grupos, nesse dia ───────────────────────
    try:
        mensagens = _buscar_mensagens_periodo(data_str, data_str)
        # Corrigido 18/08/2026: só entra no resumo mensagem de grupo cuja
        # usina ainda é reconhecida agora (catálogo permanente ou
        # Supervisão Temporária ativa) — senão usina já devolvida
        # continuava aparecendo no resumo via mensagens do grupo dela.
        grupos_validos = _grupos_ids_ativos()
        por_grupo = {}
        for m in mensagens:
            if m.get("grupo_id") not in grupos_validos:
                continue
            por_grupo.setdefault(m["nome_grupo"], []).append(m)
        resultado["mensagensPorGrupo"] = por_grupo
    except Exception as e:
        log.error(f"[ResumoDiario] Erro ao ler mensagens capturadas: {e}")
        resultado["mensagensPorGrupo"] = {}

    return resultado


def _montar_prompt_resumo_diario(dados):
    data_str = dados["data"]
    try:
        data_fmt = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        data_fmt = data_str

    prog = dados.get("programacao", {})
    cumprido = prog.get("cumprido", [])
    pendente = prog.get("pendente", [])
    extras = dados.get("extrasNaoProgramadas", [])
    progresso = dados.get("progressoDoDia", [])
    chamados = dados.get("chamadosDoDia", [])
    ocorrencias = dados.get("ocorrenciasNovasDoDia", [])
    desligamentos = dados.get("desligamentosAtivos", [])
    altas = dados.get("altaPrioridadeAberta", [])
    mensagens_por_grupo = dados.get("mensagensPorGrupo", {})

    def _fmt_lista(itens, campos):
        if not itens:
            return "(nenhum)"
        linhas = []
        for it in itens:
            linhas.append(" | ".join(f"{c}={it.get(c, '')}" for c in campos))
        return "\n".join(linhas)

    # ── Cruzamento: números de OS mencionados nas mensagens x dados já
    #    mapeados (cumprido/pendente/extras/progresso). Isso é feito aqui
    #    em Python (não só confiando na IA achar sozinha no meio de um
    #    bloco de mensagens grande) pra ser confiável — pedido do Fred
    #    em 28/07/2026: "bater as informações das OS com as conversas no
    #    WhatsApp com as equipes".
    todas_os_conhecidas = {}
    for item in cumprido:
        if item.get("os"):
            todas_os_conhecidas[str(item["os"])] = {**item, "origem": "PCM cumprido"}
    for item in pendente:
        if item.get("os"):
            todas_os_conhecidas[str(item["os"])] = {**item, "origem": "PCM pendente"}
    for item in extras:
        if item.get("numeroOS"):
            todas_os_conhecidas[str(item["numeroOS"])] = {**item, "origem": "concluída fora da programação"}
    for item in progresso:
        if item.get("numeroOS"):
            todas_os_conhecidas[str(item["numeroOS"])] = {**item, "origem": "em progresso"}

    _padrao_numero_os = re.compile(r"\b(?:os\.?\s*)?(\d{4,6})\b", re.IGNORECASE)
    confirmacoes, mencoes_sem_correspondencia = [], []
    for nome_grupo, msgs in mensagens_por_grupo.items():
        for m in msgs:
            texto_msg = m.get("texto", "")
            for numero in set(_padrao_numero_os.findall(texto_msg)):
                if numero in todas_os_conhecidas:
                    confirmacoes.append({
                        "os": numero, "grupo": nome_grupo, "remetente": m.get("remetente"),
                        "trechoMensagem": texto_msg[:200],
                        "statusNoSistema": todas_os_conhecidas[numero].get("origem"),
                    })
                elif len(numero) >= 4:  # ignora numeros curtos tipo horario/telefone parcial
                    mencoes_sem_correspondencia.append({
                        "os": numero, "grupo": nome_grupo, "remetente": m.get("remetente"),
                        "trechoMensagem": texto_msg[:200],
                    })

    def _fmt_cruzamento(itens, teto=40):
        if not itens:
            return "(nenhuma)"
        linhas = [f"OS {it['os']} | grupo={it['grupo']} | {it['remetente']}: \"{it['trechoMensagem']}\"" +
                  (f" | status no sistema: {it.get('statusNoSistema','')}" if 'statusNoSistema' in it else "")
                  for it in itens[:teto]]
        if len(itens) > teto:
            linhas.append(f"... e mais {len(itens) - teto}")
        return "\n".join(linhas)

    bloco_mensagens = []
    for nome_grupo, msgs in mensagens_por_grupo.items():
        bloco_mensagens.append(f"\n--- Grupo: {nome_grupo} ({len(msgs)} mensagens) ---")
        for m in msgs[:80]:  # teto de segurança por grupo, evita prompt gigante em dia muito movimentado
            hora = (m.get("data_hora") or "").split(" ")[-1][:5]
            bloco_mensagens.append(f"[{hora}] {m.get('remetente','?')}: {m.get('texto','')[:300]}")
    texto_mensagens = "\n".join(bloco_mensagens) if bloco_mensagens else "(nenhuma mensagem capturada nos grupos hoje)"

    # Corrigido 18/08/2026: a lista de escopo tinha só o catálogo
    # permanente, deixando de fora usina sob Supervisão Temporária ainda
    # ATIVA — o que fazia a IA tratar como "não é sua" uma usina que na
    # verdade ainda é (contraditório com os dados de programação/extras
    # que já filtram corretamente por usina_permitida).
    usinas_temp_ativas = {item["usina"] for item in _usinas_temporarias()}
    lista_usinas_fred = ", ".join(sorted(set(CATALOGO_USINAS.keys()) | usinas_temp_ativas))

    return f"""Aja como um Supervisor de O&M Sênior da Grid Co., escrevendo o resumo diário das usinas pro seu próprio controle — vai ser enviado só pra você mesmo, num grupo pessoal de gestão (não é um comunicado pra equipe nem pra cliente).

Data do resumo: {data_fmt}

USINAS SOB SUA SUPERVISÃO (lista fechada — só estas): {lista_usinas_fred}

REGRA CRÍTICA — ESCOPO DE USINA: alguns grupos de WhatsApp (principalmente os grupos de cliente, ex. "[O&M] - Grid Co. | 2C") são compartilhados com técnicos/supervisores de OUTRAS usinas do mesmo cliente que NÃO são suas. Se uma mensagem mencionar uma usina que NÃO está na lista acima (ex.: "Tupi Paulista", "Macaíba", "Santarém", "Aparecida" — nomes de exemplo, qualquer nome fora da lista se aplica), essa usina NÃO é sua responsabilidade — não mencione ela em NENHUMA seção do resumo (nem em alertas, nem em "OS sem correspondência", nem no resumo por equipe), mesmo que pareça relevante ou urgente. Trate como ruído de outro supervisor e ignore completamente. Isso vale pra toda a lista de usinas acima, e só pra essa lista.

REGRAS DE ESCRITA:
- Direto e sem enrolação, mas COMPLETO — não é pra ser um título de uma linha por item. Pra cada OS/atividade relevante, diga o que foi feito de fato (não só citar código de tarefa/OS), com o detalhe técnico disponível nos dados (equipamento, responsável, o que o histórico ou a mensagem do técnico realmente diz). Duas ou três frases por item quando o dado sustentar isso é preferível a uma linha genérica tipo "Preventiva mensal realizada" sem contexto.
- Sem saudação. Estruture em tópicos com emojis moderados pra facilitar leitura rápida no celular, mas não sacrifique informação por brevidade artificial — o Fred prefere um resumo mais longo e completo a um curto e raso.
- NUNCA invente números, nomes ou fatos que não estão nos dados abaixo. Cada dado abaixo já foi validado como evidência real do dia — não generalize nem "arredonde" a descrição da tarefa (ex.: se a descrição cita religamento mas isso é só parte de uma tarefa maior, não resuma como "fizemos religamentos" sem mais contexto). Ao mesmo tempo, USE todo campo disponível (responsável, equipamento, tipo de tarefa, trecho de mensagem) pra dar substância — não descarte detalhe só pra encurtar.
- Se uma seção não tiver nada a reportar, diga isso em uma linha curta, não pule a seção.
- IMPORTANTE — cruzamento por NÚMERO DE OS: as seções "OS confirmadas por mensagem" e "OS mencionadas sem correspondência" abaixo já foram cruzadas automaticamente por número (não precisa procurar número de OS no bloco de mensagens sozinho).
- IMPORTANTE — cruzamento por EQUIPAMENTO/USINA (faça você mesmo, é o cruzamento mais comum): o técnico raramente fala o número da OS no grupo — ele fala do EQUIPAMENTO ou da USINA ("o inversor 1.6 de Ibaté", "religamos o trafo de Araputanga", "terminamos a preventiva de Crateús"). Cruze o que foi dito nas mensagens com o campo "equipamento"/"usina" das OS listadas acima (programação, concluídas, progresso). Se uma mensagem falar de um equipamento/usina que bate com uma OS do dia, uma dessas coisas:
  (a) CONFIRMA o que já está registrado (cite isso junto da OS correspondente, incluindo um trecho literal relevante da mensagem, ex.: "Ibaté I, inversor 1.10 (OS 9781, responsável Fulano) — confirmado pelo técnico no grupo: 'terminei a inspeção do 1.10, encontrei string 3 com baixa geração'"), ou
  (b) CONTRADIZ o que está registrado (ex.: técnico diz que não terminou algo que o sistema mostra concluído, ou vice-versa) — isso é importante, aponte como um alerta de divergência pro Fred verificar, ou
  (c) é uma menção nova sem OS correspondente no sistema — mesma lógica das OS sem correspondência: vale mencionar como algo pra conferir.
- RESUMO POR EQUIPE (obrigatório, seção própria no final, antes do fechamento): pra cada grupo do WhatsApp que teve mensagem hoje, escreva um parágrafo específico do que foi tratado NAQUELE grupo — traga contexto real (o que foi reportado, por quem, qual o desfecho se houver), não uma linha genérica. Se o grupo não teve mensagem relevante hoje ("bom dia", figurinha, coisa sem conteúdo), diga isso em uma linha ("Equipe X: sem assunto relevante hoje"). Ignore mensagens só de cortesia/figurinha ao montar o resumo, mas não invente conteúdo se não houver nada de fato. Lembre da REGRA CRÍTICA DE ESCOPO acima: se um grupo compartilhado só teve mensagens sobre usina(s) fora da sua lista, trate como "sem assunto relevante hoje" pra você — não resuma o assunto de outra usina.

DADOS DO DIA:

## Programação do PCM — cumprida
{_fmt_lista(cumprido, ['usina', 'cliente', 'tarefa', 'tipo', 'os', 'equipamento', 'responsavel'])}

## Programação do PCM — pendente/não cumprida hoje
{_fmt_lista(pendente, ['usina', 'cliente', 'tarefa', 'tipo', 'os', 'equipamento', 'responsavel', 'statusReal', 'statusOSReal'])}

## Atividades CONCLUÍDAS hoje fora da programação (evidência real de trabalho no histórico, não só mudança administrativa de status)
{_fmt_lista(extras, ['usina', 'cliente', 'equipamento', 'descricao', 'numeroOS'])}

## Atividades com PROGRESSO hoje, mas ainda não concluídas (avançaram % ou mudaram de estado)
{_fmt_lista(progresso, ['usina', 'cliente', 'equipamento', 'descricao', 'numeroOS', 'percentualAtual', 'statusAtual'])}

## Chamados de fabricante abertos hoje
{_fmt_lista(chamados, ['UFV', 'Fabricante', 'Motivo da abertura do chamado', 'Status'])}

## Ocorrências novas no Painel de Falhas hoje
{_fmt_lista(ocorrencias, ['usina', 'cliente', 'falha', 'status'])}

## Desligamentos que ocorreram HOJE (não é lista de desligamentos crônicos/antigos — só hoje)
{_fmt_lista(desligamentos, ['usina', 'cliente', 'descricao'])}

## OS de alta prioridade ainda em aberto (não necessariamente de hoje)
{_fmt_lista(altas, ['usina', 'cliente', 'descricao', 'prazo'])}

## OS confirmadas por mensagem no WhatsApp (número da OS mencionado numa mensagem BATE com uma OS já mapeada acima)
{_fmt_cruzamento(confirmacoes)}

## Números mencionados nas mensagens que PARECEM ser OS, mas NÃO batem com nenhuma OS mapeada hoje (checar se é OS de outro dia, erro de digitação do técnico, ou algo que ainda não está no painel)
{_fmt_cruzamento(mencoes_sem_correspondencia)}

## Mensagens nos grupos do WhatsApp mapeados hoje, organizadas por grupo (use pro cruzamento por equipamento/usina E pro resumo por equipe)
{texto_mensagens}

FORMATO DE SAÍDA (OBRIGATÓRIO): responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes ou depois), no formato:
{{"texto": "resumo diário completo e detalhado, pronto pra enviar no WhatsApp, começando com um cabeçalho tipo '📋 RESUMO DIÁRIO — {data_fmt}', e terminando com a seção 'RESUMO POR EQUIPE' descrita acima"}}"""


def _enviar_mensagem_grupo(grupo_id, texto):
    if not WPP_SERVER_URL:
        raise RuntimeError("WPP_SERVER_URL não configurado")
    r = requests.post(
        f"{WPP_SERVER_URL}/api/enviar-mensagem",
        json={"grupoId": grupo_id, "texto": texto},
        headers={"X-Webhook-Secret": WEBHOOK_SECRET} if WEBHOOK_SECRET else {},
        timeout=20,
    )
    r.raise_for_status()
    return r.json()


def _gerar_resumo_diario_core(data_str=None, enviar=True):
    if not data_str:
        data_str = agora_br().strftime("%Y-%m-%d")
    dados = _coletar_dados_resumo_diario(data_str)
    prompt = _montar_prompt_resumo_diario(dados)
    # Corrigido em 27/08/2026 (relatado pelo Fred: "resumos muito
    # resumidos e pobres de informação"): maxOutputTokens estava em 3072
    # — baixo o bastante pra cortar o resumo no meio em dias mais
    # movimentados, forçando a IA a comprimir demais mesmo com dados ricos
    # disponíveis. thinkingBudget=0 (sem raciocínio antes de escrever)
    # também empurrava pra respostas mais rasas. Subido pra 8192 tokens de
    # saída e 1536 de thinking — timeout do Gemini também subiu de 45s
    # pra 70s pra dar espaço a essa etapa extra sem arriscar cortar a
    # chamada pela metade.
    resp = _chamar_gemini_com_retry(
        {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.3,
                "maxOutputTokens": 8192,
                "responseMimeType": "application/json",
                "thinkingConfig": {"thinkingBudget": 1536},
            },
        },
        timeout=70,
    )
    resp_data = resp.json()
    texto_bruto = resp_data["candidates"][0]["content"]["parts"][0]["text"].strip()
    texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
    texto = json.loads(texto_limpo).get("texto", "").strip()
    if not texto:
        raise ValueError("A IA não retornou nenhum texto pro resumo diário")

    resultado_envio = None
    if enviar:
        resultado_envio = _enviar_mensagem_grupo(GRUPO_GESTAO_OM_ID, texto)

    try:
        _salvar_resumo("diario", texto, data_referencia=data_str, enviado=enviar)
    except Exception as e:
        log.error(f"[ResumoDiario] Falha ao salvar no histórico do painel: {e}")

    return {"ok": True, "data": data_str, "texto": texto, "envio": resultado_envio}


@app.route("/gerar-resumo-diario", methods=["POST", "GET"])
def gerar_resumo_diario():
    """Gera (e envia por padrão) o resumo diário pro grupo Gestão O&M.
    Query params: ?data=YYYY-MM-DD (default hoje), ?enviar=false (só
    gera e devolve o texto, sem mandar pro WhatsApp — útil pra testar).
    ?debug_dados=true devolve só os dados coletados (pré-IA), pra
    depurar classificação sem gastar chamada de Gemini."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    data_str = request.args.get("data", "").strip() or None
    enviar = request.args.get("enviar", "true").lower() != "false"
    if request.args.get("debug_dados", "").lower() == "true":
        try:
            dados = _coletar_dados_resumo_diario(data_str or agora_br().strftime("%Y-%m-%d"))
            dados.pop("mensagensPorGrupo", None)  # ruído grande, não precisa pro debug de classificação
            return jsonify({"ok": True, "dados": dados}), 200
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500
    try:
        resultado = _gerar_resumo_diario_core(data_str=data_str, enviar=enviar)
        return jsonify(resultado), 200
    except Exception as e:
        log.error(f"[ResumoDiario] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _coletar_dados_resumo_semanal(data_fim_str):
    """Roda a coleta diária pra cada dia da semana (segunda até a data
    informada, tipicamente sexta) e consolida tudo numa estrutura só."""
    dt_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
    dt_inicio = dt_fim - timedelta(days=dt_fim.weekday())  # segunda-feira da mesma semana

    consolidado = {
        "dataInicio": dt_inicio.strftime("%Y-%m-%d"), "dataFim": data_fim_str,
        "programacaoCumprida": [], "programacaoPendente": [], "extrasNaoProgramadas": [],
        "progressoDaSemana": [], "chamados": [], "ocorrencias": [], "altaPrioridadeAberta": [],
        "mensagensPorGrupo": {}, "diasProcessados": [],
    }
    dia = dt_inicio
    while dia <= dt_fim:
        dia_str = dia.strftime("%Y-%m-%d")
        try:
            dados_dia = _coletar_dados_resumo_diario(dia_str)
            consolidado["diasProcessados"].append(dia_str)
            consolidado["programacaoCumprida"].extend(dados_dia.get("programacao", {}).get("cumprido", []))
            consolidado["programacaoPendente"].extend(dados_dia.get("programacao", {}).get("pendente", []))
            consolidado["extrasNaoProgramadas"].extend(dados_dia.get("extrasNaoProgramadas", []))
            consolidado["progressoDaSemana"].extend(dados_dia.get("progressoDoDia", []))
            consolidado["chamados"].extend(dados_dia.get("chamadosDoDia", []))
            consolidado["ocorrencias"].extend(dados_dia.get("ocorrenciasNovasDoDia", []))
            for nome_grupo, msgs in dados_dia.get("mensagensPorGrupo", {}).items():
                consolidado["mensagensPorGrupo"].setdefault(nome_grupo, []).extend(msgs)
        except Exception as e:
            log.error(f"[ResumoSemanal] Erro ao coletar dia {dia_str}: {e}")
        dia += timedelta(days=1)

    # alta prioridade aberta é sempre o estado ATUAL (não faz sentido somar por dia)
    try:
        consolidado["altaPrioridadeAberta"] = _coletar_dados_resumo_diario(data_fim_str).get("altaPrioridadeAberta", [])
    except Exception:
        pass

    return consolidado


def _montar_prompt_resumo_semanal(dados):
    data_ini_fmt = datetime.strptime(dados["dataInicio"], "%Y-%m-%d").strftime("%d/%m")
    data_fim_fmt = datetime.strptime(dados["dataFim"], "%Y-%m-%d").strftime("%d/%m/%Y")

    def _fmt_lista(itens, campos, teto=60):
        if not itens:
            return "(nenhum)"
        linhas = [" | ".join(f"{c}={it.get(c, '')}" for c in campos) for it in itens[:teto]]
        if len(itens) > teto:
            linhas.append(f"... e mais {len(itens) - teto} itens")
        return "\n".join(linhas)

    bloco_mensagens = []
    for nome_grupo, msgs in dados.get("mensagensPorGrupo", {}).items():
        bloco_mensagens.append(f"\n--- Grupo: {nome_grupo} ({len(msgs)} mensagens na semana) ---")
        for m in msgs[:150]:
            bloco_mensagens.append(f"[{m.get('data_hora','')}] {m.get('remetente','?')}: {m.get('texto','')[:250]}")
    texto_mensagens = "\n".join(bloco_mensagens) if bloco_mensagens else "(nenhuma mensagem capturada nos grupos essa semana)"

    return f"""Aja como um Supervisor de O&M Sênior da Grid Co., escrevendo o resumo SEMANAL das usinas pro seu próprio controle — vai ser enviado só pra você mesmo, num grupo pessoal de gestão.

Semana de {data_ini_fmt} a {data_fim_fmt}.

REGRAS DE ESCRITA:
- Direto e objetivo, com uma visão consolidada da semana (não é só empilhar os resumos diários) — destaque padrões, usinas que mais concentraram problema, equipes que mais produziram, etc.
- Estruture em tópicos curtos com emojis moderados.
- NUNCA invente números, nomes ou fatos que não estão nos dados abaixo.
- No trecho de mensagens dos grupos, sintetize os TEMAS relevantes da semana inteira — não liste mensagem por mensagem.

DADOS DA SEMANA:

## Programação do PCM cumprida na semana
{_fmt_lista(dados['programacaoCumprida'], ['usina', 'cliente', 'tarefa', 'os'])}

## Programação do PCM pendente/não cumprida na semana
{_fmt_lista(dados['programacaoPendente'], ['usina', 'cliente', 'tarefa', 'os', 'statusReal'])}

## Atividades concluídas fora da programação na semana
{_fmt_lista(dados['extrasNaoProgramadas'], ['usina', 'cliente', 'descricao', 'numeroOS'])}

## Atividades com progresso na semana, mas ainda não concluídas
{_fmt_lista(dados['progressoDaSemana'], ['usina', 'cliente', 'descricao', 'numeroOS', 'percentualAtual', 'statusAtual'])}

## Chamados de fabricante abertos na semana
{_fmt_lista(dados['chamados'], ['UFV', 'Fabricante', 'Motivo da abertura do chamado', 'Status'])}

## Ocorrências novas na semana
{_fmt_lista(dados['ocorrencias'], ['usina', 'cliente', 'falha', 'status'])}

## OS de alta prioridade ainda em aberto (estado atual)
{_fmt_lista(dados['altaPrioridadeAberta'], ['usina', 'cliente', 'descricao', 'prazo'])}

## Mensagens nos grupos do WhatsApp mapeados, na semana
{texto_mensagens}

FORMATO DE SAÍDA (OBRIGATÓRIO): responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes ou depois), no formato:
{{"texto": "resumo semanal completo, pronto pra enviar no WhatsApp, começando com um cabeçalho tipo '📊 RESUMO SEMANAL — {data_ini_fmt} a {data_fim_fmt}'"}}"""


def _gerar_resumo_semanal_core(data_fim_str=None, enviar=True):
    if not data_fim_str:
        data_fim_str = agora_br().strftime("%Y-%m-%d")
    dados = _coletar_dados_resumo_semanal(data_fim_str)
    prompt = _montar_prompt_resumo_semanal(dados)
    resp = _chamar_gemini_com_retry(
        {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.3,
                "maxOutputTokens": 4096,
                "responseMimeType": "application/json",
                "thinkingConfig": {"thinkingBudget": 0},
            },
        },
        timeout=50,
    )
    resp_data = resp.json()
    texto_bruto = resp_data["candidates"][0]["content"]["parts"][0]["text"].strip()
    texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
    texto = json.loads(texto_limpo).get("texto", "").strip()
    if not texto:
        raise ValueError("A IA não retornou nenhum texto pro resumo semanal")

    resultado_envio = None
    if enviar:
        resultado_envio = _enviar_mensagem_grupo(GRUPO_GESTAO_OM_ID, texto)

    try:
        _salvar_resumo("semanal", texto, data_referencia=dados["dataFim"],
                        data_inicio=dados["dataInicio"], data_fim=dados["dataFim"], enviado=enviar)
    except Exception as e:
        log.error(f"[ResumoSemanal] Falha ao salvar no histórico do painel: {e}")

    return {"ok": True, "dataInicio": dados["dataInicio"], "dataFim": dados["dataFim"],
            "texto": texto, "envio": resultado_envio}


@app.route("/gerar-resumo-semanal", methods=["POST", "GET"])
def gerar_resumo_semanal():
    """Gera (e envia por padrão) o resumo semanal pro grupo Gestão O&M.
    Query params: ?data=YYYY-MM-DD (default hoje, usado como fim da
    semana), ?enviar=false (só gera e devolve o texto)."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    data_str = request.args.get("data", "").strip() or None
    enviar = request.args.get("enviar", "true").lower() != "false"
    try:
        resultado = _gerar_resumo_semanal_core(data_fim_str=data_str, enviar=enviar)
        return jsonify(resultado), 200
    except Exception as e:
        log.error(f"[ResumoSemanal] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Ronda matinal (pontos de atenção e prioridades do dia) ─────────────────

def _coletar_desligamentos_ativos_agora():
    """Desligamentos ATUALMENTE ativos, não só os de hoje — cruza Falhas
    (sem status de resolvida) e Atividades (ainda não concluídas) contra
    o padrão de texto de desligamento, com o mesmo override manual
    (_DesligamentoManual) usado no resumo diário. Feito pra Ronda: pedido
    do Fred pra incluir desligamentos crônicos/de dias anteriores, não só
    os que começaram hoje."""
    try:
        overrides = {}
        ws_desl = get_desligamento_manual_sheet()
        for row in ws_desl.get_all_values()[1:]:
            if len(row) >= 3 and row[0].strip():
                overrides[f"{row[0].strip()}:{row[1].strip()}"] = row[2].strip()

        padrao_desligamento = re.compile(
            r"(?:usina|ufv)\s+(?:\w+\s+){0,3}(?:desligad[ao]|parad[ao]|sem\s+energia|desenergizad[ao]|offline|sem\s+comunica[çc][ãa]o)"
            r"|(?:desligad[ao]|parad[ao]|offline)\s+(?:\w+\s+){0,3}(?:usina|ufv)"
            r"|desligamento\s+(?:total\s+)?(?:da|de)\s+(?:usina|ufv)",
            re.IGNORECASE,
        )
        ativos = []

        falhas = _falhas_itens()
        for f in falhas:
            status = (f.get("status") or "").strip().lower()
            if any(x in status for x in ("resolvid", "conclu", "fechad")):
                continue
            override = overrides.get(f"falha:{f.get('id')}")
            texto = f.get("falha") or ""
            if override == "sim" or (override != "nao" and padrao_desligamento.search(texto)):
                if canonizar_usina(f.get("usina")) is None:
                    continue
                ativos.append({"usina": f.get("usina"), "cliente": f.get("cliente"),
                                "descricao": texto, "origem": "Falha", "desde": f.get("dataAbertura")})

        ws_ativ = get_atividades_sheet()
        for row in ws_ativ.get_all_values()[1:]:
            if len(row) < ATIV_TOTAL_COLUNAS:
                row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
            status = row[ATIV_CAMPO_COL["status"] - 1].strip()
            if _is_concluido_atividade(status):
                continue
            numero_os = row[ATIV_CAMPO_COL["numeroOS"] - 1].strip()
            descricao = row[ATIV_CAMPO_COL["descricao"] - 1].strip()
            usina_bruta = row[ATIV_CAMPO_COL["usina"] - 1].strip()
            if canonizar_usina(usina_bruta) is None:
                continue
            override = overrides.get(f"atividade:{numero_os}")
            if override == "sim" or (override != "nao" and padrao_desligamento.search(descricao)):
                ativos.append({"usina": usina_bruta, "cliente": row[ATIV_CAMPO_COL["cliente"] - 1].strip(),
                                "descricao": descricao, "origem": f"Atividade OS {numero_os}",
                                "desde": row[ATIV_CAMPO_COL["dataCriacao"] - 1].strip()})
        return ativos
    except Exception as e:
        log.error(f"[Ronda] Erro ao checar desligamentos ativos: {e}")
        return []


def _coletar_prazos_proximos(data_str, dias_janela=1):
    """Atividades pendentes com prazo vencendo hoje ou nos próximos
    `dias_janela` dias (default: hoje e amanhã)."""
    try:
        hoje = datetime.strptime(data_str, "%Y-%m-%d").date()
        limite = hoje + timedelta(days=dias_janela)
        prazos = []
        ws_ativ = get_atividades_sheet()
        for row in ws_ativ.get_all_values()[1:]:
            if len(row) < ATIV_TOTAL_COLUNAS:
                row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
            status = row[ATIV_CAMPO_COL["status"] - 1].strip()
            if _is_concluido_atividade(status):
                continue
            prazo_raw = row[ATIV_CAMPO_COL["prazo"] - 1].strip()
            if not prazo_raw:
                continue
            try:
                prazo_dt = datetime.strptime(prazo_raw, "%d/%m/%Y").date()
            except ValueError:
                continue
            if not (hoje <= prazo_dt <= limite):
                continue
            usina_bruta = row[ATIV_CAMPO_COL["usina"] - 1].strip()
            if canonizar_usina(usina_bruta) is None:
                continue
            prazos.append({
                "usina": usina_bruta, "cliente": row[ATIV_CAMPO_COL["cliente"] - 1].strip(),
                "descricao": row[ATIV_CAMPO_COL["descricao"] - 1].strip(),
                "prazo": prazo_raw, "numeroOS": row[ATIV_CAMPO_COL["numeroOS"] - 1].strip(),
                "vencesHoje": prazo_dt == hoje,
            })
        return prazos
    except Exception as e:
        log.error(f"[Ronda] Erro ao checar prazos próximos: {e}")
        return []


def _coletar_chamados_fabricante_abertos():
    """Chamados de fabricante ainda sem status de resolvido — não só os
    abertos hoje (a Ronda é sobre o que ainda precisa de atenção agora,
    não um recorte do dia)."""
    try:
        chamados = _chamados_fabricante_itens()
        return [c for c in chamados
                if not any(x in (c.get("Status") or "").strip().lower()
                            for x in ("resolvid", "conclu", "fechad", "encerrad"))]
    except Exception as e:
        log.error(f"[Ronda] Erro ao ler chamados de fabricante: {e}")
        return []


def _coletar_dados_ronda(data_str):
    """Junta o que a Ronda precisa: programação do PCM pro dia (o que
    está previsto), OS de alta prioridade em aberto, desligamentos
    ativos agora (crônicos inclusive), prazos vencendo hoje/amanhã,
    chamados de fabricante ainda sem solução, e as mensagens dos grupos
    de ONTEM (a ronda roda de manhã, antes do expediente — o que rolou
    ontem à tarde/noite ainda não foi coberto por nenhum resumo)."""
    resultado = {"data": data_str}
    resultado["programacaoDoDia"] = _pcm_linhas_do_dia(data_str)

    altas_abertas = []
    try:
        ws_ativ = get_atividades_sheet()
        for row in ws_ativ.get_all_values()[1:]:
            if len(row) < ATIV_TOTAL_COLUNAS:
                row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
            usina_bruta = row[ATIV_CAMPO_COL["usina"] - 1].strip()
            if (row[ATIV_CAMPO_COL["prioridade"] - 1].strip().lower() == "alta"
                    and not _is_concluido_atividade(row[ATIV_CAMPO_COL["status"] - 1].strip())
                    and canonizar_usina(usina_bruta) is not None):
                altas_abertas.append({
                    "usina": usina_bruta, "cliente": row[ATIV_CAMPO_COL["cliente"] - 1].strip(),
                    "descricao": row[ATIV_CAMPO_COL["descricao"] - 1].strip(),
                    "prazo": row[ATIV_CAMPO_COL["prazo"] - 1].strip(),
                })
    except Exception as e:
        log.error(f"[Ronda] Erro ao ler Atividades (alta prioridade): {e}")
    resultado["altaPrioridadeAberta"] = altas_abertas

    resultado["desligamentosAtivos"] = _coletar_desligamentos_ativos_agora()
    resultado["prazosProximos"] = _coletar_prazos_proximos(data_str)
    resultado["chamadosAbertos"] = _coletar_chamados_fabricante_abertos()

    try:
        ontem_str = (datetime.strptime(data_str, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        mensagens = _buscar_mensagens_periodo(ontem_str, ontem_str)
        # Mesma correção de 18/08/2026 aplicada ao resumo diário: filtra
        # pra só entrar mensagem de grupo cuja usina ainda é reconhecida.
        grupos_validos = _grupos_ids_ativos()
        por_grupo = {}
        for m in mensagens:
            if m.get("grupo_id") not in grupos_validos:
                continue
            por_grupo.setdefault(m["nome_grupo"], []).append(m)
        resultado["mensagensPorGrupo"] = por_grupo
        resultado["dataMensagens"] = ontem_str
    except Exception as e:
        log.error(f"[Ronda] Erro ao ler mensagens capturadas: {e}")
        resultado["mensagensPorGrupo"] = {}
        resultado["dataMensagens"] = ""

    return resultado


def _montar_prompt_ronda(dados):
    data_str = dados["data"]
    try:
        data_fmt = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        data_fmt = data_str

    programacao = dados.get("programacaoDoDia", [])
    altas = dados.get("altaPrioridadeAberta", [])
    desligamentos = dados.get("desligamentosAtivos", [])
    prazos = dados.get("prazosProximos", [])
    chamados = dados.get("chamadosAbertos", [])
    mensagens_por_grupo = dados.get("mensagensPorGrupo", {})
    data_mensagens = dados.get("dataMensagens", "")
    try:
        data_mensagens_fmt = datetime.strptime(data_mensagens, "%Y-%m-%d").strftime("%d/%m/%Y") if data_mensagens else ""
    except ValueError:
        data_mensagens_fmt = data_mensagens

    def _fmt_lista(itens, campos):
        if not itens:
            return "(nenhum)"
        linhas = []
        for it in itens:
            linhas.append(" | ".join(f"{c}={it.get(c, '')}" for c in campos))
        return "\n".join(linhas)

    # Mesma lógica de cruzamento OS↔mensagens do resumo diário (pedido do
    # Fred em 28/07/2026: bater OS com as conversas nos grupos) — aqui
    # aplicada às mensagens de ONTEM, pra achar pontos que ficaram em
    # aberto/sem resposta e precisam entrar na ronda de hoje.
    bloco_mensagens = []
    for nome_grupo, msgs in mensagens_por_grupo.items():
        bloco_mensagens.append(f"\n--- Grupo: {nome_grupo} ({len(msgs)} mensagens) ---")
        for m in msgs[:80]:
            hora = (m.get("data_hora") or "").split(" ")[-1][:5]
            texto_msg = m.get("texto", "")
            bloco_mensagens.append(f"[{hora}] {m.get('remetente','?')}: {texto_msg[:300]}")
    texto_mensagens = "\n".join(bloco_mensagens) if bloco_mensagens else "(nenhuma mensagem capturada nos grupos ontem)"

    # Corrigido 18/08/2026 (mesmo motivo do resumo diário): inclui usina
    # sob Supervisão Temporária ainda ATIVA na lista de escopo.
    usinas_temp_ativas = {item["usina"] for item in _usinas_temporarias()}
    lista_usinas_fred = ", ".join(sorted(set(CATALOGO_USINAS.keys()) | usinas_temp_ativas))

    return f"""Aja como um Supervisor de O&M Sênior da Grid Co., escrevendo a "Ronda" do dia — uma mensagem curta de PRIORIDADES E PONTOS DE ATENÇÃO pro seu próprio controle, enviada de manhã, ANTES do expediente começar. É pra você se organizar no início do dia, não é um resumo do que já aconteceu (isso é o resumo diário, que roda às 17h) — foque no que precisa de atenção HOJE.

Data da ronda: {data_fmt}

USINAS SOB SUA SUPERVISÃO (lista fechada — só estas): {lista_usinas_fred}

REGRAS DE ESCRITA:
- Direto e objetivo, sem enrolação, sem saudação.
- Estruture em tópicos curtos com emojis moderados pra facilitar leitura rápida no celular.
- NUNCA invente números, nomes ou fatos que não estão nos dados abaixo.
- Se uma seção não tiver nada a reportar, diga isso em uma linha curta, não pule a seção.
- Priorize: desligamentos ativos e prazos vencendo HOJE vêm primeiro (são os pontos mais urgentes), depois pontos críticos vindos das conversas dos grupos (ver abaixo), depois programação do dia e alta prioridade em aberto.
- Critério de PRIORIZAÇÃO por criticidade quando houver mais de um ponto concorrendo por destaque: 1º o que afeta EFICIÊNCIA/GERAÇÃO da usina (equipamento parado, string fora, etc.), 2º o que afeta SEGURANÇA (proteção, EPC, estrutura), 3º o resto.
- Feche com uma linha curta de "prioridade do dia" — se tiver desligamento ativo, prazo vencendo hoje, ou ponto crítico das conversas, essa é a prioridade; senão, aponte a atividade mais importante da programação do dia.

## Conversas dos grupos de ontem ({data_mensagens_fmt or 'sem data'})
Leia as mensagens abaixo e identifique PONTOS CRÍTICOS que ficaram em aberto/sem resposta clara de resolução — problema técnico reportado por um técnico de campo sem confirmação de solução, reclamação de cliente sem retorno, pedido de peça/autorização parado, ou qualquer coisa que pareça urgente e não tenha sido claramente encerrada na própria conversa. Adicione esses pontos numa seção própria "🗣️ PONTOS DAS CONVERSAS" na ronda, citando usina/grupo e um resumo curto do que foi dito — NUNCA invente ou complete informação que não está explícita na mensagem; se não achar nada relevante, essa seção pode dizer "Nada crítico identificado nas conversas de ontem."
{texto_mensagens}

DADOS PRA RONDA DE HOJE:

## Desligamentos ativos agora (inclui casos de dias anteriores, não só de hoje)
{_fmt_lista(desligamentos, ['usina', 'cliente', 'descricao', 'desde'])}

## Prazos vencendo hoje ou amanhã
{_fmt_lista(prazos, ['usina', 'cliente', 'descricao', 'numeroOS', 'prazo', 'vencesHoje'])}

## Programação do PCM pra hoje
{_fmt_lista(programacao, ['usina', 'cliente', 'tarefa', 'os_id', 'h_ini'])}

## OS de alta prioridade em aberto (não necessariamente de hoje)
{_fmt_lista(altas, ['usina', 'cliente', 'descricao', 'prazo'])}

## Chamados de fabricante ainda sem solução
{_fmt_lista(chamados, ['UFV', 'Fabricante', 'Motivo da abertura do chamado', 'Status'])}

FORMATO DE SAÍDA (OBRIGATÓRIO): responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes ou depois), no formato:
{{"texto": "ronda completa, pronta pra enviar no WhatsApp, começando com um cabeçalho tipo '🔔 RONDA — {data_fmt}', terminando com a linha de 'Prioridade do dia' descrita acima"}}"""


def _gerar_ronda_core(data_str=None, enviar=True):
    if not data_str:
        data_str = agora_br().strftime("%Y-%m-%d")
    dados = _coletar_dados_ronda(data_str)
    prompt = _montar_prompt_ronda(dados)
    resp = _chamar_gemini_com_retry(
        {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.3,
                "maxOutputTokens": 2048,
                "responseMimeType": "application/json",
                "thinkingConfig": {"thinkingBudget": 0},
            },
        },
        timeout=45,
    )
    resp_data = resp.json()
    texto_bruto = resp_data["candidates"][0]["content"]["parts"][0]["text"].strip()
    texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
    texto = json.loads(texto_limpo).get("texto", "").strip()
    if not texto:
        raise ValueError("A IA não retornou nenhum texto pra ronda")

    resultado_envio = None
    if enviar:
        resultado_envio = _enviar_mensagem_grupo(GRUPO_GESTAO_OM_ID, texto)

    try:
        _salvar_resumo("ronda", texto, data_referencia=data_str, enviado=enviar)
    except Exception as e:
        log.error(f"[Ronda] Falha ao salvar no histórico do painel: {e}")

    return {"ok": True, "data": data_str, "texto": texto, "envio": resultado_envio}


def _verificar_e_disparar_ronda_se_necessario():
    """Piggyback no /sync-fracttal, mesmo padrão do resumo diário/semanal
    (ver docstring de _verificar_e_disparar_resumo_diario_se_necessario):
    gera e envia a ronda todo dia na janela 08:00-08:30 (Brasília). A
    trava só é gravada DEPOIS da geração ter sucesso, pelo mesmo motivo
    documentado lá (worker morto no meio do caminho não pode travar o
    dia inteiro sem ronda)."""
    try:
        agora = agora_br()
        hoje_str = agora.strftime("%Y-%m-%d")
        if not (agora.hour == 8 and agora.minute <= 30):
            return {"disparado": False, "motivo": f"fora da janela (agora {agora.strftime('%H:%M')})"}
        ja_feito = _ler_trava("ronda_enviada_em")
        if ja_feito == hoje_str:
            return {"disparado": False, "motivo": "já enviada hoje"}
        resultado = _gerar_ronda_core(data_str=hoje_str, enviar=True)
        _gravar_trava("ronda_enviada_em", hoje_str)
        return {"disparado": True, "resultado": resultado}
    except Exception as e:
        log.error(f"[Ronda] Erro no piggyback: {e}")
        return {"disparado": False, "erro": str(e)}


@app.route("/gerar-ronda", methods=["POST", "GET"])
def gerar_ronda():
    """Gera (e envia por padrão) a ronda matinal pro grupo Gestão O&M.
    Query params: ?data=YYYY-MM-DD (default hoje), ?enviar=false (só
    gera e devolve o texto, sem mandar pro WhatsApp — útil pra testar)."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    data_str = request.args.get("data", "").strip() or None
    enviar = request.args.get("enviar", "true").lower() != "false"
    try:
        resultado = _gerar_ronda_core(data_str=data_str, enviar=enviar)
        return jsonify(resultado), 200
    except Exception as e:
        log.error(f"[Ronda] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/verificar-uma-os", methods=["POST", "OPTIONS"])
def verificar_uma_os():
    """
    Endpoint PÚBLICO (sem secret) pra forçar a checagem AO VIVO de uma
    única OS específica, furando a fila de prioridade do rodízio — útil
    pra quando o técnico acabou de concluir algo e não dá pra esperar a
    vez dela na fila (que pode ter dezenas de outras OSs mais "antigas"
    na frente, mesmo que essa seja a mais importante agora).
    """
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    id_atividade = str(body.get("id") or "").strip()
    numero_os = str(body.get("numeroOS") or "").strip()
    if not id_atividade and not numero_os:
        return jsonify({"ok": False, "error": "informe id ou numeroOS"}), 400

    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    for i, row in enumerate(todos[1:], start=2):
        if len(row) < ATIV_TOTAL_COLUNAS:
            row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
        if (id_atividade and row[0].strip() == id_atividade) or (numero_os and row[13].strip() == numero_os):
            resultado = _fracttal_verificar_e_atualizar_uma_os(ws, i, row, row[13].strip())
            if resultado is None:
                return jsonify({"ok": False, "error": "falha ao consultar a Fracttal — ver logs"}), 502
            return jsonify({"ok": True, **resultado}), 200

    return jsonify({"ok": False, "error": "atividade não encontrada"}), 404


@app.route("/atualizar-os-agora", methods=["POST", "OPTIONS"])
def atualizar_os_agora():
    """
    Endpoint PÚBLICO (sem secret) pro botão "Atualizar OS" do dashboard —
    faz uma varredura de status/estado nas OSs que JÁ estão no dashboard
    (revalida ao vivo na Fracttal, corrige status interno se precisar).
    NÃO busca OS nova — isso é o botão "Auditoria", separado.

    limite_atraso_minutos=0: um clique manual é um pedido explícito de
    dado fresco AGORA — não faz sentido aplicar o filtro de "só recheca
    se já passou de 45min" (que existe pra poupar chamadas no ciclo
    automático). Sem esse filtro, cada clique sempre processa as 35 OSs
    genuinamente mais antigas da fila, garantindo que repetir o clique
    avança de verdade pela fila inteira (bug identificado em 13/07/2026:
    uma OS checada há pouco tempo — ex.: técnico concluiu logo depois da
    última verificação — ficava presa fora da lista de elegíveis pra
    sempre, não importava quantos cliques).
    """
    if request.method == "OPTIONS":
        return ("", 204)
    resultado = _auditoria_consistencia_os_core(aplicar=True, limite_atraso_minutos=0, origem="manual (botão Atualizar OS)")
    return jsonify({"ok": True, **resultado}), 200


@app.route("/rodar-auditoria-agora", methods=["POST", "OPTIONS"])
def rodar_auditoria_agora():
    """
    Endpoint PÚBLICO (sem secret) pro botão "Auditoria" do dashboard —
    varredura COMPLETA nas usinas/equipes: busca OS nova na Fracttal
    (descoberta) e revalida ao vivo um lote amplo das já existentes,
    detectando não só mudança de percentual mas também cancelamentos e
    conclusões que possam ter passado batido. Mais pesada de propósito —
    o "Atualizar OS" (mais rápido) cuida da atualização frequente.
    """
    if request.method == "OPTIONS":
        return ("", 204)
    resultado = _auditoria_completa_core(origem="manual (botão Auditoria)")
    return jsonify({"ok": True, **resultado}), 200


@app.route("/revalidar-usinas", methods=["POST", "OPTIONS"])
def revalidar_usinas():
    """
    Revalidação ao vivo na Fracttal, filtrada por uma lista de usinas —
    diferente do "Atualizar OS" (que processa a fila inteira por ordem
    de desatualização, sem filtro). Útil quando um cluster específico
    está divergente do real e não se quer esperar o rodízio geral
    alcançar essas usinas (criado 07/08/2026, pedido do Fred pro
    cluster SP Leste 03).

    Body JSON: {"usinas": ["Usina A", "Usina B", ...]}
    Só processa OSs vinculadas à Fracttal (numeroOS preenchido) e ainda
    não finalizadas/canceladas — reaproveita a mesma função de
    revalidação usada pela auditoria automática, então o comportamento
    (o que é considerado "mudou", como o status interno é corrigido
    etc.) é idêntico ao rodízio normal.
    """
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
        usinas_filtro = set(u.strip() for u in body.get("usinas", []) if u.strip())
        if not usinas_filtro:
            return jsonify({"ok": False, "error": "campo 'usinas' é obrigatório (lista não vazia)"}), 400

        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        alvo = []
        for i, row in enumerate(todos[1:], start=2):
            if len(row) < ATIV_TOTAL_COLUNAS:
                row = row + [""] * (ATIV_TOTAL_COLUNAS - len(row))
            usina_row = row[2].strip()
            numero_os = row[13].strip()
            status_os_atual = row[14].strip()
            if usina_row not in usinas_filtro:
                continue
            if not numero_os:
                continue  # não vinculada à Fracttal, nada a revalidar
            if status_os_atual in ("Finalizada", "Cancelada"):
                continue
            alvo.append({"linha": i, "row": row, "numeroOS": numero_os})

        revalidadas = []
        erros = []
        ORCAMENTO_SEGUNDOS = 90
        inicio = time.time()
        parou_por_orcamento = False
        for a in alvo:
            if time.time() - inicio > ORCAMENTO_SEGUNDOS:
                parou_por_orcamento = True
                break
            try:
                resultado = _fracttal_verificar_e_atualizar_uma_os(ws, a["linha"], a["row"], a["numeroOS"],
                                                                     enviar_notificacao=False)
                if resultado:
                    revalidadas.append(resultado)
            except Exception as e:
                erros.append({"numeroOS": a["numeroOS"], "erro": str(e)})
            time.sleep(0.35)

        mudaram = [r for r in revalidadas if r.get("mudou")]
        if mudaram:
            try:
                def _linha_resumo_cluster(r):
                    usina = (r.get("usina") or "Usina não informada").strip()
                    tema = (r.get("descricao") or r.get("equipamento") or "sem descrição").strip()
                    if len(tema) > 35:
                        tema = tema[:35].rstrip() + "…"
                    mudanca = r.get("mudancaResumo") or r.get("statusGeralOS") or ""
                    base = f"{r['numeroOS']} · {usina} — {tema}"
                    return f"{base} ({mudanca})" if mudanca else base
                linhas = "\n".join(_linha_resumo_cluster(r) for r in mudaram[:8])
                enviar_push(
                    titulo=f"🔄 Revalidação manual — {len(mudaram)} OS(s) atualizadas",
                    corpo=f"{linhas}{chr(10) + '...' if len(mudaram) > 8 else ''}",
                    tipo="fracttal_status",
                )
            except Exception as e:
                log.error(f"[RevalidarUsinas] Falha ao enviar push resumido: {e}")

        return jsonify({
            "ok": True,
            "usinasFiltradas": sorted(usinas_filtro),
            "totalElegiveis": len(alvo),
            "totalRevalidadas": len(revalidadas),
            "totalMudaram": len(mudaram),
            "mudaram": mudaram,
            "erros": erros,
            "parouPorOrcamento": parou_por_orcamento,
        }), 200
    except Exception as e:
        log.error(f"[RevalidarUsinas] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/validar-integridade-relatorios", methods=["POST", "GET"])
def validar_integridade_relatorios():
    """
    Roda a validação de integridade dos relatórios (Painel de Falhas +
    Painel de Atividades, TODOS os clientes) sob demanda — mesma lógica
    que já roda automaticamente 3x/dia dentro da auditoria completa.
    Útil pra rodar manualmente logo antes de gerar um relatório, com
    confiança de que os dados estão íntegros.
    """
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    aplicar = request.args.get("apply", "true").lower() != "false"
    resultado = _validar_integridade_relatorios_core(aplicar=aplicar)
    return jsonify({"ok": True, **resultado}), 200


ATIV_CAMPO_LABEL = {
    "cliente": "Cliente", "usina": "Usina", "equipamento": "Equipamento", "descricao": "Descrição",
    "responsavel": "Responsável", "prazo": "Prazo", "prioridade": "Prioridade", "status": "Status",
    "numeroOS": "Nº OS",
}
ATIV_COL_HISTORICO = ATIV_CAMPO_COL["historico"]


@app.route("/atualizar-campo-atividade", methods=["POST", "OPTIONS"])
def atualizar_campo_atividade():
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido"}), 400

    atividade_id = str(body.get("id", "")).strip()
    field  = body.get("field", "").strip()
    value  = body.get("value", "")
    append = bool(body.get("append", False))
    editor = body.get("editor", "dashboard").strip()

    if not atividade_id or field not in ATIV_CAMPO_COL:
        return jsonify({"ok": False, "error": "id ou campo inválido"}), 400

    try:
        ws = get_atividades_sheet()
        _garantir_headers_atividades(ws)
        todos = _gspread_retry(lambda: ws.get_all_values())
        linha_idx = None
        linha_atual = None
        for i, row in enumerate(todos[1:], start=2):
            if row and str(row[0]).strip() == atividade_id:
                linha_idx = i
                linha_atual = row
                break
        if not linha_idx:
            return jsonify({"ok": False, "error": "atividade não encontrada"}), 404

        col = ATIV_CAMPO_COL[field]

        if field == "historico" and append:
            atual = linha_atual[ATIV_COL_HISTORICO - 1] if len(linha_atual) >= ATIV_COL_HISTORICO else ""
            novo = f"{atual}\n{value}".strip() if atual else value
            ws.update_cell(linha_idx, col, novo)
        else:
            valor_antigo = linha_atual[col - 1] if len(linha_atual) >= col else ""
            ws.update_cell(linha_idx, col, value)

            # Registra automaticamente a alteração no histórico cronológico
            if str(valor_antigo).strip() != str(value).strip():
                if field == "visualizado":
                    # Mensagem dedicada (17/07/2026): o formato genérico
                    # "visualizado alterado de '—' para 'sim'" confundia o
                    # Fred, parecendo uma edição de dado real em vez do que
                    # realmente é — só o rastreio de "já vi essa atividade"
                    # usado pro badge de não-lido.
                    entry = f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Marcado como visualizado ({_editor_legivel(editor)})."
                else:
                    label = ATIV_CAMPO_LABEL.get(field, field)
                    entry = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - {label} alterado "
                             f"de \"{valor_antigo or '—'}\" para \"{value}\" por {_editor_legivel(editor)}.")
                hist_atual = linha_atual[ATIV_COL_HISTORICO - 1] if len(linha_atual) >= ATIV_COL_HISTORICO else ""
                novo_hist = f"{hist_atual}\n{entry}".strip() if hist_atual else entry
                ws.update_cell(linha_idx, ATIV_COL_HISTORICO, novo_hist)

            if field == "status" and _is_concluido_atividade(value):
                ws.update_cell(linha_idx, 11, agora_br().strftime('%d/%m/%Y %H:%M:%S'))  # DataConclusao

        return jsonify({"ok": True})
    except Exception as e:
        log.error(f"[Atividades] Erro ao atualizar campo: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


RE_ATUALIZACAO_ATIV = re.compile(r"ATUALIZA[CÇ][AÃ]O\s+(?:DE\s+)?(OS|ATIVIDADE)", re.IGNORECASE)

def separar_blocos_atividade(texto):
    """
    Divide uma mensagem que contenha múltiplas atualizações de OS/atividade
    em blocos individuais, um por ocorrência do título "ATUALIZACAO OS/ATIVIDADE".
    Mesmo padrão de separar_blocos() usado nas ocorrências.
    """
    partes = re.split(
        r"(?=(?:^|\n)\s*ATUALIZA[CÇ][AÃ]O\s+(?:DE\s+)?(?:OS|ATIVIDADE))",
        texto, flags=re.MULTILINE | re.IGNORECASE
    )
    blocos = [p.strip() for p in partes if p.strip()]
    return blocos if blocos else [texto]


def eh_atualizacao_atividade(texto):
    return bool(RE_ATUALIZACAO_ATIV.search(texto))


def _extrair_campo_ativ(texto, nome_regex):
    # Bullet (·, *, -, •) é opcional; separador aceita ":" ou "-"/"–"; âncora por linha
    # evita capturar texto de outros campos.
    padrao = rf"^\s*[·*\-•]?\s*(?:{nome_regex})\s*[:\-–]\s*(.+)$"
    m = re.search(padrao, texto, re.IGNORECASE | re.MULTILINE)
    return m.group(1).strip() if m else ""


_STATUS_ATIV_MAP = {
    "concluido": "Concluído", "concluído": "Concluído", "concluida": "Concluído", "concluída": "Concluído",
    "finalizado": "Concluído", "finalizada": "Concluído", "resolvido": "Concluído", "resolvida": "Concluído",
    "feito": "Concluído", "ok": "Concluído",
    "em andamento": "Em Andamento", "andamento": "Em Andamento", "em execucao": "Em Andamento",
    "em execução": "Em Andamento", "executando": "Em Andamento",
    "aguardando": "Aguardando", "pendente": "Aguardando",
    "aguardando peca": "Aguardando", "aguardando peça": "Aguardando",
}

_OS_FIELD_REGEX = r"(?:N[ºo°]?\s*|N[uú]mero\s*(?:da\s*)?)?OS|Ordem\s*(?:de\s*)?Servi[cç]o"
_DESCRICAO_FIELD_REGEX = r"Descri[cç][aã]o|Obs(?:erva[cç][aã]o)?|A[cç][aã]o(?:\s+Realizada)?|Servi[cç]o\s+Realizado"
_RESPONSAVEL_FIELD_REGEX = r"Respons[aá]vel|T[eé]cnico"
_STATUS_FIELD_REGEX = r"Status|Situa[cç][aã]o"


def parse_atualizacao_atividade(texto):
    id_val = _extrair_campo_ativ(texto, "ID")
    os_val = _extrair_campo_ativ(texto, _OS_FIELD_REGEX)
    status_bruto = _extrair_campo_ativ(texto, _STATUS_FIELD_REGEX)
    status_norm = _STATUS_ATIV_MAP.get(status_bruto.strip().lower(), status_bruto.strip()) if status_bruto else ""
    return {
        "id_ou_os":    os_val or id_val,
        "status":      status_norm,
        "descricao":   _extrair_campo_ativ(texto, _DESCRICAO_FIELD_REGEX),
        "responsavel": _extrair_campo_ativ(texto, _RESPONSAVEL_FIELD_REGEX),
    }


def buscar_atividade_por_id_ou_os(todos, id_ou_os):
    alvo = str(id_ou_os).strip().lstrip("0") or "0"
    for i, row in enumerate(todos[1:], start=2):
        if not row or not row[0].strip():
            continue
        row_id = str(row[0]).strip().lstrip("0") or "0"
        row_os = str(row[13]).strip().lstrip("0") if len(row) > 13 else ""
        row_os = row_os or "0"
        if (alvo != "0" and alvo == row_os) or alvo == row_id:
            return i, row
    return None


def _aplicar_update_campo_atividade(ws, linha_idx, linha_atual, field, value, editor, append=False):
    col = ATIV_CAMPO_COL[field]
    if field == "historico" and append:
        atual = linha_atual[ATIV_COL_HISTORICO - 1] if len(linha_atual) >= ATIV_COL_HISTORICO else ""
        novo = f"{atual}\n{value}".strip() if atual else value
        ws.update_cell(linha_idx, col, novo)
        return
    valor_antigo = linha_atual[col - 1] if len(linha_atual) >= col else ""
    ws.update_cell(linha_idx, col, value)
    if str(valor_antigo).strip() != str(value).strip():
        if field == "visualizado":
            entry = f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Marcado como visualizado ({_editor_legivel(editor)})."
        else:
            label = ATIV_CAMPO_LABEL.get(field, field)
            entry = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - {label} alterado "
                     f"de \"{valor_antigo or '—'}\" para \"{value}\" por {_editor_legivel(editor)}.")
        hist_atual = linha_atual[ATIV_COL_HISTORICO - 1] if len(linha_atual) >= ATIV_COL_HISTORICO else ""
        novo_hist = f"{hist_atual}\n{entry}".strip() if hist_atual else entry
        ws.update_cell(linha_idx, ATIV_COL_HISTORICO, novo_hist)
    if field == "status" and _is_concluido_atividade(value):
        ws.update_cell(linha_idx, 11, agora_br().strftime('%d/%m/%Y %H:%M:%S'))


def _processar_um_bloco_atividade(texto, editor="tecnico-whatsapp"):
    dados = parse_atualizacao_atividade(texto)
    if not dados["id_ou_os"]:
        try:
            enviar_push(
                titulo="⚠️ Atualização de OS sem Nº OS/ID",
                corpo=f"Mensagem recebida de {editor}, mas não foi possível identificar o campo Nº OS ou ID. Confira o formato da mensagem.",
                tipo="geral",
            )
        except Exception as e:
            log.error(f"[Atividades WhatsApp] Falha ao enviar push de erro: {e}")
        return {"ok": False, "motivo": "sem ID ou Nº OS na mensagem"}

    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    encontrada = buscar_atividade_por_id_ou_os(todos, dados["id_ou_os"])
    if not encontrada:
        try:
            enviar_push(
                titulo="⚠️ Atualização de OS não vinculada",
                corpo=f"Técnico ({editor}) informou Nº OS/ID \"{dados['id_ou_os']}\" mas nenhuma atividade correspondente foi encontrada no painel.",
                tipo="geral",
            )
        except Exception as e:
            log.error(f"[Atividades WhatsApp] Falha ao enviar push de erro: {e}")
        return {"ok": False, "motivo": f"atividade {dados['id_ou_os']} não encontrada"}

    linha_idx, linha_atual = encontrada

    if dados["responsavel"]:
        _aplicar_update_campo_atividade(ws, linha_idx, linha_atual, "responsavel", dados["responsavel"], editor)
        todos = ws.get_all_values(); linha_atual = todos[linha_idx - 1]

    if dados["descricao"]:
        entry = f"{agora_br().strftime('%d/%m/%Y %H:%M')} - {_editor_legivel(editor)}: {dados['descricao']}"
        _aplicar_update_campo_atividade(ws, linha_idx, linha_atual, "historico", entry, editor, append=True)
        todos = ws.get_all_values(); linha_atual = todos[linha_idx - 1]

    if dados["status"]:
        numero_os_linha = linha_atual[13].strip() if len(linha_atual) > 13 else ""
        if numero_os_linha:
            # OS vinculada à Fracttal: o status NUNCA é escrito a partir do
            # texto da mensagem — só a Fracttal (via API e as automações
            # já existentes: rodízio, auditoria, descoberta) decide o
            # status real. A mensagem do técnico só serve de GATILHO pra
            # checar a Fracttal imediatamente, sem esperar a próxima
            # rodada de auditoria. Antes disso, o texto do WhatsApp
            # escrevia o status direto, sem nenhuma validação — foi
            # exatamente isso que causou a OS 8867 aparecer como
            # "Concluída" no painel enquanto a Fracttal ainda mostrava
            # "Em Processo" (relatado pelo Fred em 15/07/2026, que pediu
            # essa mudança de arquitetura em vez de só reconciliar depois).
            entry = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - {_editor_legivel(editor)} reportou status "
                     f"\"{dados['status']}\" pelo WhatsApp (confirmado em seguida direto com a Fracttal).")
            _aplicar_update_campo_atividade(ws, linha_idx, linha_atual, "historico", entry, editor, append=True)
            try:
                todos_frescos = ws.get_all_values()
                linha_fresca = todos_frescos[linha_idx - 1]
                _fracttal_verificar_e_atualizar_uma_os(ws, linha_idx, linha_fresca, numero_os_linha,
                                                        enviar_notificacao=False)
            except Exception as e:
                log.error(f"[Atividades WhatsApp] Falha ao checar a Fracttal pra OS {numero_os_linha}: {e}")
        else:
            # atividade manual, sem vínculo com nenhuma OS da Fracttal —
            # não existe outra fonte de verdade pra ela, então o status
            # informado pelo técnico continua sendo aceito diretamente.
            _aplicar_update_campo_atividade(ws, linha_idx, linha_atual, "status", dados["status"], editor)

    return {"ok": True, "id": linha_atual[0]}


def processar_atualizacao_atividade(texto, editor="tecnico-whatsapp"):
    blocos = separar_blocos_atividade(texto)
    resultados = [_processar_um_bloco_atividade(bloco, editor) for bloco in blocos]
    return {
        "ok": any(r.get("ok") for r in resultados),
        "total_blocos": len(resultados),
        "resultados": resultados,
    }


@app.route("/corrigir-prioridade-atividades", methods=["GET"])
def corrigir_prioridade_atividades():
    """
    Rota de manutenção pontual: corrige linhas antigas do Painel de Atividades
    cuja celula de Prioridade foi sobrescrita com um valor de Status (bug do
    mapeamento de colunas anterior a correcao). So mexe em linhas onde o valor
    atual de Prioridade nao e Alta/Media/Baixa - ou seja, so nas corrompidas.
    """
    VALORES_VALIDOS_PRIORIDADE = {"alta", "media", "média", "baixa"}
    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
        corrigidas = []
        for i, row in enumerate(todos[1:], start=2):
            if not row or not row[0].strip():
                continue
            prioridade_atual = row[7].strip() if len(row) > 7 else ""
            if prioridade_atual.lower() not in VALORES_VALIDOS_PRIORIDADE:
                ws.update_cell(i, 8, "Alta")  # coluna H = Prioridade
                hist_atual = row[11] if len(row) > 11 else ""
                entry = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Prioridade corrigida de "
                         f"\"{prioridade_atual or '—'}\" para \"Alta\" (correcao de dado legado) por sistema.")
                novo_hist = f"{hist_atual}\n{entry}".strip() if hist_atual else entry
                ws.update_cell(i, 12, novo_hist)  # coluna L = Historico
                corrigidas.append({"linha": i, "id": row[0], "de": prioridade_atual, "para": "Alta"})
        return jsonify({"ok": True, "corrigidas": corrigidas, "total": len(corrigidas)})
    except Exception as e:
        log.error(f"[corrigir-prioridade-atividades] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/converter-atividade-em-ocorrencia", methods=["POST", "OPTIONS"])
def converter_atividade_em_ocorrencia():
    """
    Converte uma Atividade em uma Ocorrência: cria uma nova linha no Painel de
    Falhas com os dados da atividade (incluindo o histórico cronológico
    transferido), e marca a atividade original como "Convertida em Ocorrência".
    """
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido"}), 400

    atividade_id = str(body.get("id", "")).strip()
    editor = body.get("editor", "dashboard").strip()
    if not atividade_id:
        return jsonify({"ok": False, "error": "id é obrigatório"}), 400

    try:
        ws_ativ = get_atividades_sheet()
        todos_ativ = ws_ativ.get_all_values()
        linha_idx = None
        linha_atual = None
        for i, row in enumerate(todos_ativ[1:], start=2):
            if row and str(row[0]).strip() == atividade_id:
                linha_idx = i
                linha_atual = row
                break
        if not linha_idx:
            return jsonify({"ok": False, "error": "atividade não encontrada"}), 404

        # linha_atual: [ID, Cliente, Usina, Equipamento, Descricao, Responsavel, Prazo,
        #               Prioridade, Status, DataCriacao, DataConclusao, Historico, Editor]
        cliente     = linha_atual[1] if len(linha_atual) > 1 else ""
        usina       = linha_atual[2] if len(linha_atual) > 2 else ""
        equipamento = linha_atual[3] if len(linha_atual) > 3 else ""
        descricao   = linha_atual[4] if len(linha_atual) > 4 else ""
        responsavel = linha_atual[5] if len(linha_atual) > 5 else ""
        prazo       = linha_atual[6] if len(linha_atual) > 6 else ""
        status_ativ = linha_atual[8] if len(linha_atual) > 8 else ""
        historico_ativ = linha_atual[11] if len(linha_atual) > 11 else ""
        numero_os_ativ = linha_atual[13] if len(linha_atual) > 13 else ""

        if not equipamento:
            equipamento = "Não informado"

        nota_conversao = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Convertida do Painel de "
                           f"Atividades (Atividade #{atividade_id}) por {_editor_legivel(editor)}.")
        historico_ocorrencia = nota_conversao
        if historico_ativ:
            historico_ocorrencia += "\n" + historico_ativ

        status_ocorrencia = status_ativ if status_ativ and status_ativ.lower() not in (
            "concluído", "concluido", "convertida em ocorrência") else "Em Aberto"

        ws_falhas = get_sheet()
        todos_falhas = carregar_planilha(ws_falhas)
        novo_id_ocorrencia = proximo_id(todos_falhas)

        dados = {
            "cliente":      cliente,
            "usina":        usina,
            "equipamento":  equipamento,
            "falha":        descricao,
            "causa":        "",
            "equip_impact": equipamento,
            "acao":         f"Responsável original: {responsavel}." if responsavel else "",
            "status":       status_ocorrencia,
            "os":           numero_os_ativ,
            "historico":    historico_ocorrencia,
        }
        gravar_nova_ocorrencia(ws_falhas, todos_falhas, dados)

        # Marca a atividade original como convertida e registra no histórico dela
        ws_ativ.update_cell(linha_idx, 9, "Convertida em Ocorrência")  # coluna I = Status
        entry = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Convertida em ocorrência "
                 f"#{novo_id_ocorrencia} por {_editor_legivel(editor)}.")
        novo_hist_ativ = f"{historico_ativ}\n{entry}".strip() if historico_ativ else entry
        ws_ativ.update_cell(linha_idx, 12, novo_hist_ativ)  # coluna L = Historico

        log.info(f"[converter-atividade] Atividade #{atividade_id} -> Ocorrência #{novo_id_ocorrencia}")
        return jsonify({"ok": True, "novaOcorrenciaId": novo_id_ocorrencia})
    except Exception as e:
        log.error(f"[converter-atividade-em-ocorrencia] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Geração de texto de OS via Gemini (gratuito), com fallback local ───────
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
# Chave separada (projeto próprio no Google AI Studio) usada só em
# diagnósticos/testes ao vivo, pra nunca disputar cota com o uso real do
# Fred. Só entra em ação quando explicitamente pedido (?diagnostico=true
# ou header X-Usar-Chave-Teste), nunca no fluxo normal do dashboard.
GEMINI_API_KEY_TESTE = os.environ.get("GEMINI_API_KEY_TESTE", "")
GEMINI_MODEL = "gemini-3.1-flash-lite"
# Migrado de gemini-2.5-flash em 23/07/2026 (2ª tentativa, dessa vez
# testada e confirmada antes de aplicar): o Gemini 3.5/3.6 Flash exigem
# OAuth2 via a Interactions API nova do Google — incompatíveis com nossa
# autenticação simples por chave, então ficam fora de cogitação por
# enquanto. Já o Gemini 3.1 Flash Lite FUNCIONA normalmente com "?key=",
# confirmado via /diag-testar-modelo-gemini, e tem cota gratuita muito
# mais generosa (500 RPD / 15 RPM / 250K TPM, conferido pelo Fred direto
# no AI Studio) do que o 2.5 Flash — e não está na lista de aposentadoria
# de 16/10/2026 (só 2.5 Flash e 2.5 Flash-Lite estão).
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"


def _montar_prompt_os(d):
    return f"""Aja como um Engenheiro e Especialista em Operação e Manutenção (O&M), com foco em Usinas Solares Fotovoltaicas, sistemas elétricos, mecânicos e atividades de facilities (limpeza, conservação, manutenções civis).

Sua tarefa é redigir Ordens de Serviço (OS) baseadas na solicitação abaixo. Transforme a solicitação em um texto objetivo, profissional, técnico e estritamente padronizado.

REGRA DE SEPARAÇÃO EM MÚLTIPLAS OS (MUITO IMPORTANTE, leia antes de tudo):
A solicitação abaixo pode descrever mais de uma frente de trabalho de uma vez (texto colado direto de anotações de campo, por exemplo). Você deve dividir em OSs SEPARADAS sempre que identificar:
- **Usinas diferentes** — cada usina distinta vira sua própria OS, SEM EXCEÇÃO (isso vale até pra câmeras/CFTV: se o mesmo assunto de câmera aparecer em duas usinas diferentes, vira duas OSs, uma por usina).
- **Equipamentos/sistemas diferentes** dentro da mesma usina — ex.: inversor e tracker são frentes diferentes, viram OSs separadas; trafo e string também.
CÂMERAS/CFTV: dentro de UMA MESMA usina, todo o conteúdo de câmeras (reposicionamento, foco, teste, instalação, várias câmeras diferentes) fica numa OS só — não precisa separar por número de câmera. Mas se o mesmo assunto de câmera envolver mais de uma usina, cada usina vira sua própria OS.
Se a solicitação já for sobre uma coisa só (uma usina, um equipamento, ou só câmeras de uma única usina), gere apenas UMA OS normalmente.

REGRAS DE FORMATAÇÃO (OBRIGATÓRIO) — aplique a cada OS individualmente:
- Esqueça introduções, conclusões, saudações, tabelas, ou seções como "Objetivo", "Descrição", "Responsáveis" ou "Evidências".
- Cada OS deve conter APENAS o "Título" e os "Comentários". Siga este modelo exato:

Título: [Nome curto e direto da atividade]
Comentários:

• [Passo 1 do procedimento]
• [Passo 2 do procedimento]
• [Passo 3...]

REGRAS DE ESCRITA E VOCABULÁRIO:
- O texto deve ser curto na estrutura (frases objetivas, sem enrolação), mas cada passo deve ser DETALHADO E ESPECÍFICO O SUFICIENTE para não deixar margem de interpretação. Escreva pensando que a equipe de campo, se o passo for vago, tende a executar de forma resumida ou pular a etapa — o texto tem que fechar essa brecha.
- Não invente informações ou equipamentos que não foram solicitados, mas garanta que o passo a passo faça sentido técnico.
- Integre orientações de segurança (EPIs, desenergização, sinalização) diretamente nos passos da atividade.
- Não repita a mesma ideia em mais de um item.
- Para atividades de acompanhamento/fiscalização, inicie os passos com verbos como: Acompanhar, verificar, conferir, registrar, avaliar, validar.
- Para atividades de execução direta, inicie os passos com verbos como: Realizar, executar, corrigir, ajustar, efetuar, acessar, inspecionar.

REGRA DE DETALHAMENTO ANTI-EXECUÇÃO-SUPERFICIAL (MUITO IMPORTANTE):
- Nunca deixe um passo genérico e solto, sem objeto claro — proibido escrever coisas como "verificar equipamento", "realizar manutenção", "checar funcionamento", "inspecionar componentes" sem dizer O QUÊ exatamente verificar/realizar/checar/inspecionar.
- Cada passo deve, sempre que aplicável, deixar explícito: (1) ONDE fazer — o ponto físico exato (ex.: entrada CC do inversor, saída CA, quadro de proteção, string X, conector MC4, cada canaleta, cada face do módulo), não só "no equipamento"; (2) COM QUE FERRAMENTA/MÉTODO — alicate amperímetro, multímetro, inspeção visual, acesso ao supervisório/datalogger, torquímetro, etc.; (3) QUAL O CRITÉRIO de aceite ou o que deve ser comparado — ex.: comparar leitura de campo com a plataforma de monitoramento, verificar se há assimetria entre strings, checar se o valor está dentro da faixa nominal, confirmar ausência de folga/oxidação/aquecimento.
- Se a atividade tiver múltiplos itens do mesmo tipo (várias strings, vários módulos, vários conectores, várias câmeras), deixe claro no passo que a verificação/ação deve ser feita EM CADA UM deles individualmente, não só "de forma geral".
- Pelo menos um dos últimos passos deve exigir registro de evidência de forma explícita e específica (fotos do antes/depois, valores medidos anotados, prints de tela) — não basta um "registrar" solto, diga o que precisa ser registrado.

REGRA PARA MEDIÇÕES ELÉTRICAS CC/CA (OBRIGATÓRIA sempre que a atividade envolver medir tensão ou corrente, ex.: strings, entrada/saída de inversor, quadros, trafo):
- Especifique sempre se a medição é de tensão ou corrente **CC** (lado dos módulos/strings, entrada do inversor) ou **CA** (lado da rede, saída do inversor) — nunca deixe implícito.
- Informe o instrumento correto: alicate amperímetro **True RMS com sonda de efeito Hall** para corrente CC (alicate comum não mede CC corretamente); alicate/multímetro comum é suficiente para grandezas CA. Inclua um passo pedindo pra selecionar a **escala/faixa correta** no instrumento antes de medir (faixa de corrente compatível com a Isc do string, faixa de tensão compatível com a Voc), e checar que as pontas de prova/garras estão com isolamento íntegro.
- Deixe claro se a medição é feita **com o sistema em operação** (string fechada, gerando) ou **em circuito aberto/desenergizado** (para medir Voc é preciso abrir a string antes) — não assuma, escreva explicitamente qual das duas.
- Se a atividade exigir desenergizar (abrir string, acessar bornes internos do inversor, quadro, trafo): inclua passos de **bloqueio e etiquetagem (LOTO)**, **teste de ausência de tensão com detector antes de tocar** (teste dos 3 pontos: testar o detector numa fonte conhecida, testar no equipamento, testar de novo na fonte conhecida) e aguardar o **tempo de descarga dos capacitores internos** do inversor antes de manusear bornes internos, seguindo a ordem de desligamento do fabricante (CA antes de CC, ou conforme manual).
- Sempre inclua o EPI adequado à classe de tensão do sistema: luva isolante de borracha na classe correta + luva de proteção mecânica por cima, óculos ou protetor facial, calçado isolante, ferramentas com cabo isolado — conforme NR-10.
- Inclua um passo pedindo pra anotar as **condições climáticas no momento da medição** (céu limpo, parcialmente nublado, nublado, chuva) — necessário pra dar contexto à comparação entre strings/inversores, já que geração varia com o clima.
- Ao medir múltiplas strings/circuitos, peça pra medir e registrar **cada um identificado pelo seu rótulo/etiqueta**, nunca uma medição genérica "do inversor".
- No passo de registro, exija o **valor medido com unidade** (não só "medir e verificar"), comparado com o valor nominal de placa/datasheet ou com a leitura da plataforma de monitoramento, além de foto do display do instrumento no ponto medido.

REGRA ESPECÍFICA DA GRID CO. (OBRIGATÓRIA, além das regras acima):
- Só inclua um passo pedindo autorização do COS (centro de operações) se a atividade envolver desligamento de inversor, desligamento da usina inteira, ou trabalho em SKID ou na Cabine de Medição Primária. Nesses casos, inclua um item pedindo autorização do COS antes da intervenção.
- Em qualquer outro caso, NÃO inclua nenhum item sobre o COS — não afirme que "não é necessário acionar o COS" nem que "a atividade não envolve manobra elétrica". Se não há necessidade de acionar o COS, simplesmente não mencione o assunto. Essa afirmação já causou erros de campo (times deixando de acionar o COS quando na verdade era necessário, confiando no texto padrão) e não deve mais ser usada.
- Sempre que a atividade envolver inspeção de trackers, estruturas de fixação/suporte de módulos fotovoltaicos, ou integridade estrutural/civil da usina de forma geral, inclua um passo avaliando as estruturas de fixação dos módulos quanto a afundamento (verificar se as bases/fundações/perfis de fixação apresentam sinais de afundamento, desnivelamento ou instabilidade no solo).

EXEMPLOS DO PADRÃO ESPERADO (cada um é o conteúdo de UMA OS):

Exemplo 1 (Atividade de Execução/Facilities)
Título: Limpeza de caixa d'água
Comentários:

• Fechar o registro de entrada de água (boia) com antecedência e isolar a área de acesso.
• Esvaziar a caixa até que reste apenas cerca de um palmo de água no fundo.
• Esfregar as paredes e o fundo utilizando escovas macias e exclusivas para este fim, sem uso de produtos químicos não homologados.
• Esvaziar a água suja, realizar o enxágue das paredes, reabrir o registro de entrada e fechar a tampa de forma hermética.

Exemplo 2 (Atividade de Diagnóstico/Elétrica)
Título: Inversor com aparente limitação de potência
Comentários:

• Acessar o sistema de monitoramento (supervisório) para verificar alarmes ativos, histórico de geração e indicação de derating.
• Realizar inspeção visual no inversor em campo, checando o funcionamento dos ventiladores e desobstrução das grades de ventilação.
• Inspecionar as medições de tensão e corrente nas entradas CC com alicate amperímetro para garantir que a queda de potência não seja causada por falha nos módulos ou sujeira.

Exemplo 3 (Atividade de Acompanhamento)
Título: Acompanhamento de roçagem
Comentários:

• Acompanhar a execução da roçagem na área designada, confirmando a delimitação do espaço.
• Verificar a sinalização e o uso correto de EPIs pela equipe terceira durante toda a atividade.
• Conferir se o serviço foi realizado conforme o planejamento, garantindo a integridade dos cabos e estruturas próximas.
• Registrar o andamento com evidências fotográficas e anotar eventuais pendências para correção.

Exemplo 4 (Atividade de Ajuste — CFTV, várias câmeras da MESMA usina ficam JUNTAS numa OS só)
Título: Reposicionamento de câmeras de CFTV
Comentários:

• Verificar a posição atual de cada câmera e o campo de visão afetado.
• Realizar o reposicionamento físico conforme a necessidade operacional, ajustando inclinação e direcionamento.
• Validar a visualização da imagem no sistema central de monitoramento para confirmar a cobertura desejada.
• Registrar a atividade e as evidências de antes e depois da intervenção.

Exemplo 5 (Visita Técnica Semanal — PADRÃO FIXO da Grid Co., use exatamente este texto sempre que a solicitação pedir "visita técnica semanal", "ronda semanal" ou equivalente, sem alterar os passos, só adaptando se algo específico for pedido a mais)
Título: Visita Técnica Semanal
Comentários:

• Realizar inspeção visual da vegetação na área da usina, avaliando a necessidade de roçagem e proximidade com os módulos e equipamentos.
• Inspecionar a sujidade dos módulos fotovoltaicos, registrando o nível de acúmulo e a necessidade de limpeza.
• Verificar as condições gerais da usina, incluindo vias de acesso, drenagem e integridade das estruturas.
• Avaliar as estruturas de fixação dos módulos fotovoltaicos quanto a sinais de afundamento, desnivelamento ou instabilidade no solo das bases/fundações.
• Conferir o cercamento perimetral, identificando pontos de vulnerabilidade ou danos.
• Inspecionar visualmente os inversores, verificando a limpeza externa, funcionamento dos ventiladores e ausência de alarmes no display.
• Coletar os dados de geração de cada inversor (dados de geração diária, de todos os dias deste mês).
• Acessar o sistema de CFTV para verificar o funcionamento das câmeras, qualidade das imagens e cobertura das áreas.
• Registrar todas as observações e evidências fotográficas para cada item inspecionado.

Exemplo 6 (Inspeção de Inversor para Abertura de Chamado — PADRÃO FIXO da Grid Co., use exatamente este texto sempre que a solicitação pedir "inspeção de inversor para abertura de chamado", "inspeção pra chamado" ou equivalente, sem alterar os passos, só adaptando se algo específico for pedido a mais)
Título: Inspeção de Inversor para Abertura de Chamado
Comentários:

• Acessar o sistema de monitoramento (supervisório) para verificar alarmes ativos, histórico de geração e indicação de derating.
• Realizar inspeção visual no inversor em campo, checando o funcionamento dos ventiladores e desobstrução das grades de ventilação.
• Inspecionar as medições de tensão e corrente nas entradas CC com alicate amperímetro para identificar possíveis anomalias.
• Inspecionar as medições de tensão e corrente nas entradas CA com alicate amperímetro para identificar possíveis anomalias.
• Coleta do número de série e posição operacional do inversor.
• Registrar todas as observações e evidências fotográficas para subsidiar a abertura de chamado.

Aplique exclusivamente este padrão. Não invente números de ticket, causas, nomes ou dados que não foram informados abaixo. Não repita a mesma OS mais de uma vez.

FORMATO DE SAÍDA (OBRIGATÓRIO): responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes ou depois), no formato:
{{"textos": ["Título: ...\\nComentários:\\n\\n• ...\\n• ...", "Título: ...\\nComentários:\\n\\n• ..."]}}
Cada item da lista é o texto completo de uma OS, no padrão exato descrito acima. Se só houver uma frente de trabalho, a lista tem um item só.

Dados da solicitação:
- Cliente: {d.get("cliente") or "não informado"}
- Usina: {d.get("usina") or "não informado"}
- Equipamento: {d.get("equipamento") or "não informado"}
- Falha/Descrição: {d.get("falha") or "não informado"}
- Causa: {d.get("causa") or "não informado"}
- Ação já realizada: {d.get("acao") or "não informado"}
- Histórico: {d.get("historico") or "não informado"}
- Responsável: {d.get("responsavel") or "não informado"}"""


def _indice_dia_util(data_str, hoje):
    """Converte uma data (dd/mm/aaaa) em 'quantos dias úteis a partir de
    amanhã' ela representa (0 = primeiro dia útil disponível). Ignora fins
    de semana na contagem. Retorna None se a data já passou ou é hoje."""
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", data_str)
    if not m:
        return None
    dt = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    d = hoje.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(d, datetime) and d.tzinfo:
        dt = dt.replace(tzinfo=d.tzinfo)
    if dt <= d:
        return None
    idx = -1
    cursor = d + timedelta(days=1)
    while cursor <= dt:
        if cursor.weekday() < 5:
            idx += 1
        cursor += timedelta(days=1)
    return idx if idx >= 0 else None


def _dia_util_por_indice(idx, hoje):
    """Inverso de _indice_dia_util: dado um índice (0 = primeiro dia útil
    disponível), devolve a data (dd/mm/aaaa) correspondente."""
    d = hoje.replace(hour=0, minute=0, second=0, microsecond=0)
    cursor = d + timedelta(days=1)
    contador = -1
    while True:
        if cursor.weekday() < 5:
            contador += 1
            if contador == idx:
                return cursor.strftime("%d/%m/%Y")
        cursor += timedelta(days=1)


def _comprimir_agenda_reprogramacao(sugestao, hoje):
    """Garantia extra e determinística, além do prompt: se a IA, mesmo com
    a lista de dias úteis explícita, deixar o primeiro dia disponível sem
    uso (ex.: começar só na terça quando segunda estava livre), desloca
    TODA a agenda sugerida pra trás em dias úteis, preservando a ordem e
    os turnos, até o dia mais cedo usado virar o primeiro dia disponível.
    Roda in-place no dict."""
    itens = sugestao.get("reprogramacoes", [])
    indices = []
    for item in itens:
        idx = _indice_dia_util((item.get("dataSugerida") or "").strip(), hoje)
        item["_idx_dia_util"] = idx
        if idx is not None:
            indices.append(idx)
    if not indices:
        return
    deslocamento = min(indices)
    if deslocamento <= 0:
        for item in itens:
            item.pop("_idx_dia_util", None)
        return
    for item in itens:
        idx = item.pop("_idx_dia_util", None)
        if idx is not None:
            item["dataSugerida"] = _dia_util_por_indice(idx - deslocamento, hoje)


def _corrigir_fins_de_semana(sugestao):
    """Garantia extra além do prompt: se a IA, mesmo assim, sugerir uma
    data em sábado ou domingo, empurra pra segunda-feira seguinte. Roda
    depois da resposta da IA, direto no dict (in-place)."""
    for item in sugestao.get("reprogramacoes", []):
        data_str = (item.get("dataSugerida") or "").strip()
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", data_str)
        if not m:
            continue
        dt = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        dias_ate_segunda = {5: 2, 6: 1}.get(dt.weekday())  # 5=sábado, 6=domingo
        if dias_ate_segunda:
            dt_corrigida = dt + timedelta(days=dias_ate_segunda)
            item["dataSugerida"] = dt_corrigida.strftime("%d/%m/%Y")


def _proximos_dias_uteis(a_partir_de, quantidade=12):
    """Retorna uma lista de (data_str, nome_dia_semana) dos próximos N dias
    úteis (seg-sex) a partir do dia seguinte a `a_partir_de`. Calculado em
    Python, não deixado por conta da IA — remove qualquer chance de erro
    de cálculo de data/dia da semana por parte do modelo."""
    nomes = ["segunda-feira", "terça-feira", "quarta-feira", "quinta-feira",
             "sexta-feira", "sábado", "domingo"]
    dias = []
    d = a_partir_de + timedelta(days=1)
    while len(dias) < quantidade:
        if d.weekday() < 5:  # 0-4 = seg a sex
            dias.append((d.strftime("%d/%m/%Y"), nomes[d.weekday()]))
        d += timedelta(days=1)
    return dias


def _chamar_gemini_com_retry(payload, timeout=45, tentativas=3, usar_chave_teste=False):
    """Chama a API do Gemini com retry automático em caso de 429 (limite
    de taxa) — espera crescente entre tentativas (2s, 5s, 10s). Um pico
    passageiro de uso (ex.: várias chamadas em sequência rápida) costuma
    se resolver sozinho em poucos segundos; isso evita expor esse erro
    direto pro usuário na maioria dos casos.

    Se todas as tentativas na chave principal falharem por 429 (limite
    de taxa esgotado — mais provável agora que várias funcionalidades de
    IA dividem a mesma cota gratuita: gerar OS, priorização, comunicados
    diários, comunicado livre e fotos de zeladoria), tenta UMA VEZ a mais
    usando GEMINI_API_KEY_TESTE como reserva, já que é um projeto separado
    no Google AI Studio com cota própria — antes disso só era usada em
    diagnósticos manuais. Só levanta a exceção se isso também falhar
    (ex.: cota diária de ambas as chaves realmente esgotada).

    usar_chave_teste=True força usar direto a chave de teste (comportamento
    original, usado em diagnósticos manuais)."""
    chave = (GEMINI_API_KEY_TESTE if usar_chave_teste and GEMINI_API_KEY_TESTE else GEMINI_API_KEY)
    esperas = [2, 5, 10]
    ultima_excecao = None
    ultimo_status = None
    teve_timeout_ou_conexao = False
    for tentativa in range(tentativas):
        try:
            resp = requests.post(f"{GEMINI_URL}?key={chave}", json=payload, timeout=timeout)
            ultimo_status = resp.status_code
            if resp.status_code == 429 and tentativa < tentativas - 1:
                time.sleep(esperas[tentativa])
                continue
            resp.raise_for_status()
            return resp
        except requests.exceptions.HTTPError as e:
            # inclui o corpo da resposta do Google na mensagem — o texto
            # padrão do HTTPError ("400 Client Error: Bad Request for
            # url: ...") não mostra o motivo real, só o código. Isso é
            # essencial pra diagnosticar erros que não são de cota (ex.:
            # parâmetro inválido, modelo não encontrado).
            try:
                detalhe = resp.text[:500]
            except Exception:
                detalhe = ""
            ultima_excecao = requests.exceptions.HTTPError(f"{e} | corpo: {detalhe}", response=resp)
            if resp.status_code == 429 and tentativa < tentativas - 1:
                time.sleep(esperas[tentativa])
                continue
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            # BUG CORRIGIDO (24/07/2026): timeout de leitura/erro de conexão
            # com a API do Gemini não era pego por nenhum except anterior
            # (só HTTPError/429 tinham retry), então a exceção crua (ex.:
            # "Read timed out") vazava direto pro usuário sem nenhuma nova
            # tentativa. Agora trata igual às outras falhas: reretenta com
            # o mesmo backoff crescente e, se esgotar, cai no fallback pra
            # chave de teste como qualquer outro tipo de falha.
            ultima_excecao = e
            teve_timeout_ou_conexao = True
            log.warning(f"[Gemini] Timeout/erro de conexão na tentativa {tentativa + 1}/{tentativas}: {e}")
            if tentativa < tentativas - 1:
                time.sleep(esperas[tentativa])
                continue

    # esgotou as tentativas na chave principal — tenta a chave de teste
    # como reserva, uma única vez, antes de desistir. Cobre 429 (cota
    # esgotada), 401/403 (chave principal inválida, revogada ou expirada
    # — identificado 23/07/2026) e agora também timeout/erro de conexão
    # (identificado 24/07/2026).
    if (ultimo_status in (429, 401, 403) or teve_timeout_ou_conexao) and not usar_chave_teste and GEMINI_API_KEY_TESTE:
        try:
            log.warning(f"[Gemini] Chave principal falhou (status={ultimo_status}, timeout={teve_timeout_ou_conexao}) — usando chave de teste como reserva")
            resp = requests.post(f"{GEMINI_URL}?key={GEMINI_API_KEY_TESTE}", json=payload, timeout=timeout)
            resp.raise_for_status()
            return resp
        except requests.exceptions.HTTPError as e:
            try:
                detalhe = resp.text[:500]
            except Exception:
                detalhe = ""
            ultima_excecao = requests.exceptions.HTTPError(f"{e} | corpo: {detalhe}", response=resp)
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            ultima_excecao = e

    raise ultima_excecao


def _montar_prompt_comunicado_livre(tema, observacoes, tem_imagem=False):
    bloco_imagem = ""
    if tem_imagem:
        bloco_imagem = """
UMA IMAGEM (PRINT) FOI ANEXADA COMO FONTE DE INFORMAÇÃO: extraia dela tudo que for relevante pro comunicado — texto, valores, nomes, datas, contexto visível. Se o tema/observações também tiverem sido preenchidos, combine as duas fontes; se estiverem vazios, baseie o comunicado inteiramente no que a imagem mostra. Não invente nada que não esteja nem no texto nem na imagem — se algo na imagem estiver ilegível ou cortado, não presuma o conteúdo.
"""
    return f"""Aja como um Supervisor de O&M da Grid Co. redigindo um comunicado TÉCNICO e DIRETO para ser enviado por WhatsApp às equipes de campo e/ou clientes.
{bloco_imagem}
REGRA MAIS IMPORTANTE — OS PONTOS ESPECÍFICOS DO USUÁRIO SÃO O NÚCLEO DO COMUNICADO: tudo que foi pedido explicitamente no tema/observações/imagem (itens a verificar, instruções, prazos, ações) tem que aparecer de forma CLARA E DESTACADA — se forem vários itens/verificações, apresente como LISTA curta (um item por linha), nunca dissolvidos dentro de um parágrafo genérico. Uma instrução prática específica (ex.: "verificar selo de calibração") NUNCA pode se perder atrás de linguagem genérica sobre riscos, normas ou importância do assunto. Se ao reler o texto pronto o ponto principal pedido pelo usuário não pular aos olhos em 2 segundos de leitura, o texto está errado.

Contexto técnico (o "porquê") é permitido, mas com moderação: no máximo UMA frase curta de enquadramento, nunca um parágrafo explicando consequências genéricas de segurança/normas regulatórias. Gaste as linhas do comunicado nos pontos que o usuário efetivamente pediu pra comunicar, não em explicações genéricas sobre o tema.

O que NÃO fazer:
- Não inventar critérios técnicos adicionais de verificação que não foram citados (ex.: tipos específicos de dano, passos extras de conferência, registro fotográfico) — isso é diferente do contexto genérico permitido; contexto genérico é só uma frase de enquadramento, nunca uma checklist extra por conta própria.
- Não inventar FATOS ESPECÍFICOS não informados — datas exatas, números de OS, prazos, quantidades, nomes.
- Não usar linguagem alarmista ou didática demais (nada de "pode resultar em consequências fatais" se isso não foi pedido) — o tom é de instrução profissional direta entre colegas de O&M, não uma aula de segurança.

Estrutura padrão (adapte ao conteúdo, mas sempre em blocos curtos com emoji de destaque):
⚠️ Linha de abertura com o assunto principal, direto ao ponto.
📋 Os pontos/itens específicos pedidos pelo usuário (do texto e/ou da imagem) — em lista, se forem vários.
📋 (opcional, só se necessário) Uma frase curta de contexto técnico ou instrução complementar.
✅ Prazo, status ou próximo passo — respeitando rigorosamente o grau de certeza da regra abaixo.

Regras gerais:
- Vá direto ao ponto, sem saudação genérica tipo "Prezados" ou assinatura formal no final.
- Tamanho: enxuto — normalmente 3 a 5 linhas curtas. Mais curto e direto é melhor do que mais longo e genérico.

REGRA CRÍTICA — PRESERVAR O GRAU DE CERTEZA DO TEXTO ORIGINAL: preste muita atenção em palavras que indicam incerteza ou expectativa, como "acredito que", "acho que", "acho possível", "acredito", "acho provável", "talvez", "devemos", "devemos conseguir". NUNCA transforme uma expectativa/crença em uma confirmação ou promessa de prazo. Se o autor disse que "acredita" que algo vai acontecer, sem data confirmada, o comunicado deve deixar claro que ainda NÃO há data definida (ex.: "ainda não temos data para retorno", "sem previsão confirmada", "assim que tivermos confirmação, avisamos") — em vez de anunciar como certo ou "em breve". Errar pra mais confiança do que o texto original tem é pior do que errar pra menos — na dúvida, seja mais conservador, não mais otimista.

Tema do comunicado: {tema or "(nenhum tema em texto — considere só a imagem, se houver)"}
Observações/detalhes: {observacoes or "nenhuma observação adicional"}

FORMATO DE SAÍDA (OBRIGATÓRIO): responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes ou depois), no formato:
{{"texto": "o comunicado pronto pra enviar"}}"""


@app.route("/diag-testar-modelo-gemini", methods=["GET"])
def diag_testar_modelo_gemini():
    """Ferramenta de diagnóstico temporária: testa se um modelo Gemini
    específico funciona com a autenticação simples por chave que este
    sistema usa. Uso: ?modelo=gemini-3.1-flash-lite&chave=teste (ou
    &chave=principal). Não faz parte do fluxo normal do app."""
    modelo = request.args.get("modelo", "").strip()
    usar_teste = request.args.get("chave", "principal") == "teste"
    if not modelo:
        return jsonify({"ok": False, "error": "informe ?modelo=nome-do-modelo"}), 400
    chave = GEMINI_API_KEY_TESTE if usar_teste else GEMINI_API_KEY
    if not chave:
        return jsonify({"ok": False, "error": "chave não configurada"}), 400
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{modelo}:generateContent"
    try:
        resp = requests.post(
            f"{url}?key={chave}",
            json={"contents": [{"parts": [{"text": "responda só a palavra: ok"}]}]},
            timeout=25,
        )
        return jsonify({"ok": resp.ok, "status": resp.status_code, "corpo": resp.text[:600]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _normalizar_tema_comunicado(texto):
    """Remove acentos e normaliza espaços/caixa pra comparar temas de forma
    tolerante (ex: 'Comunicado Padrão' == 'comunicado padrao')."""
    nfkd = _ud.normalize("NFKD", texto or "")
    sem_acento = "".join(c for c in nfkd if not _ud.combining(c))
    return re.sub(r"\s+", " ", sem_acento).strip().lower()


# Comunicados fixos por TEMA: a chave é o tema normalizado (sem acento/caixa),
# o valor é o texto pronto a ser usado. Pra usar, digite o tema normalmente
# (ex: "Solicitações de compra") e escreva "comunicado padrão" no OUTRO campo
# (tema ou observações, tanto faz) — isso funciona como gatilho dizendo
# "use o texto fixo desse tema" em vez de gerar via IA.
_COMUNICADOS_PRESET_TEMA = {
    "abastecimentos": (
        "⚠️ Pessoal, sobre abastecimentos:\n"
        "📋 Por favor, solicitem no primeiro horário da manhã. Isso evita múltiplas solicitações ao financeiro.\n"
        "✅ Após liberação, o cartão Clara fica ativo até as 17h, tempo suficiente para abastecer."
    ),
    "solicitacoes de compra": (
        "⚠️ Pessoal, sobre solicitações de compra:\n"
        "📋 Por favor, enviem o quanto antes todas as solicitações de compra necessárias.\n"
        "✅ Isso evita atrasos no processo e garante que o material chegue em tempo hábil."
    ),
}

_GATILHO_COMUNICADO_PADRAO = "comunicado padrao"  # já normalizado (sem acento)


def _resolver_preset_comunicado(tema, observacoes):
    """Se 'comunicado padrão' foi digitado num dos dois campos, usa o OUTRO
    campo (normalizado) como tema pra buscar o preset fixo correspondente.
    Retorna None se não houver gatilho ou se o tema não tiver preset."""
    tema_norm = _normalizar_tema_comunicado(tema)
    obs_norm = _normalizar_tema_comunicado(observacoes)
    if obs_norm == _GATILHO_COMUNICADO_PADRAO:
        return _COMUNICADOS_PRESET_TEMA.get(tema_norm)
    if tema_norm == _GATILHO_COMUNICADO_PADRAO:
        return _COMUNICADOS_PRESET_TEMA.get(obs_norm)
    return None


@app.route("/gerar-comunicado-livre-ia", methods=["POST", "OPTIONS"])
def gerar_comunicado_livre_ia():
    """Gera um texto de comunicado livre (tema + observações e/ou um print
    anexado) usando IA, pra ser enviado manualmente pelos grupos que o Fred
    escolher — usado pelo campo 'Gerar Comunicado' na sidebar, ao lado do
    'Gerar OS'. Se um print for anexado, a IA (Gemini, visão + texto) lê o
    conteúdo da imagem pra montar o comunicado — tema/observações em texto
    passam a ser opcionais nesse caso (mas continuam sendo combinados com
    a imagem se preenchidos). Se 'comunicado padrão' for digitado num dos
    campos de texto, busca um preset fixo pelo tema (_COMUNICADOS_PRESET_TEMA)
    e devolve na hora, sem IA — presets não usam imagem."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    tema = (body.get("tema") or "").strip()
    observacoes = (body.get("observacoes") or "").strip()
    imagem_b64 = body.get("imagemBase64") or ""
    imagem_mime = body.get("imagemMimeType") or "image/png"
    if not tema and not imagem_b64:
        return jsonify({"ok": False, "error": "informe o tema do comunicado ou anexe um print"}), 400

    preset = _resolver_preset_comunicado(tema, observacoes) if tema else None
    if preset:
        return jsonify({"ok": True, "texto": preset})

    prompt = _montar_prompt_comunicado_livre(tema, observacoes, tem_imagem=bool(imagem_b64))
    parts = [{"text": prompt}]
    if imagem_b64:
        parts.append({"inline_data": {"mime_type": imagem_mime, "data": imagem_b64}})

    diagnostico = request.args.get("diagnostico", "").lower() == "true"
    try:
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": parts}],
                "generationConfig": {
                    "temperature": 0.4,
                    "maxOutputTokens": 1024,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=45 if imagem_b64 else 20,
            usar_chave_teste=diagnostico,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        texto_bruto = candidato["content"]["parts"][0]["text"].strip()
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
        parsed = json.loads(texto_limpo)
        texto = (parsed.get("texto") or "").strip()
        if not texto:
            raise ValueError("A IA não retornou nenhum texto")
        return jsonify({"ok": True, "texto": texto})
    except Exception as e:
        log.error(f"[gerar-comunicado-livre-ia] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 502


_NOMES_GRUPOS_CONHECIDOS = {
    "120363423233716775": "[O&M] - Grid Co. | Renogrid",
    "120363423427343356": "[O&M] - Grid Co. | Thopen",
    "120363402559504115": "[O&M] - Grid Co. | 2C",
    "120363426381032089": "[O&M] - Grid Co. | Alves Lima",
    "120363423844956611": "[O&M] - Grid Co. | GD Energy",
    "120363427259899891": "[O&M] - Grid Co. | Sal Energia",
    "120363406191445169": "O&M - San. Bárb./Pirac. - SP LESTE 03",
    "120363431221706747": "Equipe Guajiru/Sol do Norte I e II",
    "120363421162420788": "COS — Técnicos O&M Centro-Oeste",
    "120363425837962709": "COS — Técnicos O&M Sul",
    "120363402176878100": "COS — Técnicos O&M Nordeste",
    "120363423533840348": "COS — Técnicos O&M Sudeste",
    "120363421052607450": "COS — Técnicos O&M Norte",
    "120363405111083249": "Equipe Crateús",
    "120363410081447469": "Equipe Crateús",
    "120363405244065477": "Equipe Bonfim/Morada Nova/Quixadá",
    "120363427839577268": "Equipe Elias Fausto",
    "120363428268426406": "Equipe Matão/Topázio",
    "120363422795399103": "Equipe Ibaté/Boa Esperança",
    "120363403858325184": "Equipe Sete Lagoas",
    "120363426700120222": "Equipe Colíder - Grid Co.",
    "120363426886851537": "Equipe Nova Xavantina",
    "120363425342949474": "Equipe Araputanga/Poconé",
    "120363406329162612": "Equipe Nobres",
    "120363424804307945": "Thopen & GridCo. | Usinas FRED ALEXANDRINO",
    "120363428178674382": "Equipe - Aquiraz/Cascavel",
    "120363406919935108": "Arquivos NVX",
    "120363423651075316": "Equipe Camila O&M",
}


@app.route("/grupos-configurados", methods=["GET"])
def listar_grupos_configurados():
    """Lista os grupos do WhatsApp configurados (GRUPOS_IDS), com nome
    amigável quando disponível — usado pra montar a lista de seleção no
    pop-up de comunicado livre."""
    itens = []
    for grupo_id in GRUPOS_FILTRO:
        grupo_id = grupo_id.strip()
        if not grupo_id:
            continue
        id_numerico = grupo_id.replace("@g.us", "")
        nome = _NOMES_GRUPOS_CONHECIDOS.get(id_numerico) or _nome_amigavel_grupo(grupo_id) or f"Grupo {grupo_id[:14]}…"
        itens.append({"id": grupo_id, "nome": nome})
    itens.sort(key=lambda x: x["nome"])
    return jsonify({"ok": True, "itens": itens})


@app.route("/disparar-comunicado-livre", methods=["POST", "OPTIONS"])
def disparar_comunicado_livre():
    """Envia um texto de comunicado livre (já gerado/editado) pra uma
    lista de grupos escolhida manualmente pelo Fred no pop-up.
    Aceita opcionalmente uma lista de imagens (prints/fotos) em base64 —
    quando presentes, a 1ª imagem é enviada com o texto como legenda e as
    demais em seguida, sem legenda. Sem imagens, comportamento igual a antes
    (mensagem de texto simples)."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    texto = (body.get("texto") or "").strip()
    grupos = body.get("grupos") or []
    imagens = body.get("imagens") or []  # [{base64, mimeType}, ...] — mimeType não usado pelo Baileys, aceito por completude
    if not texto:
        return jsonify({"ok": False, "error": "texto vazio"}), 400
    if not grupos or not isinstance(grupos, list):
        return jsonify({"ok": False, "error": "selecione ao menos um grupo"}), 400
    if not isinstance(imagens, list):
        return jsonify({"ok": False, "error": "imagens deve ser uma lista"}), 400
    if len(imagens) > 5:
        return jsonify({"ok": False, "error": "máximo de 5 imagens por comunicado"}), 400
    if not WPP_SERVER_URL:
        return jsonify({"ok": False, "error": "WPP_SERVER_URL não configurado"}), 500

    def _enviar_com_retry(payload, endpoint):
        """Mesma lógica de retry curto pra blips de conexão do WhatsApp,
        reaproveitada tanto pra texto quanto pra imagem."""
        ultimo_erro = None
        for tentativa in range(3):
            try:
                r = requests.post(
                    f"{WPP_SERVER_URL}{endpoint}",
                    json=payload,
                    headers={"X-Webhook-Secret": WEBHOOK_SECRET} if WEBHOOK_SECRET else {},
                    timeout=40,  # imagens demoram mais que texto puro
                )
                if r.ok and r.json().get("ok"):
                    return True, None
                corpo = r.text[:200]
                ultimo_erro = corpo
                if "não conectado" in corpo.lower() or "nao conectado" in corpo.lower():
                    if tentativa < 2:
                        time.sleep(4)
                        continue
                return False, ultimo_erro
            except Exception as e:
                ultimo_erro = str(e)
                return False, ultimo_erro
        return False, ultimo_erro

    enviados, erros = [], []
    for grupo_id in grupos:
        sucesso_geral = True
        erro_grupo = None

        if imagens:
            # 1ª imagem leva o texto como legenda; as demais vão sem legenda
            for i, img in enumerate(imagens):
                img_b64 = (img.get("base64") or "") if isinstance(img, dict) else ""
                if not img_b64:
                    continue
                payload_img = {"grupoId": grupo_id, "imagemBase64": img_b64}
                if i == 0:
                    payload_img["legenda"] = texto
                ok, erro = _enviar_com_retry(payload_img, "/api/enviar-imagem")
                if not ok:
                    sucesso_geral = False
                    erro_grupo = erro
                    break
            # Se por algum motivo nenhuma imagem tinha base64 válido, cai pro texto puro
            if sucesso_geral and not any((img.get("base64") if isinstance(img, dict) else None) for img in imagens):
                ok, erro = _enviar_com_retry({"grupoId": grupo_id, "texto": texto}, "/api/enviar-mensagem")
                sucesso_geral, erro_grupo = ok, erro
        else:
            ok, erro = _enviar_com_retry({"grupoId": grupo_id, "texto": texto}, "/api/enviar-mensagem")
            sucesso_geral, erro_grupo = ok, erro

        if sucesso_geral:
            enviados.append(grupo_id)
        else:
            erros.append({"grupo": grupo_id, "erro": erro_grupo})
    return jsonify({"ok": True, "enviados": enviados, "erros": erros})


def _montar_saudacao_cliente():
    """Saudação por horário (Brasília), calculada no momento em que o texto
    é gerado — usado no resumo pro cliente final (Gestão Cliente). Bom dia
    até 12h, boa tarde até 18h, boa noite depois disso."""
    hora = agora_br().hour
    if hora < 12:
        return "Bom dia"
    if hora < 18:
        return "Boa tarde"
    return "Boa noite"


def _montar_prompt_resumo_cliente(cliente, atividades, saudacao):
    """Monta o prompt pra gerar o resumo de atividades 'Em Processo'
    selecionadas manualmente pelo Fred, em linguagem polida e voltada ao
    cliente final (não ao técnico de campo) — usado pelo painel 'Gestão
    Cliente'. Diferente do comunicado técnico (objetivo/seco), aqui o tom é
    consultivo: saudação, contexto, e assinatura pessoal do Fred. A OS é
    incluída como referência ao final de cada linha (pedido do Fred em
    27/07/2026 — antes era omitida, agora é mantida pra rastreabilidade)."""
    mapa_cluster = _mapa_cluster_usina()
    por_usina = {}
    for a in atividades:
        usina = (a.get("usina") or "").strip() or "não informado"
        descricao = (a.get("descricao") or "").strip()
        numero_os = (a.get("numeroOS") or "").strip()
        if not descricao:
            continue
        linha = descricao + (f" [ref. OS {numero_os}]" if numero_os else "")
        por_usina.setdefault(usina, []).append(linha)

    blocos = []
    for usina, descricoes in por_usina.items():
        cluster = mapa_cluster.get(usina, "")
        linhas_usina = "\n".join(f"  - {d}" for d in descricoes)
        blocos.append(f"Usina: {usina}" + (f" (cluster {cluster})" if cluster else "") + f"\n{linhas_usina}")
    lista_atividades = "\n\n".join(blocos)

    return f"""Aja como Fred Alexandrino, Supervisor de O&M da Grid Co., escrevendo uma mensagem de WhatsApp pro cliente final ({cliente}) com o panorama de atividades programadas para hoje.

Esta mensagem é PRA CLIENTE, não pra equipe técnica interna — o tom deve ser polido, profissional, amigável e direto, como uma comunicação de relacionamento com cliente (não uma ordem de serviço interna).

Regras obrigatórias de formato:
- Comece com a saudação seca e direta "{saudacao}, pessoal." — sem exclamação, sem emoji, e sem mencionar o nome do cliente na saudação (o texto já vai ser enviado no grupo certo, não precisa reforçar isso).
- Uma frase curta de abertura contextualizando que segue o panorama de atividades programadas para hoje.
- Liste as atividades agrupadas por usina (use o nome da usina como pequeno destaque, ex. em negrito ou seguido de dois pontos), com marcador "•" para cada atividade daquela usina.
- Reescreva cada descrição de atividade em linguagem clara e acessível pro cliente — SEM código de ativo/equipamento cru, SEM jargão técnico interno de sistema (nada de "statusOS", "Fracttal", etc). Mantenha o conteúdo técnico real (o que será feito), só troque a forma como é dito.
- Quando a atividade tiver uma marcação "[ref. OS XXXX]" na informação fornecida, inclua o número da OS como referência ao final da linha, em formato discreto, ex.: "• Manutenção preventiva no inversor 3 (ref. OS 8508)". Não invente número de OS pra atividades que não tiverem essa marcação.
- Não invente atividades nem detalhes que não estejam na lista fornecida.
- Termine com uma frase curta de disponibilidade/cordialidade (ex. equipe de campo mobilizada, à disposição para dúvidas).
- Assine ao final com:
Atenciosamente,
Fred Alexandrino
Supervisor de O&M — Grid Co.
- Não use fontes de negrito/asterisco no WhatsApp além do já convencional (*texto* vira negrito no WhatsApp, pode usar com moderação pro nome da usina e pro nome do Fred na assinatura).

Cliente: {cliente}
Atividades selecionadas (agrupadas por usina):
{lista_atividades}

FORMATO DE SAÍDA (OBRIGATÓRIO): responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes ou depois), no formato:
{{"texto": "a mensagem pronta pra enviar, com quebras de linha \\n"}}"""


@app.route("/gerar-resumo-cliente", methods=["POST", "OPTIONS"])
def gerar_resumo_cliente():
    """Gera o texto do resumo de atividades 'Em Processo' selecionadas
    manualmente pelo Fred pra um cliente específico (painel 'Gestão
    Cliente'). Diferente dos comunicados técnicos/automáticos: a seleção
    de quais atividades entram é 100% manual (Fred marca cada uma), pra
    evitar que atividades antigas esquecidas apareçam pro cliente sem
    controle. O envio em si é feito depois via /disparar-comunicado-livre
    (mesma infraestrutura de grupos já existente), reaproveitado aqui."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    cliente = (body.get("cliente") or "").strip()
    atividades = body.get("atividades") or []
    if not cliente:
        return jsonify({"ok": False, "error": "informe o cliente"}), 400
    if not atividades or not isinstance(atividades, list):
        return jsonify({"ok": False, "error": "selecione ao menos uma atividade"}), 400

    saudacao = _montar_saudacao_cliente()
    prompt = _montar_prompt_resumo_cliente(cliente, atividades, saudacao)
    diagnostico = request.args.get("diagnostico", "").lower() == "true"
    try:
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.4,
                    "maxOutputTokens": 1024,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=20,
            usar_chave_teste=diagnostico,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        texto_bruto = candidato["content"]["parts"][0]["text"].strip()
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
        parsed = json.loads(texto_limpo)
        texto = (parsed.get("texto") or "").strip()
        if not texto:
            raise ValueError("A IA não retornou nenhum texto")
        return jsonify({"ok": True, "texto": texto})
    except Exception as e:
        log.error(f"[gerar-resumo-cliente] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 502


_EQUIP_ABREV_CHAMADO = [
    (r"invers", "INV"),
    (r"track", "TKR"),
    (r"transformador", "TRAFO"),
    (r"string\s*box|stringbox", "SB"),
    (r"m[oó]dulo", "MOD"),
    (r"otimizador", "OTM"),
    (r"medidor|medi[cç][aã]o", "MED"),
    (r"cftv|c[aâ]mera", "CAM"),
    (r"nobreak", "NBK"),
    (r"seccionadora", "SEC"),
    (r"combiner", "CMB"),
]


def _codigo_equipamento_chamado(ativo, identificacao):
    """Monta o código compacto do equipamento pro resumo de chamados pro
    cliente: INV14, TKR05, etc. — XX é o número do equipamento (inversor,
    tracker, etc.) extraído de 'Identificação do Equipamento' ou 'Ativo'.
    Se o tipo não for reconhecido, cai pra identificação bruta (sem
    espaços). Pedido do Fred em 10/08/2026."""
    fonte = f"{identificacao or ''} {ativo or ''}".strip()
    if not fonte:
        return ""
    fonte_norm = fonte.lower()
    numeros = re.findall(r"\d+", fonte)
    numero = numeros[0].zfill(2) if numeros else ""
    for padrao, abrev in _EQUIP_ABREV_CHAMADO:
        if re.search(padrao, fonte_norm):
            return f"{abrev}{numero}" if numero else abrev
    bruto = (identificacao or ativo or "").strip()
    return re.sub(r"\s+", "", bruto).upper() if bruto else ""


def _montar_resumo_chamados_cliente(cliente, chamados, saudacao):
    """Monta o resumo de chamados de fabricante (status != 'Finalizado')
    pro cliente final, agrupado por usina — botão 'Gestão Cliente' dentro
    do Painel de Chamados.

    Formato de linha fixo, pedido pelo Fred em 10/08/2026:
      INVXX - #TICKET - STATUS   (XX = nº do inversor; TKR = tracker)

    Geração 100% determinística (sem IA): o formato agora é totalmente
    estruturado/mecânico, então template evita variação indesejada de
    texto e elimina o timeout ocasional que a chamada à IA causava."""
    por_usina = {}
    for c in chamados:
        usina = (c.get("usina") or "").strip() or "não informado"
        status = (c.get("status") or "").strip() or "em andamento"
        ticket = (c.get("ticket") or "").strip()
        codigo = _codigo_equipamento_chamado(c.get("ativo"), c.get("equipamento")) or "Equipamento"
        linha = codigo
        if ticket:
            linha += f" - #{ticket}"
        linha += f" - {status}"
        por_usina.setdefault(usina, []).append((codigo, linha))

    blocos = []
    for usina in sorted(por_usina.keys()):
        itens = sorted(por_usina[usina], key=lambda x: x[0])
        linhas_usina = "\n".join(f"• {linha}" for _, linha in itens)
        blocos.append(f"*{usina}*\n{linhas_usina}")
    corpo = "\n\n".join(blocos)

    return (
        f"{saudacao}, pessoal.\n\n"
        f"Segue o panorama dos chamados de fabricante (garantia/assistência técnica) em andamento:\n\n"
        f"{corpo}\n\n"
        f"Qualquer novidade, seguimos informando. Estamos à disposição.\n\n"
        f"Atenciosamente,\n"
        f"Fred Alexandrino\n"
        f"Supervisor de O&M — Grid Co."
    )


@app.route("/gerar-resumo-chamados-cliente", methods=["POST", "OPTIONS"])
def gerar_resumo_chamados_cliente():
    """Gera o texto do resumo de chamados de fabricante (status != Finalizado)
    selecionados manualmente pelo Fred pra um cliente específico — botão
    'Gestão Cliente' dentro do Painel de Chamados. Envio feito depois via
    /disparar-comunicado-livre (mesma infraestrutura já existente).

    Geração determinística (sem IA) desde 10/08/2026 — ver
    _montar_resumo_chamados_cliente."""
    if request.method == "OPTIONS":
        return ("", 204)
    body = request.get_json(force=True, silent=True) or {}
    cliente = (body.get("cliente") or "").strip()
    chamados = body.get("chamados") or []
    if not cliente:
        return jsonify({"ok": False, "error": "informe o cliente"}), 400
    if not chamados or not isinstance(chamados, list):
        return jsonify({"ok": False, "error": "selecione ao menos um chamado"}), 400

    try:
        saudacao = _montar_saudacao_cliente()
        texto = _montar_resumo_chamados_cliente(cliente, chamados, saudacao)
        return jsonify({"ok": True, "texto": texto})
    except Exception as e:
        log.error(f"[gerar-resumo-chamados-cliente] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 502


def _montar_prompt_priorizacao(atividades, hoje_str):
    mapa_cluster = _mapa_cluster_usina()
    linhas = []
    for a in atividades:
        cluster = mapa_cluster.get((a.get("usina") or "").strip(), "não mapeado")
        linhas.append(
            f"- id={a['id']} | OS={a.get('numeroOS') or '—'} | Cliente={a['cliente']} | Usina={a['usina']} | "
            f"Cluster/Região={cluster} | Equipamento={a.get('equipamento') or '—'} | "
            f"Descrição={a.get('descricao') or '—'} | Responsável/Equipe={a.get('responsavel') or 'não informado'} | "
            f"Prioridade={a.get('prioridade') or 'Média'} | Prazo={a.get('prazo') or 'sem prazo definido'} | "
            f"Estado Fracttal={a.get('statusOS') or '—'} | % concluído={a.get('percentualOS') or '0'}"
        )
    lista_atividades = "\n".join(linhas)

    return f"""Aja como um Engenheiro(a) de O&M Sênior especialista em usinas fotovoltaicas, responsável por decidir a ordem de execução das atividades de campo do dia para múltiplas equipes técnicas espalhadas por várias usinas.

CONTEXTO:
Hoje é {hoje_str}. Há um volume grande de atividades/OS em aberto e fica difícil pro supervisor saber, à primeira vista, o que priorizar. Seu trabalho é ler a lista completa abaixo e devolver uma ordem de prioridade clara e justificada — não é uma escala de quem faz o quê, é um raio-x de "o que mais importa agora e por quê".

CRITÉRIOS DE PRIORIZAÇÃO (avalie todos, na ordem de peso abaixo):

1. IMPACTO NA GERAÇÃO/EFICIÊNCIA DA USINA — o critério mais pesado. Qualquer atividade cuja descrição indique equipamento parado, desligado, offline, string sem corrente, inversor fora, usina total ou parcialmente fora de operação, deve subir para o topo, porque isso é dinheiro de geração perdido a cada hora que passa. Uma atividade cosmética ou de rotina (limpeza, organização, inspeção sem falha) nunca deve furar a frente de uma atividade que representa geração parada.

2. CRITICIDADE + URGÊNCIA DE PRAZO — cruze o campo "Prioridade" (Alta/Média/Baixa) com o quanto falta pro prazo (ou se já está vencido). Uma "Alta" vencida ou vencendo hoje/amanhã é mais urgente que uma "Alta" com prazo confortável. Prazo vencido não é sinônimo automático de mais urgente se o impacto na geração for baixo — pese os dois juntos.

3. DEPENDÊNCIA ENTRE ATIVIDADES — isso é essencial e frequentemente ignorado numa lista simples por prazo. Antes de finalizar a ordem, procure ativamente por atividades na MESMA usina (ou no mesmo equipamento) que tenham uma relação de pré-requisito entre si, e NUNCA sugira a atividade dependente antes da que ela depende. Exemplos do tipo de raciocínio esperado (use como padrão, não como lista fechada — identifique outros casos parecidos na lista real):
   - Recomposição/emenda de cabos SEMPRE antes de amarração/organização dos mesmos cabos (não dá pra organizar o que ainda não foi consertado).
   - Diagnóstico/inspeção de uma falha SEMPRE antes do reparo dela (não dá pra consertar sem saber a causa confirmada).
   - Reparo elétrico ou troca de componente SEMPRE antes de teste/comissionamento daquele circuito.
   - Verificação de aterramento/segurança SEMPRE antes de energizar ou religar o equipamento.
   - Troca de fusível ou proteção SEMPRE antes de testar a string/circuito protegido por ele.
   - Reparo físico/estrutural numa string ou trilho SEMPRE antes de limpeza de módulos naquele mesmo trecho (limpar antes só suja de novo durante o reparo).
   - Calibração de sensor SEMPRE depois de qualquer reparo físico no equipamento monitorado por ele (calibrar antes exige recalibrar depois).
   - Controle de vegetação/acesso ao redor de um equipamento SEMPRE antes de manutenção elétrica que exija acesso seguro àquele ponto.
   Se identificar uma dependência assim na lista, a atividade pré-requisito deve aparecer com prioridade igual ou maior que a dependente, mesmo que isoladamente pareça menos urgente — porque atrasá-la atrasa a outra também.

4. AGRUPAMENTO POR EQUIPE E USINA (redução de deslocamento) — REGRA RÍGIDA, do mesmo peso que os critérios 1-3, não é só desempate: uma mesma equipe/responsável ("Responsável/Equipe" na lista) NUNCA deve aparecer com atividades de USINAS DIFERENTES intercaladas na ordem final. Se a equipe "X" tem atividades em Usina A e também em Usina B, TODAS as atividades da equipe X na Usina A devem aparecer em posições consecutivas antes de qualquer atividade da equipe X na Usina B (ou vice-versa) — nunca A, depois B, depois A de novo. Use o campo "Cluster/Região" do mesmo jeito: atividades do mesmo cluster, mesmo que de responsáveis diferentes, também devem ficar próximas na lista sempre que os critérios 1-3 permitirem. Isso existe porque a ordem de prioridade também comunica o que a equipe deve fazer em sequência no mesmo deslocamento — intercalar usinas diferentes pra mesma equipe sugere um vai-e-vem fisicamente inviável no mesmo dia.

REGRA DE OURO PRA COMBINAR OS 4 CRITÉRIOS: primeiro ordene por impacto/prazo/dependência (1-3). DEPOIS, ao montar a lista final, reagrupe mantendo blocos contíguos por equipe+usina — dentro de um mesmo bloco, a ordem interna já definida pelos critérios 1-3 se mantém; entre blocos, o bloco com a atividade mais urgente (critérios 1-3) do grupo vem primeiro.

ATIVIDADES EM ABERTO HOJE:
{lista_atividades}

FORMATO DE SAÍDA (OBRIGATÓRIO):
Responda APENAS com um JSON válido (sem markdown, sem blocos de código com crase, sem texto antes ou depois), no formato:

{{
  "resumo_executivo": "2-3 frases dando o panorama geral do dia: quantas coisas críticas existem, algum padrão de dependência ou agrupamento geográfico relevante encontrado",
  "prioridades": [
    {{
      "posicao": <número da posição na fila, 1 = mais prioritário>,
      "id": "<id da atividade, exatamente como veio na lista>",
      "numeroOS": "<número da OS, exatamente como veio na lista>",
      "usina": "<usina>",
      "cluster": "<cluster/região>",
      "equipamento": "<equipamento/descrição resumida>",
      "motivo": "<justificativa curta e direta — cite o critério principal: impacto na geração, prazo, dependência de outra atividade (cite qual), ou agrupamento geográfico>"
    }}
  ],
  "mensagem_pronta": "<texto já formatado, pronto pra copiar e enviar, começando com um cabeçalho tipo '🎯 PRIORIDADES DE HOJE — <data>', listando as atividades em ordem numerada com usina + equipamento + motivo resumido em 1 linha cada, agrupadas visualmente por cluster quando fizer sentido. Use emojis moderadamente (🔴 pra crítico/geração parada, ⚠️ pra urgente por prazo, 📍 pra agrupamento geográfico). Máximo as 15 primeiras posições — se houver mais atividades, termine com uma linha tipo 'e mais N atividades de prioridade menor no painel'.>"
}}

Não invente atividades que não estão na lista. Inclua em "prioridades" todas as atividades recebidas, ordenadas do id 1 (mais prioritário) até a última — mas em "mensagem_pronta" mostre só o topo (até 15), como instruído."""


def _montar_prompt_reprogramacao(atividades, hoje_str, proximos_dias_uteis):
    # Correção 31/07/2026: usar o mesmo _equipe_label() (cluster, com
    # fallback pro responsável normalizado) já usado no resto do sistema
    # pra identificar a equipe — em vez do campo "responsavel" cru. Duas
    # pessoas da MESMA equipe física (ex.: Rodolfo Oliveira e Andrick
    # Gouveia, que atendem Boa Esperança/Ibaté juntos) apareciam com nomes
    # diferentes no campo responsavel, fazendo a IA achar que eram equipes
    # DIFERENTES e escalando a mesma equipe em usinas distintas no mesmo
    # dia — violando a REGRA MAIS IMPORTANTE do prompt.
    mapa_cluster = _mapa_cluster_usina()
    linhas = []
    for a in atividades:
        equipe = _equipe_label(a, mapa_cluster)
        linhas.append(
            f"- id={a['id']} | OS={a.get('numeroOS') or '—'} | Cliente={a['cliente']} | Usina={a['usina']} | "
            f"Ativo/Equipamento={a.get('equipamento') or '—'} | Ação/Tarefa={a.get('descricao') or '—'} | "
            f"Responsável={a.get('responsavel') or 'não informado'} | Equipe/Cluster (grupo físico de deslocamento)={equipe or 'não informado'} | "
            f"Prioridade={a.get('prioridade') or 'Média'} | Prazo atual={a.get('prazo') or 'sem prazo definido'} | "
            f"Status={a.get('status')}"
        )
    lista_atividades = "\n".join(linhas)
    lista_dias_uteis = "\n".join(f"- {data} ({nome})" for data, nome in proximos_dias_uteis)
    primeiro_dia = proximos_dias_uteis[0][0]
    primeiro_dia_nome = proximos_dias_uteis[0][1]

    return f"""Aja como um Programador(a) de Manutenção Sênior de uma empresa de O&M de usinas solares fotovoltaicas. Você é especialista em otimizar rotas e agendas de equipes de campo, minimizando deslocamento e maximizando produtividade.

CONTEXTO:
Hoje é {hoje_str}. Abaixo está a lista de atividades/OS em aberto que precisam ser reprogramadas para datas futuras.

DIAS ÚTEIS DISPONÍVEIS PRA REPROGRAMAR (já calculados, use SOMENTE essas datas — não calcule por conta própria, não use nenhuma data fora desta lista):
{lista_dias_uteis}

O primeiro dia útil disponível é {primeiro_dia} ({primeiro_dia_nome}) — a menos que os turnos desse dia já estejam no limite do critério 3 abaixo, ele DEVE ser usado por pelo menos uma equipe. Nunca pule esse primeiro dia sem necessidade real de agenda.

REGRA MAIS IMPORTANTE (NUNCA VIOLAR):
- Cada "Equipe/Cluster (grupo físico de deslocamento)" representa uma equipe de campo fisicamente alocada — TODOS os "Responsável" que compartilham o mesmo "Equipe/Cluster" são a MESMA equipe física (podem ser pessoas diferentes revezando ou trabalhando juntas na mesma van/rota), não equipes diferentes. Use o "Equipe/Cluster" pra decidir isso, NUNCA o "Responsável" isoladamente. Uma mesma equipe (mesmo Equipe/Cluster) NUNCA pode ter atividades programadas em USINAS DIFERENTES no mesmo dia — o deslocamento entre usinas inviabiliza isso. Se a equipe tem atividades em mais de uma usina, agrupe-as em dias diferentes, dedicando um ou mais dias consecutivos inteiros a cada usina antes de mover a equipe pra próxima.
- Atividades da MESMA equipe na MESMA usina podem (e devem, quando fizer sentido) ser agrupadas no mesmo dia ou em dias consecutivos, pra reduzir viagens.

REGRA FIXA DE DIA DA SEMANA (também NUNCA VIOLAR — tem prioridade sobre TODOS os outros critérios abaixo, incluindo o limite de quantidade por turno do critério 3): a equipe do Cláudio Ferreira (cluster CE Leste 01) tem dias fixos por usina, definidos pelo Fred (regra atualizada em 07/08/2026 — as usinas da GD Energy saíram desse cluster/equipe e foram pro cluster CE Norte 01, equipe do Felipe Xavier; não têm mais dia fixo aqui):
- Segunda-feira: usina ABC Morada Nova (cliente Alves Lima).
- Terça-feira: usinas Hortina e Vitesse (cliente Sal Energia).
- Quinta-feira: usina Sítio Bonfim (cliente Thopen).
- Sexta-feira: NÃO tem usina fixa. Olhe todas as atividades em aberto dessa equipe/cluster (as da lista abaixo) e identifique qual usina do cluster tem o maior backlog (mais atividades acumuladas/atrasadas) — encaixe as atividades dessa usina na sexta-feira. Explique esse raciocínio na "justificativa".
- Se não houver nenhuma data do dia da semana correspondente disponível na lista de dias úteis fornecida, escolha a data disponível mais próxima e explique isso claramente na "justificativa".
- CRÍTICO — ISSO SOBRESCREVE O LIMITE DE QUANTIDADE DO CRITÉRIO 3 PRA ESSA EQUIPE: como a equipe só visita cada uma dessas usinas UMA vez por semana (no dia fixo dela), TODA atividade em aberto daquela usina, não importa quantas sejam, tem que ir pro mesmo dia fixo — nunca espalhe pra outro dia da semana só porque "já tem muita coisa nesse turno". O limite de 1-2 atividades por turno do critério 3 NÃO se aplica a essa equipe. Dentro do dia fixo, distribua as atividades entre manhã e tarde de forma equilibrada (turno é só organização interna do dia, não motivo pra mudar de data).
- Essa regra vale só pra essa equipe/cluster específicos — não aplique padrão parecido pra outras equipes sem instrução explícita.

OUTROS CRITÉRIOS DE PRIORIZAÇÃO (em ordem de importância):
1. Atividades com prioridade "Alta" devem ser reprogramadas para as datas mais próximas possíveis.
2. Atividades que já estão com prazo vencido ou vencendo nos próximos dias têm urgência maior que as sem prazo definido ou com prazo distante.
3. SEJA CONSERVADOR NA QUANTIDADE POR DIA — isso é crítico. Grande parte dessas atividades já está atrasada justamente porque a agenda anterior foi otimista demais e não sobrou tempo real de execução, deslocamento dentro da própria usina, imprevistos e deslocamento até o próximo compromisso. Distribua no máximo 1 atividade por turno (manhã OU tarde) por equipe — ou seja, no máximo 2 atividades por dia por equipe — a menos que sejam claramente rápidas/simples (ex.: inspeção visual, verificação de temperatura), caso em que até 2 por turno é aceitável. Nunca mais que isso. EXCEÇÃO 1: esse limite NÃO vale pra equipe do Cláudio Ferreira (cluster CE Leste 01) — ver a REGRA FIXA DE DIA DA SEMANA acima, que tem prioridade sobre este critério. EXCEÇÃO 2 (vale pra QUALQUER equipe): se uma equipe tem atividades represadas em mais de uma usina e a semana de dias úteis disponível não tem dias suficientes pra dar um ou mais dias inteiros a cada usina respeitando esse limite de 2/dia, o limite de quantidade CEDE — nunca a REGRA MAIS IMPORTANTE (não-dupla-alocação de usina no mesmo dia). Nesse caso, é preferível colocar 3, 4 ou mais atividades da MESMA usina no MESMO dia (mesmo turno inclusive) do que dividir a semana igualmente entre usinas e acabar colocando a equipe em duas usinas diferentes num mesmo dia. Dedique dias inteiros e consecutivos a cada usina, na ordem de maior urgência/backlog primeiro, até esgotar os dias úteis disponíveis — não tente encaixar todas as usinas da equipe na mesma semana só pra "distribuir bonito"; é normal e esperado que uma usina com muito backlog fique pra semana seguinte.
4. REGRA RÍGIDA, SEM NENHUMA EXCEÇÃO: a "dataSugerida" de TODA atividade precisa ser uma das datas listadas em "DIAS ÚTEIS DISPONÍVEIS" acima. Nunca use uma data que não esteja nessa lista — ela já exclui sábados e domingos pra você.
5. Preencha os dias úteis mais próximos primeiro, na ordem em que aparecem na lista — não pule um dia disponível pra frente sem necessidade. Só avance pra um dia mais distante da lista quando os turnos dos dias mais próximos já estiverem no limite do critério 3.
6. Para cada atividade, defina também um TURNO (manhã ou tarde) dentro do dia sugerido, respeitando o limite de 1-2 atividades por turno do critério 3.

CAMPO "tarefa" NA SAÍDA (regra crítica, corrigida em 24/07/2026): cada atividade da lista abaixo tem dois campos distintos — "Ativo/Equipamento" é só o código fixo do equipamento na Fracttal (ex.: "IBT100-INVR1.8", "THPN-TPZ100-SSEG1-CMRA") e NÃO diz o que precisa ser feito; "Ação/Tarefa" é a descrição real do que deve ser executado naquele ativo. O campo "tarefa" da sua resposta deve vir do "Ação/Tarefa", reescrito em português claro e natural (curto, sem jargão de código) — NUNCA copie o código bruto do "Ativo/Equipamento" pra esse campo. O MESMO código de ativo pode aparecer em duas OSs diferentes com ações completamente diferentes (ex.: uma é "recomposição de câmera", outra é "instalação de câmera") — é normal, reflita a ação real de cada OS individualmente, não generalize pelo ativo. Se "Ação/Tarefa" vier vazio ou não fizer sentido, use algo genérico como "Verificação/manutenção conforme OS" em vez de inventar detalhes.

ATIVIDADES A REPROGRAMAR:
{lista_atividades}

FORMATO DE SAÍDA (OBRIGATÓRIO):
Responda APENAS com um JSON válido (sem markdown, sem blocos de código com crase, sem texto antes ou depois), no formato:

{{
  "resumo": "1-2 frases explicando a lógica geral usada no agrupamento",
  "reprogramacoes": [
    {{
      "id": "<id da atividade, exatamente como veio na lista>",
      "numeroOS": "<número da OS, exatamente como veio na lista>",
      "cliente": "<cliente, exatamente como veio na lista>",
      "usina": "<usina>",
      "equipamento": "<código do Ativo/Equipamento, exatamente como veio na lista — mantido só como referência>",
      "tarefa": "<descrição da ação/tarefa em português legível, baseada no campo Ação/Tarefa — NUNCA o código bruto do ativo; ver regra acima>",
      "responsavel": "<valor do campo Responsável, exatamente como veio na lista — NÃO o Equipe/Cluster>",
      "dataAtual": "<prazo atual, ou 'sem prazo definido'>",
      "dataSugerida": "<nova data sugerida, formato dd/mm/aaaa, OBRIGATORIAMENTE um dia de segunda a sexta-feira>",
      "turno": "<'manhã' ou 'tarde'>",
      "justificativa": "<motivo curto da escolha dessa data/turno, mencionando o agrupamento por usina/equipe quando relevante>"
    }}
  ]
}}

Não invente atividades que não estão na lista. Não omita nenhuma atividade da lista — toda atividade precisa aparecer em "reprogramacoes" com uma data sugerida (sempre dia útil) e um turno."""


# Mesmo mapeamento usado no frontend (_RESPONSAVEL_ALIASES em index.html)
# pra manter o rótulo de "equipe" consistente entre o modal de Comunicados
# e a Priorização IA — o mesmo técnico grafado de formas diferentes na
# Fracttal (apelido x nome completo) cai numa única entrada.
RESPONSAVEL_ALIASES = {
    "deivity jhon cunha saugo": "Deivity Saugo",
    "claudio ferreira": "Cláudio Ferreira",
    "valmir junior": "Valmir Júnior",
}


def _normalizar_responsavel(nome):
    limpo = re.sub(r"\s+", " ", (nome or "")).strip()
    if not limpo:
        return ""
    return RESPONSAVEL_ALIASES.get(limpo.lower(), limpo)


def _equipe_label(item, mapa_cluster):
    """Mesmo cálculo do frontend: a.cluster || responsável normalizado ||
    'Sem cluster'. Usado tanto pra exibir quanto pra filtrar por seleção."""
    cluster = mapa_cluster.get((item.get("usina") or "").strip(), "")
    if cluster:
        return cluster
    resp = _normalizar_responsavel(item.get("responsavel"))
    return resp or "Sem cluster"


@app.route("/sugerir-priorizacao-diaria", methods=["POST", "OPTIONS"])
def sugerir_priorizacao_diaria():
    """
    Usa o Gemini pra analisar as atividades em aberto e sugerir uma ordem
    de prioridade pro dia, considerando impacto na geração, criticidade/
    prazo, dependência entre atividades (ex.: recompor cabo antes de
    amarrar) e agrupamento geográfico pra reduzir deslocamento. Usado pelo
    botão "Sugerir Priorização (IA)" dentro do modal de Comunicados —
    gera uma mensagem separada, não mistura com o comunicado normal por
    usina.

    Aceita opcionalmente {"clusters": ["SP Centro 01", ...]} no corpo do
    POST — mesma seleção de checkboxes já usada nos Comunicados — pra
    restringir a análise só às equipes marcadas, em vez de misturar todas
    as usinas numa lista só (dificultava separar por técnico na hora de
    enviar). Sem esse campo (ou lista vazia), mantém o comportamento
    antigo de considerar todas as atividades — compatibilidade com
    chamadas antigas.
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500

    dados = request.get_json(force=True, silent=True) or {}
    clusters_selecionados = set(c.strip() for c in dados.get("clusters", []) if c and c.strip())

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    mapa_cluster = _mapa_cluster_usina()
    status_excluidos = {"concluído", "concluido", "cancelado", "convertida em ocorrência", "convertida em ocorrencia"}
    atividades = []
    for row in todos[1:]:
        if len(row) < len(ATIV_HEADERS_JSON):
            row = row + [""] * (len(ATIV_HEADERS_JSON) - len(row))
        item = dict(zip(ATIV_HEADERS_JSON, row[:len(ATIV_HEADERS_JSON)]))
        if not item.get("id"):
            continue
        if (item.get("status") or "").strip().lower() in status_excluidos:
            continue
        if (item.get("statusOS") or "").strip() == "Em Revisão":
            continue  # já foi feita, aguardando confirmação — não é prioridade de execução
        if clusters_selecionados and _equipe_label(item, mapa_cluster) not in clusters_selecionados:
            continue
        atividades.append(item)

    if not atividades:
        motivo = ("Nenhuma atividade em aberto nos clusters selecionados"
                   if clusters_selecionados else
                   "Nenhuma atividade em aberto encontrada pra priorizar")
        return jsonify({"ok": False, "error": motivo}), 400

    total_original = len(atividades)
    truncado = total_original > 70
    if truncado:
        # mesmo teto de segurança usado na reprogramação — prioriza uma
        # pré-seleção grosseira (Alta primeiro, prazo mais próximo) antes
        # de mandar pra IA, que então refina de verdade considerando
        # dependência e geografia.
        def _chave_urgencia(item):
            prioridade_peso = {"alta": 0, "média": 1, "media": 1, "baixa": 2}.get((item.get("prioridade") or "").strip().lower(), 1)
            prazo_str = (item.get("prazo") or "").strip()
            m = re.match(r"(\d{2})/(\d{2})/(\d{4})", prazo_str)
            prazo_ts = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).timestamp() if m else float("inf")
            return (prioridade_peso, prazo_ts)
        atividades = sorted(atividades, key=_chave_urgencia)[:70]

    hoje_str = agora_br().strftime('%d/%m/%Y (%A)')
    prompt = _montar_prompt_priorizacao(atividades, hoje_str)

    try:
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.25,
                    "maxOutputTokens": 24576,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=50,
            usar_chave_teste=(request.args.get("diagnostico", "").lower() == "true"),
        )
        data = resp.json()
        candidato = data["candidates"][0]
        finish_reason = candidato.get("finishReason", "")
        if finish_reason == "MAX_TOKENS":
            log.error(f"[sugerir-priorizacao] Resposta cortada por limite de tokens ({len(atividades)} atividades)")
            return jsonify({"ok": False, "error": ("A resposta da IA foi cortada por ser grande demais. "
                            "Tente novamente em instantes.")}), 502
        texto = candidato["content"]["parts"][0]["text"].strip()
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto.strip())
        sugestao = json.loads(texto_limpo)
        return jsonify({"ok": True, "truncado": truncado, "total_atividades": total_original,
                         "consideradas": len(atividades), **sugestao}), 200
    except requests.exceptions.HTTPError as e:
        log.error(f"[sugerir-priorizacao] Erro HTTP do Gemini: {e}")
        return jsonify({"ok": False, "error": f"Erro ao consultar a IA: {e}"}), 502
    except (KeyError, IndexError, json.JSONDecodeError) as e:
        log.error(f"[sugerir-priorizacao] Erro ao processar resposta do Gemini: {e}")
        return jsonify({"ok": False, "error": "A IA retornou uma resposta em formato inesperado. Tente novamente."}), 502
    except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
        log.error(f"[sugerir-priorizacao] Timeout/erro de conexão com a IA mesmo após retries: {e}")
        return jsonify({"ok": False, "error": ("A IA demorou demais para responder. Tente novamente em instantes "
                        "ou com menos atividades de uma vez.")}), 504
    except Exception as e:
        log.error(f"[sugerir-priorizacao] Erro inesperado: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/sugerir-reprogramacao", methods=["POST", "OPTIONS"])
def sugerir_reprogramacao():
    """
    Analisa as atividades em aberto (vindas da Fracttal ou não) e usa o
    Gemini pra sugerir uma reprogramação otimizada, respeitando que uma
    mesma equipe não pode ser escalada em usinas diferentes no mesmo dia
    (restrição de deslocamento). Usado pela aba "Reprogramações" do
    Painel de Atividades.
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500

    try:
        body = request.get_json(force=True) or {}
    except Exception:
        body = {}

    ids_filtro = set(str(x) for x in body.get("ids", [])) if body.get("ids") else None

    try:
        ws = get_atividades_sheet()
        todos = ws.get_all_values()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    status_excluidos = {"concluído", "concluido", "cancelado", "convertida em ocorrência", "convertida em ocorrencia"}
    atividades = []
    for row in todos[1:]:
        if len(row) < len(ATIV_HEADERS_JSON):
            row = row + [""] * (len(ATIV_HEADERS_JSON) - len(row))
        item = dict(zip(ATIV_HEADERS_JSON, row[:len(ATIV_HEADERS_JSON)]))
        if not item.get("id"):
            continue
        if (item.get("status") or "").strip().lower() in status_excluidos:
            continue
        # Camada extra de proteção (17/07/2026): "Em Revisão" é o estado
        # de uma OS já concluída em campo, só aguardando confirmação da
        # Fracttal — não faz sentido sugerir uma NOVA data pra ela. O
        # frontend já filtra isso antes de mandar os ids, mas replicar
        # aqui evita que um chamado direto ao endpoint (sem passar pelo
        # filtro do modal) traga OSs que não deveriam ser reprogramadas.
        if (item.get("statusOS") or "").strip().lower() in ("em revisão", "em revisao"):
            continue
        if ids_filtro is not None and item["id"] not in ids_filtro:
            continue
        atividades.append(item)

    if not atividades:
        return jsonify({"ok": False, "error": "Nenhuma atividade em aberto encontrada para reprogramar"}), 400
    total_original = len(atividades)

    # Reformulado em 31/07/2026 — arquitetura anterior mandava TODAS as
    # atividades (de todos os clusters/equipes) num prompt só, cortando
    # num teto global de 60 (por prioridade/prazo). Isso causava dois
    # problemas sérios: (1) usinas/clusters inteiros de menor urgência
    # (ex.: Araputanga, Sol do Norte I/II) ficavam de fora silenciosamente
    # quando o total geral passava de 60; (2) com dezenas de equipes
    # diferentes disputando espaço no mesmo prompt, a IA perdia o fio e
    # violava a regra de não-dupla-alocação (ex.: Ibaté I e Ibaté II, que
    # são a MESMA equipe, caindo no mesmo dia).
    #
    # Agora cada cluster/equipe (via _equipe_label, a mesma função usada
    # em todo o resto do sistema) vira uma chamada de IA SEPARADA, em
    # paralelo. Isso garante que: nenhuma equipe passa despercebida por
    # um corte global; a IA reprogramando uma equipe só vê as atividades
    # DAQUELA equipe (contexto muito menor e mais fácil de raciocinar
    # sobre conflitos de agenda); e o teto de 60 por chamada, que já era
    # generoso pro maior cluster observado (~20 atividades), praticamente
    # nunca é atingido na prática.
    mapa_cluster = _mapa_cluster_usina()
    grupos = {}
    for item in atividades:
        chave = _equipe_label(item, mapa_cluster) or "Sem cluster"
        grupos.setdefault(chave, []).append(item)

    def _chave_urgencia(item):
        prioridade_peso = {"alta": 0, "média": 1, "media": 1, "baixa": 2}.get((item.get("prioridade") or "").strip().lower(), 1)
        prazo_str = (item.get("prazo") or "").strip()
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", prazo_str)
        prazo_ts = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1))).timestamp() if m else float("inf")
        return (prioridade_peso, prazo_ts)

    grupos_truncados = []
    for nome, itens_grupo in grupos.items():
        if len(itens_grupo) > 60:
            grupos[nome] = sorted(itens_grupo, key=_chave_urgencia)[:60]
            grupos_truncados.append(nome)

    hoje_dt = agora_br()
    hoje_str = hoje_dt.strftime('%d/%m/%Y (%A)')
    proximos_dias_uteis = _proximos_dias_uteis(hoje_dt)
    diagnostico = request.args.get("diagnostico", "").lower() == "true"

    def _processar_grupo(nome_grupo, itens_grupo):
        prompt_grupo = _montar_prompt_reprogramacao(itens_grupo, hoje_str, proximos_dias_uteis)
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": [{"text": prompt_grupo}]}],
                "generationConfig": {
                    "temperature": 0.2,
                    "maxOutputTokens": 24576,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=60,
            usar_chave_teste=diagnostico,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        finish_reason = candidato.get("finishReason", "")
        if finish_reason == "MAX_TOKENS":
            raise RuntimeError(f"Resposta cortada por limite de tokens ({len(itens_grupo)} atividades)")
        texto = candidato["content"]["parts"][0]["text"].strip()
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto.strip())
        sugestao_grupo = json.loads(texto_limpo)
        _corrigir_fins_de_semana(sugestao_grupo)
        _comprimir_agenda_reprogramacao(sugestao_grupo, hoje_dt)
        return sugestao_grupo

    reprogramacoes_combinadas = []
    resumos = []
    erros_grupos = []
    with ThreadPoolExecutor(max_workers=6) as executor:
        futuros = {executor.submit(_processar_grupo, nome, itens): nome for nome, itens in grupos.items()}
        for futuro in as_completed(futuros):
            nome_grupo = futuros[futuro]
            try:
                sugestao_grupo = futuro.result()
                reprogramacoes_combinadas.extend(sugestao_grupo.get("reprogramacoes", []))
                if sugestao_grupo.get("resumo"):
                    resumos.append(f"{nome_grupo}: {sugestao_grupo['resumo']}")
            except Exception as e:
                log.error(f"[sugerir-reprogramacao] Erro no grupo '{nome_grupo}': {e}")
                erros_grupos.append(f"{nome_grupo}: {e}")

    if not reprogramacoes_combinadas and erros_grupos:
        return jsonify({"ok": False, "error": "Falha ao gerar sugestão pra todos os grupos: " + "; ".join(erros_grupos)}), 502

    for item in reprogramacoes_combinadas:
        item["cluster"] = mapa_cluster.get((item.get("usina") or "").strip(), "")

    sugestao = {"resumo": " | ".join(resumos), "reprogramacoes": reprogramacoes_combinadas}
    avisos = []
    if grupos_truncados:
        avisos.append("Cortado em 60 atividades (limite por chamada) pra: " + ", ".join(grupos_truncados))
    if erros_grupos:
        avisos.append("Falha ao processar: " + "; ".join(erros_grupos))

    return jsonify({
        "ok": True, "sugestao": sugestao, "total_atividades": len(reprogramacoes_combinadas),
        "total_original": total_original, "truncado": bool(grupos_truncados),
        "avisos": avisos or None,
    })


@app.route("/gerar-texto-os-ia", methods=["POST", "OPTIONS"])
def gerar_texto_os_ia():
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido"}), 400

    prompt = _montar_prompt_os(body)
    diagnostico = request.args.get("diagnostico", "").lower() == "true"
    try:
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.3,
                    "maxOutputTokens": 4096,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=25,
            usar_chave_teste=diagnostico,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        finish_reason = candidato.get("finishReason", "")
        texto_bruto = candidato["content"]["parts"][0]["text"].strip()
        if not texto_bruto or len(texto_bruto) < 20:
            log.error(f"[gerar-texto-os-ia] Resposta curta/vazia (finishReason={finish_reason}): {texto_bruto!r}")
            raise ValueError(f"Resposta incompleta da IA (finishReason={finish_reason or 'desconhecido'})")

        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
        try:
            parsed = json.loads(texto_limpo)
            textos = parsed.get("textos") or []
            textos = [t.strip() for t in textos if t and t.strip()]
        except (json.JSONDecodeError, AttributeError):
            # fallback: se a IA não devolveu o JSON esperado por algum
            # motivo, trata a resposta inteira como um texto único —
            # evita quebrar a funcionalidade por causa de um formato
            # inesperado pontual.
            textos = [texto_bruto]

        if not textos:
            raise ValueError("A IA não retornou nenhum texto de OS")

        # "texto" continua existindo (primeiro item) pra não quebrar quem
        # já usava o formato antigo; "textos" é a lista completa, usada
        # quando a solicitação foi dividida em mais de uma OS.
        resultado = {"ok": True, "texto": textos[0], "textos": textos}
        if diagnostico:
            resultado["chave_teste_configurada"] = bool(GEMINI_API_KEY_TESTE)
        return jsonify(resultado)
    except requests.exceptions.HTTPError as e:
        if e.response is not None and e.response.status_code == 429:
            log.error(f"[gerar-texto-os-ia] Cota da IA esgotada mesmo apos retries: {e}")
            return jsonify({"ok": False, "error": ("A IA está temporariamente sem cota disponível (uso "
                            "excessivo em pouco tempo). Aguarde alguns minutos e tente de novo.")}), 429
        log.error(f"[gerar-texto-os-ia] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500
    except Exception as e:
        log.error(f"[gerar-texto-os-ia] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _extrair_texto_sobreaviso(file_storage, filtro_aba="FRED ALEXANDRINO"):
    """Extrai a escala de sobreaviso (planilha 'SOBREAVISOS_EQUIPES_GRID') em
    texto compacto pra usar como contexto na auditoria de ponto. A planilha
    tem uma aba por supervisor, cada uma com colunas por cluster/usina e
    linhas por semana, indicando qual técnico está de sobreaviso naquela
    semana/fim de semana/feriado.

    Por padrão processa só a(s) aba(s) do supervisor Fred Alexandrino
    (filtro_aba), já que o Controle de Ponto audita apenas as equipes dele
    — as demais abas (Vitor, Danuth, Marcelo, Camila, Pedro) são de outros
    supervisores e não devem entrar no contexto.

    IMPORTANTE: cada plantão de fim de semana ocupa 2 linhas na planilha
    (sábado + domingo), mas o nome do técnico só é preenchido na linha do
    sábado — a linha do domingo vem em branco (é o mesmo plantão, só não
    repetem o nome). Por isso, quando uma célula está vazia, herdamos o
    último valor preenchido daquela coluna (carry-forward), pra que a data
    de domingo também apareça associada ao técnico correto no texto."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(file_storage.read()), data_only=True)
    blocos = []
    filtro_norm = (filtro_aba or "").strip().upper()
    for nome_aba in wb.sheetnames:
        if filtro_norm and filtro_norm not in nome_aba.strip().upper():
            continue
        ws = wb[nome_aba]
        linhas = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, linha in enumerate(linhas):
            textos = [str(c).upper() for c in linha if c]
            if any("SEMANA" in t for t in textos) and any("DIAS" in t for t in textos):
                header_idx = i
                break
        if header_idx is None:
            continue
        header = linhas[header_idx]
        colunas = [(j, str(header[j]).replace("\n", " / ")) for j in range(3, len(header)) if header[j]]
        saida = [f"=== Supervisor: {nome_aba} ==="]
        ultimo_valor = {}
        for linha in linhas[header_idx + 1:]:
            data_val = linha[2] if len(linha) > 2 else None
            if not data_val:
                continue
            data_str = data_val.strftime("%d/%m/%Y") if hasattr(data_val, "strftime") else str(data_val)
            partes_linha = []
            for j, nome_col in colunas:
                val = linha[j] if j < len(linha) else None
                if val:
                    ultimo_valor[j] = val
                else:
                    val = ultimo_valor.get(j)
                if val:
                    partes_linha.append(f"{nome_col}: {str(val).replace(chr(10), ', ')}")
            if partes_linha:
                saida.append(f"{data_str} -> " + "; ".join(partes_linha))
        if len(saida) > 1:
            blocos.append("\n".join(saida))
    texto = "\n\n".join(blocos)
    # a escala é ordenada cronologicamente (mais antiga primeiro), então se
    # precisar truncar por tamanho, corta do INÍCIO — o período mais recente
    # (relevante pro espelho de ponto analisado) fica no fim do texto.
    limite = 250000
    if len(texto) > limite:
        texto = "[...início do texto truncado por tamanho...]\n" + texto[-limite:]
    return texto


def _montar_prompt_controle_ponto(textos_extraidos, texto_escala_sobreaviso=None, orientacao_adicional=None):
    """Monta o prompt de auditoria de ponto (setor Controle de Ponto,
    dentro do Painel Gerencial). textos_extraidos: lista de strings, uma
    por PDF enviado (cada PDF do Pontomais pode conter 1 ou vários
    colaboradores, um por página/bloco 'Colaborador: NOME'). Opcionalmente
    recebe a escala de sobreaviso (planilha) pra cruzar quem estava
    realmente escalado nos fins de semana/feriados analisados, e uma
    orientação adicional em texto livre escrita pelo Fred na hora — que
    COMPLEMENTA as regras fixas abaixo, sem substituí-las, a não ser que
    ele diga expressamente o contrário."""
    corpo = "\n\n=== NOVO ARQUIVO ===\n\n".join(textos_extraidos)
    bloco_escala = ""
    instrucao_etapa3_extra = ""
    if texto_escala_sobreaviso:
        bloco_escala = f"\n\nESCALA DE SOBREAVISO (referência — quem estava escalado em cada semana/cluster):\n{texto_escala_sobreaviso}\n"
        instrucao_etapa3_extra = (
            " Use a ESCALA DE SOBREAVISO fornecida pra cruzar: se o colaborador tem marcações em um "
            "fim de semana/feriado, verifique se ele estava de fato escalado naquela semana (pelo nome "
            "dele na escala, considerando variações de nome/apelido). Se ele trabalhou sem estar "
            "escalado, ou se um colaborador escalado não tem nenhuma marcação nem foi substituído, "
            "registre isso no campo \"anomalias_sobreaviso\" como uma observação (não necessariamente "
            "uma falta grave, apenas evidencie a divergência pro RH avaliar)."
        )
    bloco_orientacao = ""
    if orientacao_adicional and orientacao_adicional.strip():
        bloco_orientacao = (
            "\n\nORIENTAÇÃO ADICIONAL DO SUPERVISOR (Fred) PRA ESTA ANÁLISE:\n"
            "As instruções abaixo foram escritas por Fred especificamente pra esta rodada de análise. "
            "Elas COMPLEMENTAM as etapas 1 a 5 e o formato de resposta definidos acima — aplique-as "
            "junto com as regras fixas, sem descartar nenhuma etapa. Só ignore ou substitua alguma "
            "regra fixa se o texto abaixo disser isso de forma clara e explícita (ex.: \"não considere "
            "a etapa 2 desta vez\", \"ignore ajustes por esquecimento do colaborador X\"). Em qualquer "
            "outro caso, trate como um complemento às regras já definidas.\n\n"
            f"\"\"\"\n{orientacao_adicional.strip()}\n\"\"\"\n"
        )
    instrucoes = (
        "Aja como um auditor técnico de Departamento Pessoal especializado em controle de jornada. "
        "Analise o(s) espelho(s) de ponto fornecido(s) abaixo, aplicando estritamente as regras de "
        "negócio descritas. Estruture seu raciocínio de forma sequencial e entregue um diagnóstico "
        "direto, sem introduções, apenas as pendências reais.\n\n"
        "REGRA CRÍTICA (leia antes de tudo): cada linha do espelho de ponto tem colunas de horário "
        "(1ª Entrada, 1ª Saída, 2ª Entrada, 2ª Saída) e uma coluna separada \"Motivo/Observação\" que "
        "só existe quando algum ajuste foi feito. O texto da coluna Motivo/Observação descreve o que "
        "aconteceu ANTES do ajuste (ex.: \"Ajuste | Ponto não registrou no retorno da refeição\") — ele "
        "é HISTÓRICO, não o estado atual. O ESTADO ATUAL da jornada é definido apenas pelos horários "
        "que aparecem nas colunas de ponto. Se as 4 colunas de horário têm valor preenchido (mesmo que "
        "um deles tenha sido inserido via ajuste), a jornada daquele dia está COMPLETA — NÃO é uma "
        "pendência, mesmo que o texto do Motivo/Observação mencione \"não registrou\", \"faltou\" ou "
        "termo parecido. NUNCA infira uma jornada incompleta a partir do texto do Motivo/Observação — "
        "infira sempre a partir da presença ou ausência real de horário nas colunas de ponto.\n\n"
        "Etapa 1 - Verificação de dias úteis: analise as marcações de segunda a sexta-feira. Uma jornada "
        "só é \"incompleta\" quando, no estado ATUAL do espelho, falta literalmente um horário em uma "
        "das colunas de ponto (célula vazia — nenhum valor, nem original nem ajustado). Se todas as "
        "colunas relevantes do dia têm horário preenchido, o dia está regularizado e NÃO deve ser "
        "listado aqui, independentemente do que o Motivo/Observação diga. Para cada jornada realmente "
        "incompleta, registre a data e qual marcação especificamente está faltando.\n"
        "Etapa 2 - Auditoria de justificativas de ajuste: faça a varredura da coluna Motivo/Observação "
        "de dias que JÁ FORAM ajustados (ou seja, o dia está completo, mas houve um ajuste). Liste "
        "APENAS os casos em que a justificativa é literalmente sobre o colaborador ter esquecido de "
        "bater o ponto — variações de \"esqueci de bater o ponto\", \"esqueceu de bater o ponto\", "
        "\"esquecimento\". NÃO enquadre aqui justificativas de causa técnica/sistêmica como \"não "
        "registrou\", \"aplicativo travou\", \"erro do app\", \"sem internet\", \"bateria descarregada\", "
        "\"app não estava abrindo\", \"celular travou\" — essas vão exclusivamente na Etapa 4, nunca "
        "aqui, mesmo que o resultado prático (um horário faltando) seja parecido. Na dúvida sobre se "
        "é esquecimento do colaborador ou falha técnica, NÃO classifique como esquecimento.\n"
        "Etapa 3 - Validação de sobreaviso em finais de semana e feriados: a ausência total de "
        "marcações em sábados, domingos e feriados indica que o colaborador não foi acionado e "
        "DISPENSA o registro de ponto — NÃO reporte isso como pendência. Aponte anomalias nesses dias "
        "exclusivamente se houver início de jornada sem a respectiva conclusão (marcação em aberto)."
        f"{instrucao_etapa3_extra}\n"
        "Etapa 4 - Mapeamento de falhas sistêmicas: agrupe os colaboradores que registraram "
        "inconsistências técnicas (falha de aplicativo, travamento, falta de internet, bateria "
        "descarregada, \"não registrou\" sem menção a esquecimento), indicando a frequência desses "
        "eventos. Isso é apenas informativo pro RH — mesmo com frequência alta, NÃO conta como "
        "pendência do colaborador e NÃO gera comunicado sozinho.\n"
        "Etapa 5 - Um colaborador só tem \"tem_pendencia\": true e só recebe comunicado individual se "
        "tiver AO MENOS UM item real na Etapa 1 (jornada com horário literalmente faltando, sem "
        "ajuste) e/ou na Etapa 2 (ajuste com justificativa de esquecimento do próprio colaborador). "
        "As Etapas 3 e 4 são contexto informativo e NUNCA, sozinhas, tornam \"tem_pendencia\" true nem "
        "geram comunicado. Quando houver pendência real, gere um texto de comunicado individual pronto "
        "para copiar e enviar por WhatsApp diretamente a esse colaborador — tom profissional, direto e "
        "respeitoso, citando as datas e pendências específicas dele (sem citar os demais colaboradores) "
        "e solicitando a regularização. Colaboradores sem pendência real (mesmo que tenham itens nas "
        "Etapas 3/4) não recebem comunicado — deixe o campo comunicado_individual como string vazia.\n\n"
        "Responda ESTRITAMENTE em JSON, sem nenhum texto fora do JSON, no formato:\n"
        "{\"colaboradores\": [{"
        "\"nome\": string, "
        "\"tem_pendencia\": boolean, "
        "\"jornadas_incompletas\": [{\"data\": string, \"descricao\": string}], "
        "\"ajustes_indevidos\": [{\"data\": string, \"motivo\": string}], "
        "\"anomalias_sobreaviso\": [{\"data\": string, \"descricao\": string}], "
        "\"falhas_sistemicas\": {\"frequencia\": number, \"descricao\": string}, "
        "\"resumo_tecnico\": string, "
        "\"comunicado_individual\": string"
        "}]}\n\n"
        "Linguagem técnica, concisa e estruturada por colaborador, evidenciando apenas pendências "
        "reais que demandam ação do setor de Recursos Humanos."
        f"{bloco_escala}"
        f"{bloco_orientacao}\n"
        f"ESPELHOS DE PONTO:\n{corpo}"
    )
    return instrucoes


@app.route("/analisar-controle-ponto", methods=["POST", "OPTIONS"])
def analisar_controle_ponto():
    """Setor 'Controle de Ponto' do Painel Gerencial: recebe um ou mais
    PDFs de espelho de ponto (Pontomais), extrai o texto e manda pra IA
    auditar conforme as regras de negócio, retornando pendências por
    colaborador + texto de comunicado individual pronto pra copiar."""
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500

    arquivos = request.files.getlist("arquivos")
    if not arquivos:
        return jsonify({"ok": False, "error": "Nenhum arquivo enviado (campo 'arquivos')"}), 400

    try:
        import pdfplumber
    except ImportError:
        log.error("[analisar-controle-ponto] pdfplumber não instalado")
        return jsonify({"ok": False, "error": "Dependência pdfplumber não instalada no servidor"}), 500

    texto_escala = None
    arquivo_escala = request.files.get("planilha_sobreaviso")
    if arquivo_escala and arquivo_escala.filename:
        try:
            texto_escala = _extrair_texto_sobreaviso(arquivo_escala)
        except ImportError:
            log.error("[analisar-controle-ponto] openpyxl não instalado")
            return jsonify({"ok": False, "error": "Dependência openpyxl não instalada no servidor"}), 500
        except Exception as e:
            log.error(f"[analisar-controle-ponto] Erro extraindo planilha de sobreaviso: {e}")
            return jsonify({"ok": False, "error": f"Erro ao ler a planilha de sobreaviso: {e}"}), 400

    textos = []
    for f in arquivos:
        nome_arquivo = secure_filename(f.filename or "arquivo.pdf")
        try:
            with pdfplumber.open(f.stream) as pdf:
                paginas = [(p.extract_text() or "") for p in pdf.pages]
            textos.append(f"[Arquivo: {nome_arquivo}]\n" + "\n".join(paginas))
        except Exception as e:
            log.error(f"[analisar-controle-ponto] Erro extraindo '{nome_arquivo}': {e}")
            return jsonify({"ok": False, "error": f"Erro ao ler o PDF \"{nome_arquivo}\": {e}"}), 400

    prompt = _montar_prompt_controle_ponto(
        textos,
        texto_escala_sobreaviso=texto_escala,
        orientacao_adicional=request.form.get("orientacao_adicional"),
    )
    diagnostico = request.args.get("diagnostico", "").lower() == "true"
    try:
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.2,
                    "maxOutputTokens": 8192,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=55,
            usar_chave_teste=diagnostico,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        finish_reason = candidato.get("finishReason", "")
        texto_bruto = candidato["content"]["parts"][0]["text"].strip()
        if not texto_bruto:
            log.error(f"[analisar-controle-ponto] Resposta vazia (finishReason={finish_reason}): {texto_bruto!r}")
            raise ValueError(f"Resposta vazia da IA (finishReason={finish_reason or 'desconhecido'})")
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
        resultado = json.loads(texto_limpo)
        if finish_reason == "MAX_TOKENS":
            resultado["truncado"] = True
        return jsonify({"ok": True, "resultado": resultado})
    except json.JSONDecodeError as e:
        log.error(f"[analisar-controle-ponto] Resposta não é JSON válido: {e}")
        return jsonify({"ok": False, "error": ("A IA retornou um formato inesperado. Tente novamente com menos "
                        "arquivos de uma vez.")}), 502
    except requests.exceptions.HTTPError as e:
        if e.response is not None and e.response.status_code == 429:
            log.error(f"[analisar-controle-ponto] Cota da IA esgotada mesmo apos retries: {e}")
            return jsonify({"ok": False, "error": ("A IA está temporariamente sem cota disponível. Aguarde "
                            "alguns minutos e tente de novo.")}), 429
        log.error(f"[analisar-controle-ponto] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500
    except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
        log.error(f"[analisar-controle-ponto] Timeout/erro de conexão com a IA: {e}")
        return jsonify({"ok": False, "error": ("A IA demorou demais para responder. Tente novamente com "
                        "menos arquivos de uma vez.")}), 504
    except Exception as e:
        log.error(f"[analisar-controle-ponto] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/converter-ocorrencia-em-atividade", methods=["POST", "OPTIONS"])
def converter_ocorrencia_em_atividade():
    """
    Converte uma Ocorrência em uma Atividade: cria uma nova linha no Painel de
    Atividades com os dados da ocorrência (incluindo o histórico cronológico
    transferido), e marca a ocorrência original como "Convertida em Atividade".
    """
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido"}), 400

    ocorrencia_id = str(body.get("id", "")).strip()
    editor = body.get("editor", "dashboard").strip()
    if not ocorrencia_id:
        return jsonify({"ok": False, "error": "id é obrigatório"}), 400

    try:
        ws_falhas = get_sheet()
        todos_falhas = ws_falhas.get_all_values()
        linha_idx = None
        linha_atual = None
        for i, row in enumerate(todos_falhas[1:], start=2):
            if row and str(row[0]).strip() == ocorrencia_id:
                linha_idx = i
                linha_atual = row
                break
        if not linha_idx:
            return jsonify({"ok": False, "error": "ocorrência não encontrada"}), 404

        # linha_atual: [ID, Cliente, Usina, Equipamento, Falha, Causa, Impactados, Ação,
        #               Status, Ticket, NumeroOS, Historico, ...]
        cliente     = linha_atual[1]  if len(linha_atual) > 1  else ""
        usina       = linha_atual[2]  if len(linha_atual) > 2  else ""
        equipamento = linha_atual[3]  if len(linha_atual) > 3  else ""
        falha       = linha_atual[4]  if len(linha_atual) > 4  else ""
        causa       = linha_atual[5]  if len(linha_atual) > 5  else ""
        acao        = linha_atual[7]  if len(linha_atual) > 7  else ""
        status_ocorr= linha_atual[8]  if len(linha_atual) > 8  else ""
        numero_os   = linha_atual[10] if len(linha_atual) > 10 else ""
        historico_ocorr = linha_atual[11] if len(linha_atual) > 11 else ""

        descricao = falha or "Sem descrição"
        if causa:
            descricao += f" — Causa: {causa}"

        nota_conversao = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Convertida do Painel de "
                           f"Falhas (Ocorrência #{ocorrencia_id}) por {_editor_legivel(editor)}.")
        historico_atividade = nota_conversao
        if acao:
            historico_atividade += f"\nAção registrada na ocorrência: {acao}"
        if historico_ocorr:
            historico_atividade += "\n" + historico_ocorr

        status_atividade = status_ocorr if status_ocorr and status_ocorr.lower() not in (
            "concluído", "concluido", "convertida em atividade") else "Em Aberto"

        ws_ativ = get_atividades_sheet()
        todos_ativ = ws_ativ.get_all_values()
        novo_id_atividade = _proximo_id_atividade(todos_ativ)
        agora = agora_br().strftime('%d/%m/%Y %H:%M:%S')

        ws_ativ.append_row([novo_id_atividade, cliente, usina, equipamento, descricao, "", "",
                             "Média", status_atividade, agora, "", historico_atividade, editor, numero_os])

        # Marca a ocorrência original como convertida
        ws_falhas.update_cell(linha_idx, 9, "Convertida em Atividade")  # coluna I = Status
        entry = (f"{agora_br().strftime('%d/%m/%Y %H:%M')} - Convertida em atividade "
                 f"#{novo_id_atividade} por {_editor_legivel(editor)}.")
        novo_hist_ocorr = f"{historico_ocorr}\n{entry}".strip() if historico_ocorr else entry
        ws_falhas.update_cell(linha_idx, 12, novo_hist_ocorr)  # coluna L = Historico

        log.info(f"[converter-ocorrencia] Ocorrência #{ocorrencia_id} -> Atividade #{novo_id_atividade}")
        return jsonify({"ok": True, "novaAtividadeId": novo_id_atividade})
    except Exception as e:
        log.error(f"[converter-ocorrencia-em-atividade] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/gerar-relatorio-semanal", methods=["POST", "OPTIONS"])
def gerar_relatorio_semanal_route():
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
        cliente = str(body.get("cliente", "")).strip()
        data_inicio = datetime.strptime(body["dataInicio"], "%Y-%m-%d")
        data_fim = datetime.strptime(body["dataFim"], "%Y-%m-%d").replace(
            hour=23, minute=59, second=59
        )
        if not cliente:
            return jsonify({"ok": False, "error": "cliente e obrigatorio"}), 400

        ws_atividades = get_atividades_sheet()
        todos_atividades = carregar_planilha(ws_atividades)

        atividades_por_usina, desligamentos_por_usina, rondas_por_usina = coletar_atividades_e_desligamentos_por_usina(
            todos_atividades, cliente, data_inicio, data_fim)

        try:
            usinas_cliente = sorted(set(listar_usinas_cliente(todos_atividades, cliente)))
        except Exception as e:
            log.error(f"[Relatorio Semanal] Erro ao listar usinas do cliente: {e}")
            usinas_cliente = sorted(set(atividades_por_usina) | set(desligamentos_por_usina) | set(rondas_por_usina))

        if not usinas_cliente:
            return jsonify({"ok": False, "error": "Nenhuma usina encontrada para esse cliente"}), 404

        semana_num = data_fim.isocalendar()[1]
        data_label = data_fim.strftime('%d/%m/%Y')

        # Zeladoria: preenche a página com os dados reais do Painel de
        # Zeladoria. Se der qualquer erro (aba fora do ar, etc.), o
        # relatório inteiro não pode falhar por causa disso -- cai pro
        # comportamento antigo (página sai com "Em acompanhamento.").
        try:
            ws_zeladoria = get_zeladoria_sheet()
            todos_zeladoria = carregar_planilha(ws_zeladoria)
            zeladoria_status_por_usina = montar_status_zeladoria_por_usina(todos_zeladoria, cliente)
        except Exception as e:
            log.error(f"[Relatorio Semanal] Erro ao buscar dados de Zeladoria: {e}")
            zeladoria_status_por_usina = None

        # Chamados com Fabricante: preenche a página com os dados reais da
        # aba ChamadosFabricante (mesma fonte do Painel de Chamados do
        # dashboard). Se der qualquer erro, o relatório inteiro não pode
        # falhar por causa disso -- cai pro comportamento de "nenhum
        # chamado em aberto" (mesmo padrão de robustez da Zeladoria).
        try:
            chamados_fabricante_por_usina = coletar_chamados_fabricante_por_usina(
                _chamados_fabricante_itens(), cliente)
        except Exception as e:
            log.error(f"[Relatorio Semanal] Erro ao buscar Chamados com Fabricante: {e}")
            chamados_fabricante_por_usina = None

        buf = gerar_relatorio_pptx(cliente, semana_num, data_label,
                                    atividades_por_usina, desligamentos_por_usina, usinas_cliente,
                                    zeladoria_status_por_usina, rondas_por_usina,
                                    chamados_fabricante_por_usina)

        nome_arquivo = f"Apresentação {cliente} x Grid Co - O&M - Semana {semana_num}.pptx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        )
    except Exception as e:
        log.error(f"[Relatorio Semanal] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Relatório de Handover (por OS, .docx, padrão visual Grid Co.) ──────────

def _buscar_atividade_por_numero_os(numero_os):
    """
    Procura uma atividade do Painel de Atividades pelo número da OS.
    Aceita OS compostas (ex. "8467/9035") — bate se numero_os for
    qualquer um dos números separados por "/", mesma lógica usada na
    busca global do topbar do dashboard.
    """
    ws = get_atividades_sheet()
    todos = ws.get_all_values()
    for row in todos[1:]:
        if not row or not row[0].strip():
            continue
        if len(row) < len(ATIV_HEADERS_JSON):
            row = row + [""] * (len(ATIV_HEADERS_JSON) - len(row))
        item = dict(zip(ATIV_HEADERS_JSON, row[:len(ATIV_HEADERS_JSON)]))
        numeros = [n.strip() for n in (item.get("numeroOS") or "").split("/")]
        if numero_os in numeros:
            return item
    return None


def _gerar_resumo_handover_ia(atividade):
    """
    Resumo executivo formal (1-2 parágrafos) para o Relatório de Handover
    do cliente, gerado pela API da Anthropic a partir dos dados da OS.
    Nunca derruba o relatório: se a chave não estiver configurada ou a
    chamada falhar, retorna "" e a seção sai omitida no documento.
    """
    if not ANTHROPIC_API_KEY:
        return ""

    system_prompt = (
        "Você é um engenheiro de O&M de usinas solares fotovoltaicas da Grid Co., "
        "redigindo o resumo executivo de um Relatório de Handover formal para o "
        "cliente, referente ao fechamento de uma Ordem de Serviço.\n\n"
        "Escreva 1 a 2 parágrafos curtos, em português formal e técnico, terceira "
        "pessoa, sem saudações nem despedidas — apenas o corpo do resumo. Descreva "
        "o que foi identificado, o que foi executado e o resultado/status final, "
        "com base exclusivamente nos dados fornecidos. Nunca invente informações "
        "que não estejam nos dados (se um dado não vier informado, simplesmente "
        "não o mencione). Não use bullets — texto corrido, separando parágrafos "
        "com uma linha em branco."
    )
    user_content = (
        f"Cliente: {atividade.get('cliente','') or 'não informado'}\n"
        f"Usina: {atividade.get('usina','') or 'não informado'}\n"
        f"Equipamento: {atividade.get('equipamento','') or 'não informado'}\n"
        f"Nº OS: {atividade.get('numeroOS','') or 'não informado'}\n"
        f"Descrição da OS: {atividade.get('descricao','') or 'não informado'}\n"
        f"Status atual: {atividade.get('statusOS') or atividade.get('status','') or 'não informado'}\n"
        f"Observações: {atividade.get('observacoesOS','') or 'nenhuma'}\n"
        f"Histórico cronológico:\n{atividade.get('historico','') or 'sem histórico registrado'}"
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 500,
                "system": system_prompt,
                "messages": [{"role": "user", "content": user_content}],
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        for block in data.get("content", []):
            if block.get("type") == "text":
                return block.get("text", "").strip()
        return ""
    except Exception as e:
        log.error(f"[Handover] Erro ao gerar resumo IA: {e}")
        return ""


@app.route("/gerar-relatorio-handover", methods=["POST", "OPTIONS"])
def gerar_relatorio_handover_route():
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        body = request.get_json(force=True) or {}
        numero_os = str(body.get("numeroOS", "")).strip()
        if not numero_os:
            return jsonify({"ok": False, "error": "numeroOS é obrigatório"}), 400

        # Preferência: usa a atividade já carregada no dashboard (o frontend
        # já tem os dados, vindos do /atividades) — evita reler a planilha
        # inteira do Sheets a cada geração de handover, o que já causou
        # 429 (quota de leitura excedida) quando o dashboard estava com
        # bastante atividade simultânea. Só cai pro fallback via Sheets se
        # o chamador não enviar a atividade (compatibilidade).
        atividade = body.get("atividade")
        if not isinstance(atividade, dict) or not atividade:
            atividade = _buscar_atividade_por_numero_os(numero_os)
        if not atividade:
            return jsonify({"ok": False,
                             "error": f"OS {numero_os} não encontrada no Painel de Atividades."}), 404

        resumo_ia = _gerar_resumo_handover_ia(atividade)
        buf = gerar_handover_docx(atividade, resumo_ia)

        cliente_slug = re.sub(r"[^A-Za-z0-9]+", "", atividade.get("cliente", "") or "") or "GridCo"
        nome_arquivo = f"Handover_OS_{numero_os.replace('/', '-')}_{cliente_slug}_GridCo.pdf"

        log.info(f"[Relatorio Handover] Gerado para OS {numero_os} ({atividade.get('cliente','')})")
        return send_file(
            buf,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/pdf",
        )
    except Exception as e:
        log.error(f"[Relatorio Handover] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def _montar_prompt_punchlist(os_lista):
    """
    Monta o prompt pra IA sugerir itens de Punch List a partir dos dados
    já disponíveis no Painel de Atividades para as OSs selecionadas
    (descrição, histórico, observações, status por equipamento).

    IMPORTANTE — limitação conhecida: a API da Fracttal não expõe os itens
    individuais do checklist de campo (as perguntas com resposta
    Aprovou/Alerta/Falhou), só o status agregado por equipamento. Então
    isso é um RASCUNHO a partir do que está registrado no painel — não
    substitui a leitura do PDF exportado da Fracttal, que tem o checklist
    completo. Por isso a instrução é: só apontar pendência quando há
    evidência textual real (observação, nota de status diferente de
    concluído, ressalva no histórico); nunca inventar uma anormalidade
    plausível só porque o equipamento existe.
    """
    blocos = []
    for os_item in os_lista:
        bloco = (
            f"OS {os_item.get('numeroOS','?')} — {os_item.get('equipamento','?')}\n"
            f"Descrição: {os_item.get('descricao','') or 'não informada'}\n"
            f"Status: {os_item.get('statusOS') or os_item.get('status','') or 'não informado'}\n"
            f"Observações registradas: {os_item.get('observacoesOS','') or 'nenhuma'}\n"
            f"Detalhes por equipamento: {os_item.get('detalhesEquipamentosOS','') or 'não informado'}\n"
            f"Histórico:\n{os_item.get('historico','') or 'sem histórico'}"
        )
        blocos.append(bloco)
    dados_os = "\n\n---\n\n".join(blocos)

    return (
        "Você é um engenheiro de O&M de usinas solares fotovoltaicas da Grid Co., revisando "
        "Ordens de Serviço de handover para montar a Punch List (lista de pendências) de um "
        "Relatório de Handover formal para o cliente.\n\n"
        "Analise os dados abaixo, extraídos do Painel de Atividades para as OSs selecionadas. "
        "Aponte como item de punch list SOMENTE quando houver evidência textual real de "
        "pendência, anormalidade, ressalva ou item não concluído NO EQUIPAMENTO OU NA "
        "INSTALAÇÃO FÍSICA — em uma observação, nota de status, ou menção explícita no "
        "histórico. NUNCA invente uma anormalidade plausível só porque o equipamento existe ou "
        "porque handovers costumam ter pendências. Se os dados de uma OS mostram tudo "
        "concluído, sem observações, sem ressalvas, NÃO gere nenhum item de punch list para "
        "ela — isso é o resultado correto quando o handover foi limpo, não uma falha da "
        "análise.\n\n"
        "IGNORE completamente avisos administrativos/de sistema que não são sobre o "
        "equipamento em si — por exemplo, linhas de auditoria interna do painel como "
        "\"técnico não está mapeado para a usina\", alertas de roteamento/atribuição de "
        "responsável, ou qualquer nota que fale sobre cadastro/sistema em vez de sobre o "
        "estado físico do ativo. Esse tipo de nota nunca vira item de punch list, mesmo que "
        "pareça uma 'pendência' textualmente — punch list é sobre o ativo físico, não sobre "
        "o cadastro no painel.\n\n"
        "Retorne APENAS um JSON (sem markdown, sem texto fora do JSON) no formato:\n"
        '{"itens": [{"ativo": "<nome do equipamento/ativo>", '
        '"criticidade": "<Baixa|Média|Alta|Muito Alta>", '
        '"anormalidade": "<descrição objetiva da pendência, baseada só no texto fornecido>", '
        '"recomendacoes": "<ação recomendada, objetiva>", '
        '"responsavel": "<EQUIPE TÉCNICA ou o responsável mencionado, se houver>"}]}\n\n'
        "Se nenhuma OS tiver pendência real identificável, retorne {\"itens\": []}.\n\n"
        f"DADOS DAS OSs:\n\n{dados_os}"
    )


@app.route("/gerar-punchlist-ia", methods=["POST", "OPTIONS"])
def gerar_punchlist_ia():
    """
    Sugere itens de Punch List via IA a partir dos dados já registrados no
    Painel de Atividades para as OSs marcadas no formulário de Handover de
    Usina. É um RASCUNHO — o usuário revisa/edita antes de gerar o PDF
    final. Não acessa o checklist detalhado da Fracttal (ver limitação no
    docstring de _montar_prompt_punchlist).
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500
    try:
        body = request.get_json(force=True) or {}
        os_lista = body.get("os", [])
        if not os_lista:
            return jsonify({"ok": False, "error": "Nenhuma OS informada"}), 400

        prompt = _montar_prompt_punchlist(os_lista)
        resp = _chamar_gemini_com_retry(
            {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.2,
                    "maxOutputTokens": 2048,
                    "responseMimeType": "application/json",
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            },
            timeout=25,
        )
        data = resp.json()
        candidato = data["candidates"][0]
        texto_bruto = candidato["content"]["parts"][0]["text"].strip()
        texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
        parsed = json.loads(texto_limpo)
        itens = parsed.get("itens", [])

        # normaliza campos pro formato que o frontend já usa na punch list
        itens_normalizados = []
        for it in itens:
            itens_normalizados.append({
                "ativo": (it.get("ativo") or "").strip(),
                "criticidade": it.get("criticidade") if it.get("criticidade") in
                               ("Baixa", "Média", "Alta", "Muito Alta") else "Média",
                "status": "PENDENTE",
                "anormalidade": (it.get("anormalidade") or "").strip(),
                "recomendacoes": (it.get("recomendacoes") or "").strip(),
                "responsavel": (it.get("responsavel") or "EQUIPE TÉCNICA").strip(),
            })
        return jsonify({"ok": True, "itens": itens_normalizados})
    except requests.exceptions.HTTPError as e:
        if e.response is not None and e.response.status_code == 429:
            return jsonify({"ok": False, "error": "IA temporariamente sem cota. Tente novamente em instantes."}), 429
        log.error(f"[gerar-punchlist-ia] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500
    except Exception as e:
        log.error(f"[gerar-punchlist-ia] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Punch List a partir do PDF ORIGINAL da Fracttal (leitura visual) ──────
#
# Diferente de /gerar-punchlist-ia (que só analisa o resumo já salvo no
# Painel de Atividades), este pipeline lê o PDF que a própria Fracttal
# exporta pra OS — o mesmo formato "N.º: xxxx / Ordem de Trabalho" com o
# checklist de subtarefas (cada uma com uma marcação visual Aprovou/
# Alerta/Falhou) e as anotações de texto do técnico de campo.
#
# A marcação do checkbox é só visual (não existe um jeito confiável de
# extrair "qual das 3 opções foi marcada" como texto puro do PDF — os
# três rótulos "Aprovou Alerta Falhou" sempre aparecem, só muda o
# preenchimento do quadradinho) — por isso essa etapa usa IA com visão
# (Gemini multimodal), não só texto.
#
# Pipeline por ativo (equipamento):
#   1. Localizar os limites de página de cada ativo no PDF (via texto).
#   2. Dentro desse intervalo, achar só as páginas que têm o checklist
#      (contêm "Aprovou" e "Falhou") — normalmente 1-2 páginas de cada
#      ativo, o resto do intervalo é fotos de evidência que não precisam
#      ser lidas por IA.
#   3. Renderizar só essas páginas como imagem e mandar pra Gemini junto
#      com as anotações de texto do técnico (extraídas normalmente, sem
#      IA) — pede pra consolidar num item de punch list SÓ se houver
#      Alerta/Falhou real.
#   4. Ativos inteiramente "Aprovou" não geram nenhum item — resultado
#      correto, não falha da leitura.

_OCR_FRACTTAL_CACHE = {}  # md5(pdf_bytes) -> {"tem_texto": bool, "textos": {pagina_idx: texto}}
_OCR_FRACTTAL_LIMITE_PAGINAS = 40  # medido na prática: ~2s/página mesmo com 4 threads em
# paralelo — acima disso o OCR sozinho já estoura o orçamento de tempo da requisição
# (130s), sem sobrar tempo nenhum pras chamadas de visão que vêm depois.


def _fracttal_pdf_texto_paginas(pdf_bytes, indices):
    """
    Texto de cada página pedida em `indices` (0-indexed). Usa a camada
    de texto normal do PDF quando ela existe; se o documento INTEIRO
    não tiver texto nenhum — alguns exports da Fracttal saem como
    imagem pura (visto na prática: uma OS exportada assim tinha uma
    única imagem por página, sem nenhum texto por trás, enquanto a
    maioria sai com texto real) — faz OCR via tesseract (português) só
    nas páginas realmente pedidas, com cache por conteúdo do PDF pra
    não repetir trabalho entre as várias chamadas da mesma requisição
    (extrair_ativos precisa do documento inteiro; paginas_checklist e
    notas_ativo só do intervalo de cada ativo, que essa altura já
    estará cacheado pela primeira passada).

    Levanta RuntimeError se o documento não tiver texto E tiver mais
    páginas que _OCR_FRACTTAL_LIMITE_PAGINAS — OCR síncrono nesse
    volume arrisca estourar o timeout do Gunicorn (160s) numa VM com
    pouca RAM; melhor falhar rápido com mensagem clara do que travar a
    requisição sem retorno nenhum pro usuário.
    """
    import hashlib
    chave = hashlib.md5(pdf_bytes).hexdigest()

    if chave not in _OCR_FRACTTAL_CACHE:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            total = len(pdf.pages)
            amostra_texto = [(pdf.pages[i].extract_text() or "") for i in range(min(5, total))]
        tem_texto = any(len(t.strip()) > 20 for t in amostra_texto)
        if not tem_texto and total > _OCR_FRACTTAL_LIMITE_PAGINAS:
            raise RuntimeError(
                f"Esse PDF foi exportado como imagem (sem texto real) e tem {total} páginas — "
                f"acima do limite de {_OCR_FRACTTAL_LIMITE_PAGINAS} páginas que conseguimos "
                f"processar via OCR numa única requisição sem travar o servidor. Tente reexportar "
                f"essa OS da Fracttal (o export padrão costuma sair com texto real, bem mais rápido "
                f"de processar), ou peça pra dividir/reduzir esse PDF antes de anexar."
            )
        _OCR_FRACTTAL_CACHE[chave] = {"tem_texto": tem_texto, "total": total, "textos": {}}

    estado = _OCR_FRACTTAL_CACHE[chave]
    faltando = [i for i in indices if 0 <= i < estado["total"] and i not in estado["textos"]]

    if faltando:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            if estado["tem_texto"]:
                for i in faltando:
                    estado["textos"][i] = pdf.pages[i].extract_text() or ""
            else:
                log.info(f"[Fracttal PDF] OCR em {len(faltando)} página(s) (export sem texto/imagem).")
                import pytesseract
                imagens = {i: pdf.pages[i].to_image(resolution=150).original for i in faltando}

                def _ocr_pagina(i):
                    try:
                        return i, pytesseract.image_to_string(imagens[i], lang="por")
                    except Exception as e:
                        log.error(f"[Fracttal PDF OCR] Erro na página {i}: {e}")
                        return i, ""

                with ThreadPoolExecutor(max_workers=4) as ex:
                    for i, texto in ex.map(_ocr_pagina, faltando):
                        estado["textos"][i] = texto

    return {i: estado["textos"].get(i, "") for i in indices}


def _fracttal_pdf_extrair_ativos(pdf_bytes):
    """Retorna lista de {nome, pagina_inicio, pagina_fim} (0-indexed,
    pagina_fim inclusive) — um item por ativo/equipamento do PDF.

    BUG CORRIGIDO (11/08/2026): quando dois ativos começam na MESMA
    página (comum em ativos com pouco conteúdo, ex. "Fossa Séptica"
    seguido de "Infraestrutura Civil" na mesma página), a versão antiga
    só detectava o PRIMEIRO marcador "ATIVOS" de cada página (tinha um
    `break` que saía do loop de linhas assim que achava o primeiro) —
    o segundo ativo simplesmente sumia, e suas páginas ficavam
    erradamente atribuídas ao ativo anterior. Confirmado comparando a
    extração com punch lists reais: a pendência de credenciais de
    acesso (que é do ativo "Infraestrutura Civil") aparecia grudada
    nas notas da "Fossa Séptica".

    Também usa _fracttal_pdf_texto_paginas (com fallback de OCR) em vez
    de extract_text() direto — alguns exports da Fracttal saem como
    imagem pura, sem texto nenhum, e sem isso o resultado é sempre
    lista vazia (erro visto na prática: "Não consegui identificar
    nenhum ativo nesse PDF" mesmo em PDFs válidos)."""
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        total_paginas = len(pdf.pages)
    textos = _fracttal_pdf_texto_paginas(pdf_bytes, range(total_paginas))

    marcadores = []
    for i in range(total_paginas):
        text = textos[i]
        if "ATIVOS" not in text:
            continue
        lines = text.split("\n")
        for j, l in enumerate(lines):
            if l.strip() != "ATIVOS":
                continue
            for k in range(j + 1, min(j + 4, len(lines))):
                if lines[k].strip().startswith("DESCRIÇÃO:"):
                    nome = lines[k].replace("DESCRIÇÃO:", "").strip().split("{")[0].strip()
                    if nome and "Realizado com" not in nome and "Pág." not in nome:
                        marcadores.append((i, nome))
                    break

    ativos = []
    for idx, (pagina_inicio, nome) in enumerate(marcadores):
        if idx + 1 < len(marcadores):
            proximo_inicio = marcadores[idx + 1][0]
            # Se o próximo ativo começa na MESMA página, esse ativo
            # fica só com essa página (melhor granularidade possível
            # sem rastrear posição vertical dentro da página).
            pagina_fim = proximo_inicio if proximo_inicio == pagina_inicio else proximo_inicio - 1
        else:
            pagina_fim = total_paginas - 1
        ativos.append({"nome": nome, "pagina_inicio": pagina_inicio, "pagina_fim": pagina_fim})
    return ativos


def _fracttal_pdf_paginas_checklist(pdf_bytes, pagina_inicio, pagina_fim):
    """Páginas (0-indexed) dentro do intervalo que têm o checklist com
    marcação Aprovou/Alerta/Falhou, OU a tabela/nota de anexos — essas
    últimas entram também porque em PDFs sem texto (OCR) não dá pra
    confiar 100% na extração de tabela pra notas (ver
    _fracttal_pdf_notas_ativo); mandando a página pra visão, a IA
    consegue ler o texto da nota direto na imagem."""
    indices = list(range(pagina_inicio, pagina_fim + 1))
    textos = _fracttal_pdf_texto_paginas(pdf_bytes, indices)
    paginas = []
    for i in indices:
        text = textos.get(i, "")
        if ("Aprovou" in text and "Falhou" in text) or "ANEXOS DO PLANO DE MANUTENÇÃO" in text \
                or re.search(r"(?m)^\s*Nota\b", text):
            paginas.append(i)
    return paginas


def _fracttal_pdf_notas_ativo(pdf_bytes, pagina_inicio, pagina_fim):
    """Extrai as anotações de texto do técnico dentro do intervalo do
    ativo: a tabela 'ANEXOS DO PLANO DE MANUTENÇÃO' (Descrição/Detalhes)
    e o campo solto 'Observações' (aparece fora de tabela, ex.:
    "Observações UFV não tem sistema de ar-condicionado").

    BUG CORRIGIDO (11/08/2026, achado comparando com punch lists reais
    geradas via Gemini a partir dos mesmos PDFs): quando a tabela
    'ANEXOS DO PLANO DE MANUTENÇÃO' é longa, ela quebra em várias
    páginas — o pdfplumber devolve cada pedaço como uma tabela
    SEPARADA, e só a do início tem a linha de título. A versão antiga
    exigia esse título em CADA fragmento, então os fragmentos de
    continuação (sem título, só dados) eram descartados inteiros — e é
    exatamente nas continuações que ficam as notas mais importantes
    (ex.: "6 câmeras de CFTV inoperantes" só aparecia na 3ª página da
    tabela). Corrigido com um flag que persiste entre páginas dentro do
    intervalo do ativo.

    Cada página tem tabelas distintas: 'SUBTAREFAS' (pergunta do
    checklist + opções Aprovou/Alerta/Falhou — nunca é fonte de nota)
    e 'ANEXOS DO PLANO DE MANUTENÇÃO' (Descrição/Detalhes — as notas
    de verdade, só quando o técnico escreveu algo).

    IMPORTANTE: vários achados (ex. "valores implausíveis no medidor",
    "ruído anormal no ventilador") só existem como marcação VISUAL do
    checkbox Aprovou/Alerta/Falhou, sem nenhuma nota de texto — esses só
    são capturados pela leitura visual (ver _montar_prompt_punchlist_visao
    e _fracttal_pdf_paginas_checklist), não por esta função.
    """
    notas = []
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        fim = min(pagina_fim, len(pdf.pages) - 1)
        dentro_anexos = False
        for i in range(pagina_inicio, fim + 1):
            texto_pagina = pdf.pages[i].extract_text() or ""
            for linha_txt in texto_pagina.split("\n"):
                linha_txt = linha_txt.strip()
                if linha_txt.startswith("Observações") and linha_txt != "Observações":
                    obs = linha_txt.replace("Observações", "", 1).strip()
                    if obs and obs.upper() not in ("N/A", "-", "NENHUM", "NENHUMA") and obs not in notas:
                        notas.append(obs)

            for tabela in (pdf.pages[i].extract_tables() or []):
                if not tabela or not tabela[0]:
                    continue
                primeira = [str(c or "").strip() for c in tabela[0]]
                # Início de uma tabela SUBTAREFAS (3 colunas) — nunca é
                # fonte de nota, e fecha qualquer "ANEXOS" em aberto.
                if len(tabela[0]) >= 3 or primeira[0] == "SUBTAREFAS":
                    dentro_anexos = False
                    continue
                linhas = tabela
                if "ANEXOS DO PLANO DE MANUTENÇÃO" in primeira[0]:
                    dentro_anexos = True
                    linhas = tabela[1:]  # pula só a linha de título; o
                    # resto (cabeçalho + dados) pode estar na MESMA
                    # tabela quando ela não quebra de página
                if not dentro_anexos:
                    continue
                for linha in linhas:
                    if not linha or len(linha) < 2:
                        continue
                    col0 = str(linha[0] or "").strip()
                    col1 = str(linha[1] or "").strip()
                    if col0 == "Descrição" and col1 == "Detalhes":
                        continue  # cabeçalho das colunas
                    if col1 and col1.upper() not in ("N/A", "-", "NENHUM", "NENHUMA", "N/OK", "OK"):
                        if col1 not in notas:
                            notas.append(col1)
    return notas


def _fracttal_pdf_renderizar_paginas(pdf_bytes, indices_0based, dpi=110):
    """Renderiza só as páginas pedidas (1-indexed pro poppler) como PNG
    base64 — DPI moderado de propósito, a VM1 tem só 1GB de RAM."""
    imagens = []
    for idx in indices_0based:
        paginas = convert_from_bytes(pdf_bytes, dpi=dpi, first_page=idx + 1, last_page=idx + 1)
        for p in paginas:
            buf = BytesIO()
            p.save(buf, format="PNG")
            imagens.append(base64.b64encode(buf.getvalue()).decode("utf-8"))
            buf.close()
            p.close()
    return imagens


def _montar_prompt_punchlist_visao(nome_ativo, cliente, usina, cluster, notas_texto):
    notas_bloco = ("\n".join(f"- {n}" for n in notas_texto)) if notas_texto else "(nenhuma anotação de texto registrada)"
    return (
        "Você é um engenheiro de O&M de usinas solares fotovoltaicas da Grid Co., revisando "
        "o checklist de uma Ordem de Serviço de handover pra montar a Punch List (lista de "
        "pendências) de um Relatório de Handover formal pro cliente.\n\n"
        f"As imagens anexadas são páginas do checklist de subtarefas do ativo \"{nome_ativo}\" "
        f"(usina {usina}, cliente {cliente}, cluster {cluster}). Cada linha do checklist tem "
        "três opções — Aprovou / Alerta / Falhou — com um quadradinho marcado (preenchido) "
        "indicando qual foi escolhida. Leia CADA linha e identifique quais foram marcadas "
        "como Alerta ou Falhou (ignore as marcadas como Aprovou — essas estão OK, sem "
        "pendência). Algumas páginas podem mostrar a tabela de anotações do técnico "
        "('ANEXOS DO PLANO DE MANUTENÇÃO' ou linhas começando com 'Nota') em vez do "
        "checklist — leia esse texto também, ele é uma fonte de pendência tão válida quanto "
        "o checkbox marcado.\n\n"
        f"Anotações de texto que o técnico de campo registrou pra esse ativo (podem ou não "
        f"se referir a um item específico do checklist):\n{notas_bloco}\n\n"
        "Se TODOS os itens do checklist foram marcados Aprovou (nenhum Alerta/Falhou), "
        "retorne {\"itens\": []} — esse é o resultado correto, não uma falha.\n\n"
        "Se houver item(ns) Alerta/Falhou, consolide em UM item de punch list pra esse ativo "
        "(uma linha só, juntando as anormalidades encontradas numa frase objetiva — igual ao "
        "padrão Grid Co., ex.: \"Multimedidor inoperante, ruídos anormais e falha na "
        "iluminação de emergência\"). Só em casos onde há dois problemas claramente "
        "independentes e de natureza muito diferente, pode gerar mais de um item.\n\n"
        "Critérios de CRITICIDADE: Alta/Muito Alta para falhas funcionais, de segurança ou "
        "que impedem operação (equipamento fora de operação, falha de proteção, ausência de "
        "supervisório); Média para itens de manutenção/limpeza/cosmético (sujidade, pequenos "
        "reparos); Baixa para observações menores.\n\n"
        "Critérios de RESPONSÁVEL: \"EQUIPE TÉCNICA\" para reparos elétricos/mecânicos; "
        "\"EQUIPE DE CAMPO\" para limpeza/organização/civil; \"CLIENTE/SUPERVISÃO\" quando o "
        "problema depende de sistema supervisório, contratação externa, credenciais de acesso, "
        "ou está fora do escopo de campo da Grid Co.; \"FABRICANTE\" quando exige garantia ou "
        "peça/serviço do fabricante do equipamento (ex.: fonte interna de TCU/NCU); "
        "\"FABRICANTE/TÉCNICA\" quando pode precisar de qualquer um dos dois, a depender do "
        "diagnóstico.\n\n"
        "No campo \"ativo\", use uma CATEGORIA padrão (o tipo de equipamento), NUNCA o nome "
        "nem número específico da Fracttal — por exemplo, escreva \"Inversores\" mesmo que o "
        "ativo se chame \"Inversor 1.3 Huawei\", porque outros inversores da mesma usina podem "
        "gerar itens separados que serão consolidados depois numa linha só por categoria. "
        "Categorias padrão (use uma destas sempre que fizer sentido; se nenhuma servir, use uma "
        "categoria curta e clara própria): Ar Condicionado, Cabine de Medição, Caixa d'água, "
        "Estação Meteorológica, Fossa Séptica, CFTV / Segurança, Infraestrutura Civil, "
        "Infraestrutura / Supervisório, Inversores, Trackers, Módulos Fotovoltaicos, QGBT, "
        "SPDA, Nobreak e Banco de Baterias, Transformador de Potência, Relé de Proteção, "
        "Sistema de Drenagem, Sistema de Combate a Incêndio.\n\n"
        "Retorne APENAS um JSON (sem markdown, sem texto fora do JSON) no formato:\n"
        '{"itens": [{"ativo": "<categoria padrão, sem número de unidade>", '
        '"criticidade": "<Baixa|Média|Alta|Muito Alta>", '
        '"anormalidade": "<descrição objetiva e consolidada>", '
        '"recomendacoes": "<ação corretiva recomendada, objetiva>", '
        '"responsavel": "<EQUIPE TÉCNICA|EQUIPE DE CAMPO|CLIENTE/SUPERVISÃO|FABRICANTE|FABRICANTE/TÉCNICA>"}]}'
    )


def _consolidar_itens_punchlist_por_categoria(itens):
    """
    Agrupa itens de punch list que vieram de ativos diferentes mas caem
    na mesma categoria (ex.: 8 chamadas separadas pros ativos "Inversor
    1.1".."Inversor 1.8" da Fracttal, cada uma podendo gerar seu próprio
    item porque _montar_prompt_punchlist_visao já pede pra IA usar a
    categoria — "Inversores" — em vez do nome/número específico do
    ativo). Sem essa consolidação, o punch list final ficaria com uma
    linha por unidade em vez de uma linha por categoria, diferente do
    padrão real da Grid Co. (confirmado comparando com punch lists reais
    geradas via Gemini a partir dos mesmos PDFs — sempre 1 linha por
    categoria, nunca 1 por inversor/tracker individual).

    Critério de junção: chave = (usina, categoria) normalizada
    (case-insensitive). anormalidade/recomendações são concatenadas;
    criticidade fica a mais alta do grupo; responsável vira a união dos
    valores distintos (ex.: "EQUIPE TÉCNICA / FABRICANTE").
    """
    ORDEM_CRITICIDADE = {"baixa": 0, "média": 1, "media": 1, "alta": 2, "muito alta": 3}
    grupos = {}
    ordem_grupos = []
    for item in itens:
        chave = ((item.get("usina") or "").strip().lower(), (item.get("ativo") or "").strip().lower())
        if chave not in grupos:
            grupos[chave] = []
            ordem_grupos.append(chave)
        grupos[chave].append(item)

    consolidados = []
    for chave in ordem_grupos:
        grupo = grupos[chave]
        if len(grupo) == 1:
            consolidados.append(grupo[0])
            continue
        base = dict(grupo[0])
        anormalidades = [g.get("anormalidade", "").strip() for g in grupo if g.get("anormalidade", "").strip()]
        recomendacoes = [g.get("recomendacoes", "").strip() for g in grupo if g.get("recomendacoes", "").strip()]
        responsaveis = list(dict.fromkeys(g.get("responsavel", "").strip() for g in grupo if g.get("responsavel", "").strip()))
        criticidade_max = max(grupo, key=lambda g: ORDEM_CRITICIDADE.get((g.get("criticidade") or "").strip().lower(), 1))
        base["anormalidade"] = "; ".join(dict.fromkeys(anormalidades))
        base["recomendacoes"] = "; ".join(dict.fromkeys(recomendacoes))
        base["responsavel"] = " / ".join(responsaveis) if responsaveis else base.get("responsavel", "")
        base["criticidade"] = criticidade_max.get("criticidade", base.get("criticidade"))
        base["ressalva"] = any(g.get("ressalva") for g in grupo)
        consolidados.append(base)
    return consolidados


def _gerar_punchlist_ativo_via_visao(nome_ativo, cliente, usina, cluster, imagens_base64, notas_texto):
    """Chama a Gemini com as imagens do checklist + notas de um ativo,
    devolve lista de itens de punch list (pode ser vazia)."""
    if not imagens_base64:
        return []
    prompt = _montar_prompt_punchlist_visao(nome_ativo, cliente, usina, cluster, notas_texto)
    parts = [{"text": prompt}]
    for img_b64 in imagens_base64:
        parts.append({"inline_data": {"mime_type": "image/png", "data": img_b64}})
    resp = _chamar_gemini_com_retry(
        {
            "contents": [{"parts": parts}],
            "generationConfig": {
                "temperature": 0.15,
                "maxOutputTokens": 1024,
                "responseMimeType": "application/json",
                "thinkingConfig": {"thinkingBudget": 0},
            },
        },
        timeout=40,
    )
    data = resp.json()
    candidato = data["candidates"][0]
    texto_bruto = candidato["content"]["parts"][0]["text"].strip()
    texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
    parsed = json.loads(texto_limpo)
    itens = parsed.get("itens", [])
    normalizados = []
    for it in itens:
        normalizados.append({
            "ativo": (it.get("ativo") or nome_ativo).strip(),
            "criticidade": it.get("criticidade") if it.get("criticidade") in
                           ("Baixa", "Média", "Alta", "Muito Alta") else "Média",
            "status": "PENDENTE",
            "anormalidade": (it.get("anormalidade") or "").strip(),
            "recomendacoes": (it.get("recomendacoes") or "").strip(),
            "responsavel": (it.get("responsavel") or "EQUIPE TÉCNICA").strip(),
        })
    return normalizados


GEMINI_PDF_MAX_PAGINAS = 60    # limite de segurança, raramente é o fator decisivo (ver MAX_MB abaixo)
GEMINI_PDF_MAX_MB = 7          # reduzido de 14 pra 7 (11/08/2026): pedaços grandes demais faziam a
# IA perder pendências reais no meio do documento (confirmado por revisão humana página a
# página — SCADA, Relé, QGBT, EPCs, Incêndio inteiros ficaram de fora da punch list em dois
# relatórios reais). Como os PDFs da Fracttal costumam ser pesados em imagem, o limite de MB
# (não o de páginas) é o que de fato determina o tamanho de cada pedaço — reduzir aqui é o
# que realmente força pedaços menores/mais focados. Custo: mais chamadas e mais tempo total.
# 7MB de arquivo bruto vira ~9.3MB em base64, com boa margem de segurança


def _dividir_pdf_em_chunks(pdf_bytes, max_paginas=GEMINI_PDF_MAX_PAGINAS, max_mb=GEMINI_PDF_MAX_MB):
    """Divide o PDF em pedaços que respeitam o limite de payload inline
    da API da Gemini (~20MB após base64 — usamos 14MB de arquivo bruto
    de margem). Cada pedaço é um PDF válido (páginas na ordem original)."""
    reader = PdfReader(BytesIO(pdf_bytes))
    total = len(reader.pages)
    chunks = []
    inicio = 0
    while inicio < total:
        fim = min(inicio + max_paginas, total)
        while True:
            writer = PdfWriter()
            for i in range(inicio, fim):
                writer.add_page(reader.pages[i])
            buf = BytesIO()
            writer.write(buf)
            dados = buf.getvalue()
            if len(dados) <= max_mb * 1024 * 1024 or fim - inicio <= 5:
                break
            fim = inicio + max(5, (fim - inicio) // 2)
        chunks.append({"pagina_inicio": inicio, "pagina_fim": fim - 1, "bytes": dados})
        inicio = fim
    return chunks


def _montar_prompt_punchlist_pdf_nativo(cliente, usina, cluster, pagina_inicio, pagina_fim, total_paginas):
    aviso_trecho = ""
    if total_paginas > (pagina_fim - pagina_inicio + 1):
        aviso_trecho = (
            f"Este é o trecho de página {pagina_inicio + 1} a {pagina_fim + 1} de um documento "
            f"de {total_paginas} páginas no total (dividido em partes por causa do limite de "
            "tamanho da API). Só ignore um ativo se a seção 'ATIVOS' dele estiver GENUINAMENTE "
            "incompleta nesse trecho (ex.: começa nas últimas linhas da última página, sem "
            "nenhuma subtarefa visível ainda) — nesse caso raro, ele será coberto por completo "
            "em outro trecho. Não use isso como desculpa pra pular ativos só porque estão perto "
            "do início ou fim do trecho: se você consegue ver a seção 'ATIVOS' completa (nome + "
            "todas as subtarefas + notas), analise ela normalmente, mesmo que seja a primeira ou "
            "última do trecho.\n\n"
        )
    return (
        "Você é um engenheiro de O&M de usinas solares fotovoltaicas da Grid Co., analisando "
        "um PDF exportado da Fracttal (Ordem de Trabalho de handover) pra montar a Punch List "
        "(lista de pendências) de um Relatório de Handover FORMAL pro cliente — esse documento "
        "vai ser assinado e entregue de verdade. Uma punch list incompleta gera retrabalho "
        "sério (o coordenador do Fred já rejeitou relatórios duas vezes por punch list "
        "incompleta, com achados reais de SCADA, Relé de Proteção, QGBT, EPCs de segurança, "
        "Combate a Incêndio e SPDA faltando). É MUITO melhor incluir um item que depois se "
        "mostre menor do que deixar de fora uma pendência real — não subestime a quantidade "
        "de itens esperados: é normal e comum um documento ter pendências em VÁRIOS sistemas "
        "diferentes ao mesmo tempo (elétrico, civil, segurança, supervisório), não só em 1 ou 2.\n\n"
        f"Esse PDF documenta a inspeção de vários ativos/equipamentos da usina {usina} "
        f"(cliente {cliente}, cluster {cluster}). Cada ativo tem uma seção 'ATIVOS' (com o "
        "campo DESCRIÇÃO identificando o equipamento) seguida de um checklist de subtarefas "
        "com três opções — Aprovou / Alerta / Falhou — marcadas com um quadradinho preenchido "
        "indicando qual foi escolhida, e pode ter uma seção de anotações do técnico "
        "('ANEXOS DO PLANO DE MANUTENÇÃO' ou linhas 'Nota: ...'). Leia o conteúdo visual das "
        "páginas (o documento pode ser texto real ou imagem escaneada — leia do mesmo jeito).\n\n"
        f"{aviso_trecho}"
        "PROCESSO SISTEMÁTICO (siga isso pra CADA seção 'ATIVOS' que aparecer nas páginas deste "
        "trecho, uma de cada vez, sem pular nenhuma — inclusive ativos que pareçam menos "
        "críticos à primeira vista, como Relé de Proteção, QGBT, EPCs, Sistema de Combate a "
        "Incêndio, Sistema Supervisório/SCADA — esses já ficaram de fora em análises anteriores "
        "e são exatamente os que mais importam pro cliente aprovar o relatório):\n"
        "1. Identifique o ativo pelo campo DESCRIÇÃO dessa seção.\n"
        "2. Leia TODAS as subtarefas do checklist (Aprovou/Alerta/Falhou) desse ativo, uma por "
        "uma, sem pular nenhuma — mesmo que a lista seja longa (5, 7, 10+ subtarefas). Se o "
        "ativo tem MÚLTIPLAS subtarefas marcadas Falhou/Alerta, capture TODAS elas na "
        "anormalidade (ex.: se o Sistema Supervisório reprovou alarmes, memória, comunicação, "
        "medições E câmeras — as 5 entram na mesma frase, não só a primeira que você notar). "
        "Leia também as anotações do técnico associadas (ANEXOS DO PLANO DE MANUTENÇÃO / "
        "'Nota: ...').\n"
        "3. Classifique o resultado em um destes três casos:\n"
        "   a) Tudo Aprovado, sem nenhuma nota indicando problema ou ausência → NÃO gere item "
        "nenhum de punch list pra esse ativo.\n"
        "   b) Algum item Alerta/Falhou, ou nota do técnico indicando defeito/pendência real de "
        "funcionamento do equipamento → gere um item NORMAL de punch list (anormalidade "
        "objetiva, sem prefixo especial), cobrindo TODAS as subtarefas reprovadas desse ativo.\n"
        "   c) A nota do técnico indica que o ATIVO EM SI NÃO EXISTE FISICAMENTE em campo (ex.: "
        "\"não foi localizado\", \"UFV não tem sistema de X\", \"ausência completa do sistema "
        "Y\"), mesmo esse ativo estando previsto no escopo da OS → gere um item de punch list, "
        "mas comece o campo \"anormalidade\" com \"Ressalva — \" seguido da descrição da "
        "ausência, e marque \"ressalva\": true nesse item. Isso é uma constatação de ausência "
        "do ativo, não uma falha de funcionamento — o Fred vai revisar esses casos manualmente "
        "antes do relatório sair, porque pode ser um ativo fora de escopo, não instalado por "
        "decisão de projeto, ou uma pendência real a cobrar do cliente.\n\n"
        "REGRA MAIS IMPORTANTE DESTE PROMPT: só existe um ativo pra fins desta análise se você "
        "viu, com os próprios olhos nas páginas deste trecho, uma seção 'ATIVOS' com campo "
        "DESCRIÇÃO desse equipamento. A lista de categorias padrão mais abaixo é só uma "
        "referência de NOMENCLATURA pra você usar ao nomear um ativo que você JÁ CONFIRMOU que "
        "existe no documento — nunca uma lista do que \"deveria\" existir na usina. Se um tipo "
        "de ativo comum em outras usinas (ex. Estação Meteorológica, Fossa Séptica) não tiver "
        "seção própria nas páginas deste trecho, ele simplesmente não faz parte desta análise — "
        "é preferível deixar de fora do que inventar uma pendência pra um equipamento que não "
        "está no documento. (Essa regra é sobre INVENTAR ativos que não existem — não confunda "
        "com pular ativos que EXISTEM mas parecem menos importantes; esses você sempre analisa.)\n\n"
        "Critérios de CRITICIDADE — siga à risca, não subestime:\n"
        "- Alta/Muito Alta: falhas funcionais que afetam operação/eficiência (equipamento fora "
        "de operação, string parada), falhas de SEGURANÇA (proteção elétrica falhando, EPCs de "
        "segurança da equipe — luva, tapete isolante, vara de manobra — reprovados em estado/"
        "validade/calibração, sistema de combate a incêndio falhando), e falha TOTAL ou "
        "generalizada de um sistema inteiro (ex.: Sistema Supervisório/SCADA reprovando "
        "múltiplas subtarefas ao mesmo tempo — isso é o monitoramento central da usina "
        "falhando, sempre Alta, nunca Média, mesmo que cada subtarefa isolada pareça pequena).\n"
        "- Média: itens de manutenção/limpeza/cosmético isolados (ex.: limpeza de ventilador de "
        "inversor, sujidade).\n"
        "- Baixa: observações menores sem impacto funcional ou de segurança.\n"
        "- Itens de \"ressalva\" (ausência de ativo) usam Média por padrão, a menos que a "
        "ausência tenha implicação clara de segurança/operação (aí Alta).\n\n"
        "Critérios de RESPONSÁVEL: \"EQUIPE TÉCNICA\" para reparos elétricos/mecânicos; "
        "\"EQUIPE DE CAMPO\" para limpeza/organização/civil; \"CLIENTE/SUPERVISÃO\" quando o "
        "problema depende de sistema supervisório, contratação externa, credenciais de acesso, "
        "ou está fora do escopo de campo da Grid Co.; \"FABRICANTE\" quando exige garantia ou "
        "peça/serviço do fabricante do equipamento; \"FABRICANTE/TÉCNICA\" quando pode precisar "
        "de qualquer um dos dois. Itens de \"ressalva\" (ausência de ativo) normalmente são "
        "\"CLIENTE/SUPERVISÃO\", já que é uma decisão de projeto/escopo, não um reparo de campo.\n\n"
        "No campo \"ativo\", use uma CATEGORIA padrão (o tipo de equipamento), NUNCA o "
        "nome/número específico da unidade — escreva \"Inversores\" mesmo que o ativo se chame "
        "\"Inversor 1.3\", porque outras partes do documento podem ter achados do mesmo tipo de "
        "equipamento, que serão consolidados numa linha só depois. Categorias padrão (use uma "
        "destas sempre que fizer sentido pro ativo que você CONFIRMOU existir no documento; se "
        "nenhuma servir, use uma categoria curta e clara própria): Ar Condicionado, Cabine de "
        "Medição, Caixa d'água, Estação Meteorológica, Fossa Séptica, CFTV / Segurança, "
        "Infraestrutura Civil, Infraestrutura / Supervisório, Inversores, Trackers, Módulos "
        "Fotovoltaicos, QGBT, SPDA, Nobreak e Banco de Baterias, Transformador de Potência, "
        "Relé de Proteção, Sistema de Drenagem, Sistema de Combate a Incêndio.\n\n"
        "IMPORTANTE — CONSOLIDAÇÃO DENTRO DO PRÓPRIO TRECHO: se VÁRIAS unidades do MESMO TIPO "
        "de equipamento (ex.: Inversor 1.1, 1.2, 1.3...) tiverem o MESMO TIPO de achado, escreva "
        "UMA ÚNICA frase cobrindo todas elas — nunca repita a mesma frase uma vez por número de "
        "unidade. Exemplo CORRETO: \"Necessidade de limpeza ou substituição dos ventiladores em "
        "unidades específicas.\" Exemplo ERRADO (nunca faça isso): \"Necessidade de limpeza do "
        "ventilador no Inversor 1.1; Necessidade de limpeza do ventilador no Inversor 1.2; "
        "Necessidade de limpeza do ventilador no Inversor 1.3\". Só liste números de unidades "
        "específicas quando isso for realmente útil pra identificar ONDE agir (ex.: \"módulos "
        "trincados nas strings ST.03/TK.45 e ST.05/TK.36\") — nunca como lista repetitiva de "
        "achados idênticos por número. Isso vale só dentro deste trecho; se o mesmo tipo de "
        "achado aparecer em outro trecho do documento (processado separadamente), tudo bem, "
        "será consolidado depois — o que não pode é repetir a mesma frase várias vezes AQUI só "
        "trocando o número da unidade.\n\n"
        "Se não houver NENHUM item Alerta/Falhou, nota de ausência, ou pendência nesse trecho, "
        'retorne {"itens": []} — esse é o resultado correto, não uma falha.\n\n'
        "Além da punch list, procure também os campos \"Data e Hora de Início\" e \"Data e "
        "Hora de Fim\" que aparecem no cabeçalho de cada tarefa/ativo do documento (formato "
        "tipo \"2026-08-04 09:40\"). Extraia a MENOR data/hora de início e a MAIOR data/hora "
        "de fim encontradas nesse trecho — isso representa o início e o fim reais do trabalho "
        "de campo. Se não achar nenhuma, deixe os campos como string vazia.\n\n"
        "Também monte um INVENTÁRIO de ativos analisados nesse trecho, pro relatório listar "
        "quais equipamentos passaram por inspeção (a punch list só lista os com pendência — "
        "esse inventário é de TODOS, com ou sem problema). Pra cada seção 'ATIVOS' que "
        "aparecer nas páginas deste trecho, procure logo abaixo (ainda dentro da mesma seção, "
        "geralmente perto de 'TIPO DE TRABALHO: Handover') um segundo campo \"DESCRIÇÃO:\" "
        "no formato \"Handover — <nome do sistema>\" (ex.: \"Handover — Ar Condicionado\", "
        "\"Handover — Inversores\") — pegue só o \"<nome do sistema>\" (SEM o prefixo "
        "\"Handover — \", ele já é adicionado depois automaticamente), exatamente como está "
        "escrito no documento, sem reescrever ou trocar por uma categoria sua. Esse texto já "
        "vem certo e já deduplicado pela própria Fracttal (ex.: vários inversores individuais "
        "já aparecem todos com esse mesmo \"Handover — Inversores\", sem precisar juntar nada "
        "você mesmo). Se um ativo não tiver esse campo \"Handover — X\", não invente um nome "
        "pra ele — só inclua os que têm esse rótulo explícito no documento.\n\n"
        "Retorne APENAS um JSON (sem markdown, sem texto fora do JSON) no formato:\n"
        '{"itens": [{"ativo": "<categoria padrão, sem número de unidade>", '
        '"criticidade": "<Baixa|Média|Alta|Muito Alta>", '
        '"anormalidade": "<descrição objetiva e consolidada, ou \'Ressalva — ...\' pro caso c>", '
        '"recomendacoes": "<ação corretiva recomendada, objetiva>", '
        '"responsavel": "<EQUIPE TÉCNICA|EQUIPE DE CAMPO|CLIENTE/SUPERVISÃO|FABRICANTE|FABRICANTE/TÉCNICA>", '
        '"ressalva": <true se for o caso c (ativo não existe em campo), false caso contrário>}], '
        '"dataHoraInicio": "<menor \'Data e Hora de Início\' do trecho, ou \'\'>", '
        '"dataHoraFim": "<maior \'Data e Hora de Fim\' do trecho, ou \'\'>", '
        '"ativosAnalisados": ["<nome do sistema do 1º ativo, sem o prefixo \'Handover — \'>", "<2º ativo>", "..."]}'
    )


def _processar_chunk_pdf_nativo(chunk, cliente, usina, cluster, total_paginas):
    """Manda um pedaço do PDF direto pra API da Gemini como documento
    nativo (a Gemini lê texto E imagem da página sozinha — sem OCR, sem
    renderizar página por página, sem precisar achar 'ativos'
    manualmente antes). Muito mais simples e rápido que o pipeline
    antigo (OCR + visão por ativo), e funciona igual pra PDF com texto
    real ou PDF exportado como imagem (mesmo caminho, sem tratamento
    especial)."""
    prompt = _montar_prompt_punchlist_pdf_nativo(
        cliente, usina, cluster, chunk["pagina_inicio"], chunk["pagina_fim"], total_paginas)
    pdf_b64 = base64.b64encode(chunk["bytes"]).decode()

    resp = _chamar_gemini_com_retry(
        {
            "contents": [{"parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": "application/pdf", "data": pdf_b64}},
            ]}],
            "generationConfig": {
                "temperature": 0.15,
                "maxOutputTokens": 16384,
                "responseMimeType": "application/json",
                "thinkingConfig": {"thinkingBudget": 8192},
            },
        },
        timeout=150,
    )
    data = resp.json()
    candidato = data["candidates"][0]
    texto_bruto = candidato["content"]["parts"][0]["text"].strip()
    texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
    parsed = json.loads(texto_limpo)
    itens = parsed.get("itens", [])

    normalizados = []
    for it in itens:
        normalizados.append({
            "ativo": (it.get("ativo") or "").strip() or "Não identificado",
            "criticidade": it.get("criticidade") if it.get("criticidade") in
                           ("Baixa", "Média", "Alta", "Muito Alta") else "Média",
            "status": "PENDENTE",
            "anormalidade": (it.get("anormalidade") or "").strip(),
            "recomendacoes": (it.get("recomendacoes") or "").strip(),
            "responsavel": (it.get("responsavel") or "EQUIPE TÉCNICA").strip(),
            "ressalva": bool(it.get("ressalva")),
            "cliente": cliente,
            "usina": usina,
            "cluster": cluster,
        })
    return {
        "itens": normalizados,
        "dataHoraInicio": (parsed.get("dataHoraInicio") or "").strip(),
        "dataHoraFim": (parsed.get("dataHoraFim") or "").strip(),
        "ativosAnalisados": [str(a).strip() for a in (parsed.get("ativosAnalisados") or []) if str(a).strip()],
    }


@app.route("/extrair-punchlist-pdf-nativo", methods=["POST", "OPTIONS"])
def extrair_punchlist_pdf_nativo_route():
    """
    Lê o PDF original exportado da Fracttal e gera a Punch List enviando
    o PDF direto pra API da Gemini como documento nativo (a Gemini lê
    texto e imagem sozinha, sem OCR e sem pipeline de visão por ativo).
    Documentos grandes (>90 páginas ou >28MB) são divididos em pedaços e
    processados em paralelo; os itens de todos os pedaços são
    consolidados por categoria no final.

    Substitui /extrair-punchlist-fracttal-ia (que dependia de achar
    "ativos" via texto do PDF + Gemini vision por página — quebrava em
    PDFs exportados como imagem, que alguns exports da Fracttal geram).
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500
    try:
        arquivo = request.files.get("fracttalPdf")
        if not arquivo or not arquivo.filename:
            return jsonify({"ok": False, "error": "Anexe o PDF exportado da Fracttal (campo fracttalPdf)."}), 400
        if not arquivo.filename.lower().endswith(".pdf"):
            return jsonify({"ok": False, "error": "O arquivo precisa ser um PDF."}), 400

        cliente = (request.form.get("cliente") or "").strip()
        usina = (request.form.get("usina") or "").strip()
        cluster = (request.form.get("cluster") or "").strip()
        pdf_bytes = arquivo.read()

        total_paginas = len(PdfReader(BytesIO(pdf_bytes)).pages)
        chunks = _dividir_pdf_em_chunks(pdf_bytes)

        resultados = []
        erros = []
        datas_inicio = []
        datas_fim = []
        ativos_analisados = []
        # max_workers=3: cada pedaço agora é ~7MB (não mais 14MB — reduzido
        # pra melhorar a precisão da leitura, ver GEMINI_PDF_MAX_MB acima),
        # então o pico de memória por pedaço caiu bastante — dá pra rodar 3
        # em paralelo sem voltar ao risco de falta de memória que já causou
        # um "Worker (...) was sent SIGKILL!" em produção com pedaços
        # maiores. Com pedaços menores E mais chamadas, manter só 2 em
        # paralelo deixaria o tempo total alto demais.
        with ThreadPoolExecutor(max_workers=min(3, len(chunks))) as executor:
            futuros = {executor.submit(_processar_chunk_pdf_nativo, c, cliente, usina, cluster, total_paginas): c
                       for c in chunks}
            for futuro in as_completed(futuros):
                chunk = futuros[futuro]
                try:
                    resultado_chunk = futuro.result()
                    resultados.extend(resultado_chunk["itens"])
                    if resultado_chunk.get("dataHoraInicio"):
                        datas_inicio.append(resultado_chunk["dataHoraInicio"])
                    if resultado_chunk.get("dataHoraFim"):
                        datas_fim.append(resultado_chunk["dataHoraFim"])
                    ativos_analisados.extend(resultado_chunk.get("ativosAnalisados", []))
                except Exception as e:
                    log.error(f"[PunchlistPdfNativo] Erro no trecho pg{chunk['pagina_inicio']}-"
                              f"{chunk['pagina_fim']}: {e}")
                    erros.append({"trecho": f"{chunk['pagina_inicio']+1}-{chunk['pagina_fim']+1}",
                                  "erro": str(e)})

        itens_consolidados = _consolidar_itens_punchlist_por_categoria(resultados)
        # menor "Data e Hora de Início" e maior "Data e Hora de Fim" entre
        # todos os trechos — comparação por string funciona porque o
        # formato pedido no prompt é ISO ("aaaa-mm-dd hh:mm"), que ordena
        # igual a uma comparação de datas de verdade.
        data_hora_inicio = min(datas_inicio) if datas_inicio else ""
        data_hora_fim = max(datas_fim) if datas_fim else ""
        # dedup case-insensitive preservando a primeira grafia vista
        vistos = {}
        for a in ativos_analisados:
            chave = a.strip().lower()
            if chave and chave not in vistos:
                vistos[chave] = a.strip()
        ativos_analisados_dedup = sorted(vistos.values())

        return jsonify({
            "ok": True,
            "itens": itens_consolidados,
            "dataHoraInicio": data_hora_inicio,
            "dataHoraFim": data_hora_fim,
            "ativosAnalisados": ativos_analisados_dedup,
            "totalPaginas": total_paginas,
            "trechosProcessados": len(chunks) - len(erros),
            "trechosTotal": len(chunks),
            "erros": erros,
        }), 200
    except Exception as e:
        log.error(f"[PunchlistPdfNativo] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/extrair-punchlist-fracttal-ia", methods=["POST", "OPTIONS"])
def extrair_punchlist_fracttal_ia():
    """
    Lê o PDF original exportado da Fracttal (checklist com marcações
    Aprovou/Alerta/Falhou + anotações de campo) e gera a Punch List via
    IA com visão, ativo por ativo. Processa em paralelo (poucos workers
    — a VM1 tem só 1GB de RAM) e respeita um orçamento de tempo pra não
    estourar o timeout do Gunicorn (160s).
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500
    try:
        arquivo = request.files.get("fracttalPdf")
        if not arquivo or not arquivo.filename:
            return jsonify({"ok": False, "error": "Anexe o PDF exportado da Fracttal (campo fracttalPdf)."}), 400
        if not arquivo.filename.lower().endswith(".pdf"):
            return jsonify({"ok": False, "error": "O arquivo precisa ser um PDF."}), 400

        cliente = (request.form.get("cliente") or "").strip()
        usina = (request.form.get("usina") or "").strip()
        cluster = (request.form.get("cluster") or "").strip()
        pdf_bytes = arquivo.read()

        try:
            ativos = _fracttal_pdf_extrair_ativos(pdf_bytes)
        except RuntimeError as e:
            return jsonify({"ok": False, "error": str(e)}), 400
        if not ativos:
            return jsonify({"ok": False, "error": ("Não consegui identificar nenhum ativo nesse PDF — "
                            "confirme se é o formato de exportação padrão da Fracttal (Ordem de Trabalho).")}), 400

        MAX_ATIVOS_POR_CHAMADA = 40  # segurança — evita processar um PDF gigante além do razoável numa única requisição
        ativos = ativos[:MAX_ATIVOS_POR_CHAMADA]

        def _processar_ativo(ativo):
            paginas_check = _fracttal_pdf_paginas_checklist(pdf_bytes, ativo["pagina_inicio"], ativo["pagina_fim"])
            if not paginas_check:
                return {"ativo": ativo["nome"], "itens": [], "semChecklist": True}
            notas = _fracttal_pdf_notas_ativo(pdf_bytes, ativo["pagina_inicio"], ativo["pagina_fim"])
            imagens = _fracttal_pdf_renderizar_paginas(pdf_bytes, paginas_check)
            itens = _gerar_punchlist_ativo_via_visao(ativo["nome"], cliente, usina, cluster, imagens, notas)
            for it in itens:
                it["cliente"] = cliente
                it["usina"] = usina
            return {"ativo": ativo["nome"], "itens": itens, "semChecklist": False}

        ORCAMENTO_SEGUNDOS = 130
        inicio = time.time()
        resultados = []
        erros = []
        parou_por_orcamento = False

        with ThreadPoolExecutor(max_workers=4) as executor:
            futuros = {executor.submit(_processar_ativo, a): a for a in ativos}
            for futuro in as_completed(futuros):
                ativo = futuros[futuro]
                if time.time() - inicio > ORCAMENTO_SEGUNDOS:
                    parou_por_orcamento = True
                    # não cancela os já em andamento, só para de esperar novos
                try:
                    resultados.append(futuro.result())
                except Exception as e:
                    log.error(f"[PunchlistFracttalIA] Erro no ativo '{ativo['nome']}': {e}")
                    erros.append({"ativo": ativo["nome"], "erro": str(e)})

        todos_itens = []
        for r in resultados:
            todos_itens.extend(r["itens"])

        todos_itens = _consolidar_itens_punchlist_por_categoria(todos_itens)

        return jsonify({
            "ok": True,
            "itens": todos_itens,
            "ativosTotal": len(ativos),
            "ativosProcessados": len(resultados),
            "ativosComPendencia": sum(1 for r in resultados if r["itens"]),
            "erros": erros,
            "parouPorOrcamento": parou_por_orcamento,
        }), 200
    except Exception as e:
        log.error(f"[PunchlistFracttalIA] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/gerar-relatorio-handover-usina", methods=["POST", "OPTIONS"])
def gerar_relatorio_handover_usina_route():
    """
    Gera o Relatório de Handover EPC->O&M completo (usina inteira), a
    partir do MODELO REAL da Grid Co. (.docx enviado pelo Fred — capa,
    cabeçalho e estilos verdadeiros), substituindo só o conteúdo
    variável e anexando a Punch List como seção final em paisagem.

    multipart/form-data:
      - "dados": JSON (string) com os campos do formulário (ver
        relatorio_handover_usina_docx.py para o formato esperado)
      - "fracttalPdf": arquivo PDF (opcional) exportado da Fracttal.
        Se NÃO for enviado: retorna o .docx editável (modelo real).
        Se for enviado: converte o .docx pra PDF (LibreOffice) e insere
        as páginas da Fracttal logo após "3.4 Ordens de Serviço -
        Handover", retornando o PDF final já mesclado — documento +
        checklist da OS juntos, prontos pra enviar ao cliente.
    """
    if request.method == "OPTIONS":
        return ("", 204)
    try:
        content_length = request.content_length
        log.info(f"[Relatorio Handover Usina] Request recebido — Content-Length={content_length}, "
                 f"form.keys={list(request.form.keys())}, files.keys={list(request.files.keys())}")

        dados_raw = request.form.get("dados", "{}")
        try:
            dados = json.loads(dados_raw)
        except Exception:
            return jsonify({"ok": False, "error": "Campo 'dados' não é um JSON válido."}), 400

        cliente = (dados.get("cliente") or "").strip()
        usina = (dados.get("usina") or "").strip()
        if not cliente or not usina:
            return jsonify({"ok": False, "error": "cliente e usina são obrigatórios"}), 400

        fracttal_pdf_bytes = None
        arquivo = request.files.get("fracttalPdf")
        log.info(f"[Relatorio Handover Usina] arquivo fracttalPdf presente={arquivo is not None}, "
                 f"filename={getattr(arquivo, 'filename', None)}")
        if arquivo and arquivo.filename:
            if not arquivo.filename.lower().endswith(".pdf"):
                return jsonify({"ok": False, "error": "O arquivo anexado precisa ser um PDF."}), 400
            fracttal_pdf_bytes = arquivo.read()
            log.info(f"[Relatorio Handover Usina] fracttalPdf lido: {len(fracttal_pdf_bytes)} bytes")

        conteudo, tipo = gerar_handover_usina_completo(dados, fracttal_pdf_bytes)

        # Nomenclatura confirmada nos arquivos de referência da Grid Co.
        # (ex.: "UFV ABC MORADA NOVA - Relatório Handover - Grid Co.docx",
        # "UFV Hortina - Relatório Handover - Grid Co.docx").
        extensao = "pdf" if tipo == "pdf" else "docx"
        mimetype = ("application/pdf" if tipo == "pdf" else
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
        nome_arquivo = f"UFV {usina.upper()} - Relatório Handover - Grid Co.{extensao}"

        log.info(f"[Relatorio Handover Usina] Gerado ({tipo}) para {usina} ({cliente}), "
                 f"com PDF Fracttal={'sim' if fracttal_pdf_bytes else 'não'}")
        return send_file(
            BytesIO(conteudo), as_attachment=True, download_name=nome_arquivo, mimetype=mimetype,
        )
    except Exception as e:
        log.error(f"[Relatorio Handover Usina] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Ata de Reunião (Painel de Relatórios) ───────────────────────────────

def _extrair_texto_transcricao(nome_arquivo, conteudo_bytes):
    """Extrai o texto de uma transcrição de reunião (Teams), aceitando
    .docx (exportado do Teams/Stream — um parágrafo por fala, com nome do
    participante e timestamp) ou .txt simples."""
    nome = (nome_arquivo or "").lower()
    if nome.endswith(".docx"):
        doc = _docx_lib.Document(BytesIO(conteudo_bytes))
        linhas = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(linhas)
    return conteudo_bytes.decode("utf-8", errors="ignore")


def _montar_prompt_ata_reuniao(texto_transcricao, cliente_hint=""):
    contexto_cliente = (
        f'O cliente/contraparte desta reunião é "{cliente_hint}" (informado pelo usuário) — '
        f"use esse nome no subtítulo da capa e no rótulo de clientes.\n"
        if cliente_hint else
        "O nome do cliente/contraparte não foi informado — identifique-o pelo contexto da "
        "transcrição (nomes de empresas, e-mails, usinas mencionadas) e use \"Grid Co.\" como "
        "a outra parte (Grid Co. é a empresa de O&M que participa de todas essas reuniões).\n"
    )
    return f"""Você é um Supervisor de O&M da Grid Co. (empresa de operação e manutenção de usinas
fotovoltaicas no Brasil) redigindo a ATA OFICIAL de uma reunião, a partir da transcrição bruta
exportada do Microsoft Teams (um parágrafo por fala, geralmente com nome do participante e
timestamp misturados no texto).

{contexto_cliente}
TAREFA: leia a transcrição inteira com atenção e estruture uma ata de reunião profissional,
em terceiro pessoa, seguindo EXATAMENTE o schema JSON abaixo. Cada campo alimenta um componente
visual fixo do padrão Grid Co. (capa, seções numeradas, cards de tópico numerados, tabela de
ações, tabela de premissas) — não invente estrutura nova, preencha os campos do schema.

REGRAS DE CONTEÚDO (críticas):
- NUNCA invente fatos, nomes, datas, números de OS/inversor ou valores que não estejam na
  transcrição. Se algo estiver ambíguo ou incompleto, descreva como está (ex.: "data a confirmar")
  em vez de supor.
- PRESERVE O GRAU DE CERTEZA do texto original: se alguém disse "acho que", "acredito", "talvez",
  "devemos conseguir", não transforme em afirmação categórica — mantenha a incerteza na redação.
- Cada tópico em "topicos" deve corresponder a UM assunto discutido, na ordem em que apareceu na
  conversa. Uma linha "corpo" por tópico, escrita em texto corrido (não em lista), explicando o
  que foi discutido, decidido ou levantado — com os detalhes técnicos relevantes (números,
  usinas, prazos, nomes) preservados.
- Use "callout_label" + "callout_texto" SOMENTE quando o tópico tiver uma decisão clara, ação
  definida, pendência, ou prazo/próximo passo explícito — rótulos possíveis: "Decisão:",
  "Ação:", "Próximo passo:", "Pendência:", "Prazo:". Nem todo tópico precisa de callout.
  callout_tipo = "warn" só para alertas/riscos reais; use "green" pro resto.
- "responsavel" de cada tópico: quem conduziu/é dono daquele assunto (uma ou mais pessoas,
  separadas por " / "). Pode ficar vazio se não ficar claro.
- Preencha "cronograma" APENAS se a reunião tiver claramente uma lista de datas/prazos por
  usina ou item (ex.: cronograma de poda, cronograma de manutenção) que faça sentido virar
  tabela separada. Caso contrário, retorne cronograma como null — não force uma tabela.
- "acoes": consolide TODAS as ações/encaminhamentos combinados na reunião (mesmo os já citados
  dentro de algum callout de tópico) numa lista objetiva de itens únicos, cada um com um
  responsável e prazo quando existir.
- "premissas": 3 a 6 observações gerais/conclusões da reunião como um todo (não repita os
  tópicos individuais — são takeaways transversais).
- Textos SEMPRE em português do Brasil, tom técnico e direto, terceira pessoa (nunca "eu").
- Data e duração da reunião: procure no início da transcrição (o Teams normalmente inclui data/
  hora/duração antes da primeira fala). Se não encontrar duração, omita-a do objetivo.
- "cliente_nome" e "data_iso" são OBRIGATÓRIOS e usados para montar o nome do arquivo final —
  capriche na identificação: "cliente_nome" nunca deve ser "Grid Co." (ela é sempre uma das
  partes, nunca a contraparte) nem ficar vazio — se realmente não conseguir identificar o
  cliente pelo contexto, use "Cliente". "data_iso" deve ser a data em que a reunião ocorreu,
  no formato AAAA-MM-DD.

TRANSCRIÇÃO:
---
{texto_transcricao[:60000]}
---

FORMATO DE SAÍDA — responda APENAS com um JSON válido (sem markdown, sem crase, sem texto antes
ou depois), EXATAMENTE neste schema:
{{
  "titulo_capa": "Reunião Semanal" ou "Reunião de Acompanhamento" (curto, 2-4 palavras),
  "subtitulo_capa": "NOME CLIENTE & GRID CO." (maiúsculas),
  "clientes_label": "Nome Cliente & Grid Co.",
  "cliente_nome": "Nome Cliente" (Title Case, só o nome do cliente/contraparte, sem "Grid Co." —
      ex.: "Renogrid", "Alves Lima", "GD Energy", "2C Energia" — usado no nome do arquivo),
  "data_extenso": "23 de julho de 2026",
  "data_arquivo": "23-07-2026",
  "data_iso": "2026-07-23" (data da reunião em AAAA-MM-DD, pra calcular o número da semana do ano),
  "rodape_capa": "Documento de uso interno — Grid Co. / Nome Cliente",
  "objetivo": "parágrafo único: registra que o documento resume a reunião, data, duração (se souber) e participantes (nomes encontrados na transcrição, com empresa entre parênteses quando identificável).",
  "topicos": [
    {{"titulo": "título curto do assunto", "responsavel": "Nome(s)",
      "corpo": "parágrafo explicando o que foi discutido",
      "callout_label": "Ação:" ou null, "callout_texto": "..." ou null, "callout_tipo": "green"}}
  ],
  "cronograma": null OU {{"titulo": "CRONOGRAMA DE ...", "descricao": "frase intro",
      "headers": ["Usina","Data prevista","Situação"], "linhas": [["...","...","..."]]}},
  "acoes": [{{"acao": "...", "resp": "Nome — prazo se houver"}}],
  "premissas": ["...", "..."]
}}"""


def _gerar_dados_ata_com_ia(texto_transcricao, cliente_hint=""):
    prompt = _montar_prompt_ata_reuniao(texto_transcricao, cliente_hint)
    resp = _chamar_gemini_com_retry(
        {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.3,
                "maxOutputTokens": 8192,
                "responseMimeType": "application/json",
            },
        },
        timeout=90,
    )
    data = resp.json()
    candidato = data["candidates"][0]
    texto_bruto = candidato["content"]["parts"][0]["text"].strip()
    texto_limpo = re.sub(r"^```json\s*|\s*```$", "", texto_bruto.strip())
    parsed = json.loads(texto_limpo)
    if not parsed.get("topicos"):
        raise ValueError("A IA não conseguiu identificar tópicos na transcrição enviada.")
    return parsed


def _montar_nome_arquivo_ata(dados):
    """Padrão de nomenclatura fixo (definido pelo Fred, 15/08/2026), usado
    por TODOS os arquivos de Ata de Reunião gerados pelo sistema:
        "Ata de Reuniao - {Cliente} x Grid Co - O&M - Semana {N}.docx"
    (sem acento em "Reuniao" de propósito, pra manter compatibilidade com
    sistemas de arquivo/e-mail mais sensíveis a caracteres especiais)."""
    cliente = (dados.get("cliente_nome") or "").strip() or "Cliente"

    semana = None
    data_iso = (dados.get("data_iso") or "").strip()
    if data_iso:
        try:
            semana = datetime.strptime(data_iso, "%Y-%m-%d").isocalendar()[1]
        except Exception:
            semana = None
    if semana is None:
        # fallback: usa a semana atual do calendário se a IA não conseguiu
        # extrair a data da transcrição (não deveria acontecer, mas evita
        # quebrar o download por causa só do número da semana).
        semana = datetime.now(ZoneInfo("America/Fortaleza")).isocalendar()[1]

    return f"Ata de Reuniao - {cliente} x Grid Co - O&M - Semana {semana}.docx"


@app.route("/gerar-ata-reuniao", methods=["POST", "OPTIONS"])
def gerar_ata_reuniao_route():
    """
    Gera a "Ata de Reunião" (.docx, padrão visual Grid Co.) a partir de uma
    transcrição de reunião do Teams enviada pelo usuário no Painel de
    Relatórios.

    multipart/form-data:
      - "transcricao": arquivo .docx (exportado do Teams) ou .txt
      - "cliente": (opcional) nome do cliente/contraparte, pra ajudar a IA
        a rotular a capa corretamente sem precisar adivinhar do texto.
    """
    if request.method == "OPTIONS":
        return ("", 204)
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor."}), 500
    try:
        arquivo = request.files.get("transcricao")
        if not arquivo or not arquivo.filename:
            return jsonify({"ok": False, "error": "Anexe o arquivo de transcrição (.docx ou .txt)."}), 400

        cliente_hint = (request.form.get("cliente") or "").strip()
        conteudo_bytes = arquivo.read()
        texto = _extrair_texto_transcricao(arquivo.filename, conteudo_bytes)
        if not texto.strip():
            return jsonify({"ok": False, "error": "Não foi possível extrair texto da transcrição enviada."}), 400

        log.info(f"[Ata Reuniao] Transcrição recebida: {arquivo.filename} "
                 f"({len(conteudo_bytes)} bytes, {len(texto)} chars extraídos), cliente_hint={cliente_hint!r}")

        dados = _gerar_dados_ata_com_ia(texto, cliente_hint)
        conteudo = gerar_ata_reuniao_docx(dados)

        nome_arquivo = _montar_nome_arquivo_ata(dados)

        log.info(f"[Ata Reuniao] Gerada com sucesso: {nome_arquivo} ({len(dados.get('topicos', []))} tópicos)")
        return send_file(
            BytesIO(conteudo), as_attachment=True, download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
    except Exception as e:
        log.error(f"[Ata Reuniao] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


def gerar_os():
    """
    Gera o texto de solicitação de OS (Título + Comentários) a partir do
    contexto de uma falha, chamando a API da Anthropic do lado do servidor
    (a chave nunca é exposta ao navegador/dashboard).

    Body esperado (JSON):
    {
      "equipamento": "...", "usina": "...", "falha": "...", "causa": "...",
      "impactados": "...", "acao": "...", "historico": "..."
    }

    Retorna: {"ok": true, "texto": "Título:\n...\n\nComentários:\n..."}
    """
    if request.method == "OPTIONS":
        return ("", 204)

    if not ANTHROPIC_API_KEY:
        return jsonify({"ok": False, "error": "ANTHROPIC_API_KEY não configurada no servidor."}), 500

    try:
        body = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Body inválido — esperado JSON."}), 400

    equipamento = body.get("equipamento", "")
    usina       = body.get("usina", "")
    falha       = body.get("falha", "")
    causa       = body.get("causa", "")
    impactados  = body.get("impactados", "")
    acao        = body.get("acao", "")
    historico   = body.get("historico", "")

    system_prompt = (
        "Você é um engenheiro de O&M (operação e manutenção) de usinas solares, "
        "redigindo solicitações de Ordem de Serviço (OS) técnica para equipe de campo.\n\n"
        "Gere SEMPRE a saída EXATAMENTE neste formato, sem nenhum texto antes ou depois:\n\n"
        "Título:\n"
        "<uma linha, objetiva, descrevendo o diagnóstico/inspeção necessária para o "
        "equipamento e a falha específica>\n\n"
        "Comentários:\n\n"
        "• <item 1 do checklist técnico>\n"
        "• <item 2 do checklist técnico>\n"
        "• <item 3 do checklist técnico>\n"
        "• <item 4 ou mais itens, conforme necessário — geralmente entre 4 e 6 itens>\n\n"
        "Regras:\n"
        "- O checklist deve ser ESPECÍFICO ao tipo de equipamento (inversor, tracker, "
        "motor, TCU, câmera/CFTV, nobreak, transformador, chave seccionadora, "
        "piranômetro, etc.) e à causa da falha informada — nunca genérico.\n"
        "- Use linguagem técnica de campo, direta, em formato de instrução (verbos no "
        "infinitivo: verificar, conferir, inspecionar, avaliar, registrar, medir, testar).\n"
        "- Considere o histórico cronológico para não repetir verificações já feitas, e "
        "para direcionar o checklist ao que ainda falta investigar/resolver.\n"
        "- Considere os equipamentos impactados para garantir que o checklist cubra "
        "todos eles quando relevante.\n"
        "- O título deve mencionar o equipamento/local específico quando disponível.\n"
        "- Sempre que a falha envolver trackers, estruturas de fixação/suporte de módulos "
        "fotovoltaicos, ou integridade estrutural/civil de forma geral, inclua um item "
        "avaliando as estruturas de fixação dos módulos quanto a afundamento (verificar "
        "sinais de afundamento, desnivelamento ou instabilidade no solo das "
        "bases/fundações).\n"
        "- Nunca inclua explicações, saudações, ou qualquer texto fora do formato "
        "Título/Comentários especificado."
    )

    user_content = (
        f"Equipamento: {equipamento or 'não informado'}\n"
        f"Usina: {usina or 'não informado'}\n"
        f"Falha: {falha or 'não informado'}\n"
        f"Causa: {causa or 'não informado'}\n"
        f"Equipamentos impactados: {impactados or 'não informado'}\n"
        f"Ações já realizadas: {acao or 'nenhuma registrada'}\n"
        f"Histórico cronológico:\n{historico or 'sem histórico registrado'}"
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 600,
                "system": system_prompt,
                "messages": [{"role": "user", "content": user_content}],
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        texto = ""
        for block in data.get("content", []):
            if block.get("type") == "text":
                texto = block.get("text", "").strip()
                break
        if not texto:
            return jsonify({"ok": False, "error": "Resposta vazia da IA."}), 502
        return jsonify({"ok": True, "texto": texto}), 200

    except requests.exceptions.RequestException as e:
        log.error(f"[GerarOS] Erro na chamada à API Anthropic: {e}")
        return jsonify({"ok": False, "error": f"Erro ao chamar a API: {str(e)}"}), 502
    except Exception as e:
        log.error(f"[GerarOS] Erro inesperado: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/limpar-duplicatas", methods=["GET", "POST"])
def limpar_duplicatas():
    """
    Limpa duplicatas da planilha.

    Acesse direto pelo navegador (GET):
      https://api.168.138.232.237.sslip.io/limpar-duplicatas?secret=falhas2026

    Para cada grupo de linhas com mesmo fingerprint (usina+equip+falha)
    em aberto, mantém apenas a PRIMEIRA (menor ID) e remove as demais,
    consolidando as ações e o histórico na linha mantida.

    Seguro para executar múltiplas vezes (idempotente).
    Retorna: { ok, removidas, consolidadas, mantidas }
    """
    try:
        # Aceita secret via query string (GET) ou header (POST)
        secret_qs     = request.args.get("secret", "")
        secret_header = request.headers.get("X-Webhook-Secret", "")
        secret        = secret_qs or secret_header
        if WEBHOOK_SECRET and secret != WEBHOOK_SECRET:
            return jsonify({"error": "unauthorized — adicione ?secret=VALOR na URL"}), 401

        ws    = get_sheet()
        todos = carregar_planilha(ws)

        # Indexa todas as linhas abertas por fingerprint
        grupos = {}  # fingerprint → [(num_linha, row), ...]
        for i, row in enumerate(todos[1:], start=2):
            if len(row) < 9: continue
            id_val = (row[0] or "").strip()
            if not id_val: continue
            status = row[8].strip().lower()
            if "conclu" in status or "resolv" in status or "fechad" in status:
                continue
            fp = fingerprint_ocorrencia(row[2], row[3], row[4])
            if not fp: continue
            grupos.setdefault(fp, []).append((i, row))

        removidas    = 0
        consolidadas = 0
        mantidas     = 0

        for fp, linhas in grupos.items():
            if len(linhas) <= 1:
                mantidas += 1
                continue

            # Ordena por ID numérico — mantém a primeira
            linhas_ord = sorted(linhas, key=lambda x: int(x[1][0]) if x[1][0].isdigit() else 999999)
            linha_principal_num, linha_principal_row = linhas_ord[0]
            duplicatas = linhas_ord[1:]

            # Consolida ações e histórico das duplicatas na linha principal
            acao_consolidada  = (linha_principal_row[7] if len(linha_principal_row) > 7 else "").strip()
            hist_consolidado  = (linha_principal_row[11] if len(linha_principal_row) > 11 else "").strip()

            for _, dup_row in duplicatas:
                acao_dup = (dup_row[7] if len(dup_row) > 7 else "").strip()
                hist_dup = (dup_row[11] if len(dup_row) > 11 else "").strip()

                # Acrescenta ação da duplicata se tiver informação nova
                if acao_dup and acao_dup not in acao_consolidada:
                    acao_consolidada = (acao_consolidada + "\n" + acao_dup).strip()

                # Acrescenta entradas do histórico que não existem ainda
                for linha_hist in hist_dup.split("\n"):
                    linha_hist = linha_hist.strip()
                    if linha_hist and linha_hist not in hist_consolidado:
                        hist_consolidado = (hist_consolidado + "\n" + linha_hist).strip()

            # Atualiza linha principal com conteúdo consolidado
            ws.update_cell(linha_principal_num, 8,  acao_consolidada)
            ws.update_cell(linha_principal_num, 12, hist_consolidado)
            mantidas += 1
            consolidadas += 1

            # Remove duplicatas (limpa o conteúdo das células — não deleta a linha
            # para não deslocar índices; marca como removida com ID vazio)
            for dup_num, dup_row in duplicatas:
                ws.update(f"A{dup_num}:L{dup_num}", [["" for _ in range(12)]])
                removidas += 1
                log.info(f"🗑️  Removida duplicata linha {dup_num} | ID={dup_row[0]} | {dup_row[2]} / {dup_row[3]}")

        log.info(f"[Limpar] Concluído: {removidas} removidas, {consolidadas} consolidadas, {mantidas} mantidas")
        return jsonify({
            "ok":          True,
            "removidas":   removidas,
            "consolidadas": consolidadas,
            "mantidas":    mantidas,
        }), 200

    except Exception as e:
        log.error(f"[Limpar] Erro: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/test", methods=["POST"])
def test_parse():
    """Testa o parse sem gravar na planilha."""
    payload    = request.get_json(force=True) or {}
    texto      = payload.get("texto", "")
    blocos     = separar_blocos(texto)
    resultados = []
    for b in blocos:
        r = parse_bloco(b)
        if r:
            resultados.append(r)
    return jsonify({"total_blocos": len(blocos), "validos": len(resultados), "resultados": resultados}), 200


# ══════════════════════════════════════════════════════════════════════
# FV ENERGIAS RENOVÁVEIS — painel isolado (usina particular do Fred,
# integração direta com a API Pro da Solplanet/AISWEI).
#
# Este módulo é AUTOCONTIDO de propósito: não usa CATALOGO_USINAS,
# clusters, mapeamento cliente/equipe, _mapa_grupo_usina nem qualquer
# outra estrutura do restante do dashboard. As únicas peças reaproveitadas
# do sistema existente são: (1) WPP_SERVER_URL/WEBHOOK_SECRET pra enviar
# mensagem pelo mesmo bridge de WhatsApp (Baileys/VM2), e (2) as funções
# genéricas _ler_trava/_gravar_trava (aba _Sistema) só como mecanismo de
# lock pra não duplicar envio — usando chaves com prefixo "fv_energias:"
# exclusivo, sem tocar em nenhuma chave usada pelo resto do sistema.
# Adicionado 13/08/2026.
# ══════════════════════════════════════════════════════════════════════

import hmac as _fv_hmac
import hashlib as _fv_hashlib
from urllib.parse import quote as _fv_url_quote

FV_ENERGIAS_APP_KEY    = "204732162"
FV_ENERGIAS_APP_SECRET = "OK9oG0sbSeiNZAKTB5VOa8R9CQqTqLuT"
FV_ENERGIAS_APIKEY     = "75a973d76e00483c93d6e98fb79da4cc"
FV_ENERGIAS_TOKEN      = "UUdJckNFTEdjMVFTTmJvMjZyNDF0QT09"
FV_ENERGIAS_BASE_URL   = "https://ap-southeast-1-api-genergal.aisweicloud.com"
FV_ENERGIAS_GRUPO_WHATSAPP = "120363403242246431@g.us"  # "FV Energias Renováveis"

# Apelido fixo por número de série, pra ronda sempre mostrar na mesma ordem
# (INV01/INV02/INV03) independente da ordem em que a API devolve os dongles.
FV_ENERGIAS_APELIDO_INVERSOR = {
    "SP002087G2560054": "INV01",
    "SP002087G2560052": "INV02",
    "SP002087G2560056": "INV03",
}

# Horários da ronda automática (HH, hora cheia) e tamanho da janela de
# tolerância em minutos — mesmo padrão usado pelos comunicados diários
# (checagem barata a cada 5min via UptimeRobot, só age dentro da janela).
FV_ENERGIAS_HORARIOS_RONDA = ["08", "12", "15", "17"]
FV_ENERGIAS_JANELA_MINUTOS = 9

# Limiar de desbalanceamento entre inversores (geração do dia) que dispara
# alerta imediato — 15% de diferença entre o que mais gerou e o que menos
# gerou no dia, dado como fração (0.15 = 15%).
FV_ENERGIAS_LIMIAR_DESBALANCEO = 0.15

# Janela de monitoramento diurno — fora dela os inversores desligam
# naturalmente por falta de geração (sem sol), então "offline" é esperado
# e NÃO deve gerar alerta. Horário com margem folgada em torno do nascer/
# pôr do sol no Ceará.
FV_ENERGIAS_HORA_INICIO_MONITORAMENTO = 6   # 06:00
FV_ENERGIAS_HORA_FIM_MONITORAMENTO = 18     # 18:00 (após isso, offline é natural/esperado)


def _fv_energias_buscar_potencia_por_inversor():
    """Retorna {sn: potencia_kw} usando getLastTsDataPro (leitura mais
    recente por inversor) — usado só pra checagem de desbalanceamento,
    não pro painel. Diferente do e_today (acumulado do dia), aqui é a
    potência instantânea de agora."""
    resp = _fv_energias_chamar_api(
        "/pro/getLastTsDataPro",
        {"token": FV_ENERGIAS_TOKEN, "isnos": ",".join(FV_ENERGIAS_APELIDO_INVERSOR.keys())},
    )
    if resp.get("status") != 200:
        raise RuntimeError(f"Solplanet getLastTsDataPro erro: {resp}")
    potencias = {}
    for item in (resp.get("data") or []):
        sn = item.get("sn")
        try:
            potencias[sn] = float(item.get("pac") or 0) / 1000  # W -> kW
        except (TypeError, ValueError):
            pass
    return potencias


# Faixa de potência (kW) considerada "próxima do máximo" pra cada
# inversor (rated 20kW) — desbalanceamento só é avaliado quando AMBOS os
# inversores comparados estão nessa faixa, evitando ruído em horários de
# baixa geração (manhã cedo/fim de tarde), onde pequenas diferenças
# absolutas (ex.: 0.2 vs 0.1 kWh) geram percentuais enganosos.
FV_ENERGIAS_POTENCIA_MIN_COMPARACAO = 18
FV_ENERGIAS_POTENCIA_MAX_COMPARACAO = 22


def _fv_energias_detectar_desbalanceamento_potencia(potencia_por_inversor):
    """Desbalanceamento de potência instantânea entre inversores — só
    avaliado quando ambos estão perto da potência máxima (18-22kW)."""
    proximos_do_maximo = [
        (FV_ENERGIAS_APELIDO_INVERSOR.get(sn, sn), v)
        for sn, v in potencia_por_inversor.items()
        if v is not None and FV_ENERGIAS_POTENCIA_MIN_COMPARACAO <= v <= FV_ENERGIAS_POTENCIA_MAX_COMPARACAO
    ]
    if len(proximos_do_maximo) < 2:
        return []

    maior = max(proximos_do_maximo, key=lambda x: x[1])
    menor = min(proximos_do_maximo, key=lambda x: x[1])
    if maior[1] <= 0:
        return []

    diferenca = (maior[1] - menor[1]) / maior[1]
    if diferenca > FV_ENERGIAS_LIMIAR_DESBALANCEO:
        return [
            f"Desbalanceamento de potência: {maior[0]} operando a {maior[1]:.1f} kW "
            f"vs {menor[0]} a {menor[1]:.1f} kW ({diferenca*100:.0f}% de diferença), "
            f"ambos próximos da potência máxima"
        ]
    return []


def _fv_energias_traduzir_erro(e):
    """Traduz exceções técnicas (timeout, conexão, etc.) pra uma frase em
    português simples, sem jargão de programação — a mensagem crua do
    Python (tipo 'HTTPSConnectionPool... Read timed out') não faz sentido
    pra quem não é da área e aparecia direto no grupo do WhatsApp."""
    texto = str(e)
    if "timed out" in texto or "timeout" in texto.lower():
        return "o sistema do fabricante (Solplanet) demorou demais pra responder"
    if "Connection" in texto or "connection" in texto:
        return "não foi possível se conectar ao sistema do fabricante (Solplanet) agora"
    if "status" in texto.lower() and ("401" in texto or "403" in texto):
        return "o sistema do fabricante (Solplanet) recusou o acesso — pode ser um problema de credencial"
    return "o sistema do fabricante (Solplanet) não respondeu corretamente"


def _fv_energias_verificar_alertas():
    """Roda a cada checagem (5min), só dentro da janela diurna. Só envia
    mensagem quando o conjunto de problemas MUDA em relação à última
    checagem (novo problema surgiu, problema diferente, ou tudo
    normalizou) — evita spam repetindo o mesmo alerta a cada 5min
    enquanto o problema persiste.

    Offline/sem comunicação é checado em TODA chamada (a cada 5min).
    Desbalanceamento de potência só é checado nos horários pré-
    estabelecidos da ronda (08h/12h/15h/17h) — não a cada 5min."""
    agora = agora_br()
    if not (FV_ENERGIAS_HORA_INICIO_MONITORAMENTO <= agora.hour < FV_ENERGIAS_HORA_FIM_MONITORAMENTO):
        return {"ok": True, "alertou": False, "motivo": "fora da janela de monitoramento diurno"}

    problemas_atuais = []
    try:
        dados = _fv_energias_buscar_dados()
        for inv in dados["inversores"]:
            apelido = FV_ENERGIAS_APELIDO_INVERSOR.get(inv["sn"], inv["sn"])
            if not inv["online"]:
                problemas_atuais.append(f"{apelido} ({inv['sn']}) está OFFLINE / sem comunicação")
    except Exception as e:
        # Falha ao consultar a própria API da Solplanet já é, por si só,
        # um sinal de perda de comunicação com a usina. Log técnico
        # completo fica só no servidor; a mensagem do grupo é em
        # linguagem simples (ver _fv_energias_traduzir_erro).
        log.error(f"[FV Energias] Erro tecnico ao consultar Solplanet: {e}")
        problemas_atuais.append(
            f"Não conseguimos ler os dados da usina agora — {_fv_energias_traduzir_erro(e)}. "
            "Não é uma falha na usina, é só uma instabilidade temporária de consulta; "
            "o sistema vai tentar de novo automaticamente."
        )

    hora_atual = agora.strftime("%H")
    if hora_atual in FV_ENERGIAS_HORARIOS_RONDA and agora.minute < FV_ENERGIAS_JANELA_MINUTOS:
        try:
            potencias = _fv_energias_buscar_potencia_por_inversor()
            problemas_atuais.extend(_fv_energias_detectar_desbalanceamento_potencia(potencias))
        except Exception as e:
            log.error(f"[FV Energias] Erro ao checar desbalanceamento de potencia: {e}")

    assinatura_atual = "|".join(sorted(problemas_atuais))
    assinatura_anterior = _ler_trava("fv_energias:ultima_assinatura_problema") or ""

    if assinatura_atual == assinatura_anterior:
        return {"ok": True, "alertou": False, "motivo": "sem mudanca de estado"}

    if problemas_atuais:
        texto = (
            "🚨 *FV Energias Renováveis — ALERTA*\n\n"
            + "\n".join(f"⚠️ {p}" for p in problemas_atuais)
            + f"\n\nDetectado às {agora_br().strftime('%H:%M')}"
        )
    else:
        texto = (
            "✅ *FV Energias Renováveis — Normalizado*\n\n"
            "O(s) problema(s) anterior(es) não são mais detectados.\n"
            f"Verificado às {agora_br().strftime('%H:%M')}"
        )

    if not WPP_SERVER_URL:
        return {"ok": False, "alertou": True, "error": "WPP_SERVER_URL não configurado"}
    try:
        r = requests.post(
            f"{WPP_SERVER_URL}/api/enviar-mensagem",
            json={"grupoId": FV_ENERGIAS_GRUPO_WHATSAPP, "texto": texto},
            headers={"X-Webhook-Secret": WEBHOOK_SECRET} if WEBHOOK_SECRET else {},
            timeout=20,
        )
        if r.ok:
            # Só marca como "avisado" DEPOIS de confirmar o envio — se
            # gravasse antes e o envio falhasse, o sistema achava que já
            # tinha avisado e nunca mais tentava reenviar (bug real que
            # causou o alerta de "normalizado" nunca chegar).
            _gravar_trava("fv_energias:ultima_assinatura_problema", assinatura_atual)
        return {"ok": r.ok, "alertou": True, "problemas": problemas_atuais}
    except Exception as e:
        return {"ok": False, "alertou": True, "error": str(e)}


def _fv_energias_chamar_api(path, params):
    """Chamada assinada (HMAC-SHA256) à API Pro da Solplanet/AISWEI.
    Ver seção 4 da doc AISWEICloud API v1.1 pra detalhe do algoritmo.
    IMPORTANTE (confirmado testando com erro real do gateway): a
    assinatura é calculada sobre a URL CRUA (espaço, vírgula, dois-pontos
    literais, sem percent-encoding) — mas a requisição HTTP de verdade
    precisa ir codificada (senão não é uma URL válida). Por isso duas
    versões da query string: uma pra assinar, outra pra enviar."""
    items = sorted(params.items())
    query_crua = "&".join(f"{k}={v}" for k, v in items)
    query_codificada = "&".join(f"{k}={_fv_url_quote(str(v), safe='')}" for k, v in items)

    endpoint_cru = path + ("?" + query_crua if query_crua else "")
    endpoint_codificado = path + ("?" + query_codificada if query_codificada else "")

    accept = "application/json"
    headers_line = f"X-Ca-Key:{FV_ENERGIAS_APP_KEY}\n"
    string_to_sign = f"GET\n{accept}\n\n\n\n{headers_line}{endpoint_cru}"
    sig = base64.b64encode(
        _fv_hmac.new(
            FV_ENERGIAS_APP_SECRET.encode("utf-8"),
            string_to_sign.encode("utf-8"),
            _fv_hashlib.sha256,
        ).digest()
    ).decode()

    resp = requests.get(
        FV_ENERGIAS_BASE_URL + endpoint_codificado,
        headers={
            "Accept": accept,
            "X-Ca-Key": FV_ENERGIAS_APP_KEY,
            "X-Ca-Signature": sig,
            "X-Ca-Signature-Headers": "X-Ca-Key",
        },
        timeout=25,
    )
    resp.raise_for_status()
    return resp.json()


def _fv_energias_buscar_dados():
    """Busca overview da usina (potência/geração) + status dos 3
    inversores. Levanta exceção se a Solplanet retornar erro."""
    overview = _fv_energias_chamar_api(
        "/pro/getPlantOverviewPro",
        {"apikey": FV_ENERGIAS_APIKEY, "token": FV_ENERGIAS_TOKEN},
    )
    if overview.get("status") != 200:
        raise RuntimeError(f"Solplanet getPlantOverviewPro erro: {overview}")

    devices = _fv_energias_chamar_api(
        "/pro/getDeviceListPro",
        {"apikey": FV_ENERGIAS_APIKEY, "token": FV_ENERGIAS_TOKEN},
    )
    if devices.get("status") != 200:
        raise RuntimeError(f"Solplanet getDeviceListPro erro: {devices}")

    data = overview.get("data", {}) or {}
    inversores = []
    for dongle in (devices.get("data") or []):
        pstate = dongle.get("pstate")
        for inv in (dongle.get("inverters") or []):
            inversores.append({
                "sn": inv.get("isn"),
                "dongle": dongle.get("psn"),
                "online": inv.get("istate") == 1 and pstate == 1,
                "ultima_atualizacao": inv.get("ludt"),
            })

    return {
        "potencia_atual_kw": (data.get("Power") or {}).get("value"),
        "geracao_hoje_kwh":  (data.get("E-Today") or {}).get("value"),
        "geracao_mes":       data.get("E-Month") or {},
        "geracao_ano":       data.get("E-Year") or {},
        "geracao_total":     data.get("E-Total") or {},
        "co2_evitado":       data.get("CO2Avoided") or {},
        "ultima_atualizacao": data.get("ludt"),
        "inversores": inversores,
    }


def _fv_energias_montar_texto_ronda(dados, hora_label):
    inversores_ordenados = sorted(
        dados["inversores"],
        key=lambda inv: FV_ENERGIAS_APELIDO_INVERSOR.get(inv["sn"], inv["sn"]),
    )
    linhas_inv = []
    for inv in inversores_ordenados:
        apelido = FV_ENERGIAS_APELIDO_INVERSOR.get(inv["sn"], inv["sn"])
        emoji = "✅" if inv["online"] else "⚠️"
        estado = "Em Geração" if inv["online"] else "Offline/Atenção"
        linhas_inv.append(f"{emoji} {apelido} — {estado}")
    inversores_txt = "\n".join(linhas_inv) if linhas_inv else "Sem dados de inversores."

    geracao = dados.get("geracao_hoje_kwh")
    geracao_txt = f"{geracao} kWh" if geracao is not None else "indisponível"
    potencia = dados.get("potencia_atual_kw")
    potencia_txt = f"{potencia} kW" if potencia is not None else "indisponível"

    return (
        f"☀️ *FV Energias Renováveis — Ronda {hora_label}*\n\n"
        f"Geração até agora: {geracao_txt}\n"
        f"Potência atual: {potencia_txt}\n\n"
        f"*Inversores:*\n{inversores_txt}\n\n"
        f"Atualizado às {dados.get('ultima_atualizacao', '—')}"
    )


def _fv_energias_enviar_ronda(hora_label):
    if not WPP_SERVER_URL:
        return {"ok": False, "error": "WPP_SERVER_URL não configurado"}
    try:
        dados = _fv_energias_buscar_dados()
    except Exception as e:
        return {"ok": False, "error": f"Erro ao buscar dados da Solplanet: {e}"}

    texto = _fv_energias_montar_texto_ronda(dados, hora_label)
    try:
        r = requests.post(
            f"{WPP_SERVER_URL}/api/enviar-mensagem",
            json={"grupoId": FV_ENERGIAS_GRUPO_WHATSAPP, "texto": texto},
            headers={"X-Webhook-Secret": WEBHOOK_SECRET} if WEBHOOK_SECRET else {},
            timeout=20,
        )
        if r.ok and r.json().get("ok"):
            return {"ok": True, "dados": dados}
        return {"ok": False, "error": r.text[:300]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _fv_energias_buscar_curva_potencia(data_str):
    """Curva de potência (kW) ao longo do dia, somando os 3 inversores
    por horário — via getInverterDataPagePro, paginado. data_str no
    formato YYYY-MM-DD."""
    isnos = ",".join(FV_ENERGIAS_APELIDO_INVERSOR.keys())
    start = f"{data_str} 00:00:00"
    end = f"{data_str} 23:59:59"

    agregados = {}
    page_num = 1
    page_size = 500
    while True:
        resp = _fv_energias_chamar_api(
            "/pro/getInverterDataPagePro",
            {
                "apikey": FV_ENERGIAS_APIKEY,
                "token": FV_ENERGIAS_TOKEN,
                "isnos": isnos,
                "startDate": start,
                "endDate": end,
                "pageNum": page_num,
                "pageSize": page_size,
            },
        )
        if resp.get("status") != 200:
            raise RuntimeError(f"Solplanet getInverterDataPagePro erro: {resp}")
        data = resp.get("data") or {}
        for item in (data.get("result") or []):
            for ponto in (item.get("dataList") or []):
                tim = ponto.get("tim")  # "YYYY-MM-DD HH:MM:SS"
                pac = ponto.get("pac")  # potência ativa, unidade 1W
                if not tim or pac is None:
                    continue
                try:
                    pac_val = float(pac)
                    momento = datetime.strptime(tim, "%Y-%m-%d %H:%M:%S")
                except (TypeError, ValueError):
                    continue
                # Cada inversor reporta em instantes levemente diferentes
                # (ex.: um às 11:01:28, outro às 11:04:18) — agrupar por
                # janela de 5min pra somar os 3 inversores no mesmo ponto,
                # em vez de criar pontos separados que nunca se somam.
                bucket = momento - timedelta(minutes=momento.minute % 5, seconds=momento.second, microseconds=momento.microsecond)
                agregados[bucket] = agregados.get(bucket, 0.0) + pac_val

        total_pages = data.get("totalPages") or 1
        if page_num >= total_pages:
            break
        page_num += 1

    serie = sorted(agregados.items())
    dados_por_bucket = {b: round(soma / 1000, 3) for b, soma in serie}

    # Preenche a grade completa de 5 em 5 minutos do dia inteiro (mesmo
    # onde não há leitura) — sem isso, o eixo X do gráfico fica com
    # espaçamento irregular (pontos existentes nem sempre caem em
    # intervalos uniformes), o que deixa os rótulos de hora sem lógica
    # aparente. Com a grade completa, cada rótulo representa sempre o
    # mesmo intervalo de tempo real.
    dia_base = datetime.strptime(data_str, "%Y-%m-%d")
    grade_completa = []
    cursor = dia_base
    fim_dia = dia_base + timedelta(days=1)
    while cursor < fim_dia:
        grade_completa.append({"hora": cursor.strftime("%H:%M"), "potencia_kw": dados_por_bucket.get(cursor)})
        cursor += timedelta(minutes=5)
    return grade_completa


def _fv_energias_buscar_curva_irradiacao(data_str):
    """Irradiação horária (W/m²) do dia — fonte Open-Meteo, sem atraso
    (funciona até pro dia de hoje)."""
    url = (
        "https://api.open-meteo.com/v1/forecast"
        "?latitude=-4.034966&longitude=-38.4098811"
        "&hourly=shortwave_radiation"
        "&timezone=America%2FSao_Paulo"
        f"&start_date={data_str}&end_date={data_str}"
    )
    resp = requests.get(url, timeout=20)
    resp.raise_for_status()
    payload = resp.json()
    horas = payload.get("hourly", {}).get("time", [])
    valores = payload.get("hourly", {}).get("shortwave_radiation", [])
    return [
        {"hora": h.split("T")[1][:5] if "T" in h else h, "irradiancia_w_m2": v}
        for h, v in zip(horas, valores)
        if v is not None
    ]


@app.route("/fv-energias/curva-dia", methods=["GET"])
def fv_energias_curva_dia():
    """Curva intradiária: potência (kW) por horário + irradiação (W/m²)
    por horário, pro gráfico estilo 'Dia' do painel — pra comparar se
    quedas de geração acompanham quedas de irradiação (nuvem) ou têm
    outra causa (falha de equipamento)."""
    data_str = request.args.get("data") or agora_br().strftime("%Y-%m-%d")
    try:
        datetime.strptime(data_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"ok": False, "error": "data invalida, use YYYY-MM-DD"}), 400

    try:
        potencia = _fv_energias_buscar_curva_potencia(data_str)
    except Exception as e:
        return jsonify({"ok": False, "error": f"erro ao buscar potencia: {e}"}), 502

    try:
        irradiancia = _fv_energias_buscar_curva_irradiacao(data_str)
    except Exception as e:
        irradiancia = []
        log.error(f"[FV Energias] Erro ao buscar irradiacao horaria: {e}")

    return jsonify({"ok": True, "data": data_str, "potencia": potencia, "irradiancia": irradiancia}), 200


@app.route("/fv-energias/status", methods=["GET"])
def fv_energias_status():
    """Leitura pro painel visual isolado — só devolve dados já
    processados, nunca as credenciais da Solplanet."""
    try:
        dados = _fv_energias_buscar_dados()
        return jsonify({"ok": True, **dados}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502


def _fv_energias_buscar_historico_geracao(dias=7):
    """Geração diária real dos últimos N dias (kWh/dia), somando os 3
    inversores via getInverterETodayPro — esse endpoint devolve energia
    (kWh) por data, diferente do getPlantOutputPro (que devolve potência
    de pico, não geração, confirmado por teste real)."""
    resultado = []
    hoje = agora_br().date()
    for i in range(dias - 1, -1, -1):
        dia = hoje - timedelta(days=i)
        data_str = dia.strftime("%Y-%m-%d")
        try:
            resp = _fv_energias_chamar_api(
                "/pro/getInverterETodayPro",
                {
                    "apikey": FV_ENERGIAS_APIKEY,
                    "token": FV_ENERGIAS_TOKEN,
                    "date": data_str,
                },
            )
            valor = 0.0
            if resp.get("status") == 200:
                itens = (resp.get("data") or {}).get("result") or []
                for item in itens:
                    for sn, v in item.items():
                        try:
                            valor += float(v or 0)
                        except (TypeError, ValueError):
                            pass
        except Exception:
            valor = None
        resultado.append({"data": data_str, "geracao_kwh": valor})
    return resultado


@app.route("/fv-energias/historico-proprio", methods=["GET"])
def fv_energias_historico_proprio():
    """Lê o histórico construído dia a dia pelo snapshot automático
    (chaves fv_energias:historico_diario:YYYY-MM-DD em _Sistema) — usado
    pra períodos longos (ano) sem precisar reconsultar a Solplanet uma
    vez por dia histórico. Dias anteriores ao início da captura (ou dias
    sem captura, ex.: sistema fora do ar às 18h) simplesmente não
    aparecem na resposta."""
    try:
        inicio = datetime.strptime(request.args.get("inicio"), "%Y-%m-%d").date()
        fim = datetime.strptime(request.args.get("fim"), "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "parametros inicio e fim (YYYY-MM-DD) sao obrigatorios"}), 400

    if (fim - inicio).days > 400:
        return jsonify({"ok": False, "error": "intervalo maximo de 400 dias"}), 400

    dias_lista = []
    d = inicio
    while d <= fim:
        dias_lista.append(d.strftime("%Y-%m-%d"))
        d += timedelta(days=1)

    chaves = [f"fv_energias:historico_diario:{d}" for d in dias_lista]
    mapa = _ler_travas(chaves)

    historico = []
    for d in dias_lista:
        chave = f"fv_energias:historico_diario:{d}"
        valor = mapa.get(chave)
        historico.append({"data": d, "geracao_kwh": float(valor) if valor else None})

    return jsonify({"ok": True, "historico": historico}), 200


@app.route("/fv-energias/historico-geracao", methods=["GET"])
def fv_energias_historico_geracao():
    """Série diária pro gráfico do painel. Aceita ?dias=N (padrão 7,
    máximo 90 — cada dia é 1 chamada à Solplanet, então não escala pra
    período anual sob demanda) OU ?inicio=YYYY-MM-DD&fim=YYYY-MM-DD pra
    um intervalo customizado (também limitado a 90 dias)."""
    inicio_str = request.args.get("inicio")
    fim_str = request.args.get("fim")

    if inicio_str and fim_str:
        try:
            inicio = datetime.strptime(inicio_str, "%Y-%m-%d").date()
            fim = datetime.strptime(fim_str, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"ok": False, "error": "datas invalidas, use YYYY-MM-DD"}), 400
        dias = (fim - inicio).days + 1
        if dias < 1:
            return jsonify({"ok": False, "error": "fim deve ser depois de inicio"}), 400
        dias = min(dias, 90)
    else:
        try:
            dias = min(int(request.args.get("dias", 7)), 90)
        except (TypeError, ValueError):
            dias = 7

    try:
        historico = _fv_energias_buscar_historico_geracao(dias)
        return jsonify({"ok": True, "historico": historico}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502


@app.route("/fv-energias/irradiacao", methods=["GET"])
def fv_energias_irradiacao():
    """Irradiação solar diária (kWh/m²/dia) dos últimos N dias (incluindo
    hoje) pra comparar com a geração da usina — fonte pública Open-Meteo
    (sem chave, sem custo), usando as coordenadas reais da usina
    (Aquiraz/CE: lat -4.034966, lon -38.4098811). Diferente da NASA POWER,
    não tem atraso de dias — os dados de hoje já vêm disponíveis (modelo
    de previsão/reanálise quase em tempo real)."""
    try:
        dias = min(int(request.args.get("dias", 7)), 90)
    except (TypeError, ValueError):
        dias = 7

    url = (
        "https://api.open-meteo.com/v1/forecast"
        "?latitude=-4.034966&longitude=-38.4098811"
        "&daily=shortwave_radiation_sum"
        "&timezone=America%2FSao_Paulo"
        f"&past_days={dias}"
        "&forecast_days=1"
    )
    try:
        resp = requests.get(url, timeout=20)
        resp.raise_for_status()
        payload = resp.json()
        datas = payload.get("daily", {}).get("time", [])
        valores_mj = payload.get("daily", {}).get("shortwave_radiation_sum", [])
        # Open-Meteo devolve em MJ/m² — convertendo pra kWh/m² (÷3.6) pra
        # ficar na mesma unidade usada antes e mais intuitivo de comparar.
        historico = [
            {"data": d, "irradiacao_kwh_m2": round(v / 3.6, 3)}
            for d, v in zip(datas, valores_mj)
            if v is not None
        ]
        return jsonify({"ok": True, "historico": historico, "fonte": "Open-Meteo (shortwave_radiation_sum)"}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502


FV_ENERGIAS_HORA_SNAPSHOT = 18  # captura o total do dia às 18h (fim da janela de geração)
FV_ENERGIAS_SNAPSHOT_JANELA_MINUTOS = 9


def _fv_energias_capturar_snapshot_diario():
    """Guarda o total de geração do dia (uma vez por dia, perto das 18h)
    em chave própria _Sistema — constrói um histórico anual real ao
    longo do tempo, sem depender de reconsultar a Solplanet dia a dia
    (que não escala pra período de 1 ano)."""
    agora = agora_br()
    if not (agora.hour == FV_ENERGIAS_HORA_SNAPSHOT and agora.minute < FV_ENERGIAS_SNAPSHOT_JANELA_MINUTOS):
        return {"capturado": False, "motivo": "fora da janela de captura"}

    hoje_str = agora.strftime("%Y-%m-%d")
    chave = f"fv_energias:historico_diario:{hoje_str}"
    if _ler_trava(chave):
        return {"capturado": False, "motivo": "ja capturado hoje"}

    try:
        dados = _fv_energias_buscar_dados()
        geracao = dados.get("geracao_hoje_kwh")
        if geracao is not None:
            _gravar_trava(chave, str(geracao))
            return {"capturado": True, "geracao_kwh": geracao}
    except Exception as e:
        return {"capturado": False, "erro": str(e)}
    return {"capturado": False, "motivo": "sem dado de geracao"}


def _fv_energias_processar_ciclo():
    """Roda a checagem de alertas (sempre) + a ronda informativa. Isolado
    de propósito — só é acoplado ao restante do sistema pelo PONTO DE
    DISPARO (reaproveita o monitor UptimeRobot que já bate em /health a
    cada 5min), não pela lógica em si.

    A ronda usa lógica de RECUPERAÇÃO: em vez de só tentar dentro de uma
    janela estreita (ex.: 08:00-08:09) e desistir até o dia seguinte se
    perder essa janela (ex.: por instabilidade do WhatsApp bem naquele
    momento), verifica TODOS os horários de hoje que já passaram e ainda
    não foram confirmados como enviados, e tenta o mais atrasado deles —
    até um limite de 3h de atraso, pra não mandar uma "ronda das 8h"
    describida à tardezinha."""
    resultado_alerta = _fv_energias_verificar_alertas()
    resultado_snapshot = _fv_energias_capturar_snapshot_diario()

    agora = agora_br()
    hoje_str = agora.strftime("%Y-%m-%d")

    ronda_disparada = False
    ronda_resultado = None
    LIMITE_ATRASO_HORAS = 3

    for hora_str in FV_ENERGIAS_HORARIOS_RONDA:
        hora_agendada = agora.replace(hour=int(hora_str), minute=0, second=0, microsecond=0)
        atraso = agora - hora_agendada
        if timedelta(0) <= atraso <= timedelta(hours=LIMITE_ATRASO_HORAS):
            chave_trava = f"fv_energias:ronda:{hora_str}:{hoje_str}"
            if _ler_trava(chave_trava) != "enviado":
                ronda_resultado = _fv_energias_enviar_ronda(f"{hora_str}h")
                if ronda_resultado.get("ok"):
                    _gravar_trava(chave_trava, "enviado")
                    ronda_disparada = True
                break  # só uma tentativa de envio por ciclo, mesmo que haja mais de uma pendente

    return {
        "ronda_disparada": ronda_disparada,
        "ronda_resultado": ronda_resultado,
        "alerta": resultado_alerta,
        "snapshot": resultado_snapshot,
    }


@app.route("/fv-energias/ronda-check", methods=["GET", "POST"])
def fv_energias_ronda_check():
    """Endpoint manual/redundante — a automação de verdade roda embutida
    no /health (ver _fv_energias_processar_ciclo), que já é pingado a
    cada 5min pelo monitor UptimeRobot existente. Esse endpoint continua
    aqui só pra testes e disparo manual sob demanda."""
    if WEBHOOK_SECRET:
        secret = request.headers.get("X-Webhook-Secret", "") or request.args.get("secret", "")
        if secret != WEBHOOK_SECRET:
            return jsonify({"ok": False, "error": "unauthorized"}), 401

    resultado = _fv_energias_processar_ciclo()
    return jsonify({"ok": True, **resultado}), 200


@app.route("/fv-energias/ronda-disparar", methods=["POST"])
def fv_energias_ronda_disparar():
    """Disparo manual, sem depender da janela de horário — pro botão no
    painel FV Energias. Sem exigência de WEBHOOK_SECRET de propósito
    (mesmo padrão do /disparar-comunicado-cluster): chamado direto do
    navegador, não tem como o frontend guardar o secret com segurança."""
    hora_label = agora_br().strftime("%H:%M")
    resultado = _fv_energias_enviar_ronda(hora_label)
    status_code = 200 if resultado.get("ok") else 502
    return jsonify(resultado), status_code


# ── Assistente IA do dashboard (/chat-ia) ────────────────────────────────
# Balão de chat flutuante, visível só pro Fred (roles admin/manager) no
# frontend, que responde perguntas sobre os dados operacionais do Central
# O&M usando function calling do Gemini: o modelo escolhe qual ferramenta
# chamar (atividades, zeladoria, chamados, programação PCM) e a gente
# executa a consulta de verdade nos dados (Sheets/PCM) — o Gemini nunca
# inventa números, só interpreta o que a consulta trouxe.

_CHAT_IA_MAX_RODADAS = 5

_CHAT_IA_TOOLS = [{
    "functionDeclarations": [
        {
            "name": "consultar_atividades",
            "description": "Consulta o Painel de Atividades (ordens de serviço / atividades de campo). Retorna lista de atividades filtradas.",
            "parameters": {
                "type": "object",
                "properties": {
                    "usina": {"type": "string", "description": "Nome da usina, ex: 'Nova Xavantina I'. Deixe vazio para todas."},
                    "cliente": {"type": "string", "description": "Nome do cliente, ex: 'RENOGRID'. Deixe vazio para todos."},
                    "status": {"type": "string", "description": "Status da atividade, ex: 'Em Aberto', 'Concluído', 'Em Andamento'. Deixe vazio para todos."},
                    "numeroOS": {"type": "string", "description": "Número da OS específica, se souber."},
                    "cluster": {"type": "string", "description": "Código do cluster/equipe regional, ex: 'CE Leste 01'. Use quando a pergunta for sobre um coordenador de equipe/cluster (veja a tabela de clusters no prompt do sistema) em vez de uma usina específica. Deixe vazio se não aplicável."},
                    "responsavel": {"type": "string", "description": "Nome (ou parte do nome) do técnico responsável pela atividade, campo 'Responsável' — use pra perguntas sobre um técnico de campo específico. Deixe vazio se não aplicável."},
                },
            },
        },
        {
            "name": "consultar_zeladoria",
            "description": "Consulta a situação de zeladoria (roçada/supressão vegetal, poda química, lavagem de módulos, controle de pragas) por usina.",
            "parameters": {
                "type": "object",
                "properties": {
                    "usina": {"type": "string", "description": "Nome da usina. Deixe vazio para todas."},
                },
            },
        },
        {
            "name": "consultar_chamados",
            "description": "Consulta chamados/protocolos abertos com fabricantes de equipamentos (inversores, etc).",
            "parameters": {
                "type": "object",
                "properties": {
                    "usina": {"type": "string", "description": "Nome da usina (campo UFV). Deixe vazio para todas."},
                },
            },
        },
        {
            "name": "consultar_programacao_pcm",
            "description": "Consulta a programação semanal do PCM (Power Automate) para um dia específico — o que está agendado para execução.",
            "parameters": {
                "type": "object",
                "properties": {
                    "data": {"type": "string", "description": "Data no formato YYYY-MM-DD. Se vazio, usa hoje."},
                },
            },
        },
        {
            "name": "consultar_ocorrencias",
            "description": "Consulta o Painel de Falhas (ocorrências/falhas de equipamento detectadas via monitoramento ou ronda: inversores, trackers, strings, CFTV, comunicação, etc). É uma base DIFERENTE do Painel de Atividades — 'ocorrência' ou 'falha' aqui, 'atividade' ou 'OS de manutenção' em consultar_atividades. Use esta ferramenta quando perguntarem sobre ocorrências/falhas em aberto, andamento de uma ocorrência, causa, ação tomada, chamado de fabricante vinculado, etc.",
            "parameters": {
                "type": "object",
                "properties": {
                    "usina": {"type": "string", "description": "Nome da usina. Deixe vazio para todas."},
                    "cliente": {"type": "string", "description": "Nome do cliente, ex: 'RENOGRID'. Deixe vazio para todos."},
                    "status": {"type": "string", "description": "Status da ocorrência, ex: 'Em Aberto', 'Concluído', 'Em Andamento'. Deixe vazio para todos."},
                },
            },
        },
        {
            "name": "consultar_anotacoes",
            "description": "Consulta o Sketchbook — anotações rápidas do Fred que não são atividades/ocorrências formais: observações, lembretes, contexto histórico sobre uma usina/cliente, ou regras/decisões que ele quer que fiquem registradas. Use quando a pergunta pedir contexto, histórico informal, observações ou 'o que você sabe sobre' algo que não estaria em atividades/ocorrências.",
            "parameters": {
                "type": "object",
                "properties": {
                    "usina": {"type": "string", "description": "Nome da usina. Deixe vazio para todas."},
                    "cliente": {"type": "string", "description": "Nome do cliente. Deixe vazio para todos."},
                    "categoria": {"type": "string", "description": "Categoria/tag da anotação, se souber. Deixe vazio para todas."},
                },
            },
        },
    ]
}]


# Trunca campos de texto longos (histórico, observações) antes de mandar
# pra Gemini — descoberto em 03/09/2026 que perguntas que somam várias
# ferramentas (ex.: "pontos de atenção de hoje") geravam um payload de
# retorno gigante (até 60 itens x campos verbosos x 3-4 ferramentas),
# deixando a geração da resposta lenta o bastante pra estourar o timeout
# do Gunicorn (160s) e cair em erro 500/502 real, não só cosmético.
def _ia_trunc(texto, limite=180):
    texto = (texto or "").strip()
    if len(texto) <= limite:
        return texto
    return texto[:limite].rstrip() + "…"


def _ia_consultar_atividades(usina="", cliente="", status="", numeroOS="", cluster="", responsavel=""):
    ws = get_atividades_sheet()
    todos = _gspread_retry(lambda: ws.get_all_values())
    mapa_cluster = _mapa_cluster_usina()
    usina_norm = canonizar_usina(usina) if usina else None
    cluster_norm = _normalizar_tema_comunicado(cluster) if cluster else None
    out = []
    for row in todos[1:]:
        if len(row) < len(ATIV_HEADERS_JSON):
            row = row + [""] * (len(ATIV_HEADERS_JSON) - len(row))
        if not row[0].strip():
            continue
        item_full = dict(zip(ATIV_HEADERS_JSON, row[:len(ATIV_HEADERS_JSON)]))
        if not usina_permitida(item_full.get("usina", "")):
            continue
        if usina_norm and canonizar_usina(item_full.get("usina", "")) != usina_norm:
            continue
        if cliente and cliente.strip().lower() not in item_full.get("cliente", "").strip().lower():
            continue
        if status and status.strip().lower() not in item_full.get("status", "").strip().lower():
            continue
        if numeroOS and numeroOS.strip() != item_full.get("numeroOS", "").strip():
            continue
        if responsavel and responsavel.strip().lower() not in item_full.get("responsavel", "").strip().lower():
            continue
        item_cluster = mapa_cluster.get(item_full.get("usina", "").strip(), "")
        if cluster_norm and _normalizar_tema_comunicado(item_cluster) != cluster_norm:
            continue
        # campos compactos — descricao/observacoesOS truncados, histórico
        # completo e demais campos verbosos omitidos (não essenciais pra
        # síntese e infláveis o bastante pra deixar a Gemini lenta)
        out.append({
            "id": item_full.get("id"), "cliente": item_full.get("cliente"), "usina": item_full.get("usina"),
            "cluster": item_cluster, "equipamento": item_full.get("equipamento"),
            "descricao": _ia_trunc(item_full.get("descricao"), 150), "responsavel": item_full.get("responsavel"),
            "prazo": item_full.get("prazo"), "status": item_full.get("status"),
            "numeroOS": item_full.get("numeroOS"), "statusOS": item_full.get("statusOS"),
            "percentualOS": item_full.get("percentualOS"),
        })
    # limita pra não estourar o contexto do Gemini em consultas amplas
    limitado = out[:25]
    return {"total_encontrado": len(out), "mostrando": len(limitado), "atividades": limitado}


def _ia_consultar_zeladoria(usina=""):
    ws = get_zeladoria_sheet()
    todos = _gspread_retry(lambda: ws.get_all_values())
    indice_cols = _gspread_retry(lambda: _zel_montar_indice_colunas(ws))
    usina_norm = canonizar_usina(usina) if usina else None
    out = []
    for row in todos[2:]:
        if len(row) < 2 or not row[1].strip():
            continue
        nome_usina = row[1].strip()
        if usina_norm and canonizar_usina(nome_usina) != usina_norm:
            continue
        if not usina_permitida(nome_usina):
            continue
        grupos = {}
        for grupo, subcols in indice_cols.items():
            dados_grupo = {}
            for subcol, col_idx in subcols.items():
                valor = row[col_idx - 1].strip() if col_idx - 1 < len(row) else ""
                if valor:
                    dados_grupo[subcol] = valor
            if dados_grupo:
                grupos[grupo] = dados_grupo
        out.append({"usina": nome_usina, "grupos": grupos})
    return {"total": len(out), "zeladoria": out}


def _ia_consultar_chamados(usina=""):
    itens = _chamados_fabricante_itens()
    usina_norm = canonizar_usina(usina) if usina else None
    if usina_norm:
        itens = [it for it in itens if canonizar_usina(it.get("UFV", "")) == usina_norm]
    # campos compactos — "Observações" e afins podem ter texto bem longo
    compactos = [{
        "ufv": it.get("UFV"), "cliente": it.get("Cliente"), "ativo": it.get("Ativo"),
        "fabricante": it.get("Fabricante"), "ticket": it.get("Ticket/RMA"),
        "motivo": _ia_trunc(it.get("Motivo da abertura do chamado"), 150),
        "causa": _ia_trunc(it.get("Causa da Falha"), 120),
        "status": it.get("Status"), "statusOS": it.get("Status OS"),
        "diasCorridos": it.get("Dias corridos"), "numeroOS": it.get("N° da Solicitação de OS"),
    } for it in itens]
    limitado = compactos[:25]
    return {"total_encontrado": len(compactos), "mostrando": len(limitado), "chamados": limitado}


def _ia_consultar_programacao_pcm(data=""):
    data_filtro = (data or "").strip() or datetime.now(_TZ_BR).strftime("%Y-%m-%d")
    try:
        dt = datetime.strptime(data_filtro, "%Y-%m-%d").date()
    except ValueError:
        return {"erro": "data inválida, use YYYY-MM-DD"}
    cache = _buscar_programacao_pcm_core()
    dados = cache.get("dados") or {}
    semanas = dados.get("semanas", [])
    iso_year, iso_week, _ = dt.isocalendar()
    semana_alvo = f"{iso_year}-W{iso_week:02d}"
    semana = next((s for s in semanas if s.get("week") == semana_alvo), None)
    if semana is None:
        return {"aviso": f"Sem programação publicada pelo PCM pra semana {semana_alvo}."}
    dia_pt = _DIA_SEMANA_PT[dt.weekday()]
    usinas_temp_nomes = {item["usina"] for item in _usinas_temporarias()}
    linhas_dia = [
        r for r in semana.get("rows", [])
        if (r.get("responsavel") == _PCM_RESPONSAVEL or r.get("usina") in usinas_temp_nomes) and r.get("dia") == dia_pt
    ]
    itens = [{
        "usina": r.get("usina"), "cliente": r.get("cliente"), "tipo": r.get("tipo"),
        "tarefa": r.get("tarefa"), "status": r.get("status"),
        "hIni": r.get("h_ini"), "hFim": r.get("h_fim"),
    } for r in linhas_dia]
    limitado = itens[:40]
    return {"data": data_filtro, "diaSemana": dia_pt, "total_encontrado": len(itens), "mostrando": len(limitado), "programacao": limitado}


# Layout de colunas do Painel de Falhas (0-indexed), confirmado via
# gravar_nova_ocorrencia() e CAMPO_COL — é uma aba DIFERENTE do Painel de
# Atividades, com sua própria numeração de coluna.
_FALHAS_HEADERS_JSON = [
    "id", "cliente", "usina", "equipamento", "falha", "causa", "impactados",
    "acao", "status", "ticketFabricante", "numeroOS", "historico", "dataAbertura",
]


def _ia_consultar_ocorrencias(usina="", cliente="", status=""):
    ws = get_sheet()
    todos = _gspread_retry(lambda: ws.get_all_values())
    usina_norm = canonizar_usina(usina) if usina else None
    out = []
    for row in todos[1:]:
        if len(row) < len(_FALHAS_HEADERS_JSON):
            row = row + [""] * (len(_FALHAS_HEADERS_JSON) - len(row))
        if not row[0].strip():
            continue
        item_full = dict(zip(_FALHAS_HEADERS_JSON, row[:len(_FALHAS_HEADERS_JSON)]))
        if not usina_permitida(item_full.get("usina", "")):
            continue
        if usina_norm and canonizar_usina(item_full.get("usina", "")) != usina_norm:
            continue
        if cliente and cliente.strip().lower() not in item_full.get("cliente", "").strip().lower():
            continue
        if status and status.strip().lower() not in item_full.get("status", "").strip().lower():
            continue
        # campos compactos — "historico" pode ser um timeline inteiro de
        # texto; "acao" truncada mantém o essencial pro "andamento"
        out.append({
            "id": item_full.get("id"), "cliente": item_full.get("cliente"), "usina": item_full.get("usina"),
            "equipamento": item_full.get("equipamento"), "falha": _ia_trunc(item_full.get("falha"), 150),
            "causa": _ia_trunc(item_full.get("causa"), 120), "acao": _ia_trunc(item_full.get("acao"), 150),
            "status": item_full.get("status"), "ticketFabricante": item_full.get("ticketFabricante"),
            "numeroOS": item_full.get("numeroOS"), "dataAbertura": item_full.get("dataAbertura"),
        })
    limitado = out[:25]
    return {"total_encontrado": len(out), "mostrando": len(limitado), "ocorrencias": limitado}


_CHAT_IA_FERRAMENTAS_PYTHON = {
    "consultar_atividades": _ia_consultar_atividades,
    "consultar_zeladoria": _ia_consultar_zeladoria,
    "consultar_chamados": _ia_consultar_chamados,
    "consultar_programacao_pcm": _ia_consultar_programacao_pcm,
    "consultar_ocorrencias": _ia_consultar_ocorrencias,
    "consultar_anotacoes": _ia_consultar_anotacoes,
}


def _chat_ia_system_prompt():
    hoje = agora_br().strftime("%d/%m/%Y (%A)")
    try:
        mapa_coord = _mapa_coordenador_cluster()
    except Exception as e:
        log.error(f"[chat-ia] Falha ao carregar coordenador_cluster: {e}")
        mapa_coord = {}
    if mapa_coord:
        linhas_tabela = "\n".join(f"- {cluster}: {nome}" for cluster, nome in sorted(mapa_coord.items()))
        bloco_clusters = f"""TABELA DE CLUSTERS E COORDENADORES/TÉCNICOS PRINCIPAIS (lida ao vivo da configuração — se a organização mudar, é atualizada na aba _Sistema, chave "coordenador_cluster:<Cluster>", sem precisar de deploy):
{linhas_tabela}

Alguns clusters têm mais de um nome listado (separados por "/") porque a vistoria de 26/08/2026 encontrou mais de um técnico com volume relevante de atividades e não achou um responsável único óbvio — considere ambos ao investigar."""
    else:
        bloco_clusters = "TABELA DE CLUSTERS E COORDENADORES: não disponível no momento (falha ao ler configuração) — não presuma nomes de coordenador, só responda com base no que as ferramentas retornarem."

    return f"""Você é o assistente de IA embutido no dashboard Central O&M da Grid Co., empresa de operação e manutenção de usinas solares fotovoltaicas. Você conversa com Fred Alexandrino, Supervisor de O&M, respondendo perguntas sobre os dados operacionais do painel: atividades/OS, ocorrências/falhas, zeladoria, chamados de fabricante e programação do PCM.

Hoje é {hoje}, horário de Brasília.

IMPORTANTE — ATIVIDADES x OCORRÊNCIAS SÃO BASES DIFERENTES: "Painel de Atividades" (ferramenta consultar_atividades) tem as OS de manutenção — preventivas, corretivas, rondas. "Painel de Falhas" (ferramenta consultar_ocorrencias) tem as ocorrências/falhas de equipamento detectadas por monitoramento ou ronda (inversor, tracker, string, CFTV, comunicação, etc.), cada uma com falha/causa/ação/status próprios e às vezes um chamado de fabricante e/ou uma OS vinculados. Se a pergunta usar as palavras "ocorrência(s)" ou "falha(s)", use consultar_ocorrencias. Se usar "atividade(s)" ou "OS" no sentido de manutenção programada, use consultar_atividades. Em caso de dúvida real (a pergunta poderia ser sobre qualquer uma), chame as duas.

{bloco_clusters}

IMPORTANTE SOBRE NOMES DE PESSOAS: uma vistoria cruzando atividades reais com clusters (26/08/2026) mostrou que coordenadores de cluster GERALMENTE também aparecem como "responsavel" em várias atividades (eles executam campo também, não só coordenam). Por isso, ao perguntarem sobre "atividades do Fulano": (1) primeiro chame consultar_atividades com responsavel="Fulano" pra pegar o que está diretamente atribuído a ele; (2) SE Fulano for um coordenador de cluster (está na tabela acima), chame TAMBÉM consultar_atividades com cluster="<cluster dele>" pra não perder atividades de outros técnicos da equipe dele que ele também acompanha; (3) apresente os dois resultados de forma clara, deixando explícito o que é "atribuído diretamente a ele" vs "da equipe/cluster dele".

REGRAS OBRIGATÓRIAS:
- Use SEMPRE as ferramentas disponíveis para consultar dados reais antes de responder qualquer pergunta sobre números, status, OS, usinas, prazos ou pendências. NUNCA invente ou estime dados que não vieram de uma chamada de ferramenta.
- Se uma pergunta puder ser respondida com mais de uma ferramenta (ex: "o que está pendente na usina X"), chame todas as ferramentas relevantes.
- Se a ferramenta não retornar nada relevante, diga claramente que não encontrou, em vez de supor. Antes de concluir que não há nada sobre uma pessoa, confira se ela é coordenadora de cluster (ver acima) e tente também pelo cluster.
- Responda em português, de forma direta e objetiva — sem rodeios, sem saudações desnecessárias. Fred prefere respostas curtas e factuais, com números e nomes específicos.
- Em perguntas amplas que exigem várias ferramentas (ex.: "pontos de atenção de hoje", "resumo geral"): seja SELETIVO — destaque só o que realmente precisa de atenção (atrasado, pausado, aguardando algo há muito tempo), não liste item por item de tudo que veio das ferramentas. Respostas mais enxutas geram mais rápido e são mais úteis.
- Nomes de usina usam numeração romana (ex: Matão I, Sol do Norte I) — normalize antes de comparar.
- Se a pergunta não tiver relação com os dados do painel (ex: pergunta genérica), pode responder normalmente sem usar ferramentas."""


@app.route("/chat-ia", methods=["POST"])
def chat_ia():
    """Assistente de IA do dashboard: recebe uma pergunta em linguagem
    natural e responde consultando os dados reais do painel (Atividades,
    Zeladoria, Chamados, Programação PCM) via function calling do Gemini
    — nunca gera números "de memória". Uso interno (widget flutuante,
    visível só pro Fred no frontend)."""
    if not GEMINI_API_KEY:
        return jsonify({"ok": False, "error": "GEMINI_API_KEY não configurada no servidor"}), 500

    body = request.get_json(force=True, silent=True) or {}
    pergunta = (body.get("pergunta") or "").strip()
    historico = body.get("historico") or []  # [{role:'user'|'model', texto:'...'}]
    if not pergunta:
        return jsonify({"ok": False, "error": "informe 'pergunta'"}), 400

    contents = []
    for turno in historico[-10:]:  # limita histórico pra não estourar contexto
        role = turno.get("role")
        texto = turno.get("texto", "")
        if role in ("user", "model") and texto:
            contents.append({"role": role, "parts": [{"text": texto}]})
    contents.append({"role": "user", "parts": [{"text": pergunta}]})

    ferramentas_chamadas = []
    try:
        for _rodada in range(_CHAT_IA_MAX_RODADAS):
            payload = {
                "system_instruction": {"parts": [{"text": _chat_ia_system_prompt()}]},
                "contents": contents,
                "tools": _CHAT_IA_TOOLS,
                "generationConfig": {
                    "temperature": 0.3,
                    "maxOutputTokens": 1200,
                    "thinkingConfig": {"thinkingBudget": 0},
                },
            }
            resp = _chamar_gemini_com_retry(payload, timeout=45)
            data = resp.json()
            candidatos = data.get("candidates") or []
            if not candidatos:
                return jsonify({"ok": False, "error": "Gemini não retornou resposta"}), 502
            content = candidatos[0].get("content") or {}
            parts = content.get("parts") or []
            chamadas_funcao = [p["functionCall"] for p in parts if "functionCall" in p]

            if not chamadas_funcao:
                texto_final = "".join(p.get("text", "") for p in parts).strip()
                return jsonify({
                    "ok": True,
                    "resposta": texto_final or "Não consegui gerar uma resposta.",
                    "ferramentasUsadas": ferramentas_chamadas,
                }), 200

            contents.append(content)
            partes_resposta = []
            for chamada in chamadas_funcao:
                nome = chamada.get("name")
                args = chamada.get("args") or {}
                fn = _CHAT_IA_FERRAMENTAS_PYTHON.get(nome)
                if fn is None:
                    resultado = {"erro": f"ferramenta '{nome}' não existe"}
                else:
                    try:
                        resultado = fn(**args)
                    except Exception as e:
                        resultado = {"erro": str(e)}
                    ferramentas_chamadas.append({"nome": nome, "args": args})
                partes_resposta.append({"functionResponse": {"name": nome, "response": resultado}})
            contents.append({"role": "user", "parts": partes_resposta})

        return jsonify({"ok": False, "error": "Limite de rodadas de consulta atingido sem resposta final"}), 500
    except requests.exceptions.HTTPError as e:
        log.error(f"[chat-ia] Erro HTTP do Gemini: {e}")
        return jsonify({"ok": False, "error": "Erro ao consultar IA"}), 502
    except Exception as e:
        log.error(f"[chat-ia] Erro: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500



try:
    carregar_push_subscriptions()
except Exception as e:
    log.error(f"[Push] Erro na carga inicial de subscriptions: {e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=False)

